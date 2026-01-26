import frappe
import requests
from datetime import date
import erpnext
import json
import frappe
from itertools import groupby
from operator import itemgetter
import frappe
import io
import re
from frappe.utils import now_datetime
from datetime import time
from frappe.utils import money_in_words
import urllib.parse
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from io import BytesIO
from frappe.utils import now, fmt_money
from frappe import throw,_, bold
from frappe.utils import flt
from dateutil.relativedelta import relativedelta
from frappe.utils import now_datetime
from frappe.utils import format_time, formatdate, now
from frappe.model.naming import make_autoname
from erpnext.setup.utils import get_exchange_rate
from frappe.model.workflow import apply_workflow
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment, Border, Side,Font
from frappe.utils import (
	add_days,
	ceil,
	cint,
	comma_and,
	flt,
	get_link_to_form,
	getdate,
	now_datetime,
	datetime,get_first_day,get_last_day,
	nowdate,
	today,
)
from frappe.utils import cstr, cint, getdate, get_last_day, get_first_day, add_days,date_diff
from datetime import date, datetime, timedelta
import datetime as dt
from datetime import datetime, timedelta
from typing import Dict, Optional, Tuple, Union
from datetime import datetime
from onegene.mark_attendance import check_holiday
from frappe.utils.password import update_password
from frappe.utils.background_jobs import enqueue
import datetime as dt
from datetime import datetime, timedelta
from frappe.contacts.doctype.address.address import (
	get_address_display
)
from frappe.utils.file_manager import get_file
import csv
from onegene.onegene.doctype.logistics_request.logistics_request import create_workflow_notification
class OperationSequenceError(frappe.ValidationError):
	pass

import datetime
@frappe.whitelist()
def return_total_schedule(doc,method):
	# method to restrict the excess and low order scheduling
	total = frappe.db.sql(""" select `tabSales Order Schedule Item`.item_code, sum(`tabSales Order Schedule Item`.schedule_qty) as qty from `tabSales Order`
	left join `tabSales Order Schedule Item` on `tabSales Order`.name = `tabSales Order Schedule Item`.parent where `tabSales Order`.name = '%s' group by `tabSales Order Schedule Item`.item_code"""%(doc.name),as_dict = 1)

	item_total = frappe.db.sql(""" select `tabSales Order Item`.item_code, sum(`tabSales Order Item`.qty) as qty from `tabSales Order`
	left join `tabSales Order Item` on `tabSales Order`.name = `tabSales Order Item`.parent where `tabSales Order`.name = '%s' group by `tabSales Order Item`.item_code"""%(doc.name),as_dict = 1)
	for t in total:
		for i in item_total:
			if i.item_code == t.item_code:
				if t.qty > i.qty:
					frappe.throw(
						_(
							"Schedule Qty {2} is Greater than -  {0} for - {1}."
						).format(
							frappe.bold(i.qty),
							frappe.bold(i.item_code),
							frappe.bold(t.qty),
						)
					)
					frappe.validated = False
				if t.qty < i.qty:
					frappe.throw(
						_(
							"Schedule Qty {2} is Less than -  {0} for - {1}."
						).format(
							frappe.bold(i.qty),
							frappe.bold(i.item_code),
							frappe.bold(t.qty),
						)
					)
					frappe.validated = False

@frappe.whitelist()
# method to delete the order schedule during so cancel
def cancel_order_schedule_on_so_cancel(doc,method):
	inter_office_memo = None
	if frappe.db.exists("Approval for Schedule Increase", {"sales_order": doc.name, "parenttype": "Inter Office Memo"}):
		inter_office_memo = frappe.db.get_value("Approval for Schedule Increase", {"sales_order": doc.name, "parenttype": "Inter Office Memo"}, "parent")
		frappe.errprint(inter_office_memo)
		docstatus = frappe.db.get_value("Inter Office Memo", inter_office_memo, "docstatus")
		if docstatus == 1:
			frappe.errprint(inter_office_memo)
			frappe.db.sql("update `tabInter Office Memo` set docstatus = 2 where name = '%s'"%(inter_office_memo))
	# Delete Sales Open Order
	if frappe.db.exists("Open Order", {"sales_order_number": doc.name}):
		open_order = frappe.get_doc("Open Order", {"sales_order_number": doc.name})
		open_order.delete()

	# Cancel Sales Order Schedule
	exists = frappe.db.exists("Sales Order Schedule",{"sales_order_number":doc.name, "docstatus": 1})
	if exists:
		os = frappe.db.get_all("Sales Order Schedule",{"sales_order_number":doc.name, "docstatus": 1},'name')
		for o in os:
			order_schedule = frappe.get_doc('Sales Order Schedule',o.name)
			order_schedule.cancel()

	if inter_office_memo and docstatus == 1:
		frappe.db.sql("update `tabInter Office Memo` set docstatus = 1 where name = '%s'"%(inter_office_memo))


@frappe.whitelist()
def get_so_details(sales):
	dict_list = []
	so = frappe.get_doc("Sales Order",sales)
	for i in so.items:
		dict_list.append(frappe._dict({"name":i.name,"item_code":i.item_code,"pending_qty":i.qty,"bom":i.bom_no,"description": i.description,"warehouse":i.warehouse,"rate":i.rate,"amount":i.amount}))
	return dict_list

@frappe.whitelist()
def sample_check():
	item_code = "333QRJLA-EC03"
	sf = frappe.db.sql("""select `tabMaterial Request Item`.qty as qty from `tabMaterial Request`
		left join `tabMaterial Request Item` on `tabMaterial Request`.name = `tabMaterial Request Item`.parent
		where `tabMaterial Request Item`.item_code = '%s' and `tabMaterial Request`.tatus != 2docs and `tabMaterial Request`.transaction_date = CURDATE() """%(item_code),as_dict = 1)[0].qty or 0
	print(sf)

def get_exploded_items(bom, data, indent=0, qty=1):

	exploded_items = frappe.get_all(
		"BOM Item",
		filters={"parent": bom},
		fields=["qty", "bom_no", "qty", "item_code", "item_name", "description", "uom"],
	)

	for item in exploded_items:
		item["indent"] = indent
		data.append(
			{
				"item_code": item.item_code,
				"item_name": item.item_name,
				"indent": indent,
				"bom_level": indent,
				"bom": item.bom_no,
				"qty": item.qty * qty,
				"uom": item.uom,
				"description": item.description,
			}
		)
		if item.bom_no:
			get_exploded_items(item.bom_no, data, indent=indent + 1, qty=item.qty)

@frappe.whitelist()
def get_open_order(name, item_code, delivery_date, item_name, qty, rate, warehouse, amount):
	new_doc = frappe.new_doc('Open Order')
	new_doc.sales_order_number = name
	new_doc.set('open_order_table', [])
	new_doc.append("open_order_table", {
		"item_code": item_code,
		"delivery_date": delivery_date,
		"item_name": item_name,
		"rate": float(rate),
		"warehouse": warehouse,
		"amount": amount,
		"qty": qty,   
	})
	new_doc.save(ignore_permissions=True)
	return "ok"

@frappe.whitelist()
def delete_sales_order_schedule(doc, method):
	if doc.order_schedule:
		if frappe.db.exists('Sales Order Schedule', doc.order_schedule):
			schedule = frappe.get_doc('Sales Order Schedule', doc.order_schedule)
			if schedule.order_type == "Open":
				schedule.cancel()
				doc.save(ignore_permissions=True)

@frappe.whitelist()
def create_sales_order_schedule_from_so(doc,method):
	if doc.customer_order_type == "Fixed" and not doc.custom_schedule_table:
		frappe.throw("Schedule not Created")
		frappe.throw("Schedule உருவாக்கப்படவில்லை")
	if doc.customer_order_type == "Fixed" and doc.custom_schedule_table:
		for schedule in doc.custom_schedule_table:
			if frappe.db.exists('Sales Order Schedule', {'sales_order_number': doc.name, 'item_code': schedule.item_code, 'schedule_date': schedule.schedule_date}):
				frappe.throw("Schedule already exists for this item code and schedule date")
				frappe.throw("இந்த item code மற்றும் schedule date க்கான Schedule ஏற்கனவே உள்ளது")
			new_doc = frappe.new_doc('Sales Order Schedule') 
			new_doc.customer_code = doc.custom_customer_code
			new_doc.sales_order_number = doc.name
			new_doc.item_code = schedule.item_code
			new_doc.schedule_date = schedule.schedule_date
			new_doc.qty = schedule.schedule_qty
			for item in doc.items:
				if item.item_code == schedule.item_code:
					new_doc.child_name = schedule.name
					new_doc.schedule_amount = schedule.schedule_qty * item.rate
					new_doc.order_rate = item.rate
					new_doc.pending_qty = schedule.schedule_qty
					new_doc.pending_amount = schedule.schedule_qty * item.rate
			new_doc.save(ignore_permissions=True) 
			new_doc.submit()

	if doc.customer_order_type == "Open" and doc.custom_schedule_table:
		for schedule in doc.custom_schedule_table:
			# date_obj = datetime.strptime(schedule.schedule_date, '%Y-%m-%d')
			month = schedule.schedule_date.strftime('%b') 
			new_doc = frappe.new_doc('Sales Order Schedule') 
			new_doc.customer_code = doc.custom_customer_code
			new_doc.sales_order_number = doc.name
			new_doc.item_code = schedule.item_code
			new_doc.schedule_date = schedule.schedule_date
			new_doc.schedule_month = month.upper()
			new_doc.qty = schedule.schedule_qty
			for item in doc.items:
				if item.item_code == schedule.item_code:
					new_doc.child_name = schedule.name
					new_doc.schedule_amount = schedule.schedule_qty * item.rate
					new_doc.order_rate = item.rate
					new_doc.pending_qty = schedule.schedule_qty
					new_doc.pending_amount = schedule.schedule_qty * item.rate
			new_doc.save(ignore_permissions=True) 
			new_doc.submit()

@frappe.whitelist()
def create_order_schedule_from_so_for_open(item_code, schedule_date, schedule_qty, customer_code, name, rate, month):
	rate = float(rate)
	schedule_qty = float(schedule_qty)
	new_doc = frappe.new_doc('Sales Order Schedule') 
	new_doc.customer_code = customer_code
	new_doc.customer_order_type = "Open"
	new_doc.sales_order_number = name
	new_doc.item_code = item_code
	new_doc.schedule_date = schedule_date
	new_doc.qty = schedule_qty
	new_doc.schedule_amount = schedule_qty * rate
	new_doc.order_rate = rate
	new_doc.schedule_month = month.upper()
	new_doc.pending_qty = schedule_qty
	new_doc.pending_amount = schedule_qty * rate
	new_doc.save(ignore_permissions=True) 
	frappe.db.commit()
	return("ok")


@frappe.whitelist()
def generate_production_plan():
	# create production plan report list based daily based on the scheduled date from order schedule
	from frappe.utils import getdate
	from datetime import datetime
	start_date = datetime.today().replace(day=1).date()
	work_order = frappe.db.sql("""
		SELECT item_code, item_name, item_group, SUM(pending_qty) AS qty
		FROM `tabSales Order Schedule`
		WHERE MONTH(schedule_date) = MONTH(CURRENT_DATE())
		GROUP BY item_code, item_name, item_group
	""", as_dict=1)
	for j in work_order:
		rej_allowance = frappe.get_value("Item",j.item_code,['rejection_allowance'])
		pack_size = frappe.get_value("Item",j.item_code,['pack_size'])
		fg_plan = frappe.get_value("Kanban Quantity",{'item_code':j.item_code},['fg_kanban_qty']) or 0
		sfg_days = frappe.get_value("Kanban Quantity",{'item_code':j.item_code},['sfg_days']) or 0
		today_plan = frappe.get_value("Kanban Quantity",{'item_code':j.item_code},['today_production_plan']) or 0
		tent_plan_i= frappe.get_value("Kanban Quantity",{'item_code':j.item_code},['tentative_plan_i']) or 0
		tent_plan_ii = frappe.get_value("Kanban Quantity",{'item_code':j.item_code},['tentative_plan_ii']) or 0
		stock = frappe.db.sql(""" select sum(actual_qty) as actual_qty from `tabBin` where item_code = '%s' """%(j.item_code),as_dict = 1)[0]
		if not stock["actual_qty"]:
			stock["actual_qty"] = 0
		pos = frappe.db.sql("""select `tabDelivery Note Item`.item_code as item_code,`tabDelivery Note Item`.qty as qty from `tabDelivery Note`
		left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
		where `tabDelivery Note Item`.item_code = '%s' and `tabDelivery Note`.docstatus = 1 and `tabDelivery Note`.posting_date = CURDATE() """%(j.item_code),as_dict = 1)
		del_qty = 0
		if len(pos)>0:
			for l in pos:
				del_qty = l.qty
		delivery = frappe.db.sql("""select `tabDelivery Note Item`.item_code as item_code,`tabDelivery Note Item`.qty as qty from `tabDelivery Note`
		left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
		where `tabDelivery Note Item`.item_code = '%s' and `tabDelivery Note`.docstatus = 1 and `tabDelivery Note`.posting_date between '%s' and '%s' """%(j.item_code,start_date,today()),as_dict = 1)
		del_qty_as_on_date = 0
		if len(delivery)>0:
			for d in delivery:
				del_qty_as_on_date = d.qty
		produced = frappe.db.sql("""select `tabStock Entry Detail`.item_code as item_code,`tabStock Entry Detail`.qty as qty from `tabStock Entry`
		left join `tabStock Entry Detail` on `tabStock Entry`.name = `tabStock Entry Detail`.parent
		where `tabStock Entry Detail`.item_code = '%s' and `tabStock Entry`.docstatus = 1 and `tabStock Entry`.posting_date = CURDATE() and `tabStock Entry`.stock_entry_type = "Manufacture"  """%(j.item_code),as_dict = 1)
		prod = 0
		if len(produced)>0:
			for l in produced:
				prod = l.qty
		produced_as_on_date = frappe.db.sql("""select `tabStock Entry Detail`.item_code as item_code,`tabStock Entry Detail`.qty as qty from `tabStock Entry`
		left join `tabStock Entry Detail` on `tabStock Entry`.name = `tabStock Entry Detail`.parent
		where `tabStock Entry Detail`.item_code = '%s' and `tabStock Entry`.docstatus = 1 and `tabStock Entry`.posting_date between '%s' and '%s' and `tabStock Entry`.stock_entry_type = "Manufacture" """%(j.item_code,start_date,today()),as_dict = 1)
		pro_qty_as_on_date = 0
		if len(produced_as_on_date)>0:
			for d in produced_as_on_date:
				pro_qty_as_on_date = d.qty
		work_days = frappe.db.get_single_value("Production Plan Settings", "working_days")
		with_rej = (j.qty * (rej_allowance/100)) + j.qty
		per_day = j.qty / int(work_days)
		if pack_size > 0:
			cal = per_day/ pack_size
		total = ceil(cal) * pack_size
		today_balance = 0
		reqd_plan = 0
		balance = 0
		if with_rej and fg_plan:
			balance = (int(with_rej) + int(fg_plan))
			reqd_plan = (float(total) * float(sfg_days)) + float(fg_plan)
			today_balance = int(today_plan)-int(prod)
		td_balance = 0
		if today_balance > 0:
			td_balance = today_balance
		else:
			td_balance = 0
		exists = frappe.db.exists("Production Plan Report",{"date":today(),'item':j.item_code})
		if exists:
			doc = frappe.get_doc("Production Plan Report",{"date":today(),'item':j.item_code})
		else:
			doc = frappe.new_doc("Production Plan Report")
		doc.item = j.item_code
		doc.item_name = j.item_name
		doc.item_group = j.item_group
		doc.date = today()
		doc.rej_allowance = rej_allowance
		doc.monthly_schedule = with_rej
		doc.bin_qty = pack_size
		doc.per_day_plan = total
		doc.fg_kanban_qty = fg_plan
		doc.sfg_days = sfg_days
		doc.stock_qty = stock["actual_qty"]
		doc.delivered_qty = del_qty
		doc.del_as_on_yes = del_qty_as_on_date
		doc.produced_qty = prod
		doc.pro_as_on_yes = pro_qty_as_on_date
		doc.monthly_balance = balance
		doc.today_prod_plan = today_plan
		doc.today_balance = td_balance
		doc.required_plan = reqd_plan
		doc.tent_prod_plan_1 = tent_plan_i
		doc.tent_prod_plan_2 = tent_plan_ii
		doc.save(ignore_permissions=True)

@frappe.whitelist()
# error will be thrown when actice employee have relieving date
def inactive_employee(doc,method):
	if doc.status=="Active":
		if doc.relieving_date:
			throw(_("Please remove the relieving date for the Active Employee."))

@frappe.whitelist()
def list_all_raw_materials(order_schedule, scheduleqty):
	doc_list = []
	consolidated_items = {}

	self = frappe.get_doc("Sales Order Schedule", order_schedule)
	data = []
	bom_list = []

	bom = frappe.db.get_value("BOM", {'item': self.item_code}, ['name'])
	bom_list.append(frappe._dict({"bom": bom, "qty": scheduleqty}))

	for k in bom_list:
		get_exploded_items(k["bom"], data, k["qty"], bom_list)

	unique_items = {}
	for item in data:
		item_code = item['item_code']
		qty = item['qty']
		if item_code in unique_items:
			unique_items[item_code]['qty'] += qty
		else:
			unique_items[item_code] = item
	combined_items_list = list(unique_items.values())
	doc_list.append(combined_items_list)

	for i in doc_list:
		for h in i:
			item_code = h["item_code"]
			qty = h["qty"]
			if item_code in consolidated_items:
				consolidated_items[item_code] += qty
			else:
				consolidated_items[item_code] = qty
	return consolidated_items

def get_exploded_items(bom, data, qty, skip_list):
	exploded_items = frappe.get_all("BOM Item", filters={"parent": bom},
									fields=["qty", "bom_no", "item_code", "item_name", "description", "uom"])
	for item in exploded_items:
		item_code = item['item_code']
		if item_code in skip_list:
			continue
		item_qty = float(item['qty']) * float(qty)
		stock = frappe.db.get_value("Bin", {'item_code': item_code, 'warehouse': "SFS Store - O"},
									['actual_qty']) or 0
		to_order = item_qty - stock if item_qty > stock else 0
		data.append({
			"item_code": item_code,
			"qty": item_qty,
		})
		if item['bom_no']:
			get_exploded_items(item['bom_no'], data, qty=item_qty, skip_list=skip_list)


# The below two methods are called in MRP Test Report, Material Requirement Planning, Internal Material Request Plan
@frappe.whitelist()
def return_print(item_type,based_on):
	from frappe.utils import cstr, add_days, date_diff, getdate,today,gzip_decompress
	pr_name = frappe.db.get_value('Prepared Report', {'report_name': 'Material Requirements Planning','status':'Completed'}, 'name')
	attached_file_name = frappe.db.get_value("File",{"attached_to_doctype": 'Prepared Report',"attached_to_name": pr_name},"name",)
	attached_file = frappe.get_doc("File", attached_file_name)
	compressed_content = attached_file.get_content()
	uncompressed_content = gzip_decompress(compressed_content)
	dos = json.loads(uncompressed_content.decode("utf-8"))
	doc = frappe.new_doc("Material Request")
	doc.material_request_type = "Purchase"
	doc.transaction_date = frappe.utils.today()
	doc.schedule_date = frappe.utils.today()
	doc.set_warehouse = "Stores - O"
	if based_on == "Highlighted Rows":
		for i in dos['result']:
			if float(i['safety_stock']) > float(i['actual_stock_qty']):
				uom = frappe.db.get_value("Item",i['item_code'],'stock_uom')
				pps = frappe.db.sql("""select sum(actual_qty) as qty from `tabBin`
										where item_code = %s and warehouse != 'SFS Store - O' """, (i['item_code']), as_dict=1)[0].qty or 0
				sfs = frappe.db.sql("""select sum(actual_qty) as qty from `tabBin`
										where item_code = %s and warehouse = 'SFS Store - O' """, (i['item_code']), as_dict=1)[0].qty or 0

				if i['to_order'] > 0:
					doc.append("items", {
						'item_code': i['item_code'],
						'custom_item_type': i['item_type'],
						'schedule_date': frappe.utils.today(),
						'qty': i['to_order'],
						'custom_mr_qty': i['to_order'],
						'custom_total_req_qty': i['to_order'],
						'custom_current_req_qty': i['to_order'],
						'custom_stock_qty_copy': pps,
						'custom_shop_floor_stock': sfs,
						'custom_expected_date': i['expected_date'],
						# 'custom_today_req_qty': today_req,
						'uom': uom
					})
		doc.save()
		name = [
			"""<a href="/app/Form/Material Request/{0}">{1}</a>""".format(doc.name, doc.name)
		]
		frappe.msgprint(_("Material Request - {0} created").format(", ".join(name)))
	if based_on == "Item Type":
		for i in dos['result']:
			if i['item_type'] in item_type:
				uom = frappe.db.get_value("Item",i['item_code'],'stock_uom')
				pps = frappe.db.sql("""select sum(actual_qty) as qty from `tabBin`
										where item_code = %s and warehouse != 'SFS Store - O' """, (i['item_code']), as_dict=1)[0].qty or 0
				sfs = frappe.db.sql("""select sum(actual_qty) as qty from `tabBin`
										where item_code = %s and warehouse = 'SFS Store - O' """, (i['item_code']), as_dict=1)[0].qty or 0

				if i['to_order'] > 0:
					doc.append("items", {
						'item_code': i['item_code'],
						'custom_item_type': i['item_type'],
						'schedule_date': frappe.utils.today(),
						'qty': i['to_order'],
						'custom_mr_qty': i['to_order'],
						'custom_total_req_qty': i['to_order'],
						'custom_current_req_qty': i['to_order'],
						'custom_stock_qty_copy': pps,
						'custom_shop_floor_stock': sfs,
						'custom_expected_date': i['expected_date'],
						# 'custom_today_req_qty': today_req,
						'uom': uom
					})
		doc.save()
		name = [
			"""<a href="/app/Form/Material Request/{0}">{1}</a>""".format(doc.name, doc.name)
		]
		frappe.msgprint(_("Material Request - {0} created").format(", ".join(name)))

@frappe.whitelist()
def return_item_type():
	dict = []
	dict_list = []
	from frappe.utils import cstr, add_days, date_diff, getdate,today,gzip_decompress
	pr_name = frappe.db.get_value('Prepared Report', {'report_name': 'Material Requirements Planning','status':'Completed'}, 'name')
	attached_file_name = frappe.db.get_value("File",{"attached_to_doctype": 'Prepared Report',"attached_to_name": pr_name},"name",)
	attached_file = frappe.get_doc("File", attached_file_name)
	compressed_content = attached_file.get_content()
	uncompressed_content = gzip_decompress(compressed_content)
	dos = json.loads(uncompressed_content.decode("utf-8"))
	doc = frappe.new_doc("Material Request")
	doc.material_request_type = "Purchase"
	doc.transaction_date = frappe.utils.today()
	doc.schedule_date = frappe.utils.today()
	doc.set_warehouse = "Stores - O"
	for i in dos['result']:
		if i['item_type'] not in dict:
			dict.append(i['item_type'])
			dict_list.append(frappe._dict({'item_type':i['item_type']}))

	return dict_list

@frappe.whitelist()
def return_mr_details(mr):
	# method to set the expected delivery date in purchase order
	doc = frappe.get_doc("Material Request",mr)
	return doc.items

# The below two methods are called in MRP Test Report, Material Requirement Planning
@frappe.whitelist()
def stock_details_mpd_report(item):
	w_house = frappe.db.get_value("Warehouse",['name'])
	data = ''
	stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin where item_code = '%s' order by warehouse """%(item),as_dict=True)
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;color:white;background-color:#f68b1f" colspan = 10><center>Stock Availability</center></th></tr>'
	data += '''
	<tr><td style="padding:1px;border: 1px solid black" colspan = 4><b>Item Code</b></td>
	<td style="padding:1px;border: 1px solid black" colspan = 6>%s</td></tr>
	<tr><td style="padding:1px;border: 1px solid black" colspan = 4><b>Item Name</b></td>
	<td style="padding:1px;border: 1px solid black" colspan = 6>%s</td></tr>'''%(item,frappe.db.get_value('Item',item,'item_name'))
	data += '''
	<td style="padding:1px;border: 1px solid black;background-color:#f68b1f;color:white"  colspan = 4><b>Warehouse</b></td>
	<td style="padding:1px;border: 1px solid black;background-color:#f68b1f;color:white" colspan = 3><b>Stock Qty</b></td>
	</tr>'''
	i = 0
	for stock in stocks:
		if stock.warehouse != w_house:
			if stock.actual_qty > 0:
				data += '''<tr><td style="padding:1px;border: 1px solid black" colspan = 4 >%s</td><td style="padding:1px;border: 1px solid black" colspan = 3>%s</td></tr>'''%(stock.warehouse,stock.actual_qty)
	i += 1
	stock_qty = 0
	for stock in stocks:
		stock_qty += stock.actual_qty
	data += '''<tr><td style="background-color:#909e8a;padding:1px;border: 1px solid black;color:white;font-weight:bold" colspan = 4 >%s</td><td style="background-color:#909e8a;padding:1px;border: 1px solid black;color:white;font-weight:bold" colspan = 3>%s</td></tr>'''%("Total     ",stock_qty)
	data += '</table>'

	return data



@frappe.whitelist()
def stock_details_mpd(item,quantity):
	w_house = frappe.db.get_value("Warehouse",['name'])
	data = ''
	stocks = frappe.db.sql("""select actual_qty,warehouse,stock_uom,stock_value from tabBin where item_code = '%s' order by warehouse """%(item),as_dict=True)
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;color:white;background-color:#f68b1f" colspan = 10><center>Stock Availability</center></th></tr>'
	data += '''
	<tr><td style="padding:1px;border: 1px solid black" colspan = 4><b>Item Code</b></td>
	<td style="padding:1px;border: 1px solid black" colspan = 6>%s</td></tr>
	<tr><td style="padding:1px;border: 1px solid black" colspan = 4><b>Item Name</b></td>
	<td style="padding:1px;border: 1px solid black" colspan = 6>%s</td></tr>'''%(item,frappe.db.get_value('Item',item,'item_name'))
	data += '''
	<td style="padding:1px;border: 1px solid black;background-color:#f68b1f;color:white"  colspan = 4><b>Warehouse</b></td>
	<td style="padding:1px;border: 1px solid black;background-color:#f68b1f;color:white" colspan = 3><b>Stock Qty</b></td>
	<td style="padding:1px;border: 1px solid black;background-color:#f68b1f;color:white" colspan = 3><b>Required Qty</b></td>

	</tr>'''
	req_qty = 0
	qty = frappe.get_doc("Material Planning Details",quantity)
	for q in qty.material_plan:
		req_qty += q.required_qty
	for stock in stocks:
		if stock.warehouse == w_house:
			if stock.actual_qty > 0:
				comp = frappe.get_value("Warehouse",stock.warehouse,['company'])
				data +=''' <tr><td style="padding:1px;border: 1px solid black;color:black;font-weight:bold" colspan = 4>%s</td><td style="padding:1px;border: 1px solid black;color:black;font-weight:bold" colspan = 3>%s</td><td style="padding:1px;border: 1px solid black;color:black;font-weight:bold" colspan = 3>%s</td></tr>'''%(stock.warehouse,stock.actual_qty,'')
	i = 0
	for stock in stocks:
		if stock.warehouse != w_house:
			if stock.actual_qty > 0:
				data += '''<tr><td style="padding:1px;border: 1px solid black" colspan = 4 >%s</td><td style="padding:1px;border: 1px solid black" colspan = 3>%s</td><td style="padding:1px;border: 1px solid black;color:black;font-weight:bold" colspan = 3>%s</td></tr>'''%(stock.warehouse,stock.actual_qty,"")
	i += 1
	stock_qty = 0
	for stock in stocks:
		stock_qty += stock.actual_qty
	data += '''<tr><td style="background-color:#909e8a;padding:1px;border: 1px solid black;color:white;font-weight:bold" colspan = 4 >%s</td><td style="background-color:#909e8a;padding:1px;border: 1px solid black;color:white;font-weight:bold" colspan = 3>%s</td><td style="background-color:#909e8a;color:white;padding:1px;border: 1px solid black;font-weight:bold" colspan = 3>%s</td></tr>'''%("Total     ",stock_qty,req_qty)
	data += '</table>'

	return data

@frappe.whitelist()
def previous_purchase(item_table):
	item_table = json.loads(item_table)
	data = []
	for item in item_table:
		try:
			item_name = frappe.get_value('Item',{'name':item['item_code']},"item_name")
			pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,sum(`tabPurchase Order Item`.qty) as qty from `tabPurchase Order`
			left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
			where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 """%(item["item_code"]),as_dict=True)
			for po in pos:
				data.append([item['item_code'],item_name,po.qty])
		except:
			pass
	return data


@frappe.whitelist()
def previous_po_html(item_code):
	# method get the previous po details of the item on click
	if item_code:
		if not frappe.db.exists("Item", item_code):
			return "Item not found"
		data = ""
		item_name = frappe.get_value('Item',{'item_code':item_code},"item_name") or ""
		pos = frappe.db.sql("""select `tabPurchase Order Item`.item_code as item_code,`tabPurchase Order Item`.item_name as item_name,`tabPurchase Order`.supplier as supplier,`tabPurchase Order Item`.qty as qty,`tabPurchase Order Item`.rate as rate,`tabPurchase Order Item`.amount as amount,`tabPurchase Order`.transaction_date as date,`tabPurchase Order`.name as po from `tabPurchase Order`
		left join `tabPurchase Order Item` on `tabPurchase Order`.name = `tabPurchase Order Item`.parent
		where `tabPurchase Order Item`.item_code = '%s' and `tabPurchase Order`.docstatus != 2 order by date"""%(item_code),as_dict=True)


		data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;color:white;background-color:#f68b1f" colspan=6><center>Previous Purchase Order</center></th></tr>'
		data += '''
		<tr><td colspan =2 style="padding:1px;border: 1px solid black;width:300px" ><b>Item Code</b></td>
		<td style="padding:1px;border: 1px solid black;width:200px" colspan =4>%s</td></tr>
		<tr><td colspan =2 style="padding:1px;border: 1px solid black" ><b>Item Name</b></td>
		<td style="padding:1px;border: 1px solid black" colspan =4>%s</td></tr>

		<tr><td style="padding:1px;border: 1px solid black" colspan =1><b>Supplier Name</b></td>
		<td style="padding:1px;border: 1px solid black" colspan=1><b>Previous Purchase Order</b></td>
		<td style="padding:1px;border: 1px solid black" colspan=1><b>PO Date</b></td>
		<td style="padding:1px;border: 1px solid black" colspan=1><b>PO Rate</b></td>
		<td style="padding:1px;border: 1px solid black" colspan=1><b>PO Quantity</b></td>
		<td style="padding:1px;border: 1px solid black" colspan=1><b>PO Amount</b>
		</td></tr>'''%(item_code,item_name)
		for po in pos:
			data += '''<tr>
				<td style="padding:1px;border: 1px solid black" colspan =1>%s</td>
				<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
				<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
				<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
				<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
				<td style="padding:1px;border: 1px solid black" colspan=1>%s</td></tr>'''%(po.supplier,po.po,po.date,po.rate,po.qty,po.amount)

		data += '</table>'
		return data

# The below method is called in MRP Test Report, Material Requirement Planning
@frappe.whitelist()
def mpd_details(name):
	data = ""
	pos = frappe.db.sql("""select `tabMaterial Planning Item`.item_code,`tabMaterial Planning Item`.item_name,`tabMaterial Planning Item`.uom,`tabMaterial Planning Item`.order_schedule_date,sum(`tabMaterial Planning Item`.required_qty) as qty from `tabMaterial Planning Details`
		left join `tabMaterial Planning Item` on `tabMaterial Planning Details`.name = `tabMaterial Planning Item`.parent
		where `tabMaterial Planning Details`.name = '%s' group by `tabMaterial Planning Item`.order_schedule_date """%(name),as_dict = 1)
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;color:white;background-color:#f68b1f" colspan=6><center>Order Schedule Details</center></th></tr>'
	data += '''
	<tr><td style="padding:1px;border: 1px solid black" colspan =1><b>Item Code</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>Item Name</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>UOM</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>Schedule Date</b></td>
	<td style="padding:1px;border: 1px solid black" colspan=1><b>Quantity</b></td>
	</td></tr>'''
	for po in pos:
		data += '''<tr>
			<td style="padding:1px;border: 1px solid black" colspan =1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td></tr>'''%(po.item_code,po.item_name,po.uom,po.order_schedule_date,po.qty)
	data += '</table>'
	return data


@frappe.whitelist()
def list_raw_mat():
	qty = 120
	skip_list = []
	data = []
	bom = "BOM-742-HWFAB-002"
	exploded_items = frappe.get_all("BOM Item", filters={"parent": bom},fields=["qty", "bom_no as bom", "item_code", "item_name", "description", "uom"])
	for item in exploded_items:
		item_code = item['item_code']
		if item_code in skip_list:
			continue
		item_qty = flt(item['qty']) * qty
		data.append({"item_code": item_code,"item_name": item['item_name'],"bom": item['bom'],"uom": item['uom'],"qty": item_qty,"description": item['description']})



@frappe.whitelist()
# method to fetch the operation item list linked to the bom and append in the table
def get_bom_details(bo, child):
	dict_list = []
	seen_items = set()

	# Get BOM doc
	bom_doc = frappe.get_doc("BOM", bo)

	# Get items linked to THIS operation only
	current_op_items = frappe.get_all(
		"Operation Item List",
		filters={"operation_name": child, "document_name": bo},
		fields=["item", "selected_field", "req_tot_qty", "uom"]
	)

	# Add items already assigned to this operation (with check_box ticked)
	for row in current_op_items:
		item_name = frappe.db.get_value("Item",row.item,'item_name')
		item_qty = frappe.db.get_value("BOM Item",{"parent": bo,"item_code":row.item},["qty"]) or ""
		dict_list.append(frappe._dict({
			"check_box": 1,
			"name": row.selected_field,
			"item_code": row.item,
			"req_tot_qty": item_qty,
			"uom": row.uom,
			'item_name':item_name
		}))
		seen_items.add(row.item)

	# Add only BOM items that are NOT assigned to ANY other operation
	items_used_in_any_op = set(frappe.get_all(
		"Operation Item List",
		filters={"document_name": bo},
		pluck="item"
	))

	for i in bom_doc.items:
		# Add only if:
		# 1. Not already in THIS operation (handled above)
		# 2. Not used in ANY other operation
		if i.item_code not in seen_items and i.item_code not in items_used_in_any_op:
			item_name = frappe.db.get_value("Item",i.item_code,'item_name')
			dict_list.append(frappe._dict({
				"item_code": i.item_code,
				"req_tot_qty": i.qty,
				"uom": i.uom,
				"item_name":item_name
			}))
			seen_items.add(i.item_code)

	return dict_list

@frappe.whitelist()
# method to create operation item list on button click from BOM
def table_multiselect(docs,item,item_code,child,uom,req_tot_qty):
	op = frappe.db.get_value("Operation Item List",{"document_name":docs,"item":item_code,"operation_name":child},["name"])
	if not op:
		bom_child = frappe.new_doc("Operation Item List")
		bom_child.document_name = docs
		bom_child.item = item_code
		bom_child.operation_name = child
		bom_child.selected_field = item
		bom_child.req_tot_qty = req_tot_qty
		bom_child.uom = uom
		bom_child.save()

@frappe.whitelist()
def bday_allocate():
	# create additional salary automatically when employee have birthday on that month
	employee_query = """
	SELECT *
	FROM `tabEmployee`
	WHERE
		status = 'Active'
		AND employee_category IN ('Staff', 'Operator', 'Sub Staff')
		AND MONTH(date_of_birth) = MONTH(CURDATE())
		AND date_of_joining < CURDATE()
	"""
	employee = frappe.db.sql(employee_query, as_dict=True)
	pay =  get_first_day(nowdate())
	for emp in employee:
		if frappe.db.exists("Salary Structure Assignment",{'employee':emp.name,'docstatus':1}):
			if not frappe.db.exists('Additional Salary',{'employee':emp.name,'payroll_date':pay,'salary_component':"Birthday Allowance",'docstatus':('!=',2)}):
				bday_amt = frappe.new_doc("Additional Salary")
				bday_amt.employee = emp.name
				bday_amt.payroll_date = pay
				bday_amt.company = emp.company
				bday_amt.salary_component = "Birthday Allowance"
				bday_amt.currency = "INR"
				bday_amt.amount = 1000
				bday_amt.save(ignore_permissions = True)
				bday_amt.submit()




@frappe.whitelist()
def overtime_hours(doc,method):
	ot_hours=frappe.db.sql("""select sum(custom_overtime_hours) from `tabAttendance` where employee = '%s' and attendance_date between '%s' and '%s'"""%(doc.employee,doc.start_date,doc.end_date),as_dict=True)[0]
	doc.custom_overtime_hours = ot_hours['sum(custom_overtime_hours)']

@frappe.whitelist()
# method to set the fixed salary in salary slip
def fixed_salary(doc,method):
	if doc.salary_structure!='Operators' and doc.salary_structure!='Apprentice':
		base_amount=frappe.db.get_value("Salary Structure Assignment",{'employee':doc.employee,'docstatus':1},['base'], order_by="modified desc")
		earned_basic=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Basic"},["amount"]) or 0
		da=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Dearness Allowance"},["amount"]) or 0
		hra=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"House Rent Allowance"},["amount"]) or 0
		wa=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Washing Allowance"},["amount"]) or 0
		ca=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Conveyance Allowance"},["amount"]) or 0
		ea=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Education Allowance"},["amount"]) or 0
		pa=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Performance Allowance"},["amount"]) or 0
		sa=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Special Allowance"},["amount"]) or 0
		stipend=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Stipend"},["amount"]) or 0
		att_inc=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Attendance Incentive"},["amount"]) or 0
		basic_da=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Basic & DA"},["amount"]) or 0
		lta=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Leave Travel Allowance"},["amount"]) or 0
		mnc=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Medical & Conveyance Allowance"},["amount"]) or 0
		sp=frappe.db.get_value("Salary Detail",{"parent":doc.name,"salary_component":"Special Pay"},["amount"]) or 0

		if doc.designation=='Asst General Manager':
			basic=base_amount*0.4
			hra= base_amount*0.25
			pa=base_amount * 0.10
		else:
			basic=base_amount*0.5
			hra= base_amount*0.2
			pa=base_amount * 0.05
		lta=base_amount * 0.05
		mnc=base_amount * 0.03
		sa=base_amount * 0.17
		total = basic+hra+pa+lta+mnc+sa
		doc.custom_basic = basic
		doc.custom_dearness_allowance = da
		doc.custom_house_rent_allowance = hra
		doc.custom_performance_allowance = pa
		doc.custom_special_allowance = sa
		doc.custom_basic_da = basic
		doc.custom_leave_travel_allowance = lta
		doc.custom_medical_conveyance_allowance = mnc
		doc.custom_total_fixed_amount = round(total)
		doc.save(ignore_permissions = True)
		frappe.db.commit()
		
	if doc.salary_structure=='Operators':
		base=frappe.db.get_value("Salary Structure Assignment",{'employee':doc.employee,'docstatus':1},['base'], order_by="modified desc")
		basic=base*0.25
		da=base * 0.25
		hra=base * 0.2
		wa=base * 0.05
		ca=base * 0.1
		ea=base * 0.1
		pa=base * 0.05
		epf=basic+da
		frappe.db.set_value("Salary Slip",doc.name,'custom_basic',basic)
		frappe.db.set_value("Salary Slip",doc.name,'custom_dearness_allowance',da)
		frappe.db.set_value("Salary Slip",doc.name,'custom_house_rent_allowance',hra)
		frappe.db.set_value("Salary Slip",doc.name,'custom_washing_allowance',wa)
		frappe.db.set_value("Salary Slip",doc.name,'custom_conveyance_allowance',ca)
		frappe.db.set_value("Salary Slip",doc.name,'custom_education_allowance',ea)
		frappe.db.set_value("Salary Slip",doc.name,'custom_performance_allowance',pa)
		# frappe.db.set_value("Salary Slip",doc.name,'custom_attendance_incentive',1000)
	if doc.salary_structure=='Apprentice':
		base=frappe.db.get_value("Salary Structure Assignment",{'employee':doc.employee,'docstatus':1},['base'], order_by="modified desc")
		basic=base*0.5
		da=base * 0.5
		frappe.db.set_value("Salary Slip",doc.name,'custom_basic',basic)
		frappe.db.set_value("Salary Slip",doc.name,'custom_dearness_allowance',da)
		# frappe.db.set_value("Salary Slip",doc.name,'custom_attendance_incentive',1000)
		frappe.db.set_value("Salary Slip",doc.name,'custom_total_fixed_amount',basic+da)
	if doc.salary_structure=='Trainee':
		base=frappe.db.get_value("Salary Structure Assignment",{'employee':doc.employee,'docstatus':1},['base'], order_by="modified desc")
		basic=base*0.5
		da=base * 0.5
		frappe.db.set_value("Salary Slip",doc.name,'custom_basic',basic)
		frappe.db.set_value("Salary Slip",doc.name,'custom_dearness_allowance',da)
		# frappe.db.set_value("Salary Slip",doc.name,'custom_attendance_incentive',1000)
		frappe.db.set_value("Salary Slip",doc.name,'custom_total_fixed_amount',basic+da)

@frappe.whitelist()
def sick_leave_allocation():
	# sl will be added 0.5 every month
	import datetime
	from datetime import date
	today = date.today()
	year_start_date = date(today.year, 1, 1)
	year_end_date = date(today.year, 12, 31)
	employees=frappe.db.get_all("Employee",{"Status":"Active"},['*'])
	for emp in employees:
		# if emp.employee_category=='Apprentice':
		if emp.employee_category=='Staff' or emp.employee_category=='Sub Staff' or emp.employee_category=='Operator':
			la=frappe.db.exists("Leave Allocation",{"employee":emp.name,"leave_type":"Sick Leave","from_date":year_start_date,"to_date":year_end_date,'docstatus':1})
			if la:
				leave_all=frappe.get_doc("Leave Allocation",la)
				leave_all.new_leaves_allocated +=0.5
				leave_all.total_leaves_allocated +=0.5
				leave_all.save(ignore_permissions=True)
				leave_all.submit()

			else:
				leave_all=frappe.new_doc("Leave Allocation")
				leave_all.employee=emp.name
				leave_all.leave_type="Sick Leave"
				leave_all.from_date=year_start_date
				leave_all.to_date=year_end_date
				leave_all.new_leaves_allocated=0.5
				leave_all.total_leaves_allocated =0.5
				# leave_all.carry_forward=1
				leave_all.save(ignore_permissions=True)
				leave_all.submit()
	
def update_leave_policy():
	# earned leave will be allocated automatically based on the present and half days
	pre_year = date.today().year - 1
	start_of_year = date(pre_year, 1, 1)
	end_of_year = date(pre_year, 12, 31)
	current_year = date.today().year
	start = date(current_year, 1, 1)
	end = date(current_year, 12, 31)
	leave = frappe.get_all("Leave Policy Detail", ["leave_type", "annual_allocation"])
	for i in leave:
		if i.leave_type =="Earned Leave":
			employees = frappe.get_all("Employee",{"status": "Active",'employee_category':('!=','Contractor')},["name","company"])
			for emp in employees:
				present = frappe.db.count("Attendance",{"employee":emp.name,"status":"Present","attendance_date": ["between", [start_of_year, end_of_year]]})
				half_day = frappe.db.count("Attendance",{"employee":emp.name,"status":"Half Day","attendance_date": ["between", [start_of_year, end_of_year]]})
				half = half_day/2
				attendance = present + half
				earned_leave = round(attendance /20)
				if earned_leave:
					allow = frappe.new_doc("Leave Allocation")
					allow.employee = emp.name
					allow.company = emp.company
					allow.leave_type = "Earned Leave"
					allow.from_date = start
					allow.to_date = end
					allow.new_leaves_allocated = earned_leave
					allow.total_leaves_allocated = earned_leave
					allow.save(ignore_permissions=True)
					allow.submit()
	frappe.db.commit()



@frappe.whitelist()
def update_shift(employee,from_date,to_date):
	shift_3 = frappe.db.count("Attendance",{"employee":employee,"attendance_date": ["between", [from_date, to_date]],"status":"Present","shift":"3"})
	shift_3_half = frappe.db.count("Attendance",{"employee":employee,"attendance_date": ["between", [from_date, to_date]],"status":"Half Day","shift":"3"})
	half_3 = shift_3_half/2
	shift3 = shift_3 + half_3
	shift_5 = frappe.db.count("Attendance",{"employee":employee,"attendance_date": ["between", [from_date, to_date]],"status":"Present","shift":"5"})
	shift_5_half = frappe.db.count("Attendance",{"employee":employee,"attendance_date": ["between", [from_date, to_date]],"status":"Half Day","shift":"5"})
	half_5 = shift_5_half/2
	shift5 = shift_5 + half_5
	shift = shift3 + shift5
	return shift


from frappe.utils import cstr, cint, getdate,get_first_day, get_last_day, today, time_diff_in_hours
@frappe.whitelist()
# return the time diff for the permission time
def att_req_hours(f_time,t_time,custom_session,custom_shift):
	if custom_session == "Flexible":
		if f_time and t_time:
			time_diff = time_diff_in_hours(t_time,f_time)
			return time_diff
	elif custom_session == "Full Day":
		return "8"
	else :
		return "4"

@frappe.whitelist()
def od_hours_update(doc, method):
	# update the attendance status based on the sessions from od
	if doc.workflow_state=='Approved':
		dates = get_dates(doc.from_date, doc.to_date)
		for date in dates:
			# update attendance with present when session is full day
			if doc.reason == "On Duty" and doc.custom_session == "Full Day":
				if frappe.db.exists("Attendance", {'employee': doc.employee, 'attendance_date': date, 'docstatus': ('!=', 2)}):
					att = frappe.get_doc("Attendance", {'employee': doc.employee, 'attendance_date': date, 'docstatus': ('!=', 2)})
				else:
					att = frappe.new_doc("Attendance")
					att.employee = doc.employee
				att.company = doc.company
				att.status = "Present"
				att.working_hours = 8
				att.attendance_request = doc.name
				att.save(ignore_permissions=True)
				att.submit()
				frappe.db.commit()
			# update attendance status based on session and existing working hours
			if doc.reason == "On Duty" and doc.custom_session in ["First Half", "Second Half"]:
				if frappe.db.exists("Attendance", {'employee': doc.employee, 'attendance_date': date, 'docstatus': ('!=', 2)}):
					att = frappe.get_doc("Attendance", {'employee': doc.employee, 'attendance_date': date, 'docstatus': ('!=', 2)})
					if att.working_hours >= 4:
						att.working_hours += 4
						att.status = "Present"
					else:
						att.working_hours += 4
						att.status = "Half Day"
					att.company = doc.company
				else:
					att = frappe.new_doc("Attendance")
					att.employee = doc.employee
					att.working_hours = 4
					att.company = doc.company
					att.status = "Half Day"
				att.attendance_request = doc.name
				att.save(ignore_permissions=True)
				att.submit()
				frappe.db.commit()
			if doc.reason == "On Duty" and doc.custom_session == "Flexible":
				if frappe.db.exists("Attendance", {'employee': doc.employee, 'attendance_date': date, 'docstatus': ('!=', 2)}):
					att = frappe.get_doc("Attendance", {'employee': doc.employee, 'attendance_date': date, 'docstatus': ('!=', 2)})
					if att.in_time and att.out_time:
						st = datetime.strptime(str(doc.custom_from_time), '%H:%M:%S').time()
						start_time = dt.datetime.combine(att.attendance_date,st)
						if att.in_time > start_time :
							att.in_time = start_time
						et = datetime.strptime(str(doc.custom_to_time), '%H:%M:%S').time()
						end_time = dt.datetime.combine(att.attendance_date,et)
						if att.out_time < end_time :
							att.out_time = end_time
						att.save(ignore_permissions=True)
						frappe.db.commit()
						
					

def get_dates(from_date,to_date):
	# method to get the dates from the given range
	no_of_days = date_diff(add_days(to_date, 1), from_date)
	dates = [add_days(from_date, i) for i in range(0, no_of_days)]
	return dates

@frappe.whitelist()
def update_birthday_alowance(doc,method):
	# add birthday allowance for left employee , birthday is on same as relieving month
	if doc.status == "Left":
		if doc.date_of_birth > doc.relieving_date:
			if doc.date_of_birth.month == doc.relieving_date.month:
				first_day = get_first_day(doc.relieving_date)
				if frappe.db.exists("Additional Salary", {'employee': doc.name, 'salary_component': "Birthday Allowance", 'payroll_date': first_day, 'docstatus': 1}):
					ad = frappe.get_doc("Additional Salary", {'employee': doc.name, 'salary_component': "Birthday Allowance", 'payroll_date': first_day, 'docstatus': 1})
					ad.update({
						'docstatus': 2
					})
					ad.save()


@frappe.whitelist()
def create_lwf():
	# add lwf deduction amount 20 rupees by default on december.
	def is_december_1(date_to_check):
		return date_to_check.month == 12 and date_to_check.day == 1
	employee_query = """
	SELECT *
	FROM `tabEmployee`
	WHERE
		status = 'Active'  """
	employee = frappe.db.sql(employee_query, as_dict=True)
	date_to_check = date.today()
	if is_december_1(date_to_check):
		for emp in employee:
			if frappe.db.exists("Salary Structure Assignment", {'employee': emp.name, 'docstatus': 1}):
				if not frappe.db.exists('Additional Salary', {'employee': emp.name, 'payroll_date': date_to_check, 'salary_component': "Labour Welfare Fund", 'docstatus': ('!=', 2)}):
					lwf = frappe.new_doc("Additional Salary")
					lwf.employee = emp.name
					lwf.payroll_date = date_to_check
					lwf.company = emp.company
					lwf.salary_component = "Labour Welfare Fund"
					lwf.currency = "INR"
					lwf.amount = 20
					lwf.save(ignore_permissions=True)
					lwf.submit()
	else:
		print("The date is not December 1st.")

@frappe.whitelist()
def renamed_doc(doc,method):
	# method used to rename employee
	name = doc.name
	employee_number = doc.employee_number
	emp = frappe.get_doc("Employee",name)
	emps=frappe.get_all("Employee",{"status":"Active"},['*'])
	for i in emps:
		if emp.employee_number == employee_number:
			pass
		elif i.employee_number == employee_number:
			frappe.throw(f"Employee Number already exists for {i.name}")
		else:
			frappe.db.set_value("Employee",name,"employee_number",employee_number)
			frappe.rename_doc("Employee", name, employee_number, force=1)


@frappe.whitelist(allow_guest=True)
# code for live attendance page in home
def get_live_attendance():
	nowtime = datetime.now()
	att_details = {}
	att_details['nowtime'] = datetime.strftime(nowtime, '%d-%m-%Y %H:%M:%S')
	max_out = datetime.strptime('06:30', '%H:%M').time()

	if nowtime.time() > max_out:
		date1 = nowtime.date()
	else:
		date1 = (nowtime - timedelta(days=1)).date()

	staff_count = frappe.db.sql("""
		SELECT COUNT(*) AS count
		FROM `tabAttendance`
		WHERE attendance_date = %s
		AND custom_employee_category IN ("Staff", "Sub Staff", "Director")
		AND in_time IS NOT NULL
		AND out_time IS NULL
	""", (date1,), as_dict=True)

	att_details['staff_count'] = staff_count[0].count if staff_count else 0
	trainee_count = frappe.db.sql("""
		SELECT COUNT(*) AS count
		FROM `tabAttendance`
		WHERE attendance_date = %s
		AND custom_employee_category IN ("Trainee")
		AND in_time IS NOT NULL
		AND out_time IS NULL
	""", (date1,), as_dict=True)

	att_details['trainee_count'] = trainee_count[0].count if trainee_count else 0
	ops_count = frappe.db.sql("""
		SELECT COUNT(*) AS count
		FROM `tabAttendance`
		WHERE attendance_date = %s
		AND custom_employee_category IN ("Operator")
		AND in_time IS NOT NULL
		AND out_time IS NULL
	""", (date1,), as_dict=True)

	att_details['ops_count'] = ops_count[0].count if ops_count else 0
	aps_count = frappe.db.sql("""
		SELECT COUNT(*) AS count
		FROM `tabAttendance`
		WHERE attendance_date = %s
		AND custom_employee_category IN ("Apprentice")
		AND in_time IS NOT NULL
		AND out_time IS NULL
	""", (date1,), as_dict=True)

	att_details['aps_count'] = aps_count[0].count if aps_count else 0
	cl_count = frappe.db.sql("""
		SELECT COUNT(*) AS count
		FROM `tabAttendance`
		WHERE attendance_date = %s
		AND custom_employee_category IN ("Contractor")
		AND in_time IS NOT NULL
		AND out_time IS NULL
	""", (date1,), as_dict=True)

	att_details['cl_count'] = cl_count[0].count if cl_count else 0
	tot_count = frappe.db.sql("""
		SELECT COUNT(*) AS count
		FROM `tabAttendance`
		WHERE attendance_date = %s
		AND in_time IS NOT NULL
		AND out_time IS NULL
	""", (date1,), as_dict=True)

	att_details['tot_count'] = tot_count[0].count if tot_count else 0
	return att_details

@frappe.whitelist()
def update_leave_ledger():
	leave_updates = [
	"""update `tabAttendance` set 
		late_entry = 0,
		early_exit = 0,
		custom_late_entry_time = NULL,
		custom_early_out_time = NULL 
	where status = "Half Day" and attendance_date between "2024-03-01" and "2024-05-31"
	"""
]

	for query in leave_updates:
		leave = frappe.db.sql(query, as_dict=True)

@frappe.whitelist()
def check_pf_type(name):
	# method to return print based on the PF eligility
	if frappe.db.exists("Salary Detail",{"parent":name,"salary_component":"Provident Fund"}):
		return "With PF"
	else:
		return "Without PF"


@frappe.whitelist()
def mark_disable(doc,method):
	# disable user if employee status is left
	if doc.status=='Left':
		frappe.db.set_value("User",doc.user_id,"enabled",0)  


@frappe.whitelist()
def update_role(id):
	# If employee category is staff/sub staff then it enables the below role in user list.
	usr=frappe.get_doc("User",id)
	usr.append("roles",{
		"role":"Staff/Sub Staff"
	})
	usr.save(ignore_permissions=True)
	frappe.db.commit()

@frappe.whitelist()
def remove_system_manager_role(doc,method):
	# remove system manager from the roles after creating user document
	usr=frappe.get_doc("User",doc.name)
	usr.remove_roles("System Manager")
	usr.save(ignore_permissions=True)
	frappe.db.commit()

@frappe.whitelist()
def create_user_id(doc,method):
	# method to create user after the employee mis created against the employee
	user_id=doc.name.lower()+'@onegeneindia.in'
	password = "wonjin@321"
	if frappe.db.exists("User",{"email":user_id}):
		frappe.throw("User ID already exists")
	else:
		user=frappe.new_doc("User")
		user.first_name=doc.first_name
		user.middle_name=doc.middle_name
		user.last_name=doc.last_name
		user.username=doc.employee
		user.full_name=doc.employee_name
		user.email=user_id
		
		user.save(ignore_permissions=True)
		frappe.db.commit()
		from frappe.utils.password import update_password
		update_password(user=user_id, pwd=password)
		frappe.db.set_value("Employee",doc.name,'user_id',user_id)
		frappe.db.set_value('Employee',doc.name, 'create_user_permission', 1)


@frappe.whitelist()
def get_deleted_automatically():
	# delete the night shift planning document when the overtime is not done
	yesterday = add_days(today(), -1)
	planning = frappe.db.exists("Night Shift Auditors Planning List", {'attendance_date': yesterday})
	if planning:
		attendance_exists = frappe.db.exists("Attendance", {'employee': planning.emp, 'attendance_date': yesterday, 'docstatus': ('!=', 2)})
		if attendance_exists:
			attendance = frappe.get_doc("Attendance", {'employee': planning.emp, 'attendance_date': yesterday, 'docstatus': ('!=', 2)})
			date1 = dt.datetime.strptime(yesterday, "%Y-%m-%d").date()
			shift_end_time = datetime.strptime("05:00:00", '%H:%M:%S').time()
			start_time = dt.datetime.combine(add_days(date1,1), shift_end_time)
			if attendance.out_time :
				if attendance.out_time > start_time:
					status = "Eligible"
				else:
					status = "Not-Eligible"
			else:
				status = "Not-Eligible"
			if status == "Not-Eligible":
				frappe.delete_doc("Night Shift Auditors Planning List", planning.name, ignore_permissions=True)



@frappe.whitelist()
# method returns the live attendance shift and department wise
def get_data_system(date):
	data =""
	shift=frappe.get_all("Shift Type",{'name':('!=',"4")},['*'],order_by='name ASC')
	shift2=4
	for i in shift:
		shift2+=1
	ec1=0
	ec_count=frappe.get_all("Employee Category",{'name':('not in',['Sub Staff','Director','Trainee'])},['*'])
	for i in ec_count:
		ec1 +=1 
	data = "<table class='table table-bordered=1'>"
	data += "<tr><td colspan ={}  style='border: 1px solid black;background-color:#f6d992;text-align:center'><b>Live Attendance</b></td><td colspan ={} style='border: 1px solid black;background-color:#f6d992;text-align:center'><b>Date {}  </b></td><tr>" .format(shift2,ec1,date)
	shift1=1
	for i in shift:
		shift1+=1
	data += "<tr><td rowspan=2 style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center;'>Parent Department</td><td rowspan=2 style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center;'>Department</td><td colspan={} style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center'>Shift</td><td colspan={} style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center'>Category</td><td rowspan=2 style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center'>CheckOut</td></tr>".format(shift1,ec1)        
	data += "<tr>"
	for i in shift:
		data += "<td style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center'>{}</td>".format(i.name)
	data += "<td style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center'>Total Present</td>"        
	
	ec=frappe.get_all("Employee Category",{'name':('not in',['Sub Staff','Director'])},['*'])
	for i in ec:
		data += "<td style='border: 1px solid black;background-color:#FFA500;font-weight:bold;text-align:center'>{}</td>".format(i.name)
	data +="</tr>"

	total = 0
	department = frappe.get_all("Department", {'disabled': ('!=', 1),"parent_department":"All Departments"}, ['name'])        
	for d in department:
		length=2
		department1 = frappe.get_all("Department", {'disabled': ('!=', 1),"parent_department":d.name}, ['name'])
		for dep in department1:
			length+=1
		parent_dep=d.name
		total_pre=0
		total_cl=0
		total_trainee=0
		total_ops=0
		total_staff=0
		totl_ch_out=0
		data += "<tr><td rowspan={} style='border: 1px solid black;text-align:left'>{}</td><td style='border: 1px solid black;text-align:center'></td>".format(length,d.name)
		for i in shift:
			shift_attendance_count = frappe.db.sql("""
				SELECT COUNT(*) AS count
				FROM `tabAttendance`
				WHERE attendance_date = %s
				AND shift = %s
				AND department = %s
				AND in_time IS NOT NULL

			""", (date, i.name, d.name), as_dict=True)
			shift_attendance = shift_attendance_count[0].count if shift_attendance_count else 0
			data += "<td style='border: 1px solid black;text-align:center'>{}</td>".format(shift_attendance)
		staff_count = frappe.db.sql("""
			SELECT COUNT(*) AS count
			FROM `tabAttendance`
			WHERE attendance_date = %s
			AND custom_employee_category IN ("Staff", "Sub Staff", "Director")
			AND department = %s
			AND in_time IS NOT NULL
		""", (date,d.name), as_dict=True)
		staff = staff_count[0].count if staff_count else 0
		ops_count = frappe.db.sql("""
			SELECT COUNT(*) AS count
			FROM `tabAttendance`
			WHERE attendance_date = %s
			AND custom_employee_category IN ("Operator")
			AND department = %s
			AND in_time IS NOT NULL
		""", (date,d.name), as_dict=True)
		ops = ops_count[0].count if ops_count else 0
		aps_count = frappe.db.sql("""
			SELECT COUNT(*) AS count
			FROM `tabAttendance`
			WHERE attendance_date = %s
			AND custom_employee_category IN ("Apprentice")
			AND department = %s
			AND in_time IS NOT NULL
		""", (date,d.name), as_dict=True)
		trainee = aps_count[0].count if aps_count else 0
		cl_count = frappe.db.sql("""
			SELECT COUNT(*) AS count
			FROM `tabAttendance`
			WHERE attendance_date = %s
			AND custom_employee_category IN ("Contractor")
			AND department = %s
			AND in_time IS NOT NULL
		""", (date,d.name), as_dict=True)
		cl = cl_count[0].count if cl_count else 0
		
		checkout_count = frappe.db.sql("""
			SELECT COUNT(*) AS count
			FROM `tabAttendance`
			WHERE attendance_date = %s
			AND department = %s
			AND in_time IS NOT NULL
			AND out_time IS NOT NULL
		""", (date,d.name), as_dict=True)
		ch_out = checkout_count[0].count if checkout_count else 0
		total += (staff+ops+trainee+cl)
		total_pre+=(staff+ops+trainee+cl)
		total_cl+=cl
		total_trainee+=trainee
		total_ops+=ops
		total_staff+=staff
		totl_ch_out+=ch_out
		data += "<td style='border: 1px solid black;text-align:center;background-color:#ADD8E6'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center;background-color:#BACC81'>%s</td>" % ((staff+ops+trainee+cl),cl,trainee,ops,staff,ch_out)
		data += '</tr>'
		department = frappe.get_all("Department", {'disabled': ('!=', 1),"parent_department":d.name}, ['name'])
		for d in department:
			data += "<tr><td style='border: 1px solid black;text-align:center'>%s</td>"%(d.name)
			for i in shift:
				shift_attendance_count = frappe.db.sql("""
					SELECT COUNT(*) AS count
					FROM `tabAttendance`
					WHERE attendance_date = %s
					AND shift = %s
					AND department = %s
					AND in_time IS NOT NULL
				""", (date, i.name, d.name), as_dict=True)
				shift_attendance = shift_attendance_count[0].count if shift_attendance_count else 0
				data += "<td style='border: 1px solid black;text-align:center'>{}</td>".format(shift_attendance)
			staff_count = frappe.db.sql("""
				SELECT COUNT(*) AS count
				FROM `tabAttendance`
				WHERE attendance_date = %s
				AND custom_employee_category IN ("Staff", "Sub Staff", "Director")
				AND department = %s
				AND in_time IS NOT NULL
			""", (date,d.name), as_dict=True)
			staff = staff_count[0].count if staff_count else 0
			ops_count = frappe.db.sql("""
				SELECT COUNT(*) AS count
				FROM `tabAttendance`
				WHERE attendance_date = %s
				AND custom_employee_category IN ("Operator")
				AND department = %s
				AND in_time IS NOT NULL
			""", (date,d.name), as_dict=True)
			ops = ops_count[0].count if ops_count else 0
			aps_count = frappe.db.sql("""
				SELECT COUNT(*) AS count
				FROM `tabAttendance`
				WHERE attendance_date = %s
				AND custom_employee_category IN ("Apprentice")
				AND department = %s
				AND in_time IS NOT NULL
			""", (date,d.name), as_dict=True)
			trainee = aps_count[0].count if aps_count else 0
			cl_count = frappe.db.sql("""
				SELECT COUNT(*) AS count
				FROM `tabAttendance`
				WHERE attendance_date = %s
				AND custom_employee_category IN ("Contractor")
				AND department = %s
				AND in_time IS NOT NULL
			""", (date,d.name), as_dict=True)
			cl = cl_count[0].count if cl_count else 0
			checkout_count = frappe.db.sql("""
				SELECT COUNT(*) AS count
				FROM `tabAttendance`
				WHERE attendance_date = %s
				AND department = %s
				AND in_time IS NOT NULL
				AND out_time IS NOT NULL
			""", (date,d.name), as_dict=True)
			ch_out = checkout_count[0].count if checkout_count else 0
			total += (staff+ops+trainee+cl)
			total_pre+=(staff+ops+trainee+cl)
			total_cl+=cl
			total_trainee+=trainee
			total_ops+=ops
			total_staff+=staff
			totl_ch_out+=ch_out
			data += "<td style='border: 1px solid black;text-align:center;background-color:#ADD8E6'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center;background-color:#BACC81'>%s</td></tr>" % ((staff+ops+trainee+cl),cl,trainee,ops,staff,ch_out)
		data += "<tr style='border: 1px solid black;text-align:center;background-color:#C0C0C0'><td style='border: 1px solid black;text-align:center'>Total</td>"
		for i in shift:
			shift_count=0
			shift_attendance_count = frappe.db.sql("""
				SELECT COUNT(*) AS count
				FROM `tabAttendance`
				WHERE attendance_date = %s
				AND shift = %s
				AND department = %s
				AND in_time IS NOT NULL

			""", (date, i.name, parent_dep), as_dict=True)
			shift_attendance = shift_attendance_count[0].count if shift_attendance_count else 0
			shift_count+=shift_attendance
			department = frappe.get_all("Department", {'disabled': ('!=', 1),"parent_department":parent_dep}, ['name'])
			for d in department:
				shift_attendance_count = frappe.db.sql("""
					SELECT COUNT(*) AS count
					FROM `tabAttendance`
					WHERE attendance_date = %s
					AND shift = %s
					AND department = %s
					AND in_time IS NOT NULL

				""", (date, i.name, d.name), as_dict=True)
				shift_attendance = shift_attendance_count[0].count if shift_attendance_count else 0
				shift_count+=shift_attendance
			data += "<td style='border: 1px solid black;text-align:center'>{}</td>".format(shift_count)
		data+="<td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td><td style='border: 1px solid black;text-align:center'>%s</td></tr>" % (total_pre,total_cl,total_trainee,total_ops,total_staff,totl_ch_out)
	colspan=(shift2)-2
	data += "<tr><td colspan = {} style='border: 1px solid black;text-align:left'>Total Present</td><td colspan=6 style='border: 1px solid black;text-align:left'>{}</td></tr>" .format(colspan,total)
	data += "</table>"
	return data

import frappe
from datetime import date
@frappe.whitelist()
# when already leave application present in draft status for existing balance, then error will be thrown

def restrict_for_zero_balance(doc, method):
	today = date.today()

	
	start_of_year = date(today.year, 1, 1)

	end_of_year = date(today.year, 12, 31)
	if doc.is_new() and doc.leave_type!='Leave Without Pay': 
		total_leave_days_present=0
		total_lbalance=doc.leave_balance
		draft_leave_applications = frappe.get_all("Leave Application", {"employee": doc.employee,"workflow_state":('in',['Draft','Pending For HOD']),"leave_type": doc.leave_type,'from_date':('between',(start_of_year,end_of_year)),'to_date':('between',(start_of_year,end_of_year))},["*"])
		for i in draft_leave_applications:
			total_leave_days_present+=i.total_leave_days
		total_leave_days_present += doc.total_leave_days
		available=total_lbalance-total_leave_days_present
		if available < 0 :
			frappe.throw("Insufficient leave balance for this leave type")

@frappe.whitelist()
def att_request_cancel(doc, method):
	# empty the field attendance request during cancel
	att=frappe.db.get_value("Attendance",{'attendance_request':doc.name},['name'])
	if att:
		attendance = frappe.db.get_value('Attendance', {
			'employee': doc.employee,
			'attendance_date': doc.from_date,
			'docstatus': ("!=", 2)
		}, ['name'])
		frappe.db.set_value('Attendance',att,'attendance_request','')


# restriction to apply leave after a leave day
@frappe.whitelist()
def condition_for_la(doc, method):
	if doc.workflow_state == "Draft" or doc.is_new():
		diff = date_diff(today(), doc.from_date)
		role = frappe.db.get_value("Has Role",{"parent":frappe.session.user,"role":["in",["HR User","HR Manager"]]})
		if not role:
			employee_category = frappe.db.get_value("Employee",doc.employee,"employee_category")
			if employee_category in ["Staff","Sub Staff"]:
				if diff > 3:
					frappe.throw("The Leave Application must be apply within 3 days from the leave date")
			if employee_category in ["Apprentice","Operator", "Contractor", "Trainee"]:
				if diff > 0:
					frappe.throw("Leave applications must be applied on or before the same day.")

@frappe.whitelist()
def return_items(doctype,docname):
	doc = frappe.get_doc(doctype,docname)
	return doc.items
@frappe.whitelist()
# restriction to apply attendance request when days exceeded 3 
def condition_for_ar(doc,method):
	diff = date_diff(today(), doc.from_date)
	role = frappe.db.get_value("Has Role",{"parent":frappe.session.user,"role":["in",["HR User","HR Manager","HOD"]]})
	if not role:
		if diff > 3:
			frappe.throw("The Attendance Request must be apply within 3 days")

@frappe.whitelist()
# restriction to apply comp leave request when days exceeded 3 
def condition_for_compoff_lr(doc,method):
	diff = date_diff(today(), doc.work_from_date)
	role = frappe.db.get_value("Has Role",{"parent":frappe.session.user,"role":["in",["HR User","HR Manager","HOD"]]})
	if not role:
		if diff > 3:
			frappe.throw("The Compensatory Leave Request must be apply within 3 days")

@frappe.whitelist()
# restriction to apply attendance permission when days exceeded 3 
def condition_for_ap(doc,method):
	diff = date_diff(today(), doc.permission_date)
	role = frappe.db.get_value("Has Role",{"parent":frappe.session.user,"role":["in",["HR User","HR Manager","HOD"]]})
	if not role:
		if diff > 3:
			frappe.throw("The Attendance Permission must be apply within 3 days")

@frappe.whitelist()
# restriction to apply night shift auditors plan swapping when days exceeded 3 
def condition_for_nsaps(doc,method):
	diff = date_diff(today(), doc.requesting_date)
	role = frappe.db.get_value("Has Role",{"parent":frappe.session.user,"role":["in",["HR User","HR Manager","HOD"]]})
	if not role:
		if diff > 3:
			frappe.throw("The Night Shift Auditors Plan Swapping must be apply within 3 days")

@frappe.whitelist()
# method to return the table in the leave application with the ot value and balance
def get_ot_balance(custom_employee,custom_from_date,custom_to_date):
	data = ''
	OTBalance = frappe.qb.DocType("OT Balance")
	ot_balance = (
		frappe.qb.from_(OTBalance)
		.select(OTBalance.employee, OTBalance.total_ot_hours, OTBalance.comp_off_pending_for_approval,OTBalance.comp_off_used,OTBalance.ot_balance)
		.where(
			(OTBalance.employee == custom_employee)
			& ((custom_from_date >= OTBalance.from_date) & (custom_to_date <= OTBalance.to_date))
		)
	).run(as_dict=True)
	if ot_balance and ot_balance[0]:
		data += '<br><br>'
		data += '<table border=1 width=100%>'
		data += '<tr style="text-align:center;background-color:#ff9248;color:#FFFFFF"><td>Total OT Hours</td><td>C-OFF (Pending for Approval) in Day(s)</td><td>C-OFF Used in Day(s)</td><td>OT Balance Hours</td></tr>'
		data += '<tr style="text-align:center;"><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>'%(ot_balance[0].total_ot_hours,ot_balance[0].comp_off_pending_for_approval,ot_balance[0].comp_off_used,ot_balance[0].ot_balance)
		data += '</table><br><br>'
	else:
		data += '<p style="text-align:center;"><b>OT balance is not available</b></p>'
	return data

@frappe.whitelist()
#checks the OT balance is sufficient to apply C-OFF from OT
def validate_ot(employee,total_leave_days,from_date,to_date,employee_category):
	result=2
	OTBalance = frappe.qb.DocType("OT Balance")
	ot_balance = (
		frappe.qb.from_(OTBalance)
		.select(OTBalance.ot_balance)
		.where(
			(OTBalance.employee == employee)
			& ((from_date >= OTBalance.from_date) & (to_date <= OTBalance.to_date))
		)
	).run(as_dict=True)
	if ot_balance and ot_balance[0]:
		if float(total_leave_days)*float(8) > float(ot_balance[0].ot_balance):
			result=frappe.throw("Insufficient OT Balance to apply for C-OFF")
			return result
		else:
			if employee_category not in["Staff","Sub Staff"]:
				if frappe.db.exists("Leave Allocation",{'employee':employee,'leave_type':"Compensatory Off",'from_date':['<=', from_date],'to_date':['>=', to_date],'docstatus':("!=",2)}):
					lal=frappe.get_doc("Leave Allocation",{'employee':employee,'leave_type':"Compensatory Off",'from_date':['<=', from_date],'to_date':['>=', to_date],'docstatus':("!=",2)})
					lal.new_leaves_allocated = lal.new_leaves_allocated + float(total_leave_days)
					# frappe.db.set_value('Leave Allocation',lal.name,'new_leaves_allocated',lal.new_leaves_allocated + float(total_leave_days))
					# frappe.db.set_value('Leave Allocation',lal.name,'total_leaves_allocated',lal.total_leaves_allocated + float(total_leave_days))
					lal.save(ignore_permissions=True)
					lal.submit()
				else:
					from_date = datetime.strptime(from_date, "%Y-%m-%d")
					last_date_of_year = date(from_date.year, 12, 31)
					first_date_of_year = date(from_date.year, 1, 1)
					lal=frappe.new_doc("Leave Allocation")
					lal.employee=employee
					lal.leave_type='Compensatory Off'
					lal.from_date=first_date_of_year
					lal.to_date=last_date_of_year
					lal.new_leaves_allocated=total_leave_days
					lal.save(ignore_permissions=True)
					lal.submit()

@frappe.whitelist()
#Returns the number of days between the custom from date and to date in Leave Application
def get_number_of_leave_days(
	custom_employee: str,
	custom_from_date: datetime.date,
	custom_to_date: datetime.date,
	custom_half_day: Union[int, str, None] = None,
	custom_half_day_date: Union[datetime.date, str, None] = None,
	holiday_list: Optional[str] = None,
) -> float:
	"""Returns number of leave days between 2 dates after considering half day and holidays
	(Based on the include_holiday setting in Leave Type)"""
	number_of_days = 0
	if cint(custom_half_day) == 1:
		if getdate(custom_from_date) == getdate(custom_to_date):
			number_of_days = 0.5
		elif custom_half_day_date and getdate(custom_from_date) <= getdate(custom_half_day_date) <= getdate(custom_to_date):
			number_of_days = date_diff(custom_to_date, custom_from_date) + 0.5
		else:
			number_of_days = date_diff(custom_to_date, custom_from_date) + 1
	else:
		number_of_days = date_diff(custom_to_date, custom_from_date) + 1

	return number_of_days

@frappe.whitelist()
#returns the Leave Type in the custom feild based on their Leave Balance in Leave Ledger Entry 
def return_select_options(employee):
	from datetime import datetime
	select_option = []
	date=today()
	current_datetime = datetime.now()
	current_year = current_datetime.year
	employee_category = frappe.db.get_value('Employee',{'name':employee},'employee_category')
	leave = frappe.db.sql("""
		SELECT leave_type, SUM(leaves) AS total_leaves
		FROM `tabLeave Ledger Entry`
		WHERE docstatus != '2'
		AND employee = %s
		AND from_date <= %s
		AND to_date >= %s
		GROUP BY leave_type
		HAVING total_leaves > 0
		ORDER BY leave_type
	""", (employee, date, date), as_dict=1)

	if employee_category not in ["Staff","Sub Staff"]:
		select_option = ["Comp-off from OT","Leave Without Pay"]
		if leave:
			for l in leave:
				select_option.append(l['leave_type'])
	else:
		select_option = ["Leave Without Pay"]
		if leave:
			for l in leave:
				select_option.append(l['leave_type'])
	return select_option

@frappe.whitelist()
#returns the Leave Type in the custom feild based on their Leave Balance in Leave Ledger Entry 
def get_the_employee_category(employee):
	employee_category = frappe.db.get_value('Employee',{'name':employee},'employee_category')
	user_id = frappe.db.get_value('Employee',{'name':employee},'user_id')
	designation = frappe.db.get_value('Employee',{'name':employee},'designation')
	employee_name = frappe.db.get_value('Employee',{'name':employee},'employee_name')
	department = frappe.db.get_value('Employee',{'name':employee},'department')
	company = frappe.db.get_value('Employee',{'name':employee},'company')
	data =[employee_category,user_id,designation,employee_name,department,company]
	return data

@frappe.whitelist()
#update the draft c-off application count and approved c-off from OT in OT Balance
def otbalance(doc, method):
	month_start = get_first_day(doc.from_date)
	month_end = get_last_day(doc.from_date)
	draft_leave_applications = frappe.get_all(
		"Leave Application",
		filters={
			'employee': doc.employee,
			'from_date': ('between', [month_start, month_end]),
			'to_date': ('between', [month_start, month_end]),
			'workflow_state': 'Pending For HOD',
			'custom_select_leave_type':'Comp-off from OT'
		},
		fields=["total_leave_days"]
	)
	approved_leave_applications = frappe.get_all(
		"Leave Application",
		filters={
			'employee': doc.employee,
			'from_date': ('between', [month_start, month_end]),
			'to_date': ('between', [month_start, month_end]),
			'workflow_state': 'Approved',
			'custom_select_leave_type':'Comp-off from OT'
		},
		fields=["total_leave_days"]
	)
	total_draft_leave_days = sum([i['total_leave_days'] for i in draft_leave_applications])
	total_approved_leave_days = sum([i['total_leave_days'] for i in approved_leave_applications])
	if frappe.db.exists("OT Balance", {'employee': doc.employee, 'from_date': month_start, 'to_date': month_end}):
		otb = frappe.get_doc("OT Balance", {'employee': doc.employee, 'from_date': month_start, 'to_date': month_end})
		otb.comp_off_pending_for_approval = float(total_draft_leave_days)
		otb.comp_off_used = float(total_approved_leave_days)
		otb.ot_balance =float(otb.total_ot_hours)-((float(total_draft_leave_days)*float(8)) + (float(total_approved_leave_days)*float(8)))
		otb.save(ignore_permissions=True)

@frappe.whitelist()
def cancel_leave_application(doc, method):
	if doc.custom_select_leave_type=="Comp-off from OT":
		leave_allocation = frappe.get_doc("Leave Allocation", {
			'employee': doc.custom_employee2,
			'leave_type': "Compensatory Off",
			'from_date': ['<=', doc.from_date],
			'docstatus': ("!=", 2)
		})
		
		if leave_allocation:
			leave_allocation.new_leaves_allocated += float(doc.custom_total_leave_days)
			leave_allocation.save(ignore_permissions=True)
			frappe.db.commit()

		OTBalance = frappe.get_doc("OT Balance", {
			'employee': doc.employee,
			'from_date': ['<=', doc.from_date],
			'to_date': ['>=', doc.to_date]
		})

		if OTBalance:
			OTBalance.ot_balance += float(doc.custom_total_leave_days) * 8
			OTBalance.comp_off_used-=doc.custom_total_leave_days
			OTBalance.save(ignore_permissions=True)
			frappe.db.commit()

@frappe.whitelist()
#send a mail alert to the users in the Custom Settings if the Item is below the Safety Stock
def mail_alert_for_safety_stock():
	item = frappe.get_all("Item",{"disabled":0,"safety_stock":("!=",0)},["name","safety_stock"])
	data = ""
	count=0
	data += '<table class="table table-bordered"><tr><th style="padding:1px;border: 1px solid black;color:white;background-color:#f68b1f" colspan=3><center>Stock Details</center></th></tr>'
	data += '<tr><td style="padding:1px;border: 1px solid black" colspan =1><b>Item Code</b></td><td style="padding:1px;border: 1px solid black" colspan=1><b>Safety Stock</b></td><td style="padding:1px;border: 1px solid black" colspan=1><b>Available Qty</b></td></tr>'
	for i in item:
		stockqty = frappe.db.sql(""" select item_code,sum(actual_qty) as qty from `tabBin` where item_code = '%s' """%(i.name),as_dict = 1)[0]
		if stockqty['qty']:
			stockqty['qty'] = stockqty['qty']
		else:
			stockqty['qty'] =0
		if i.safety_stock >= stockqty['qty']:
			count+=1
			data += '''  
			<tr><td style="padding:1px;border: 1px solid black" colspan =1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td>
			<td style="padding:1px;border: 1px solid black" colspan=1>%s</td></tr>'''%(i.name,i.safety_stock,stockqty['qty'] or 0)
	data += '</table>'
	doc = frappe.get_doc("Custom Settings", "Custom Settings")
	emails = []
	cc_mail=[]
	for d in doc.mail_settings:
		if d.report_name == "Stock Details":
			emails.extend(d.recipients.split('\n'))
			cc_mail.extend(d.cc.split('\n'))
	if count > 0:
		frappe.sendmail(
			recipients=emails,
			cc = cc_mail if d.cc else '',
			subject='Stock Details',
			message="""Dear Sir/Mam,<br><br>
				Kindly Check below Item list qty<br>{0}
				""".format(data)
		)
	# user = frappe.db.sql("""
	#     SELECT `tabUser`.name as name
	#     FROM `tabUser`
	#     LEFT JOIN `tabHas Role` ON `tabHas Role`.parent = `tabUser`.name 
	#     WHERE `tabHas Role`.Role = "Stock Manager" AND `tabUser`.enabled = 1
	# """, as_dict=True)
	# if data:
	#     for i in user:
	#         frappe.sendmail(
	#             recipients=["jenisha.p@groupteampro.com"],
	#             subject='Stock Details',
	#             message="""Dear Sir/Mam,<br><br>
	#                 Kindly Check below Item list qty<br>{0}
	#                 """.format(data)
	#         )


@frappe.whitelist()
def return_options():
	from datetime import datetime
	select_option = []
	current_datetime = datetime.now()
	current_year = current_datetime.year
	leave = frappe.db.sql("""
		SELECT leave_type, SUM(leaves) AS total_leaves
		FROM `tabLeave Ledger Entry`
		WHERE docstatus != '2'
		AND employee = 'H0002'
		AND YEAR(from_date) <= '2024-01-01'
		AND YEAR(to_date) >= '2024-12-31'
		GROUP BY leave_type
		HAVING total_leaves > 0
		ORDER BY leave_type
	""", as_dict=1)
	employee_category=frappe.db.get_value("Employee",{'name':'H0002'},['employee_category'])
	if employee_category not in ["Staff","Sub Staff"]:
		select_option = ["Comp-off from OT","Leave Without Pay"]
		if leave:
			for l in leave:
				select_option.append(l['leave_type'])
	else:
		select_option = ["Leave Without Pay"]
		if leave:
			for l in leave:
				select_option.append(l['leave_type'])
	return select_option

# @frappe.whitelist()
# def emp_category_update_in_leaveallocation():
# 	leaves=frappe.db.get_all("Leave Allocation",{"leave_type":"Sick Leave","docstatus":"1","from_date":"01-01-2024","to_date":"31-12-2024"},["*"])
# 	ind=0
# 	for i in leaves:
# 		emp=frappe.db.get_value("Employee",{"name":i.employee},["employee_category"])
# 		frappe.db.set_value("Leave Allocation",i.name,"custom_employee_category",emp)
# 		ind+=1

@frappe.whitelist()
#return the last execution time of attendance cron
def update_last_execution():
	doc=frappe.db.get_value("Scheduled Job Log",{"scheduled_job_type":"mark_attendance.mark_att","status":"Complete"},["creation"])
	if doc:
		return doc
	

# @frappe.whitelist()
# #use to create user after saving the employee MIS
# def create_user(employee, first_name, employee_name,date_of_birth,gender):
# # def create_user():
#     # employee = "jjjjj1234"
#     # first_name ="test" 
#     # employee_name ="employee_name"
#     password = "Wonjin@2024"
#     # gender ="Female"
#     # date_of_birth ="2003-04-29"
#     email_str = str(employee) + '@onegeneindia.in'

#     # Check if user already exists
#     existing_user = frappe.get_all('User', filters={'email': email_str})
	
#     if existing_user:
#         return {"message": f"User with email {email_str} already exists. User has been updated."}
#     else:
#         # If user does not exist, create a new user
#         user_doc = frappe.new_doc("User")
#         user_doc.email = email_str
#         user_doc.first_name = first_name
#         user_doc.username = employee
#         user_doc.user_type = "System User"
#         user_doc.birth_date = date_of_birth
#         user_doc.gender = gender
#         user_doc.insert(ignore_permissions=True)
#         # frappe.utils.password.update_password(user_doc.name, password)
#         from frappe.utils.password import update_password
#         update_password(user=user_doc.name, pwd=password)
#         # frappe.db.get_value('')
#         # update_password(user_doc.name, password)
#         # frappe.db.set_value('User',user_doc.name,'new_password',password)
#         # frappe.db.set_value('User',user_doc.name,'password_for_admin',password)
#         # Set create_user_permission for the employee
#     frappe.db.set_value('Employee', {'name': employee}, 'user_id', email_str)
#     frappe.db.set_value('Employee', {'name': employee}, 'create_user_permission', 1)
	
#     # user_id =frappe.db.get_value('User', {'name': email_str}, 'password_for_admin')
#     # print(user_id)
#     return email_str

@frappe.whitelist()
#send a mail alert to PS and API if any scheduled job failed
def schedule_log_fail(doc,method):
	if doc.status=='Failed':
		message = """
		The schedule Job type <b>{}</b> is failed.<br> Kindly check the log <b>{}</b>
		""".format(doc.scheduled_job_type,doc.name)
		frappe.sendmail(
				recipients=["pavithra.s@groupteampro.com","abdulla.pi@groupteampro.com"],
				subject='Scheduled Job type failed(ONEGENE)',
				message=message
			)
		
@frappe.whitelist()
#send a mail alert to PS and API if any scheduled job failed
def update_early():
	frappe.db.set_value('Attendance','HR-ATT-2024-420334','custom_early_out_time','00:00:00')
import datetime
import frappe

@frappe.whitelist()
# def get_absent_count():
def get_absent_count(start_date,end_date,employee):
	import datetime
	# start_date='2025-01-01'
	# end_date='2025-01-31'
	# employee='H0004'
	start_date=str(start_date)
	end_date=str(end_date)
	start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
	end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')
	absent_count = 0
	current_week_start = start_date
	while current_week_start <= end_date:
		current_week_end = current_week_start + datetime.timedelta(days=(6 - current_week_start.weekday()))
		if current_week_end>=end_date:
			current_week_end=end_date
		sdate=current_week_start.date()
		edate=current_week_end.date()
		absent_attendance = frappe.db.get_all("Attendance",{'attendance_date': ('between', (sdate, edate)),'status': "Absent",'employee': employee,'docstatus':['!=',2]},['*'])
		woff=0
		for ab in absent_attendance:
			hh=check_holiday(ab.attendance_date,employee)
			if hh is None and ab.leave_application is None and ab.attendance_request is None:
				print(ab.attendance_date)
				woff+=1
		if woff>0:
			absent_count+=1
		current_week_start = current_week_end + datetime.timedelta(days=1)
	return absent_count
from datetime import datetime
import calendar
@frappe.whitelist()
def allocate_el_automatically():
	# frappe.db.set_value("Employee",'4321','employee_category','Staff')
	emp=frappe.db.get_all("Employee",{'Status':'Active'},['name','date_of_joining','employee_category'])
	current_date = datetime.now().date()  # Get the current date
	current_year = current_date.year
	current_month = current_date.month
	# if current_month > 1:
	#     previous_month = current_month-1
	#     start_date_month = datetime(current_year, previous_month, 1).date()
	#     end_date_month = datetime(current_year, previous_month, calendar.monthrange(current_year, previous_month)[1]).date()
	#     start_date = datetime(current_year+1, 1, 1).date()
	#     end_date = datetime(current_year+1, 12, 31).date()
	# else:
	#     previous_month = 12
	#     start_date_month = datetime(current_year - 1, previous_month, 1).date()
	#     end_date_month = datetime(current_year - 1, previous_month, calendar.monthrange(current_year - 1, previous_month)[1]).date()
	#     start_date = datetime(current_year, 1, 1).date()
	#     end_date = datetime(current_year, 12, 31).date()
	
	# for e in emp:
	#     count=0
	#     current_date = datetime.now().date()
	#     doj = e.date_of_joining
	#     diff = current_date - doj
	#     years = diff.days / 365.25 
	#     if int(years)>0:
	#         attendance_filters = {
	#             "attendance_date": ("between", (start_date_month, end_date_month)),
	#             'employee': e['name'],
	#             'status': 'Present',
	#             'docstatus':['!=',2]
	#         }
	#         att = frappe.db.count("Attendance", filters=attendance_filters)
	#         act=32
	#         if e.employee_category in ['Sub Staff','Staff','Operator']:
	#             act=20
	#         elif e.employee_category in ['Apprentice','Trainee']:
	#             act=30
	#         else:
	#             act=32
	#         if att>=act:
	#             count=1
	#         if count>0:
	#             if not frappe.db.exists('Leave Allocation',{'leave_type':'Earned Leave','from_date':start_date,'to_date':end_date,'docstatus':['!=',2],'employee':e.name}):
	#                 allow = frappe.new_doc("Leave Allocation")
	#                 print("EL")
	#                 allow.employee = e.name
	#                 allow.company = e.company
	#                 allow.leave_type = "Earned Leave"
	#                 allow.from_date = start_date
	#                 allow.to_date = end_date
	#                 allow.new_leaves_allocated = count
	#                 allow.total_leaves_allocated = count
	#                 allow.insert()
	#                 allow.save(ignore_permissions=True)
	#                 allow.submit()	
	#                 frappe.db.commit
	#             else:
	#                 allow=frappe.db.get_value('Leave Allocation',{'leave_type':'Earned Leave','from_date':start_date,'to_date':end_date,'docstatus':['!=',2],'employee':e['name']},['name'])
	#                 leaves=frappe.db.get_value('Leave Allocation',{'leave_type':'Earned Leave','from_date':start_date,'to_date':end_date,'docstatus':['!=',2],'employee':e['name']},['new_leaves_allocated'])
	#                 tot=frappe.db.get_value('Leave Allocation',{'leave_type':'Earned Leave','from_date':start_date,'to_date':end_date,'docstatus':['!=',2],'employee':e['name']},['total_leaves_allocated'])
	#                 tot+=count
	#                 leaves+=count
	#                 frappe.db.set_value('Leave Allocation',allow,'new_leaves_allocated',leaves)
	#                 frappe.db.set_value('Leave Allocation',allow,'total_leaves_allocated',leaves)


@frappe.whitelist()
def create_job_fail():
	job = frappe.db.exists('Scheduled Job Type', 'lwf_creation')
	if not job:
		emc = frappe.new_doc("Scheduled Job Type")
		emc.update({
			"method": 'onegene.onegene.custom.create_lwf',
			"frequency": 'Monthly',
		})
		emc.save(ignore_permissions=True)


@frappe.whitelist()
def cron_failed_method():
	cutoff_time = datetime.now() - timedelta(minutes=5)
	failed_jobs = frappe.get_all(
		"Scheduled Job Log",
		filters={
			"status": "Failed",
			"creation": [">=", cutoff_time]
		},
		fields=["scheduled_job_type"]
	)
	unique_job_types = set()
	for job in failed_jobs:
		unique_job_types.add(job['scheduled_job_type'])

	for job_type in unique_job_types:
		frappe.sendmail(
			recipients = ["erp@groupteampro.com","jenisha.p@groupteampro.com","pavithra.s@groupteampro.com","gifty.p@groupteampro.com"],
			subject = 'Failed Cron List - Wonjin',
			message = 'Dear Sir / Mam <br> Kindly find the below failed Scheduled Job  %s'%(job_type)
		)

@frappe.whitelist()
def update_coff():
	count=0
	tot=0
	att= frappe.db.get_all("Attendance",{'attendance_date':'2025-01-12','status':'Present','docstatus':1},['*'])
	for a in att:
		# count+=1
		if frappe.db.exists("Attendance",{'attendance_date':'2025-01-16','status':'Absent','employee': a.employee}):
			count+=1
			print(a.employee)
			# if a.employee not in ['AN4594','S0373']:
			# 		la = frappe.new_doc('Leave Application')
			# 		print(a.employee)
			# 		la.employee = a.employee
			# 		la.leave_type = 'Compensatory Off'
			# 		la.custom_employee2 = a.employee
			# 		la.custom_leave_type = 'Compensatory Off'
			# 		la.from_date = '2025-01-15'
			# 		la.to_date = '2025-01-15'
			# 		la.description = 'Created automatically Via Bulk compensation allocation document'
			# 		la.company = 'WONJIN AUTOPARTS INDIA PVT.LTD.'
			# 		la.status = 'Approved'  
			# 		la.insert(ignore_permissions=True)
			# 		la.submit()
			# 		print(la.name)
			# 		print(a.employee)

	return count

from io import BytesIO
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import frappe
from frappe.utils.background_jobs import enqueue

@frappe.whitelist()
def manpower_cost_download():
	filename = "MANPOWER COST"
	args = frappe.local.form_dict
	frappe.msgprint("Report is generating in the background,kindly check after few mins in the same page.")
	enqueue(build_xlsx_response, queue='long', timeout=36000, event='build_xlsx_response',filename=filename,args=args)
	# build_xlsx_response(filename)

def build_xlsx_response(filename,args):
	xlsx_file = make_xlsx(filename,args)
	ret = frappe.get_doc({
			"doctype": "File",
			"attached_to_name": ' ',
			"attached_to_doctype": 'Reports Dashboard',
			"attached_to_field": 'attach',
			"file_name": filename + '.xlsx',
			"is_private": 0,
			"content": xlsx_file.getvalue(),
			"decode": False
		})
	ret.save(ignore_permissions=True)
	frappe.db.commit()
	attached_file = frappe.get_doc("File", ret.name)
	# frappe.db.set_value('Reports Dashboard',None,'manpower_cost_report',attached_file.file_url)
	doc = frappe.get_single("Reports Dashboard")
	doc.reload()
	doc.manpower_cost_report = attached_file.file_url
	doc.save()


def apply_border_to_merged_cells(ws, start_row, start_col, end_row, end_col, border):
	for row in range(start_row, end_row + 1):
		for col in range(start_col, end_col + 1):
			cell = ws.cell(row=row, column=col)
			cell.border = border

def make_xlsx(data,args, sheet_name="MANPOWER COST", wb=None, column_widths=None):
	if wb is None:
		wb = openpyxl.Workbook()

	ws = wb.active
	ws.title = sheet_name

	from_date = datetime.strptime(args.get('from_date'), '%Y-%m-%d')
	to_date = datetime.strptime(args.get('to_date'), '%Y-%m-%d')
	fromdate = datetime.strptime(args.get('from_date'), '%Y-%m-%d').date()
	from_date_formated = fromdate.strftime('%d-%m-%Y')
	todate = datetime.strptime(args.get('to_date'), '%Y-%m-%d').date()
	to_date_formated = todate.strftime('%d-%m-%Y')
	month_year_format = from_date.strftime('%B %Y')
	# if isinstance(from_date, str):
	#     month_year_format_str = datetime.strptime(from_date, '%Y-%m-%d')
	# else:
	#     month_year_format_str = from_date
	# month_year_format = month_year_format_str.strftime('%B %Y')
	# Header
	header_value = f"Manpower Cost Report ({from_date_formated}  to  {to_date_formated})" 
	# Get the last column index (based on date columns)
	num_days = (to_date - from_date).days + 1
	last_column_index = 6 + 1 * 5 - 1  # 5 columns per date (Attn, Amount, OT, Amount OT, Total)

	# Merge the header cells across all columns
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column_index)
	
	thin_border = Border(
		left=Side(style="thin"),
		right=Side(style="thin"),
		top=Side(style="thin"),
		bottom=Side(style="thin"),
	)
	apply_border_to_merged_cells(ws, 1, 1, 1, last_column_index, thin_border)
	first_cell = ws.cell(row=1, column=1)
	first_cell.value = header_value
	first_cell.font = Font(bold=True, size=14, color="0000FF")  # Blue color for header_value
	first_cell.alignment = Alignment(horizontal="center", vertical="center")  # Center align header

	# Table Headers
	headers = ["Sl No", "Department", "Block Name", "Sal / month", "MP Plan"]
	bg_fill = PatternFill(start_color="e2efd9", end_color="e2efd9", fill_type="solid")
	
	for i, header in enumerate(headers, 1):
		ws.merge_cells(start_row=2, start_column=i, end_row=4, end_column=i)
		apply_border_to_merged_cells(ws, 2, i, 4, i, thin_border)
		cell = ws.cell(row=2, column=i)
		cell.value = header
		if header in ["Sl No", "MP Plan"]:
			cell.font = Font(bold=True, color="FF0000")  # Red color for headers
		if header in ["Department", "Block Name"]:
			cell.fill = bg_fill
			cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center", vertical="center")
	
	department_hierarchy = get_department_hierarchy()
	shift_filter = args.get('shift')  # The shift filter is optional
	
	ws.column_dimensions['A'].width = 5
	ws.column_dimensions['B'].width = 23
	ws.column_dimensions['C'].width = 25
	ws.column_dimensions['D'].width = 12

	row_index = 5
	serial_number = 1
	date_columns = []
	# Dates Row
	start_column = 6  # Starting column for dates

	# for day_offset in range(num_days):
	current_date = from_date + timedelta(days=1)
	date_str = current_date.strftime("%d-%m-%Y")

	# Merge and set date
	# ws.merge_cells(start_row=2, start_column=start_column, end_row=2, end_column=start_column + 4)
	# apply_border_to_merged_cells(ws, 2, start_column, 2, start_column + 4, thin_border)
	# date_cell = ws.cell(row=2, column=start_column)
	# date_cell.value = ''
	# date_cell.font = Font(bold=True)
	# date_cell.alignment = Alignment(horizontal="center", vertical="center")
	# date_cell.font = Font(bold=True)

	# Employee Categories Row
	employee_category = args.get('employee_category')
	ws.merge_cells(start_row=2, start_column=start_column, end_row=3, end_column=start_column + 4)
	apply_border_to_merged_cells(ws, 2, start_column, 3, start_column + 4, thin_border)
	cat_cell = ws.cell(row=2, column=start_column, value=employee_category)
	cat_cell.alignment = Alignment(horizontal="center", vertical="center")
	cat_cell.border = thin_border
	cat_cell.font = Font(bold=True)

	# Sub-columns below categories
	col_type = ["Attn", "Amount", "OT", "Amount", "Total"]
	for col_offset, col_name in enumerate(col_type):
		col_cell = ws.cell(row=4, column=start_column + col_offset)
		col_cell.value = col_name
		col_cell.font = Font(bold=True)
		col_cell.alignment = Alignment(horizontal="center", vertical="center")
		col_cell.border = thin_border

	# Store the starting column index for attendance
	# date_columns.append((current_date, start_column))
	# start_column += 5
	
	total_att = 0
	total_att_amt = 0
	total_ot = 0
	total_ot_amt = 0
	grand_total = 0
	data_column = 6
	
	for department, blocks in department_hierarchy.items():
		start_row = row_index
		
		dept_total_att = 0
		dept_total_att_amt = 0
		dept_total_ot = 0
		dept_total_ot_amt = 0
		dept_total = 0
		
		cell = ws.cell(row=row_index, column=2, value=department)
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center", vertical="center")
		for block in blocks:
			
			s_no = ws.cell(row=row_index, column=1, value=serial_number).border = thin_border
			s_no.font = Font(bold=True)
			child = ws.cell(row=row_index, column=3, value=block).border = thin_border
			child.font = Font(bold=True)
			
			base_result = frappe.db.sql("""
				SELECT SUM(base) AS base 
				FROM `tabSalary Structure Assignment` 
				WHERE custom_employee_category = %s 
				AND department = %s
				AND docstatus = 1
			""", (employee_category, block), as_dict=True)

			if base_result and base_result[0].get("base") is not None:
				base = base_result[0].get("base", 0)
				base_percent = base / 30  # Assuming 30 days in a month for calculation
			else:
				base_percent = 0
				
			round_of_base = round(base_percent)
			
			base_per = ws.cell(row=row_index, column=4, value= round_of_base).border = thin_border
			base_per.font = Font(bold=True)
			start_col =6
		# for date, start_col in date_columns:
			# from_date = datetime.strptime(args.get('from_date'), '%Y-%m-%d')
			# to_date = datetime.strptime(args.get('to_date'), '%Y-%m-%d')
			# current_date = from_date + timedelta(days=1)
			# to_date_str = to_date
			# date_str = current_date.strftime('%Y-%m-%d')
			month_str = current_date.strftime("%b")
			year = current_date.year
			date_str = datetime.strptime(args.get('from_date'), "%Y-%m-%d")
			to_date_str = datetime.strptime(args.get('to_date'), "%Y-%m-%d")
			manpower_plan = frappe.db.sql("""
											SELECT SUM(business_plan) as plan
											FROM `tabManpower Plan`
											WHERE department = %s
											AND month = %s
											AND year = %s
											""", (block, month_str, year))
			
			manpower = manpower_plan[0][0] if manpower_plan and manpower_plan[0][0] is not None else 0
			manpower = manpower*num_days
			man = ws.cell(row=row_index, column=5, value= manpower).border = thin_border
			man.font = Font(bold=True)

			# Attendance Calculation
			if shift_filter:
				attendance = frappe.db.sql("""
					SELECT COUNT(*) as count 
					FROM `tabAttendance` 
					WHERE 
						attendance_date BETWEEN %s AND %s
						AND department = %s 
						AND custom_employee_category = %s 
						AND docstatus != 2 
						AND status = 'Present'
						AND shift = %s
				""", (date_str,to_date_str, block, employee_category, shift_filter), as_dict=True)
			else:
				attendance = frappe.db.sql("""
					SELECT COUNT(*) as count 
					FROM `tabAttendance` 
					WHERE 
						attendance_date BETWEEN %s AND %s 
						AND department = %s 
						AND custom_employee_category = %s 
						AND docstatus != 2 
						AND status = 'Present'
				""", (date_str,to_date_str, block, employee_category), as_dict=True)

			att_value = attendance[0]["count"] if attendance else 0

			# Ensure base_percent is rounded properly
			base_percent = round(base / 30) if base_result and base_result[0].get("base") else 0

			# Calculate amount_value using integer multiplication
			att_amount = att_value * base_percent
			
			att = ws.cell(row=row_index, column=start_col, value= att_value).border = thin_border
			att.font = Font(bold=True)
			
			amount_att = ws.cell(row=row_index, column=start_col+1, value= att_amount).border = thin_border
			amount_att.font = Font(bold=True)
			
			if shift_filter:
				ot_result = frappe.db.sql("""
					SELECT SUM(custom_overtime_hours) as ot
					FROM `tabAttendance` 
					WHERE 
						attendance_date BETWEEN %s AND %s
						AND department = %s 
						AND custom_employee_category = %s 
						AND docstatus != 2 
						AND status = 'Present'
						AND shift = %s
				""", (date_str,to_date_str, block, employee_category ,shift_filter), as_dict=True)
			else:
				ot_result = frappe.db.sql("""
					SELECT SUM(custom_overtime_hours) as ot
					FROM `tabAttendance` 
					WHERE 
						attendance_date BETWEEN %s AND %s 
						AND department = %s 
						AND custom_employee_category = %s 
						AND docstatus != 2 
						AND status = 'Present'
				""", (date_str,to_date_str, block, employee_category), as_dict=True)

			ot_value = ot_result[0]["ot"] if ot_result and ot_result[0]["ot"] else 0

			ot = ws.cell(row=row_index, column=start_col+2, value= round(ot_value, 2)).border = thin_border
			ot.font = Font(bold=True)
			
			if shift_filter:
				attendance_ot = frappe.db.sql("""
					SELECT SUM(custom_overtime_hours) as ot
					FROM `tabAttendance` 
					WHERE 
						attendance_date BETWEEN %s AND %s
						AND department = %s 
						AND custom_employee_category = %s 
						AND docstatus != 2 
						AND status = 'Present'
						AND shift = %s
				""", (date_str,to_date_str, block, employee_category, shift_filter), as_dict=True)
			else:
				attendance_ot = frappe.db.sql("""
					SELECT SUM(custom_overtime_hours) as ot
					FROM `tabAttendance` 
					WHERE 
						attendance_date BETWEEN %s AND %s
						AND department = %s 
						AND custom_employee_category = %s 
						AND docstatus != 2 
						AND status = 'Present'
				""", (date_str,to_date_str, block, employee_category), as_dict=True)

			sum_ot = attendance_ot[0]["ot"] if attendance_ot and attendance_ot[0]["ot"] is not None else 0
			ot_amount = sum_ot * base_percent
			
			amount_ot = ws.cell(row=row_index, column=start_col+3, value= ot_amount).border = thin_border
			amount_ot.font = Font(bold=True)
			
			total = att_amount + ot_amount  # Total remains precise with integer arithmetic
			
			tot = ws.cell(row=row_index, column=start_col+4, value= total).border = thin_border
			tot.font = Font(bold=True)
			
			dept_total_att += att_value
			dept_total_att_amt += att_amount
			dept_total_ot += ot_value
			dept_total_ot_amt += ot_amount
			dept_total += total
			
			serial_number += 1
			row_index += 1
			# data_column += 5
			
		total_att += dept_total_att
		total_att_amt += dept_total_att_amt
		total_ot += dept_total_ot
		total_ot_amt += dept_total_ot_amt
		grand_total += dept_total
		
		if blocks:
			ws.merge_cells(start_row=start_row, start_column=2, end_row=row_index - 1, end_column=2)
			
		ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
		cell = ws.cell(row=row_index, column=1, value="SUB - TOTAL")
		cell.font = Font(bold=True, color="FF0000")
		cell.fill = bg_fill
		cell.alignment = Alignment(horizontal="right", vertical="center")
		cell.border = thin_border
		
		if department != list(department_hierarchy.keys())[-1]:  
			ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
			cell = ws.cell(row=row_index, column=1, value="SUB - TOTAL")
			cell.font = Font(bold=True, color="FF0000")
			cell.fill = bg_fill
			cell.alignment = Alignment(horizontal="right", vertical="center")

			col_index = 6
			while col_index <= last_column_index:
				data_cell = ws.cell(row=row_index, column=col_index, value=dept_total_att)
				data_cell.font = Font(bold=True, color="FF0000")
				data_cell.fill = bg_fill
				data_cell.border = thin_border
				col_index += 1

				data_cell = ws.cell(row=row_index, column=col_index, value=dept_total_att_amt)
				data_cell.font = Font(bold=True, color="FF0000")
				data_cell.fill = bg_fill
				data_cell.border = thin_border
				col_index += 1

				data_cell = ws.cell(row=row_index, column=col_index, value=dept_total_ot)
				data_cell.font = Font(bold=True, color="FF0000")
				data_cell.fill = bg_fill
				data_cell.border = thin_border
				col_index += 1

				data_cell = ws.cell(row=row_index, column=col_index, value=dept_total_ot_amt)
				data_cell.font = Font(bold=True, color="FF0000")
				data_cell.fill = bg_fill
				data_cell.border = thin_border
				col_index += 1

				data_cell = ws.cell(row=row_index, column=col_index, value=dept_total)
				data_cell.font = Font(bold=True, color="FF0000")
				data_cell.fill = bg_fill
				data_cell.border = thin_border
				col_index += 1
			
			row_index += 1
			
		if department ==  list(department_hierarchy.keys())[-2]:
			green = PatternFill(start_color="385723", end_color="385723", fill_type="solid")
			ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
			cell = ws.cell(row=row_index, column=1, value="VARIABLE SALARY TOTAL")
			cell.font = Font(bold=True, color="FFFFFF")
			cell.fill = green
			cell.alignment = Alignment(horizontal="right", vertical="center")
			
			var_col = 6
			while var_col <= last_column_index:
				cell = ws.cell(row=row_index, column=var_col, value=total_att)
				cell.font = Font(bold=True, color="FFFFFF")
				cell.fill = green
				cell.alignment = Alignment(horizontal="right", vertical="center")
				cell.border = thin_border
				var_col += 1
				
				cell = ws.cell(row=row_index, column=var_col, value=total_att_amt)
				cell.font = Font(bold=True, color="FFFFFF")
				cell.fill = green
				cell.alignment = Alignment(horizontal="right", vertical="center")
				cell.border = thin_border
				var_col += 1
				
				cell = ws.cell(row=row_index, column=var_col, value=total_ot)
				cell.font = Font(bold=True, color="FFFFFF")
				cell.fill = green
				cell.alignment = Alignment(horizontal="right", vertical="center")
				cell.border = thin_border
				var_col += 1
				
				cell = ws.cell(row=row_index, column=var_col, value=total_att_amt)
				cell.font = Font(bold=True, color="FFFFFF")
				cell.fill = green
				cell.alignment = Alignment(horizontal="right", vertical="center")
				cell.border = thin_border
				var_col += 1
				
				cell = ws.cell(row=row_index, column=var_col, value=grand_total)
				cell.font = Font(bold=True, color="FFFFFF")
				cell.fill = green
				cell.alignment = Alignment(horizontal="right", vertical="center")
				cell.border = thin_border
				var_col += 1
				
			row_index += 1
			
		if department == list(department_hierarchy.keys())[-1]: 

			ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
			brown = PatternFill(start_color="833c0b", end_color="833c0b", fill_type="solid")

			cell = ws.cell(row=row_index, column=1, value="FIXED SALARY TOTAL")
			cell.font = Font(bold=True, color="FFFFFF")
			cell.fill = brown
			cell.alignment = Alignment(horizontal="center", vertical="center")

			col_idx_last = 6
			while col_idx_last <= last_column_index:
				grand = ws.cell(row=row_index, column=col_idx_last, value=total_att)
				grand.font = Font(bold=True, color="FFFFFF")
				grand.fill = brown
				grand.border = thin_border
				col_idx_last += 1

				grand = ws.cell(row=row_index, column=col_idx_last, value=total_att_amt)
				grand.font = Font(bold=True, color="FFFFFF")
				grand.fill = brown
				grand.border = thin_border
				col_idx_last += 1

				grand = ws.cell(row=row_index, column=col_idx_last, value=total_ot)
				grand.font = Font(bold=True, color="FFFFFF")
				grand.fill = brown
				grand.border = thin_border
				col_idx_last += 1

				grand = ws.cell(row=row_index, column=col_idx_last, value=total_ot_amt)
				grand.font = Font(bold=True, color="FFFFFF")
				grand.fill = brown
				grand.border = thin_border
				col_idx_last += 1

				grand = ws.cell(row=row_index, column=col_idx_last, value=grand_total)
				grand.font = Font(bold=True, color="FFFFFF")
				grand.fill = brown
				grand.border = thin_border
				col_idx_last += 1

			for col in range(1, last_column_index + 1):
				ws.cell(row=row_index, column=col).fill = brown
				ws.cell(row=row_index, column=col).border = thin_border

			row_index += 1

			white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
			ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
			cell = ws.cell(row=row_index, column=1, value="FIXED SALARY TOTAL")
			cell.font = Font(bold=True, color="000000")
			cell.fill = white
			cell.alignment = Alignment(horizontal="center", vertical="center")

			last_col = 6
			while last_col <= last_column_index:
				last = ws.cell(row=row_index, column=last_col, value=total_att)
				last.font = Font(bold=True, color="000000")
				last.fill = white
				last.alignment = Alignment(horizontal="center", vertical="center")
				last_col += 1

				last = ws.cell(row=row_index, column=last_col, value=total_att_amt)
				last.font = Font(bold=True, color="000000")
				last.fill = white
				last_col += 1

				last = ws.cell(row=row_index, column=last_col, value=total_ot)
				last.font = Font(bold=True, color="000000")
				last.fill = white
				last_col += 1

				last = ws.cell(row=row_index, column=last_col, value=total_ot_amt)
				last.font = Font(bold=True, color="000000")
				last.fill = white
				last_col += 1

				last = ws.cell(row=row_index, column=last_col, value=grand_total)
				last.font = Font(bold=True, color="000000")
				last.fill = white
				last_col += 1

			for col in range(1, last_column_index + 1):
				ws.cell(row=row_index, column=col).fill = white
				ws.cell(row=row_index, column=col).border = thin_border

			row_index += 1

	ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)
	apply_border_to_merged_cells(ws, ws.max_row, 1, ws.max_row, 2, thin_border)

	# Adjust column widths
	if column_widths:
		for col_idx, width in enumerate(column_widths, 1):
			ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

	# Save to BytesIO
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	xlsx_file.seek(0)

	return xlsx_file

def get_department_hierarchy():
	departments = frappe.db.get_all(
		"Department",
		filters={"disabled": 0, "name": ["!=", "All Departments"]},
		fields=["name", "parent_department"],
		order_by="parent_department ASC, name ASC"
	)
	department_hierarchy = {}
	for department in departments:
		parent = department["parent_department"]
		child = department["name"]

		if parent not in department_hierarchy:
			department_hierarchy[parent] = []
		department_hierarchy[parent].append(child)

	return department_hierarchy

from datetime import datetime, date
@frappe.whitelist()
def allocate_el():
	emp=frappe.db.get_all("Employee",{'status':'Active','employee_category':('in',('Staff','Sub Staff','Operator','Apprentice','Trainee'))},['name','date_of_joining','employee_category'])
	current_date = datetime.now().date()
	pcount=0
	if current_date.day == 1 and current_date.month == 1:  
		next_year = current_date.year
		first_date_next_year = datetime(next_year, 1, 1).date()
		last_date_next_year = datetime(next_year, 12, 31).date()
		current_year = current_date.year-1
	else:
		next_year = current_date.year+1
		first_date_next_year = datetime(next_year, 1, 1).date()
		last_date_next_year = datetime(next_year, 12, 31).date()
		current_year = current_date.year
	# next_year = current_date.year
	# first_date_next_year = datetime(next_year, 1, 1).date()
	# last_date_next_year = datetime(next_year, 12, 31).date()
	# current_year = current_date.year-1
	start_date=datetime(current_year, 1, 1).date()
	yesterday=add_days(current_date,-1)
	# print(emp)
	for e in emp:
		print(e.name)
		count=0
		current_date = datetime.now().date()
		doj = e.date_of_joining
		diff = current_date - doj
		years = diff.days / 365.25 
		if e.employee_category in ['Sub Staff','Staff']:
			act=20
			start_date=datetime(current_year, 1, 1).date()
			print(start_date)
			dates=get_dates(start_date,yesterday)
			for date in dates:
				hh=check_holiday(date,e.name)
				if not hh:
					if frappe.db.exists("Attendance",{'docstatus':['!=',2],'status': 'Present','employee': e.name,"attendance_date":date}):
						count+=1
			print(count)
			if count>0:
				leave=count/act
				if leave>=1:
					if not frappe.db.exists('Leave Allocation',{'leave_type':'Earned Leave','from_date':('between',(first_date_next_year,last_date_next_year)),'to_date':('between',(first_date_next_year,last_date_next_year)),'docstatus':['!=',2],'employee':e.name}):
						allow = frappe.new_doc("Leave Allocation")
						allow.employee = e.name
						allow.company = e.company
						allow.leave_type = "Earned Leave"
						allow.from_date = first_date_next_year
						allow.to_date = last_date_next_year
						allow.new_leaves_allocated = int(leave)
						allow.total_leaves_allocated = int(leave)
						allow.insert()
						allow.save(ignore_permissions=True)
						allow.submit()	
						frappe.db.commit
					else:
						allow=frappe.db.get_value('Leave Allocation',{'leave_type':'Earned Leave','from_date':first_date_next_year,'to_date':last_date_next_year,'docstatus':['!=',2],'employee':e.name},['name'])
						alloc = frappe.get_doc("Leave Allocation",allow)
						alloc.new_leaves_allocated = int(leave)
						alloc.total_leaves_allocated = int(leave)
						alloc.save(ignore_permissions=True)
						frappe.db.commit
		else:
			if e.employee_category == "Operator":
				act=20
			else:
				act=30
		
			if int(years)>0:
				if years < 2 and doj.year==current_year-1:
					doj_date=e.date_of_joining.day
					doj_month=e.date_of_joining.month
					doj=datetime(current_year, doj_month, doj_date).date()
					start_date=doj
				else:
					start_date=datetime(current_year, 1, 1).date()
				print(start_date)
				dates=get_dates(start_date,yesterday)
				for date in dates:
					hh=check_holiday(date,e.name)
					if not hh:
						if frappe.db.exists("Attendance",{'docstatus':['!=',2],'status': 'Present','employee': e.name,"attendance_date":date}):
							count+=1
				print(count)
				if count>0:
					leave=count/act
					if leave>=1:
						if not frappe.db.exists('Leave Allocation',{'leave_type':'Earned Leave','from_date':('between',(first_date_next_year,last_date_next_year)),'to_date':('between',(first_date_next_year,last_date_next_year)),'docstatus':['!=',2],'employee':e.name}):
							allow = frappe.new_doc("Leave Allocation")
							allow.employee = e.name
							allow.company = e.company
							allow.leave_type = "Earned Leave"
							allow.from_date = first_date_next_year
							allow.to_date = last_date_next_year
							allow.new_leaves_allocated = int(leave)
							allow.total_leaves_allocated = int(leave)
							allow.insert()
							allow.save(ignore_permissions=True)
							allow.submit()	
							frappe.db.commit
						else:
							allow=frappe.db.get_value('Leave Allocation',{'leave_type':'Earned Leave','from_date':first_date_next_year,'to_date':last_date_next_year,'docstatus':['!=',2],'employee':e.name},['name'])
							alloc = frappe.get_doc("Leave Allocation",allow)
							alloc.new_leaves_allocated = int(leave)
							alloc.total_leaves_allocated = int(leave)
							alloc.save(ignore_permissions=True)
							frappe.db.commit

@frappe.whitelist()
def enqueue_el_allocation():
	enqueue(allocate_el, queue='long', timeout=6000)

import requests
import json
import frappe
from frappe import _

@frappe.whitelist()
def update_issue_status_from_teampro():
	issues = frappe.get_all(
		"Issue",
		filters={
			"custom_mail_sent_to_teampro": 1,
			"status": ["not in", ["Resolved", "Closed"]],
		},
		fields=["subject", "priority",'name']
	)

	for i in issues:
		issue_doc=frappe.get_doc("Issue",i.name)
		url = "https://erp.teamproit.com/api/resource/Issue"
		headers = {
			"Content-Type": "application/json",
			"Authorization": "token daa4a43f429c844:3b0d3fbc3c5e4ce"
		}
		params = {
			"filters": json.dumps([
				["subject", "=", i.subject],
				["raised_by", "=", "wonjin_corporate@onegeneindia.in"]
			]),
			"fields": json.dumps(["name","task","custom_issue_status"]),
			"limit_page_length": 1000
		}

		try:
			response = requests.get(url, headers=headers, params=params, verify=False)
			res = response.json()
			for issue_data in res.get("data", []):
				issue_doc.custom_issue_id = issue_data.get("name")
				issue_doc.custom_issue_status = issue_data.get("custom_issue_status")
				task_name = issue_data.get("task")
				if not task_name:
					continue
				task_url = "https://erp.teamproit.com/api/resource/Task"
				task_params = {
					"filters": json.dumps([
						["name", "=", task_name]
					]),
					"fields": json.dumps(["creation", "custom_allocated_to", "status", "description","exp_end_date"]),
				}

				try:
					task_response = requests.get(task_url, headers=headers, params=task_params, verify=False)
					task_data = task_response.json()
					for task in task_data.get("data", []):
						print(task_response.text)
						issue_doc.custom_task_id = task_name
						issue_doc.custom_task_status = task.get("status")
						creation_datetime = task.get("creation")  # Example: '2025-04-09 10:38:26.403313'
						if creation_datetime:
							creation_date = datetime.strptime(creation_datetime, "%Y-%m-%d %H:%M:%S.%f").date()
							issue_doc.custom_task_creation_date = creation_date.strftime("%d-%m-%Y")
						issue_doc.custom_target_date = task.get("exp_end_date")
						issue_doc.custom_task_allocated_to = task.get("custom_allocated_to")
						issue_doc.custom_task_description = task.get("description")
				except Exception:
					frappe.log_error(frappe.get_traceback(), "Teampro Task API Call Failed")
			issue_doc.save(ignore_permissions=True)
		except Exception as e:
			frappe.log_error(frappe.get_traceback(), "Teampro API Call Failed")

@frappe.whitelist()
def create_job_fail1():
	job = frappe.db.exists('Scheduled Job Type', 'create_lwf')
	if not job:
		emc = frappe.new_doc('Scheduled Job Type')
		emc.update({
			"method": 'onegene.onegene.custom.create_lwf',
			"frequency": 'Cron',
			"cron_format": '0 0 1 * *'
		})
		emc.save(ignore_permissions=True)

@frappe.whitelist()
def issue_closing_mail(name,subject, message, recipients, sender):
	"""
	Send an email to the specified recipients with the given subject and message.
	"""
	if not frappe.db.exists("Email Account", {"email_id": sender}):
		frappe.throw(_("Sender email not configured in Email Account"))

	# Add additional content to the message body
	message_body = f"""
		<div style="font-family: 'Times New Roman', Times, serif; font-size: 14px;">
			<p>Dear Employee,</p>

			<p>A Ticket : {name} with the below mentioned Subject and Description has been resolved</p>

			<p><strong>Subject:</strong> {subject}</p>
			<p><strong>Description:</strong>{message}</p>


			<p>Thanks & Regards,<br>
			Wonjin Team</p>
		</div>
	"""

	try:
		frappe.sendmail(
			recipients=recipients,
			sender=sender,
			subject=subject,
			message=message_body,
			now=True
		)
		return "Email sent successfully"
	except Exception as e:
		frappe.log_error(frappe.get_traceback(), 'Email Sending Failed')
		return "Failed to send email"
	
@frappe.whitelist()
def create_purchase_open_order(doc, method):
	if doc.custom_order_type == "Open":
		new_doc = frappe.new_doc('Purchase Open Order')
		new_doc.purchase_order = doc.name
		new_doc.set('open_order_table', [])
		for po in doc.items:
			new_doc.append("open_order_table", {
				"item_code": po.item_code,
				"delivery_date": po.schedule_date,
				"item_name": po.item_name,
				"qty": po.qty,
				"rate": po.rate,
				"warehouse": po.warehouse,
				"amount": po.amount,
				"docname": po.name,
			})
		new_doc.save(ignore_permissions=True)

@frappe.whitelist()
def create_purchase_order_schedule_from_po(doc,method):
	if doc.custom_order_type == "Fixed" and not doc.custom_schedule_table:
		frappe.throw("Schedule not Created")
		frappe.throw("Schedule உருவாக்கப்படவில்லை")
	if doc.custom_order_type == "Fixed" and doc.custom_schedule_table:
		for schedule in doc.custom_schedule_table:
			if frappe.db.exists('Purchase Order Schedule', {'purchase_order_number': doc.name, 'item_code': schedule.item_code, 'schedule_date': schedule.schedule_date}):
				frappe.throw("Schedule already exists for this item code and schedule date")
				frappe.throw("இந்த Item Code மற்றும் Schedule Date க்கான schedule ஏற்கனவே உள்ளது")
			po_type = frappe.db.get_value("Purchase Order",{"name":doc.name},"custom_po_type")
			new_doc = frappe.new_doc('Purchase Order Schedule')
			new_doc.po_type = po_type
			new_doc.supplier_code = doc.supplier_code
			new_doc.supplier_name = doc.supplier
			new_doc.purchase_order_number = doc.name
			new_doc.item_code = schedule.item_code
			new_doc.schedule_date = schedule.schedule_date
			new_doc.qty = schedule.schedule_qty
			for item in doc.items:
				if item.item_code == schedule.item_code:
					new_doc.child_name = schedule.name
					new_doc.schedule_amount = schedule.schedule_qty * item.rate
					new_doc.order_rate = item.rate
					new_doc.pending_qty = schedule.schedule_qty
					new_doc.pending_amount = schedule.schedule_qty * item.rate
			new_doc.save(ignore_permissions=True) 
			new_doc.submit()

	if doc.custom_order_type == "Open" and doc.custom_schedule_table:
		for schedule in doc.custom_schedule_table:
			date_obj = datetime.strptime(schedule.schedule_date, '%Y-%m-%d')
			month = date_obj.strftime('%b') 
			new_doc = frappe.new_doc('Purchase Order Schedule') 
			new_doc.supplier_code = doc.supplier_code
			new_doc.supplier_name = doc.supplier
			po_type = frappe.db.get_value("Purchase Order",{"name":doc.name},"custom_po_type")
			new_doc.po_type = po_type
			new_doc.purchase_order_number = doc.name
			new_doc.item_code = schedule.item_code
			new_doc.schedule_date = schedule.schedule_date
			new_doc.schedule_month = month.upper()
			new_doc.qty = schedule.schedule_qty
			for item in doc.items:
				if item.item_code == schedule.item_code:
					new_doc.child_name = schedule.name
					new_doc.schedule_amount = schedule.schedule_qty * item.rate
					new_doc.order_rate = item.rate
					new_doc.pending_qty = schedule.schedule_qty
					new_doc.pending_amount = schedule.schedule_qty * item.rate
			new_doc.save(ignore_permissions=True) 
			new_doc.submit()

@frappe.whitelist()
def	reload_po(doc,method):
	doc.reload()

@frappe.whitelist()
def delete_purchase_order_schedule(doc, method):
	if doc.order_schedule:
		if frappe.db.exists('Purchase Order Schedule', doc.order_schedule):
			schedule = frappe.get_doc('Purchase Order Schedule', doc.order_schedule)
			if schedule.order_type == "Open":
				schedule.cancel()
				doc.save(ignore_permissions=True)

@frappe.whitelist()
def create_sales_open_order(doc,method):
	if doc.customer_order_type == "Open":
		new_doc = frappe.new_doc('Open Order')
		new_doc.sales_order_number = doc.name
		new_doc.set('open_order_table', [])
		for so in doc.items:
			new_doc.append("open_order_table", {
				"item_code": so.item_code,
				"delivery_date": so.delivery_date,
				"item_name": so.item_name,
				"qty": so.qty,
				"rate": so.rate,
				"warehouse": so.warehouse,
				"amount": so.amount,
				"docname": so.name,
			})
		new_doc.save(ignore_permissions=True)

@frappe.whitelist()
def get_previous_purchase_rate(item_code):
	pr = frappe.db.sql("""
		SELECT 
			lci.rate
		FROM 
			`tabPurchase Order Item` lci
		LEFT JOIN 
			`tabPurchase Order` lcv
		ON 
			lci.parent = lcv.name
		JOIN 
			(SELECT 
				item_code, MAX(creation) as latest_creation
			FROM 
				`tabPurchase Order Item`
			WHERE 
				rate IS NOT NULL
			GROUP BY 
				item_code) latest_lci
		ON 
			lci.item_code= latest_lci.item_code
			AND lci.creation = latest_lci.latest_creation
		WHERE 
			lci.item_code = %s and lcv.docstatus !=2
	""", item_code, as_dict=True)

	if pr:
		return pr[0].get('rate')
	else:
		return '0.0'

@frappe.whitelist()
def set_lead_time(item,supplier):
	item_code = frappe.get_doc("Item",item)
	for i in item_code.supplier_items:
		if supplier == i.supplier:
			return i.custom_lead_time_in_days


@frappe.whitelist()
def create_order_schedule_from_po_for_open(item_code, schedule_date, schedule_qty, supplier_code, name, rate, month):
	rate = float(rate)
	schedule_qty = float(schedule_qty)
	new_doc = frappe.new_doc('Purchase Order Schedule') 
	new_doc.supplier_name = frappe.db.get_value("Supplier",{"supplier_code": supplier_code},"name")
	new_doc.supplier_code = supplier_code
	po_type = frappe.db.get_value("Purchase Order",{"name":name},"custom_po_type")
	new_doc.po_type = po_type
	new_doc.order_type = "Open"
	new_doc.purchase_order_number = name
	new_doc.item_code = item_code
	new_doc.schedule_date = schedule_date
	new_doc.qty = schedule_qty
	new_doc.schedule_amount = schedule_qty * rate
	new_doc.order_rate = rate
	new_doc.schedule_month = month.upper()
	new_doc.pending_qty = schedule_qty
	new_doc.pending_amount = schedule_qty * rate
	new_doc.save(ignore_permissions=True) 
	new_doc.submit()
	return("ok")

@frappe.whitelist()
def update_address_in_supplier():
	address_data = frappe.db.get_all('Address',{'address_type':"Billing"},['name'])
	count =0
	for s in address_data:
		address = frappe.get_value("Dynamic Link", {'parent': s.get('name'), 'link_doctype': 'Supplier'}, 'link_name')
		# print(address)
		if address:
			doc = frappe.get_doc('Address',s.get('name'))
			address_display = get_address_display(doc.as_dict())
			# frappe.db.set_value('Customer',{'customer_primary_address':doc.address_title,'customer_primary_address':['is','not set']},'customer_primary_address',doc.name)
			# frappe.db.set_value('Customer',{'customer_primary_address':doc.address_title,'customer_primary_address':['is','not set']},'primary_address',address_display)
			frappe.db.set_value('Supplier',{'name':doc.address_title,'supplier_primary_address':['is','not set']},'supplier_primary_address',doc.name)
			frappe.db.set_value('Supplier',{'name':doc.address_title,'supplier_primary_address':['is','not set']},'primary_address',address_display)        
			# break
	# doc =frappe.get_doc('Address','M.S.N. ENGINEERING ENTERPRISES-Billing')
	# address_display = get_address_display(doc.as_dict())
	# frappe.db.set_value('Supplier','M.S.N. ENGINEERING ENTERPRISES','supplier_primary_address','M.S.N. ENGINEERING ENTERPRISES-Billing')
	# frappe.db.set_value('Supplier','M.S.N. ENGINEERING ENTERPRISES','primary_address',address_display)

@frappe.whitelist()
def get_last_date_of_fiscal():
	last_date = frappe.db.get_value("Fiscal Year", "2025-2026", "year_end_date")
	return last_date

@frappe.whitelist()
def return_conversion(currency,price_list_currency):
	conv_rate = get_exchange_rate(currency, price_list_currency)
	return conv_rate

@frappe.whitelist()
def set_naming_series_po(name, new_name):
	if name != new_name:
		frappe.rename_doc("Purchase Order", name, new_name, force=True)
		frappe.db.commit()
		return new_name

import math
@frappe.whitelist()
def set_box_weight(item,box,item_qty):
	box_qty = frappe.db.get_value("Box Table",{'parent':item,'box':box},['qty'])
	no_of_boxes = 0
	if box_qty:
		no_of_boxes = math.ceil(float(item_qty) / float(box_qty))
	box_weight = frappe.db.get_value("Box",{'name':box},['weight'])
	total_weight_of_boxes = no_of_boxes * box_weight
	total_box_length = frappe.db.get_value("Box",{'name':box},['length']) * no_of_boxes
	total_box_breadth = frappe.db.get_value("Box",{'name':box},['breadth']) * no_of_boxes
	total_box_height = frappe.db.get_value("Box",{'name':box},['height']) * no_of_boxes
	return no_of_boxes, total_weight_of_boxes, total_box_length, total_box_breadth, total_box_height

@frappe.whitelist()
def set_pallet_weight(item,box,pallet,box_qty):
	pallet_qty = frappe.db.get_value("Pallet Table",{'parent':item,'box':box,'pallet':pallet},['qty'])
	no_of_pallets = 0
	if pallet_qty:
		no_of_pallets = math.ceil(float(box_qty) / float(pallet_qty))
	pallet_weight = frappe.db.get_value("Pallet",{'name':pallet},['weight'])
	total_weight_of_pallets = no_of_pallets * pallet_weight
	total_pallet_length = frappe.db.get_value("Pallet",{'name':pallet},['length']) 
	total_pallet_breadth = frappe.db.get_value("Pallet",{'name':pallet},['breadth']) 
	total_pallet_height = frappe.db.get_value("Pallet",{'name':pallet},['height']) 
	return no_of_pallets, total_weight_of_pallets, total_pallet_length, total_pallet_breadth, total_pallet_height

@frappe.whitelist()
def calculate_total_cbm(doc, method):
	total_cbm = 0
	for row in doc.items:
		if row.custom_cbm:
			total_cbm += row.custom_cbm
	doc.custom_total_cbm = total_cbm

@frappe.whitelist()
def validate_on_trash(doc, method):
	if "System Manager" not in frappe.get_roles(frappe.session.user):
		if doc.workflow_state and doc.workflow_state not in ["Draft", "Cancelled"]:
			if doc.docstatus == 0:
				frappe.throw(
					"Cannot delete this document as the workflow has moved to the next level.",
					title="Not Permitted"
				)
@frappe.whitelist()
def get_items_for_sii(doctype, txt, searchfield, start, page_len, filters):
	if isinstance(filters, str):
		filters = json.loads(filters)

	customer = filters.get("customer")
	invoice_type = filters.get("invoice_type")
	if not customer:
		return []

	search_text = f"%{txt.strip()}%" if txt else "%"
	start = int(start)
	page_len = int(page_len)
	items = list(frappe.db.sql("""
		SELECT DISTINCT soi.item_code, soi.item_name 
		FROM `tabSales Order Item` soi
		LEFT JOIN `tabSales Order` so ON soi.parent = so.name
		WHERE so.customer = %s
		  AND so.docstatus = 1
		  AND soi.custom_disabled = 0
		  AND (soi.item_code LIKE %s OR soi.item_name LIKE %s)
		  AND (so.status IN ('To Deliver', 'To Deliver and Bill', 'Overdue') OR so.status IS NULL)
	""", (customer, search_text, search_text)))

	if invoice_type == "Credit Note":
		items.append(("Cash Discount", "Cash Discount"))

	return items

@frappe.whitelist()
def get_so_for_sii(doctype, txt, searchfield, start, page_len, filters):
	if isinstance(filters, str):
		filters = json.loads(filters)

	item_code = filters.get("item_code")
	customer = filters.get("customer")
	if not item_code and not customer:
		return []

	start = int(start)
	page_len = int(page_len)

	search_text = f"%{txt.strip()}%" if txt else "%"

	return frappe.db.sql("""
		SELECT DISTINCT so.name, so.customer_name, so.status
		FROM `tabSales Order Item` soi
		LEFT JOIN `tabSales Order` so ON soi.parent = so.name
		WHERE soi.item_code = %s
		  AND so.customer = %s
		  AND so.docstatus = 1
		  AND soi.custom_disabled = 0
		  ANd so.name LIKE %s	
		  AND (so.status IN ('To Deliver', 'To Deliver and Bill') OR so.status IS NULL)
	""", (item_code, customer, search_text))

@frappe.whitelist()
def get_so_item_details(sales_order, item_code):
	
	qty_to_be_invoiced, so_detail,rate = get_so_qty(sales_order, item_code)
	return qty_to_be_invoiced, so_detail,rate


@frappe.whitelist()
def validate_qty_in_sales_invoice(item_code, sales_order, qty, idx, delivery_note=None):
	if delivery_note:
		delivered_qty = frappe.db.get_value("Delivery Note Item", {"parent": delivery_note, "item_code": item_code}, "qty") or 0
		if flt(qty) > delivered_qty:
			return delivered_qty
	else:	
		qty_to_be_invoiced, so_detail, rate = get_so_qty(sales_order, item_code)
		if flt(qty) > flt(qty_to_be_invoiced):
			return qty_to_be_invoiced

@frappe.whitelist()
def mark_stock_entry_as_cancelled(doc, method):
	"""Mark the Stock Entry workflow as cancelled when it is cancelled."""
	frappe.db.set_value("Stock Entry", doc.name, "workflow_state", "Cancelled")
	 
@frappe.whitelist()
def get_so_qty(sales_order, item_code):
	data = frappe.db.sql("""
		SELECT 
			IF(billed_amt != 0 AND rate != 0, 
				(amount - billed_amt) / rate, 
				qty - returned_qty
			) AS qty, 
			name, rate
		FROM `tabSales Order Item`
		WHERE parent = %s AND item_code = %s
	""", (sales_order, item_code), as_dict=True)
	return data[0].qty if data and data[0].qty else 0, data[0].name if data and data[0].name else None, data[0].rate if data and data[0].rate else 0

@frappe.whitelist()
def validate_sales_invoice(doc, method):
	# Validate Sales Order Item quantities against Sales Order
	for row in doc.items:
		if row.qty > 0:
			qty_to_be_invoiced, so_detail,rate = get_so_qty(row.sales_order, row.item_code)
			if row.qty > qty_to_be_invoiced:
				frappe.throw(
					_("<b>Row {0}</b>: Cannot invoice more than the quantity available in Sales Order {1} for Item {2}. Available Quantity: {3}").format(
						row.idx, bold(row.sales_order), bold(row.item_code), bold(qty_to_be_invoiced)
					),
					title=_("Quantity Exceeds Sales Order Limit")
				)
				frappe.throw(
					_("<b>Row {0}</b>: Sales Order {1} இல் உள்ள அளவை விட அதிகமாக Item {2} க்கு invoice செய்ய முடியாது. கிடைக்கும் அளவு: {3}")
					.format(row.idx, bold(row.sales_order), bold(row.item_code), bold(qty_to_be_invoiced)),
					title=_("Quantity Exceeds Sales Order Limit")
				)

	# Update SO in SI
	seen_so = set()
	doc.set('custom_sales_orders', [])
	for row in doc.items:
		so = row.sales_order
		if so and so not in seen_so:
			doc.append("custom_sales_orders", {
				"sales_order": so
			})
			seen_so.add(so)

@frappe.whitelist()
def qty_check(sales,item,customer,qty):
	if sales:
		order_type = frappe.db.get_value("Sales Order",sales,'customer_order_type')
		if order_type == "Fixed":
			s_qty = 0
			so_exist = frappe.db.exists('Sales Order Schedule',{"sales_order_number":sales,"customer_code":customer,"item_code":item})
			if so_exist:
				exist_so = frappe.get_all("Sales Order Schedule",{"sales_order_number":sales,"customer_code":customer,"item_code":item},["*"])
				for i in exist_so:
					old_qty = i.qty
					s_qty += old_qty
				if order_type == "Fixed":
					sales_order = frappe.get_all("Sales Order Item",{"parent": sales, "item_code": item},["qty","item_code"])
					if sales_order and len(sales_order) > 0:
						total_os_qty = float(s_qty)+float(qty)
						sales_order_qty = sales_order[0].get("qty")
						item_cde = sales_order[0].get("item_code")
						idx = sales_order[0].get("idx")
						if sales_order_qty < total_os_qty:
							frappe.throw(f"Validation failed: Quantity <b>{total_os_qty}</b> exceeds Sales Order quantity <b>{sales_order_qty}</b> in Line Item <b>{item_cde}</b>.")
							frappe.throw(f"Validation failed: Line Item <b>{item_cde}</b> க்கான, Quantity <b>{total_os_qty}</b> Sales Order quantity <b>{sales_order_qty}</b> ஐ விட அதிகமாக உள்ளது.")

					else:
						frappe.throw(f"Item <b>{item}</b> not found in the Sales Order - <b>{sales}</b>")
						frappe.throw(f"Item <b>{item}</b> Sales Order - <b>{sales}</b> இல் கண்டுபிடிக்கப்படவில்லை")


@frappe.whitelist()
def get_customer_po_number(customer, item_code):
	if not customer or not item_code:
		return []

	po_nos = frappe.db.sql("""
		SELECT DISTINCT so.po_no
		FROM `tabSales Order` so
		JOIN `tabSales Order Item` soi ON soi.parent = so.name
		WHERE so.customer = %s
		  AND soi.item_code = %s
		  AND so.docstatus = 1
		  AND so.po_no IS NOT NULL
		  AND (so.status IN ('To Deliver', 'To Deliver and Bill', 'Overdue') OR so.status IS NULL)
	""", (customer, item_code), as_list=1)

	return [f"{row[0]}" for row in po_nos]

@frappe.whitelist()
def get_customer_po_query(doctype, txt, searchfield, start, page_len, filters):
	customer = filters.get("customer")
	item_code = filters.get("item_code")

	if not customer or not item_code:
		return []

	return frappe.db.sql("""
		SELECT DISTINCT so.po_no
		FROM `tabSales Order` so
		JOIN `tabSales Order Item` soi ON soi.parent = so.name
		WHERE so.customer = %s
		  AND soi.item_code = %s
		  AND so.docstatus = 1
		  AND soi.custom_disabled = 0
		  AND (so.status IN ('To Deliver', 'To Deliver and Bill', 'Overdue') OR so.status IS NULL)
		  AND so.po_no IS NOT NULL
		  AND so.po_no LIKE %s
		LIMIT %s OFFSET %s
	""", (customer, item_code, f"%{txt}%", page_len, start))


@frappe.whitelist()
def create_cutomer_po(doc, method):
	if frappe.db.exists("Customer Purchase Order", doc.po_no):
		return
	else:
		cpo = frappe.new_doc("Customer Purchase Order")
		cpo.customers_po = doc.po_no
		cpo.insert(ignore_permissions=True)




def customer_address_validate(doc, method):
	if any(link.link_doctype == "Customer" for link in doc.links):
		# if doc.country != "India":
			
		#     doc.state = "Other Country"
		frappe.db.set_value("Customer",doc.address_title,"custom_country",doc.country)
		frappe.db.set_value("Customer",doc.address_title,"default_currency",doc.custom_currency)

@frappe.whitelist()
def update_pr_in_sdn(doc, method):
	frappe.db.sql("""
		UPDATE `tabAdvance Shipping Note`
		SET purchase_receipt = %s
		WHERE name = %s
	""", (doc.name, doc.supplier_delivery_note))

@frappe.whitelist()
def remove_pr_in_sdn(doc, method):
	frappe.db.sql("""
		UPDATE `tabAdvance Shipping Note`
		SET purchase_receipt = ''
		WHERE name = %s
	""", (doc.supplier_delivery_note))

@frappe.whitelist()
def update_received_stock_in_order_schedule(doc, method):
	if doc.supplier_delivery_note:
		frappe.db.sql("""
			UPDATE `tabAdvance Shipping Note`
			SET workflow_state = "Material Received"
			WHERE name = %s
		""", (doc.supplier_delivery_note))
		sdn = frappe.get_doc("Advance Shipping Note", doc.supplier_delivery_note)
		for sdn_row in sdn.item_table:
			for row in doc.items:
				if row.item_code == sdn_row.item_code:
					sdn_row.received_qty = (row.qty)
					sdn_row.pending_after_received = flt(sdn_row.pend_qty) - flt(row.qty)
					sdn_row.remarks = row.custom_sr_qty
		sdn.save(ignore_permissions=True)
		from onegene.onegene.doctype.advance_shipping_note.advance_shipping_note import on_submit_action
		on_submit_action(doc.supplier_delivery_note)

@frappe.whitelist()
def revert_received_stock_in_order_schedule(doc, method):
	frappe.db.sql("""
		UPDATE `tabAdvance Shipping Note`
		SET workflow_state = "Gate Received"
		WHERE name = %s
	""", (doc.supplier_delivery_note))
	# sdn = frappe.get_doc("Advance Shipping Note", doc.supplier_delivery_note)
	# for sdn_row in sdn.item_table:
	# 	for row in doc.items:
	# 		if row.item_code == sdn_row.item_code:
	# 			sdn_row.received_qty = 0
	# 			sdn_row.pending_after_received = 0
	# 			sdn_row.remarks = 0
	# sdn.save(ignore_permissions=True)
	from onegene.onegene.doctype.advance_shipping_note.advance_shipping_note import on_cancel_action
	on_cancel_action(doc.supplier_delivery_note)

def rename_docname_field_in_po(doc, method, old_name, new_name, merge=False):
	doc.custom_docname = new_name
	frappe.db.set_value("Purchase Order", doc.name, "custom_docname", new_name)


def update_machine_age():
	today = date.today()
	workstations = frappe.get_all("Workstation", fields=["name", "custom_machine_purchase_date"])
	
	for ws in workstations:
		if ws.custom_machine_purchase_date:
			purchase_date = ws.custom_machine_purchase_date
			rd = relativedelta(today, purchase_date)

			custom_age = f"{rd.years} Years {rd.months} Months {rd.days} Days"

			frappe.db.set_value("Workstation", ws.name, "custom_age", custom_age)
			
			
def update_machine_age_scheduler():
	job = frappe.db.exists('Scheduled Job Type', 'update_machine_age')
	if not job:
		sjt = frappe.new_doc("Scheduled Job Type")
	sjt.update({
		"method": 'onegene.onegene.custom.update_machine_age',
		"frequency": 'Daily'
		
	})
	sjt.save(ignore_permissions=True)
	
@frappe.whitelist()
def update_ot_hrs(doc, method):
	shifts = frappe.db.get_all('Shift Type', filters={'custom_disabled': 0}, fields=['name'])

	existing_shifts = [d.shift for d in doc.custom_ot_details]

	for s in shifts:
		if s.name not in existing_shifts:
			row = doc.append('custom_ot_details', {})
			row.shift = s.name
			row.allowed_ot = 0

	


def sales_order_item_on_update(doc, method):
	

	if not doc.item_code:
		frappe.throw("Item Code cannot be empty in row.")

	if not doc.parent:
		return

	parent_doc = frappe.get_doc("Sales Order", doc.parent)

   
	for row in parent_doc.items:
		if row.name != doc.name and row.item_code == doc.item_code and row.gst_hsn_code == doc.gst_hsn_code:
			frappe.db.sql("DELETE FROM `tabSales Order Item` WHERE name=%s", doc.name)
			frappe.msgprint(f"Removed new duplicate row for Item {doc.item_code} with HSN {doc.gst_hsn_code}.")
			return  

	
	for row in parent_doc.items:
		if row.name != doc.name and row.gst_hsn_code and doc.gst_hsn_code:
			if row.gst_hsn_code != doc.gst_hsn_code:
				frappe.db.sql("DELETE FROM `tabSales Order Item` WHERE name=%s", doc.name)
				frappe.msgprint(
					f"Removed new row of Item {doc.item_code} due to HSN mismatch ({doc.gst_hsn_code} vs {row.gst_hsn_code})."
				)
				return
				

@frappe.whitelist()
def create_work_order_from_production_plan(doc, method):
	"""On submission of the Production Plan, the Work Order will be
	created and submitted. The Work Order will submit after creation,
	if it is linked with the Production Plan"""

	doc.make_work_order()
	work_orders = frappe.db.get_all("Work Order", {"production_plan": doc.name, "docstatus": 0}, "name")
	for work_order in work_orders:
		wo = frappe.get_doc("Work Order", work_order.name)
		if wo.production_plan:
			wo.submit()


@frappe.whitelist()
def get_assembly_items_update(item_code,rej_allowance,month,qty=None):
	
	
	
 
	year = datetime.datetime.strptime(today(), "%Y-%m-%d").year
	month = datetime.datetime.strptime(today(), "%Y-%m-%d").month
	
	start_date = date(year, month, 1)          
	to_date = today()                     
	
	
	pos = frappe.db.sql("""select `tabDelivery Note Item`.item_code as item_code,`tabDelivery Note Item`.qty as qty from `tabDelivery Note`
	left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
	where `tabDelivery Note Item`.item_code = '%s' and `tabDelivery Note`.docstatus = 1 and `tabDelivery Note`.posting_date between '%s' and '%s'"""%(item_code,start_date,to_date),as_dict = 1)
	
	del_qty = 0
	if len(pos)>0:
		for l in pos:
			del_qty += l.qty

	si_qty_ = frappe.db.sql("""select `tabSales Invoice Item`.item_code as item_code,`tabSales Invoice Item`.qty as qty from `tabSales Invoice`
	left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent
	where `tabSales Invoice`.update_stock = 1 and `tabSales Invoice Item`.item_code = '%s' and `tabSales Invoice`.docstatus = 1 and `tabSales Invoice`.posting_date between '%s' and '%s'"""%(item_code, start_date,to_date),as_dict = 1)
	if len(si_qty_)>0:
		for l in si_qty_:
			del_qty += l.qty
	
	if qty:
		qty = float(qty)
		rej_allowance = float(rej_allowance) /100
	
		work_days = frappe.db.get_single_value("Production Plan Settings", "working_days")
		with_rej = (qty * rej_allowance) + qty
		per_day = with_rej / int(work_days)
		
		pack_size = frappe.get_value("Item",item_code,['pack_size'])
		
		if pack_size > 0:
			cal = per_day/ pack_size

			total = ceil(cal) * pack_size
		else:
			total = ceil(per_day)
		
	else:
		total = 0
  
	
	return {
		"per_day": total,
		"del_qty": del_qty
	}

@frappe.whitelist()
def get_sub_assembly_items_update(item_code,qty,rej_allowance):
	
	qty = float(qty)
	rej_allowance = float(rej_allowance) /100
	
	

	
	work_days = frappe.db.get_single_value("Production Plan Settings", "working_days")
	with_rej = (qty * rej_allowance) + qty
	per_day = with_rej / int(work_days)
	
	
	pack_size = frappe.get_value("Item",item_code,['pack_size'])
		
	if pack_size > 0:
		cal = per_day/ pack_size

		total = ceil(cal) * pack_size
	else:
		total = ceil(per_day)

	
	
	
	year = datetime.datetime.strptime(today(), "%Y-%m-%d").year
	month = datetime.datetime.strptime(today(), "%Y-%m-%d").month
	
	start_date = date(year, month, 1)          
	to_date = today()                     
	
	
	pos = frappe.db.sql("""select `tabDelivery Note Item`.item_code as item_code,`tabDelivery Note Item`.qty as qty from `tabDelivery Note`
	left join `tabDelivery Note Item` on `tabDelivery Note`.name = `tabDelivery Note Item`.parent
	where `tabDelivery Note Item`.item_code = '%s' and `tabDelivery Note`.docstatus = 1 and `tabDelivery Note`.posting_date between '%s' and '%s'"""%(item_code,start_date,to_date),as_dict = 1)
	del_qty = 0
	if len(pos)>0:
		for l in pos:
			del_qty += l.qty

	si_qty_ = frappe.db.sql("""select `tabSales Invoice Item`.item_code as item_code,`tabSales Invoice Item`.qty as qty from `tabSales Invoice`
	left join `tabSales Invoice Item` on `tabSales Invoice`.name = `tabSales Invoice Item`.parent
	where `tabSales Invoice`.update_stock = 1 and `tabSales Invoice Item`.item_code = '%s' and `tabSales Invoice`.docstatus = 1 and `tabSales Invoice`.posting_date between '%s' and '%s'"""%(item_code, start_date,to_date),as_dict = 1)
	if len(si_qty_)>0:
		for l in si_qty_:
			del_qty += l.qty
	
	
	
	
	return {
	"per_day": total,
	"del_qty": del_qty,
	
	}


@frappe.whitelist()
def upload_kanban_qty(file, production_plan):
	file = get_file(file)
	
	content = file[1]
	if isinstance(content, bytes):
		content = content.decode('utf-8')

	csv_rows = csv.reader(content.splitlines())
	header_skipped = False

	for row in csv_rows:
		# Skip header row
		if not header_skipped:
			header_skipped = True
			continue

		item_code = row[0]
		kanban_qty = row[2]

		doc = frappe.get_doc("Production Plan", production_plan)
		for row in doc.po_items:
			if row.item_code == item_code:
				row.custom_production_kanban_qty = kanban_qty
		for row in doc.sub_assembly_items:
			planned_qty = frappe.db.get_value("Production Plan Item", {"parent": production_plan, "item_code": row.parent_item_code}, "planned_qty")
			if row.parent_item_code == item_code:
				row.custom_production_kanban_qty = flt(row.qty/planned_qty) * flt(kanban_qty)
		doc.save(ignore_permissions=True)
	return 'OK'

@frappe.whitelist(allow_guest=False)
def download_kqty_template():
	import json
	from frappe.utils.csvutils import UnicodeWriter
	from frappe.utils import cstr

	data = frappe.form_dict.get('data')
	if isinstance(data, str):
		data = json.loads(data)

	w = UnicodeWriter()
	w = add_header(w, data)

	frappe.response['result'] = cstr(w.getvalue())
	frappe.response['type'] = 'csv'
	frappe.response['doctype'] = "Production Kanban Qty"
	frappe.response['filename'] = "kanban_qty_template.csv"

def add_header(w,data):
	w.writerow(['Item Code','Item Name','Production Kanban Qty'])
	for i in data:
		item_name = frappe.db.get_value("Item", i.get("item_code"), "item_name")
		w.writerow([i.get("item_code"),item_name,''])
	return w




@frappe.whitelist()
def sales_invoice_custom_autoname(doc, method):
	prefix_map = {
		"Tax Invoice": "D",
		"Export Invoice": "E",
		"Supplementary Invoice": "SI",
		"Non Commercial Invoice": "NCI",
		"Tooling Invoice": "TI",
		"Scrap Invoice": "SCR",
		"Tooling Invoice - Exp": "TI",
	}
	prefix = prefix_map.get(doc.custom_invoice_type, "")

	year = now_datetime().strftime("%y")
	next_year = str(int(year) + 1).zfill(2)

	series = make_autoname(prefix + year + next_year + ".#####")

	doc.name = f"{prefix}{year}{next_year}{series[-5:]}"

@frappe.whitelist()
def update_transfer_material_against_and_operation(doc, method):
	"""Update the Transfer Material Against from Work Order to Job Card, 
	and Operations in the Raw Materials in Work Order document"""

	doc.transfer_material_against = "Job Card"
	doc.source_warehouse = "Shop Floor - WAIP"
	idx = 0
	operation = None
	for row in doc.operations:
		idx+=1
		if idx == 1:
			operation = row.operation
		if not row.workstation:
			ws = frappe.db.get_value("Operation",row.operation,'workstation')
			row.workstation = ws if ws and ws is not None else "Workstation - 1" # Temporary
		if row.time_in_mins == 0:
			row.time_in_mins = flt(1) # Temporary
	for row in doc.required_items:
		if frappe.db.exists("Operation Item List",{'document_name':doc.bom_no,'item':row.item_code}):
			row.operation = frappe.db.get_value("Operation Item List",{'document_name':doc.bom_no,'item':row.item_code},['operation'])
		else:
			row.operation = operation
		if not row.source_warehouse:
			row.source_warehouse = "Shop Floor - WAIP"

@frappe.whitelist()
def update_actual_fg_in_work_order(doc, method):
	"""Update actual FG and parent SFG from linked Production Plan items."""

	# Try from Production Plan Item first
	if doc.get("production_plan_item"):
		actual_fg, parent_sfg = frappe.db.get_value(
			"Production Plan Item", 
			doc.production_plan_item, 
			["custom_actual_fg", "item_code"]
		) or (None, None)

		if actual_fg:
			doc.custom_actual_fg = actual_fg
			doc.custom_parent_sfg = parent_sfg
			return

	# Try from Sub Assembly Item if first fails
	if doc.get("production_plan_sub_assembly_item"):
		actual_fg, parent_sfg = frappe.db.get_value(
			"Production Plan Sub Assembly Item", 
			doc.production_plan_sub_assembly_item, 
			["custom_actual_fg", "parent_item_code"]
		) or (None, None)

		doc.custom_actual_fg = actual_fg
		doc.custom_parent_sfg = parent_sfg
		rejection_allowance = 0
		if actual_fg:
			rejection_allowance = frappe.db.get_value("Item", actual_fg, "rejection_allowance") or 0
		doc.custom_rejection_allowance = rejection_allowance
		doc.custom_rejection_qty = flt(doc.qty) * flt(rejection_allowance) / 100

def queue_the_work_order(doc, method):
	work_orders = frappe.db.get_all(
		"Work Order",
		filters={
			"production_plan": ["!=", doc.production_plan],
			"docstatus": 1,
			"status": ["!=", 'Completed'],
			"production_item":doc.production_item
		},
		fields=["name", "production_item", "production_plan"]
	)

	for row in work_orders:
		if row.production_plan and doc.production_plan:
			doc.custom_is_queued = 1
			break

@frappe.whitelist()
def unmark_work_order(doc, method):
	if doc.custom_order_schedule:
		if frappe.db.exists("Sales Order Schedule", doc.custom_order_schedule):
			frappe.db.set_value("Sales Order Schedule", doc.custom_order_schedule, "work_order_created", 0)

@frappe.whitelist()
def update_title_in_mr(doc, method):
	"""Before Insert: if source warehouse is filled,
	update the title as value of source warehouse"""
	if doc.set_from_warehouse:
		doc.title = doc.set_from_warehouse

@frappe.whitelist()
def update_item_name_in_jobcard(doc, method):
	""""To update the item name before inserting
	the document in the DB"""
	if doc.production_item:
		doc.item_name = frappe.db.get_value("Item", doc.production_item, "item_name")
		doc.custom_item_group = frappe.db.get_value("Item", doc.production_item, "item_group")

@frappe.whitelist()
def queue_the_job_card(doc, method):
	is_queued = frappe.db.get_value("Work Order", doc.work_order, "custom_is_queued")
	if is_queued:
		doc.custom_is_queued = 1

@frappe.whitelist()
def get_raw_materials_for_jobcard(doc, method):
	if doc.is_new():
		datetime = frappe.db.get_value("Work Order", doc.work_order, "creation")
		doc.custom_work_order_date = datetime.date()
		doc.custom_shift_start_time = ''
		doc.custom_shift_end_time = ''
		doc.custom_waiting_qty = doc.for_quantity - doc.total_completed_qty
		# Setup Time & Cycle Time
		setup_time = frappe.db.get_value("BOM Operation", {"parent": doc.bom_no, "operation": doc.operation}, "custom_setup_time")
		doc.custom_setup_time = setup_time
		doc.custom_cycle_time = 1
	# Raw Materials
	if doc.bom_no:
		exploded_items = explode_bom(doc.bom_no)
		doc.set('custom_rm_availability', [])
		doc.set('custom_required_material_for_operation', [])
		possible = 0
		for item in doc.items:
			previous_possible = doc.for_quantity - doc.total_completed_qty
			item_warehouse = frappe.db.get_value("Item", item.item_code, "custom_warehouse")
			if item_warehouse in ["Semi Finished Goods - WAIP", "Finished Goods - WAIP"]:
				warehouse = item_warehouse
			else:
				warehouse = "Shop Floor - WAIP"

			required_qty = flt(item.required_qty) / flt(doc.for_quantity)
			total_required_qty = required_qty * doc.for_quantity
   
			stock_qty = flt(frappe.db.get_value("Bin", {
				"item_code": item.item_code,
				"warehouse": warehouse
			}, "actual_qty")) or 0

			available_qty = flt(frappe.db.get_value("Bin", {
				"item_code": item.item_code,
				"warehouse": "Work In Progress - WAIP",
			}, "actual_qty")) or 0
			
			# Limit available_qty to total_required_qty
			available_qty = min(available_qty, total_required_qty)
			actual_qty = available_qty + stock_qty
			actual_possible = min(total_required_qty, actual_qty) # in Required Qty
			actual_possible = actual_possible / required_qty # in Manufacturing Qty
   
			doc.append("custom_rm_availability", {
				"item_code": item.item_code,
				"item_name": frappe.db.get_value("Item", item.item_code, "item_name"),
				"actual_qty": actual_qty,
				"available_qty": available_qty,
				"stock_qty": stock_qty,
				"rate": 0,
				"warehouse": warehouse,
			})
			doc.append("custom_required_material_for_operation", {
				"item_code": item.item_code,
				"item_name": frappe.db.get_value("Item", item.item_code, "item_name"),
				"required_qty": required_qty,
				"total_required_qty": total_required_qty,
				"stock_qty": stock_qty,
				"available_qty": actual_qty,
				"possible_production": actual_qty / required_qty,
				"consumption_qty": 0,
				"actual_possible": actual_possible
			})
			possible = min(previous_possible, (actual_qty / required_qty))
		# if len(doc.items)>0:
		#     doc.custom_possible_qty = math.floor(possible) - (flt(doc.custom_rework_qty) - flt(doc.custom_rework_accepted_qty))
		# else:
		#     possible_qty = doc.custom_possible_qty or 0
		#     processed_qty = doc.custom_processed_qty or 0
		#     doc.custom_possible_qty = possible_qty if (possible_qty) > 0 else 0

def zero_ot():
	departments = frappe.db.get_all("Department", "name")
	for row in departments:
		frappe.db.set_value("OT Details", {"parent": row.name}, "allowed_ot", 0)

	
@frappe.whitelist()
def update_quantity_in_job_card(doc, method):
	operation_count = frappe.db.count("Work Order Operation", {"parent": doc.work_order})
	doc.custom_accepted_qty = doc.total_completed_qty - (doc.custom_rework_accepted_qty or 0)
	item_weight = frappe.db.get_value("Item", doc.production_item, "weight_per_unit") or 0
	doc.custom_scrap_weight = flt(doc.custom_rejected_qty) * flt(item_weight)
	doc.custom_accepted_weight = flt(doc.custom_accepted_qty) * flt(item_weight)
	doc.custom_rejected_weight = flt(doc.custom_rejected_qty) * flt(item_weight)
	doc.custom_rework_weight = flt(doc.custom_rework_qty) * flt(item_weight)
	doc.custom_waiting_qty = flt(doc.for_quantity) - flt(doc.custom_processed_qty)
	
	possible_qty = doc.custom_possible_qty or 0
	if len(doc.custom_required_material_for_operation) > 0:
		possible_qty = 0
		for row in doc.items:
			required_qty = (flt(row.required_qty) / flt(doc.for_quantity))
			row.custom_consumption_qty = flt(required_qty) * flt(doc.custom_processed_qty)
			
		for row in doc.custom_required_material_for_operation:
			total_required_qty = flt(row.required_qty) * flt(doc.custom_waiting_qty)
			row.total_required_qty = total_required_qty
			actual_possible = min(row.available_qty, total_required_qty)
			row.actual_possible = actual_possible / row.required_qty
			row.consumption_qty = flt(row.required_qty) * flt(doc.custom_processed_qty)
			possible_qty = min(row.possible_production, possible_qty) if possible_qty > 0 else row.possible_production
		
		possible_qty = min(possible_qty, doc.for_quantity)
		possible_qty = math.floor(possible_qty) - flt(doc.custom_rework_waiting_qty_details) # if rework qty is there, reduce from possible qty
		possible_qty = possible_qty - (flt(doc.custom_rejected_qty) - flt(doc.custom_rework_rejected_qty)) # if there is rejected qty in process, reduce from possible qty
		
		if operation_count == 1:
			if possible_qty > (flt(doc.for_quantity) - flt(doc.custom_processed_qty)):
				yet_to_process = (flt(doc.for_quantity) - flt(doc.custom_processed_qty)) # Waiting Qty
				possible_qty = yet_to_process

		if operation_count > 1:
			possible_qty = possible_qty + (flt(doc.custom_rejected_qty) + flt(doc.custom_rework_rejected_qty))
			final_job_card_accepted_qty = frappe.db.get_value("Job Card", {"work_order": doc.work_order, "sequence_id": operation_count}, "total_completed_qty") or 0
			if final_job_card_accepted_qty > 0:
				waiting_qty = (flt(doc.for_quantity) - flt(doc.custom_processed_qty))
				if possible_qty < waiting_qty:
					if (possible_qty + final_job_card_accepted_qty) > waiting_qty:
						possible_qty = waiting_qty
					else:
						possible_qty = possible_qty + final_job_card_accepted_qty - doc.custom_processed_qty
				else:
					possible_qty = waiting_qty
			else:
				possible_qty = possible_qty - doc.custom_processed_qty
	else:
		""""Get Total Completed Quantity from the previous Job Card 
		for the Possible Quantity"""
		rework_accepted = frappe.db.get_value("Job Card", {"work_order": doc.work_order, "sequence_id": int(doc.sequence_id) - 1}, "custom_rework_accepted_qty") or 0
		direct_accepted = frappe.db.get_value("Job Card", {"work_order": doc.work_order, "sequence_id": int(doc.sequence_id) - 1}, "custom_accepted_qty") or 0
		possible_qty = direct_accepted + rework_accepted
		possible_qty = possible_qty - flt(doc.custom_processed_qty) 
	
	doc.custom_possible_qty = possible_qty
	
	# Bin count calculation
	if doc.custom_pack_size > 0:
		bin_count = int(round(doc.total_completed_qty / doc.custom_pack_size,2))
	else:
		bin_count = 0
	doc.custom_no_of_bins = bin_count if bin_count > 0 else 1
	
	sequence_id = int(doc.sequence_id) + 1
	total_completed_qty = frappe.db.get_value("Job Card", {"work_order": doc.work_order, "sequence_id": sequence_id}, ['custom_processed_qty']) or 0
	frappe.db.set_value("Job Card", {"work_order": doc.work_order, "sequence_id": sequence_id}, "custom_possible_qty",doc.total_completed_qty-total_completed_qty)


@frappe.whitelist()
def get_tool_data(doc, method):
	if doc.is_new():
		if doc.operation and doc.workstation:
			# Mutliple Tools
			tools = frappe.get_all("Tool", {"workstation": doc.workstation, "operation": doc.operation, "status": ["!=", "Scrap"]}, ["name", "tool_name", "tool_life", "status", "total_processed_qty"])
			if tools:
				doc.set("custom_tools", [])
				for tl in tools:
					doc.append("custom_tools", {
						"tool": tl.name,
						"tool_name": tl.tool_name,
						"self_life": tl.tool_life,
						"status": tl.status,
						"total_processed_qty": tl.total_processed_qty,
						"balance_life": flt(tl.tool_life) - flt(tl.total_processed_qty),
					})


@frappe.whitelist()
def make_url_for_scan_barcode(doc, method):

	params = {
		"inspection_type": "In Process",
		"reference_type": doc.doctype,
		"reference_name": doc.name,
		"item_code": doc.production_item,
	}
	query_string = urllib.parse.urlencode(params)
	# full_url = f"https://erp.onegeneindia.in/api/method/onegene.onegene.custom.create_quality_inspection_from_qr?{query_string}"
	full_url = f"https://erp.onegeneindia.in/app/quality-inspection-tool?{query_string}"
	encoded_url = urllib.parse.quote(full_url, safe="")
	doc.custom_scan_barcode = encoded_url

	# query_string = urllib.parse.urlencode(params)
	# url = "https://erp.onegeneindia.in/app/quality-inspection/new-quality-inspection-akadwslxkf"
	# full_url = f"{url}?{query_string}"
	# encoded_url = urllib.parse.quote(full_url, safe="")
	# doc.custom_scan_barcode = encoded_url

import frappe
from frappe.utils import flt

@frappe.whitelist()
def update_the_queued_status(doc, method):
	wo_sequence = frappe.db.count("Work Order Operation", {"parent": doc.work_order})
	if doc.sequence_id == wo_sequence:
		if flt(doc.for_quantity) == flt(doc.custom_processed_qty):
			production_plan = frappe.db.get_value("Work Order", {"name": doc.work_order}, "production_plan")
			next_wo = frappe.db.get_all(
				"Work Order",
				filters={
					"creation": [">", doc.creation],
					"production_item": doc.production_item,
					"production_plan": ["!=", production_plan] 
				},
				fields=["name", "production_item", "creation"],
				order_by="creation asc",
				limit=1
			)
			if next_wo:
				next_wo_name = next_wo[0].name
				frappe.db.set_value("Work Order", next_wo_name, "custom_is_queued", 0)
				frappe.db.set_value("Job Card", {"work_order": next_wo_name}, "custom_is_queued", 0)

@frappe.whitelist()
def revert_processed_qty_in_tool(doc, method):
	if not doc.total_completed_qty:
		return
   
	# if doc.docstatus == 0 and method == "on_trash":
	# 	return

	if doc.docstatus == 2 and method == "on_trash":
		return


	tool_name = frappe.db.get_value("Job Card Tool", {"parent": doc.name}, "tool")
	if not tool_name:
		return

	if frappe.db.exists("Tool", tool_name):
		tool = frappe.get_doc("Tool", tool_name)

	   
		tool.total_processed_qty = max(0, tool.total_processed_qty - doc.total_completed_qty)

		
		for row in tool.tool_life_logs[:]:  
			if row.job_card == doc.name:
				tool.remove(row)
				break  

		tool.save(ignore_permissions=True)
		frappe.msgprint(f"Tool updated and Job Card '{doc.name}' log removed.")
		frappe.msgprint(f"Tool புதுப்பிக்கப்பட்டது மற்றும் Job Card '{doc.name}' நீக்கப்பட்டது.")


@frappe.whitelist()
def job_card_rework_entry(args):
	if isinstance(args, str):
		args = json.loads(args)
	args = frappe._dict(args)
	job_card = frappe.get_doc("Job Card", args.job_card_id)
	data = tool_life(args.job_card_id, args.processed_qty)
	if data:
		frappe.throw(data, title = "Invalid Quantity")
		frappe.throw(data, title=_("தவறான அளவு"))


	create_stock_entry_for_rework(job_card, args)
	update_tool_and_log(args.processed_qty, job_card)

	job_card.custom_rework_waiting_qty -= args.processed_qty
	job_card.custom_rework_waiting_qty_details -= args.processed_qty
	job_card.custom_rework_processed_qty += args.processed_qty
	job_card.custom_rework_accepted_qty += args.accepted_qty
	job_card.custom_rework_rejected_qty += args.rejected_qty
	job_card.process_loss_qty += args.accepted_qty
	job_card.custom_shift_type = args.shift
	job_card.workstation = args.workstation
	job_card.custom_department = args.department
	job_card.custom_supervisor = args.supervisor
	job_card.set("employee", [])
	job_card.append("employee", {
		"employee": args.employee,
	})

	for row in job_card.custom_tools:
		row.balance_life = 100
		if row.total_processed_qty >= row.self_life:
			row.status = "Inactive"
			
	if args.rejected_qty > 0:
		if len(job_card.custom_required_material_for_operation) > 0:
			for rm_row in job_card.custom_required_material_for_operation:
				found = False
				for row in job_card.custom_rejected_items:
					if row.item_code == rm_row.item_code:
						row.stock_qty += args.rejected_qty * rm_row.required_qty
						found = True
						break

				if not found:
					new_row = job_card.append('custom_rejected_items', {})
					new_row.item_code = rm_row.item_code
					new_row.stock_qty = args.rejected_qty * rm_row.required_qty
		else:
			if job_card.sequence_id == 1:
				frappe.throw("There is no Raw Materials to process", title="Not Permitted")
				frappe.throw("செயலாக்கதற்கான மூலப்பொருட்கள் இல்லை", title="Not Permitted")

			else:
				result = frappe.db.sql("""
					SELECT jc.name as job_card, COUNT(jcrm.name) as rm_count
					FROM `tabJob Card` jc
					INNER JOIN `tabJob Card Required Material Operation` jcrm
					ON jc.name = jcrm.parent
					WHERE jc.sequence_id < %s
						AND jc.work_order = %s
					GROUP BY jc.name
					HAVING COUNT(jcrm.name) > 0
					ORDER BY jc.sequence_id desc
					""", (job_card.sequence_id, job_card.work_order,), as_dict=1)
				if result:
					rm_jc = result[0]["job_card"]
					rm_jc_doc = frappe.get_doc("Job Card", rm_jc)
					for rm_row in rm_jc_doc.custom_required_material_for_operation:
						found = False
						for row in job_card.custom_rejected_items:
							if row.item_code == rm_row.item_code:
								row.stock_qty += args.rejected_qty * rm_row.required_qty
								found = True
								break

						if not found:
							new_row = job_card.append('custom_rejected_items', {})
							new_row.item_code = rm_row.item_code
							new_row.stock_qty = args.rejected_qty * rm_row.required_qty
	
		rm_required_qty = frappe.db.sql("""
			SELECT SUM(required_qty) as total_required_qty
			FROM `tabJob Card Required Material Operation`
			WHERE parent = %s
		""", (job_card.name,), as_dict=1)[0].total_required_qty or 0
	
		job_card.append("custom_rejection_logs", {
			"datetime": now(),
			"rejected_qty": args.rejected_qty,
			"employee": args.employee,
			"rejection_category": args.rejection_category,
			"rejection_remarks": args.rejection_remarks,
			"shift_type": args.shift,
			"entry_type": "Rework",
			"quantity": rm_required_qty * args.rejected_qty,
		}) 
							
	shift_start_time = frappe.db.get_value("Shift Type", args.shift, "start_time")
	shift_end_time = frappe.db.get_value("Shift Type", args.shift, "end_time")
	shift_end_date = add_days(job_card.posting_date, 1) if shift_start_time > shift_end_time else job_card.posting_date
	jc_from_time = cstr(job_card.posting_date) + " " + cstr(shift_start_time)
	jc_to_time = cstr(shift_end_date)+ " " + cstr(shift_end_time)
	job_card.append("time_logs", {
		"employee": args.employee,
		"from_time": jc_from_time,
		"to_time": jc_to_time,
		"time_in_mins": 3,
		"completed_qty": args.accepted_qty,
		"custom_rejected_qty": args.rejected_qty,
		"custom_rejection_category": args.rejection_category,
		"custom_rejection_reason": args.rejection_remarks,
		"employee": args.employee,
		"custom_created_by": frappe.session.user,
		"custom_entry_type": "Rework",
		"custom_shift_type": args.shift,
	})

	job_card.append("custom_rework_logs", {
		"datetime": now(),
		"reworked_qty": args.processed_qty,
		"accepted_qty": args.accepted_qty,
		"rejected_qty": args.rejected_qty,
		"employee": args.employee,
		"rejection_category": args.rejection_category,
		"rejection_remarks": args.rejection_remarks,
		"shift_type": args.shift,
	})

	job_card.save(ignore_permissions=True)
	update_rejected_qty_in_next_job_card(job_card.work_order, job_card.sequence_id, args.rejected_qty)
	create_quality_pending(args, job_card.production_item)

@frappe.whitelist()
def tool_life(name, processed_qty):
	doc = frappe.get_doc("Job Card", name)
	smallest_row = None
	smallest_diff = None

	for row in doc.custom_tools:
		diff = row.self_life - row.total_processed_qty
		if smallest_diff is None or diff < smallest_diff:
			smallest_diff = diff
			smallest_row = row
	if smallest_row and smallest_diff < processed_qty:
		if smallest_diff > 0 :
			return (
				f"Tool <a href='/app/tool/{smallest_row.tool}'><b style='color: #941f1f;'>{smallest_row.tool}</b></a> has only <b>{smallest_diff}</b> remaining life, "
				f"which is less than the required <b>{doc.custom_waiting_qty}</b>"
			)

@frappe.whitelist()
def create_stock_entry_for_rework(doc, args):
	"""If all the Job Cards are done and on completion of final
	Job Card, the completed qty will be updated in the Work Order.
	And Stock Entry will be created."""
	actual_fg = frappe.db.get_value("Work Order", doc.work_order, "custom_actual_fg") or None
	# if doc.production_item != actual_fg:
	completed_qty = args.accepted_qty
	process_loss_qty = args.rejected_qty
	process_loss_percentage = (process_loss_qty / completed_qty) * 100 if completed_qty > 0 else 0
	produced_qty = completed_qty - process_loss_qty
	last_sequence = frappe.db.count("Work Order Operation", {"parent": doc.work_order})
	frappe.db.sql("""
			UPDATE `tabWork Order Operation`
			SET completed_qty = completed_qty + %s,
				process_loss_qty = process_loss_qty + %s,
				custom_waiting_qty = %s - completed_qty
			WHERE name = %s
		""", (flt(completed_qty), flt(doc.process_loss_qty), flt(doc.for_quantity), doc.operation_id))
	if doc.sequence_id == last_sequence:
		source_warehouse = frappe.db.get_value("Work Order", doc.work_order, "wip_warehouse")
		target_warehouse = frappe.db.get_value("Item", doc.production_item, "custom_warehouse")
		item_billing_type = frappe.db.get_value("Item", {"item_code": doc.production_item}, "item_billing_type")
		if completed_qty > 0:
			se = frappe.new_doc("Stock Entry")
			se.disable_auto_set_process_loss_qty = 1
			se.company = "WONJIN AUTOPARTS INDIA PVT.LTD."
			se.stock_entry_type = "Manufacture"
			se.work_order = doc.work_order
			se.from_bom = 1
			se.fg_completed_qty = completed_qty
			se.process_loss_qty = completed_qty
			se.process_loss_percentage = process_loss_percentage
			se.bom_no = doc.bom_no
			se.set('items', [])
			if len(doc.custom_required_material_for_operation):
				for row in doc.custom_required_material_for_operation:
					converted_raw_material_qty = (row.required_qty) * completed_qty
					se.append('items', {
						"s_warehouse": source_warehouse,
						"item_code": row.item_code,
						"qty": converted_raw_material_qty,
						"basic_rate": 12,
						"expense_account": "Stock Adjustment - WAIP",
						"valuation_rate": 12,
						"allow_zero_valuation_rate": 1,
					})
			else:
				update_job_card_raw_material_for_operation(doc.sequence_id, doc.work_order, se, source_warehouse, completed_qty)
			se.append('items', {
				"t_warehouse": target_warehouse if item_billing_type != "Billing" else "Quality Inspection Pending - WAIP", # WIP
				"item_code": doc.production_item,
				"qty": produced_qty,
				"basic_rate": 0.4766,
				"expense_account": "Stock Adjustment - WAIP",
				"valuation_rate": 0.4766,
				"allow_zero_valuation_rate": 1,
			})
			se.insert()
			se.submit()
			frappe.db.commit()

@frappe.whitelist()
def update_tool_and_log(processed_qty, job_card):
	for row in job_card.custom_tools:
		tool = frappe.get_doc("Tool", row.tool)
		tool.total_processed_qty += processed_qty
		tool.append('tool_life_logs', {
			"job_card": job_card.name,
			"date": today(),
			"tool_life": tool.custom_balance_life,
			"processed": processed_qty,
			"balance_life": flt(tool.custom_balance_life) - flt(processed_qty),
		})
		tool.save()


@frappe.whitelist()
def update_rejected_qty_in_next_job_card(work_order, sequence_id, rejected_qty):
	"""Update the rejected qty in next job card if exists"""
	next_job_card = frappe.db.get_all(
		"Job Card",
		filters={
			"work_order": work_order,
			"sequence_id": [">", sequence_id],
		},
		fields=["name", "process_loss_qty"],
		order_by="sequence_id asc",
		limit=1
	)

	if next_job_card:
		next_job_card_name = next_job_card[0].name
		current_process_loss_qty = next_job_card[0].process_loss_qty or 0
		new_process_loss_qty = flt(current_process_loss_qty) + flt(rejected_qty)
		
		frappe.db.set_value("Job Card", next_job_card_name, "process_loss_qty", new_process_loss_qty)
		
	return "ok"

@frappe.whitelist()
def get_shift_for_current_time(doctype=None, txt=None, searchfield=None, start=None, page_len =None, filters=None):
	
	current_datetime = now()
	current_time = current_datetime.split(" ")[1]
	return frappe.db.sql("""
		SELECT name
		FROM `tabShift Type`
		WHERE (
				(start_time <= end_time AND %(now)s BETWEEN start_time AND end_time)
			 OR (start_time > end_time AND (%(now)s >= start_time OR %(now)s <= end_time))
			)
			AND docstatus < 2
			AND custom_disabled = 0
	""", {"now": current_time})

@frappe.whitelist()
def get_supervisor(doctype, txt, searchfield, start, page_len, filters):
	if isinstance(filters, str):
		try:
			filters = json.loads(filters)
		except Exception:
			filters = {}

	department = filters.get("department")

	start = int(start) if start else 0
	page_len = int(page_len) if page_len else 20

	engineers = frappe.db.sql("""
		SELECT e.name, e.employee_name, d.engineer
		FROM `tabDepartment Engineers` d
		INNER JOIN `tabEmployee` e ON e.user_id = d.engineer
		WHERE d.parent = %s
		  AND d.engineer LIKE %s
		LIMIT %s, %s
	""", (department, f"%{txt}%", start, page_len))

	return engineers

@frappe.whitelist()
def get_raw_materials_for_jobcard_on_mt(name, disable_save=False):
	doc = frappe.get_doc('Job Card',name)
	possible = 0
	if doc.bom_no:
		exploded_items = explode_bom(doc.bom_no)
		doc.set('custom_rm_availability', [])
		doc.set('custom_required_material_for_operation', [])
		for item in doc.items:
			previous_possible = possible
			item_warehouse = frappe.db.get_value("Item", item.item_code, "custom_warehouse")
			if item_warehouse in ["Semi Finished Goods - WAIP", "Finished Goods - WAIP"]:
				warehouse = item_warehouse
			else:
				warehouse = "Shop Floor - WAIP"

			required_qty = flt(item.required_qty) / flt(doc.for_quantity)
			total_required_qty = flt(required_qty) * flt(doc.for_quantity)

			stock_qty = flt(frappe.db.get_value("Bin", {
				"item_code": item.item_code,
				"warehouse": warehouse
			}, "actual_qty")) or 0

			available_qty = flt(frappe.db.get_value("Bin", {
				"item_code": item.item_code,
				"warehouse": "Work In Progress - WAIP",
			}, "actual_qty")) or 0

			# Limit available_qty to total_required_qty
			available_qty = min(available_qty, total_required_qty)
			actual_qty = available_qty + stock_qty
			actual_possible = min(actual_qty, total_required_qty) # in Required Qty
			actual_possible = actual_possible / required_qty # in Manufacturing Qty
			doc.append("custom_rm_availability", {
				"item_code": item.item_code,
				"item_name": frappe.db.get_value("Item", item.item_code, "item_name"),
				"actual_qty": actual_qty,
				"available_qty": available_qty,
				"stock_qty": stock_qty,
				"rate": 0,
				"warehouse": warehouse,
			})
			
			doc.append("custom_required_material_for_operation", {
				"item_code": item.item_code,
				"item_name": frappe.db.get_value("Item", item.item_code, "item_name"),
				"required_qty": required_qty,
				"total_required_qty": total_required_qty,
				"stock_qty": stock_qty,
				"available_qty": actual_qty,
				"possible_production": actual_qty / required_qty,
				"actual_possible": actual_possible
			})
			possible = min(previous_possible, (actual_qty / required_qty))
		doc.custom_possible_qty = possible
		if not disable_save:
			doc.save(ignore_permissions=True)

@frappe.whitelist()
def explode_bom(bom):
	bom_qty = frappe.db.get_value("BOM", bom, "quantity") or 1
	bom_items = frappe.get_all("BOM Item", filters={"parent": bom}, fields=["item_code", "qty"])
	exploded_items = []
	for item in bom_items:
		exploded_items.append({
			"item_code": item.item_code,
			"qty": (item.qty * bom_qty) or 0
		})
	return exploded_items

@frappe.whitelist()
def job_card_process_entry(args):
	if isinstance(args, str):
		args = json.loads(args)
	args = frappe._dict(args)
	job_card = frappe.get_doc("Job Card", args.job_card_id)
	validate_job_card_sequence_id(job_card, args.processed_qty)
	data = tool_life(args.job_card_id, args.processed_qty)
	if data:
		frappe.throw(data, title = "Invalid Quantity")
		frappe.throw(data, title=_("தவறான அளவு"))

	create_stock_entry(job_card, args)
	update_tool_and_log(args.processed_qty, job_card)

	job_card.custom_processed_qty += args.processed_qty
	job_card.custom_accepted_qty += args.accepted_qty
	job_card.custom_rejected_qty += args.rejected_qty
	job_card.custom_rework_qty += args.rework_qty
	job_card.custom_rework_waiting_qty += args.rework_qty
	job_card.custom_rework_waiting_qty_details += args.rework_qty
	job_card.custom_shift_type = args.shift
	job_card.workstation = args.workstation
	job_card.custom_department = args.department
	job_card.custom_supervisor = args.supervisor
	job_card.set("employee", [])
	job_card.append("employee", {
		"employee": args.employee,
	})

	for row in job_card.custom_tools:
		row.balance_life = 100
		# row.total_processed_qty += job_card.custom_processed_qty
		if row.total_processed_qty >= row.self_life:
			row.status = "Inactive"

	if args.rejected_qty > 0:
		if len(job_card.custom_required_material_for_operation) > 0:
			for rm_row in job_card.custom_required_material_for_operation:
				found = False
				for row in job_card.custom_rejected_items:
					if row.item_code == rm_row.item_code:
						row.stock_qty += args.rejected_qty * rm_row.required_qty
						found = True
						break

				if not found:
					new_row = job_card.append('custom_rejected_items', {})
					new_row.item_code = rm_row.item_code
					new_row.stock_qty = args.rejected_qty * rm_row.required_qty
		else:
			if job_card.sequence_id == 1:
				frappe.throw("There is no Raw Materials to process", title="Not Permitted")
				frappe.throw("செயலாக்கதற்கான மூலப்பொருட்கள் இல்லை", title="Not Permitted")

			else:
				result = frappe.db.sql("""
					SELECT jc.name as job_card, COUNT(jcrm.name) as rm_count
					FROM `tabJob Card` jc
					INNER JOIN `tabJob Card Required Material Operation` jcrm
					ON jc.name = jcrm.parent
					WHERE jc.sequence_id < %s
						AND jc.work_order = %s
					GROUP BY jc.name
					HAVING COUNT(jcrm.name) > 0
					ORDER BY jc.sequence_id desc
					""", (job_card.sequence_id, job_card.work_order,), as_dict=1)
				if result:
					rm_jc = result[0]["job_card"]
					rm_jc_doc = frappe.get_doc("Job Card", rm_jc)
					for rm_row in rm_jc_doc.custom_required_material_for_operation:
						found = False
						for row in job_card.custom_rejected_items:
							if row.item_code == rm_row.item_code:
								row.stock_qty += args.rejected_qty * rm_row.required_qty
								found = True
								break

						if not found:
							new_row = job_card.append('custom_rejected_items', {})
							new_row.item_code = rm_row.item_code
							new_row.stock_qty = args.rejected_qty * rm_row.required_qty
							
		rm_required_qty = frappe.db.sql("""
			SELECT SUM(required_qty) as total_required_qty
			FROM `tabJob Card Required Material Operation`
			WHERE parent = %s
		""", (job_card.name,), as_dict=1)[0].total_required_qty or 0

		job_card.append("custom_rejection_logs", {
			"datetime": now(),
			"rejected_qty": args.rejected_qty,
			"employee": args.employee,
			"rejection_category": args.rejection_category,
			"rejection_remarks": args.rejection_remarks,
			"shift_type": args.shift,
			"entry_type": "Direct",
			"quantity": rm_required_qty * args.rejected_qty,

		})

	shift_start_time = frappe.db.get_value("Shift Type", args.shift, "start_time")
	shift_end_time = frappe.db.get_value("Shift Type", args.shift, "end_time")
	shift_end_date = add_days(job_card.posting_date, 1) if shift_start_time > shift_end_time else job_card.posting_date
	jc_from_time = cstr(job_card.posting_date) + " " + cstr(shift_start_time)
	jc_to_time = cstr(shift_end_date)+ " " + cstr(shift_end_time)
	job_card.append("time_logs", {
		"employee": args.employee,
		"from_time": jc_from_time,
		"to_time": jc_to_time,
		"time_in_mins": 3,
		"completed_qty": args.accepted_qty,
		"custom_rejected_qty": args.rejected_qty,
		"custom_rework_qty": args.rework_qty,
		"custom_rejection_category": args.rejection_category,
		"custom_rejection_reason": args.rejection_remarks,
		"custom_rework_category": args.rework_category,
		"custom_rework_reason": args.rework_remarks,
		"employee": args.employee,
		"custom_created_by":frappe.session.user,
		"custom_entry_type": "Direct",
		"custom_shift_type": args.shift,
	})
	job_card.save(ignore_permissions=True)
	update_rejected_qty_in_next_job_card(job_card.work_order, job_card.sequence_id, args.rejected_qty)
	create_quality_pending(args, job_card.production_item)

@frappe.whitelist()
def validate_job_card_sequence_id(self, processed_qty):
	if self.is_corrective_job_card:
			return

	if not (self.work_order and self.sequence_id):
		return

	current_operation_qty = 0.0
	data = self.get_current_operation_data()
	current_operation_qty = flt(self.custom_processed_qty) + flt(processed_qty)
	data = frappe.get_all(
			"Work Order Operation",
			fields=["operation", "status", "completed_qty", "sequence_id"],
			filters={"docstatus": 1, "parent": self.work_order, "sequence_id": ("<", self.sequence_id)},
			order_by="sequence_id, idx",
		)

	# message = "Job Card {0}: ".format(
	# 	bold(self.name), bold(get_link_to_form("Work Order", self.work_order))
	# )

	for row in data:
		if row.completed_qty < current_operation_qty:
			frappe.throw(
				_("The previous operation {0} has only been completed for {1} units. " \
				"So, you cannot proceed with the current operation {2} for more than {1} units.").format(
					bold(row.operation), bold(row.completed_qty), bold(self.operation)
				),
				OperationSequenceError,
			)

		if row.completed_qty < current_operation_qty:
			msg = f"""The completed quantity {bold(current_operation_qty)}
				of an operation {bold(self.operation)} cannot be greater
				than the completed quantity {bold(row.completed_qty)}
				of a previous operation
				{bold(row.operation)}.
			"""

			return frappe.throw(_(msg))

@frappe.whitelist()
def create_stock_entry_after_qc(doc, method):
	""""On submission of Quality Inspection, 
	the stock will be moved from Quality Inspection Pending to Quality Outward warehouse"""

	if flt(doc.custom_accepted_qty) > 0 and doc.workflow_state != "Rejected":
		se = frappe.new_doc("Stock Entry")
		se.company = "WONJIN AUTOPARTS INDIA PVT.LTD."
		se.stock_entry_type = "Material Transfer"
		se.from_warehouse = "Quality Inspection Pending - WAIP"
		se.to_warehouse = "Store Receipt Pending - WAIP"
		se.inspection_required = 1
		# se.posting_date = today()
		# se.posting_time = now()

		se.custom_issued_by = frappe.session.user
		se.custom_issued_by_name = frappe.db.get_value(
			"Employee",
			{"user_id": frappe.session.user},
			"employee_name"
		)

		se.append('items', {
			"s_warehouse": "Quality Inspection Pending - WAIP",
			"t_warehouse": "Store Receipt Pending - WAIP",
			"item_code": doc.item_code,
			"qty": flt(doc.custom_accepted_qty) or 0,
			"basic_rate": 0.4766,
			"valuation_rate": 0.4766,
			"allow_zero_valuation_rate": 1,
			"quality_inspection": doc.name
		})

		se.insert()
		se.submit()
		frappe.db.commit()

@frappe.whitelist()
def create_stock_entry(doc, args):
	"""If all the Job Cards are done and on completion of final
	Job Card, the completed qty will be updated in the Work Order.
	And Stock Entry will be created."""
	
	actual_fg = frappe.db.get_value("Work Order", doc.work_order, "custom_actual_fg") or None
	# if doc.production_item != actual_fg:
	completed_qty = args.accepted_qty
	process_loss_qty = args.rejected_qty
	process_loss_percentage = (process_loss_qty / completed_qty) * 100 if completed_qty > 0 else 0
	produced_qty = completed_qty - process_loss_qty
	last_sequence = frappe.db.count("Work Order Operation", {"parent": doc.work_order})
	frappe.db.sql("""
			UPDATE `tabWork Order Operation`
			SET completed_qty = completed_qty + %s,
				process_loss_qty = process_loss_qty + %s,
				custom_waiting_qty = %s - completed_qty
			WHERE name = %s
		""", (flt(completed_qty), flt(doc.process_loss_qty), flt(doc.for_quantity), doc.operation_id))
	if doc.sequence_id == last_sequence:
		source_warehouse = frappe.db.get_value("Work Order", doc.work_order, "wip_warehouse")
		target_warehouse = frappe.db.get_value("Item", doc.production_item, "custom_warehouse")
		item_billing_type = frappe.db.get_value("Item", {"item_code": doc.production_item}, "item_billing_type")
		if completed_qty > 0:
			se = frappe.new_doc("Stock Entry")
			se.disable_auto_set_process_loss_qty = 1
			se.company = "WONJIN AUTOPARTS INDIA PVT.LTD."
			se.stock_entry_type = "Manufacture"
			se.work_order = doc.work_order
			se.from_bom = 1
			se.fg_completed_qty = completed_qty
			se.process_loss_qty = process_loss_qty
			se.process_loss_percentage = process_loss_percentage
			se.bom_no = doc.bom_no
			se.set('items', [])
			if len(doc.custom_required_material_for_operation):
				for row in doc.custom_required_material_for_operation:
					converted_raw_material_qty = (row.required_qty) * completed_qty
					se.append('items', {
						"s_warehouse": source_warehouse,
						"item_code": row.item_code,
						"qty": converted_raw_material_qty,
						"basic_rate": 12,
						"expense_account": "Stock Adjustment - WAIP",
						"valuation_rate": 12,
						"allow_zero_valuation_rate": 1,
					})
			else:
				update_job_card_raw_material_for_operation(doc.sequence_id, doc.work_order, se, source_warehouse, completed_qty)
			se.append('items', {
				"t_warehouse": target_warehouse if item_billing_type != "Billing" else "Quality Inspection Pending - WAIP", # WIP
				"item_code": doc.production_item,
				"qty": produced_qty,
				"basic_rate": 0.4766,
				"expense_account": "Stock Adjustment - WAIP",
				"valuation_rate": 0.4766,
				"allow_zero_valuation_rate": 1,
			})
			se.insert()
			se.submit()
			frappe.db.commit()


@frappe.whitelist()
def job_card_dashboard_comment(name):
	doc = frappe.get_doc("Job Card", name)
	smallest_row = None
	smallest_diff = None
	for row in doc.custom_tools:
		diff = row.self_life - row.total_processed_qty
		if smallest_diff is None or diff < smallest_diff:
			smallest_diff = diff
			smallest_row = row

	if smallest_row and smallest_diff < doc.custom_waiting_qty:
		if smallest_diff <= 0:
			return f"Tool <a href='/app/tool/{smallest_row.tool}'><b style='color: #941f1f;'>{smallest_row.tool}</b></a> has completely exhausted its life. Please replace or refurbish the tool before proceeding"
		elif smallest_diff < doc.custom_waiting_qty:
			return (
				f"Tool <a href='/app/tool/{smallest_row.tool}'><b style='color: #941f1f;'>{smallest_row.tool}</b></a> has only <b>{smallest_diff}</b> remaining life, "
				f"which is less than the required <b>{doc.custom_waiting_qty}</b>"
			)

@frappe.whitelist()
def get_quality_inspection_data(quality_inspection):
	accepted_qty = frappe.db.get_value("Quality Inspection", quality_inspection, "custom_accepted_qty") or 0
	item_code = frappe.db.get_value("Quality Inspection", quality_inspection, "item_code")
	rack = frappe.db.get_value("Item Rack", {"parent": item_code, "is_default": 1}, "rack") or ""
	location = frappe.db.get_value("Item Rack Location", {"parent": item_code, "is_default": 1}, "location") or ""
	uom = frappe.db.get_value("Item", item_code, "stock_uom")
	item_name = frappe.db.get_value("Quality Inspection", quality_inspection,['item_name'])
	return item_code, accepted_qty, rack, location, uom, item_name


@frappe.whitelist()
def get_previous_purchase_rate(item_code):
	pr = frappe.db.sql("""
		SELECT 
			lci.rate
		FROM 
			`tabPurchase Order Item` lci
		LEFT JOIN 
			`tabPurchase Order` lcv
		ON 
			lci.parent = lcv.name
		JOIN 
			(SELECT 
				item_code, MAX(creation) as latest_creation
			FROM 
				`tabPurchase Order Item`
			WHERE 
				rate IS NOT NULL
			GROUP BY 
				item_code) latest_lci
		ON 
			lci.item_code= latest_lci.item_code
			AND lci.creation = latest_lci.latest_creation
		WHERE 
			lci.item_code = %s and lcv.docstatus !=2
	""", item_code, as_dict=True)

	if pr:
		return pr[0].get('rate')
	else:
		return '0.0'


@frappe.whitelist()
def get_pmr_data(item_code, name=None, warehouse=None):
	if not warehouse:
		return
	current_datetime = now_datetime()
	if current_datetime.time() > time(8, 30):
		start_datetime = f"{today()} 08:31:00"
		end_datetime = f"{add_days(today(), 1)} 08:30:00"
	else:
		start_datetime = f"{add_days(today(), -1)} 08:31:00"
		end_datetime = f"{today()} 08:30:00"

	start_datetime = frappe.utils.get_datetime(start_datetime)
	end_datetime = frappe.utils.get_datetime(end_datetime)

	result = frappe.db.sql("""
		SELECT name FROM `tabProduction Material Request`
		WHERE creation BETWEEN %s AND %s
	""", (start_datetime, end_datetime))
	mr_qty = frappe.db.sql("""SELECT sum(mri.custom_requesting_qty)
						FROM `tabMaterial Request` mr
						INNER JOIN `tabMaterial Request Item` mri
						ON mri.parent = mr.name 
						WHERE mri.item_code = %s AND mri.from_warehouse = %s AND
							mr.material_request_type = 'Material Transfer' AND parent != %s
							AND mr.creation between %s AND %s""",
						(item_code, warehouse, name, start_datetime, end_datetime))[0][0] or 0
	if result:
		production_material_request = result[0][0]
	else:
		# production_material_request = "PMR-00013"
		frappe.throw("No Production Material Request found in the given time range.")
		frappe.throw("கொடுக்கப்பட்ட காலப்பகுதியில் எந்த Production Material Request-ம் கண்டுபிடிக்கப்படவில்லை.")

	if not frappe.db.exists("Sub Assembly Item", {"parent": production_material_request, "item_code": item_code, "warehouse": warehouse}) and not frappe.db.exists("Raw Materials", {"parent": production_material_request, "item_code": item_code, "warehouse": warehouse}):
		return "item not found"
		# frappe.throw(f"Item {item_code} not found in Production Material Request of {production_material_request}")
	pmr = frappe.db.sql("""
		SELECT 
			item_code,
			item_name,
			(SUM(required_plan) - %s) AS total_required_qty,
			SUM(stock_in_shop_floor) AS shop_floor_qty
		FROM (
			SELECT 
				rm.item_code, 
				rm.item_name, 
				rm.required_plan, 
				rm.stock_in_shop_floor
			FROM `tabRaw Materials` rm
			WHERE 
				rm.parent = %s AND 
				rm.item_code = %s AND 
				rm.warehouse = %s

			UNION ALL

			SELECT 
				sai.item_code, 
				sai.item_name, 
				sai.required_plan, 
				sai.stock_in_shop_floor
			FROM `tabSub Assembly Item` sai
			WHERE 
				sai.parent = %s AND 
				sai.item_code = %s AND 
				sai.warehouse = %s
		) AS combined
		GROUP BY item_code, item_name
	""", (
		flt(mr_qty), 
		production_material_request, item_code, warehouse,  # for Raw Materials
		production_material_request, item_code, warehouse   # for Sub Assembly Item
	), as_dict=1)

	data = []

	for item in pmr:
		data.append({
			"item_code": item.item_code,
			"item_name": item.item_name,
			"pack_size": frappe.db.get_value("Item", item.item_code, "pack_size"),
			"uom": frappe.db.get_value("Item", item.item_code, "stock_uom"),
			"total_required_qty": item.total_required_qty or 0,
			"shop_floor_qty": item.shop_floor_qty or 0,
			"actual_qty": flt(frappe.db.get_value("Bin", {
				"item_code": item.item_code,
				"warehouse": warehouse
			}, "actual_qty")) or 0
		})
	return data


# @frappe.whitelist()
# def update_material_req(doc, method):
# 	tot_qty=0
# 	completed=0
# 	if doc.custom_material_request:
# 		mr=frappe.get_doc('Material Request',doc.custom_material_request)
		
# 		if mr.material_request_type=='Material Transfer':
# 			for i in mr.items:
# 				tot_qty+=i.qty
# 				completed+=i.ordered_qty
# 			if completed==0:
# 				status='Not Yet Received'
# 			elif tot_qty==completed:
# 				status='Fully Received'
# 			else:
# 				status='Partially Received'
# 			frappe.db.set_value('Material Request',doc.custom_material_request,'custom_received_item_status',status)


@frappe.whitelist()
def update_material_req(doc, method):
	tot_qty=0
	completed=0
	if doc.custom_material_request:
		mr=frappe.get_doc('Material Request',doc.custom_material_request)
		
		if mr.material_request_type=='Material Transfer':
			for i in mr.items:
				tot_qty+=i.qty
				completed+=i.ordered_qty
			if completed==0:
				status='Not Yet Received'
			elif tot_qty==completed:
				status='Fully Received'
			else:
				status='Partially Received'
			frappe.db.set_value('Material Request',doc.custom_material_request,'custom_received_item_status',status)


from erpnext.stock.stock_ledger import NegativeStockError, get_previous_sle, get_valuation_rate
@frappe.whitelist()
def validate_stock_qty(item_code,warehouse,posting_date,posting_time,qty,idx):
	from erpnext.stock.stock_ledger import is_negative_stock_allowed
	allow_negative_stock = is_negative_stock_allowed(item_code=item_code)
	previous_sle = get_previous_sle(
		{
			"item_code": item_code,
			"warehouse": warehouse,
			"posting_date": posting_date,
			"posting_time": posting_time,
		}
	)

	# get actual stock at source warehouse
	actual_qty = previous_sle.get("qty_after_transaction") or 0

	# validate qty during submit
	if (
		warehouse
		and not allow_negative_stock
		and flt(actual_qty,)
		< flt(qty,)
	):
		frappe.throw(
			_(
				"Row {0}: Quantity not available for {4} in warehouse {1} at posting time of the entry ({2} {3})"
			).format(
				idx,
				frappe.bold(warehouse),
				formatdate(posting_date),
				format_time(posting_time),
				frappe.bold(item_code),
			)
			+ "<br><br>"
			+ _("Available quantity is {0}, you need {1}").format(
				frappe.bold(flt(actual_qty)), frappe.bold(qty)
			),
			NegativeStockError,
			title=_("Insufficient Stock"),
		)
	return actual_qty


@frappe.whitelist()
def validate_qi_on_submission(doc, method):
	if flt(doc.custom_inspected_qty) == 0:
		frappe.throw("Cannot submit the document with 0 Inspected Quantity", title="Mandatory Error")
	for row in doc.readings:
		for i in range(1, 6):
			fieldname = f"reading_{i}"
			if not getattr(row, fieldname, None):
				frappe.throw(
					f"Cannot submit the document without filling Reading {i} for Row {row.idx}.",
					title="Not Permitted"
				)
			if row.status == "Rejected":
				if flt(doc.custom_accepted_qty) > 0:
					if not row.custom_remarks:
						frappe.throw(
							f"Cannot submit the document without filling Rejection Reason for Row {row.idx}.",
							title="Not Permitted"
						)
					else: 
						row.status = "Accepted"
	doc.status = "Accepted"

@frappe.whitelist()
def get_items_from_production_material_request(doctype, txt, searchfield, start, page_len, filters):
	current_datetime = now_datetime()
	if current_datetime.time() > time(8, 30):
		start_datetime = f"{today()} 08:31:00"
		end_datetime = f"{add_days(today(), 1)} 08:30:00"
	else:
		start_datetime = f"{add_days(today(), -1)} 08:31:00"
		end_datetime = f"{today()} 08:30:00"

	start_datetime = frappe.utils.get_datetime(start_datetime)
	end_datetime = frappe.utils.get_datetime(end_datetime)

	result = frappe.db.sql("""
		SELECT name FROM `tabProduction Material Request`
		WHERE creation BETWEEN %s AND %s
	""", (start_datetime, end_datetime))

	if result:
		production_material_request = result[0][0]
	else:
		# production_material_request = "PMR-00013"
		frappe.throw("No Production Material Request found in the given time range.")
		frappe.throw("கொடுக்கப்பட்ட காலப்பகுதியில் எந்த Production Material Request-ம் கண்டுபிடிக்கப்படவில்லை.")


	if isinstance(filters, str):
		filters = json.loads(filters)

	source_warehouse = filters.get("source_warehouse")
	if not source_warehouse:
		return []

	search_text = f"%{txt.strip()}%" if txt else "%"
	start = int(start)
	page_len = int(page_len)
	return frappe.db.sql("""
		SELECT DISTINCT item_code, item_name
		FROM (
			SELECT rm.item_code, rm.item_name
			FROM `tabRaw Materials` rm
			WHERE rm.parent = %s AND rm.warehouse = %s

			UNION ALL

			SELECT sai.item_code, sai.item_name
			FROM `tabSub Assembly Item` sai
			WHERE sai.parent = %s AND sai.warehouse = %s
		) AS merged_items
		WHERE (item_code LIKE %s OR item_name LIKE %s)
	""", (production_material_request, source_warehouse, production_material_request, source_warehouse, search_text, search_text))

@frappe.whitelist()
def update_jobcard_posting_date():
	now = today()
	
	job_cards = frappe.get_all("Job Card",filters = {"docstatus": 0},fields = ["name"])
	
	for jc in job_cards:
		
		frappe.db.set_value("Job Card", jc.name, "posting_date",now) 

@frappe.whitelist()
def get_quality_inspection(item_code):
	qc = frappe.db.get_value("Item", item_code, "quality_inspection_template")
	if qc:
		return "ok"

@frappe.whitelist()
def get_warehouses(doctype, txt, searchfield, start, page_len, filters):
	operation = frappe.get_doc("Operation",filters.get("name"))
	ws = []
	if operation.custom_other_workstations:
		for wh in operation.custom_other_workstations:
			ws.append((wh.workstation,))
	if operation.workstation:
		ws.append((operation.workstation,))
	return ws

@frappe.whitelist()
def update_possible_qty_onload(name):
	doc = frappe.get_doc("Job Card", name)

	if doc.status in ["Submitted", "Cancelled", "Completed"]:
		return {"updated": False}

	if not doc.bom_no:
		return {"updated": False}

	possible = 0
	custom_rm_availability = []
	custom_required_material_for_operation = []

	for item in doc.items:
		previous_possible = doc.for_quantity - doc.total_completed_qty
		item_warehouse = frappe.db.get_value("Item", item.item_code, "custom_warehouse")
		warehouse = (
			item_warehouse
			if item_warehouse in ["Semi Finished Goods - WAIP", "Finished Goods - WAIP"]
			else "Shop Floor - WAIP"
		)

		required_qty = flt(item.required_qty) / flt(doc.for_quantity)
		total_required_qty = flt(required_qty) * flt(doc.for_quantity)

		stock_qty = flt(frappe.db.get_value("Bin", {
			"item_code": item.item_code,
			"warehouse": warehouse
		}, "actual_qty")) or 0

		available_qty = flt(frappe.db.get_value("Bin", {
			"item_code": item.item_code,
			"warehouse": "Work In Progress - WAIP",
		}, "actual_qty")) or 0

		available_qty = min(available_qty, total_required_qty)
		actual_qty = available_qty + stock_qty

		custom_rm_availability.append({
			"item_code": item.item_code,
			"item_name": frappe.db.get_value("Item", item.item_code, "item_name"),
			"actual_qty": actual_qty,
			"available_qty": available_qty,
			"stock_qty": stock_qty,
			"rate": 0,
			"warehouse": warehouse,
		})
		custom_required_material_for_operation.append({
			"item_code": item.item_code,
			"item_name": frappe.db.get_value("Item", item.item_code, "item_name"),
			"required_qty": required_qty,
			"total_required_qty": required_qty * flt(doc.for_quantity),
			"stock_qty": stock_qty,
			"available_qty": actual_qty,
			"possible_production": actual_qty / required_qty,
			"consumption_qty": 0,
		})
		possible = min(previous_possible, (actual_qty / required_qty))

	if len(doc.items) > 0:
		possible_qty = round(possible)
	else:
		possible_qty = (
			doc.custom_possible_qty
			if doc.custom_possible_qty > 0
			else 0
		)
	return {
		"updated": True,
		"possible_qty": possible_qty,
		"rm_availability": custom_rm_availability,
		"required_materials": custom_required_material_for_operation
	}

@frappe.whitelist()
@frappe.whitelist()
def update_so_workflow():
	so=frappe.db.get_all('Sales Order',{'docstatus':1},['name'])
	for s in so:
		frappe.db.set_value('Sales Order',s.name,'workflow_state','Approved')

@frappe.whitelist()
def update_job_card_raw_material_for_operation(sequence_id, work_order, se_doc, source_warehouse, produced_qty):
	wo_doc = frappe.get_doc("Work Order", work_order)
	for row in wo_doc.required_items:
		actual_required = row.required_qty / wo_doc.qty
		converted_raw_material_qty = actual_required * produced_qty
		se_doc.append('items', {
			"s_warehouse": source_warehouse,
			"item_code": row.item_code,
			"qty": converted_raw_material_qty,
			"basic_rate": 12,
			"expense_account": "Stock Adjustment - WAIP",
			"valuation_rate": 12,
			"allow_zero_valuation_rate": 1,
		})

@frappe.whitelist()
def get_workstation_details(bom, operation):
	dict_list = []
	checked_workstations = []
	operation = frappe.get_doc("Operation", operation)
	defaut_workstation = operation.workstation
	default_workstation_code = frappe.db.get_value("Workstation", operation.workstation, "custom_machine_code")
	for row in operation.custom_other_workstations:
		workstation_code = frappe.db.get_value("Workstation", row.workstation, "custom_machine_code")
		department = frappe.db.get_value("Workstation", row.workstation, "custom_department")
		make = frappe.db.get_value("Workstation", row.workstation, "custom_make")
		dict_list.append(frappe._dict({
			"check_box": 1,
			"workstation": row.workstation,
			"workstation_code": workstation_code,
			"make": make,
			"department": department
		}))
		checked_workstations.append(row.workstation)
	checked_workstations.append(defaut_workstation)
	unchecked_workstations = frappe.db.get_all("Workstation", {"name": ["not in", checked_workstations]}, ["name", "custom_machine_code", "custom_department", "custom_make"])
	for row in unchecked_workstations:
		dict_list.append(frappe._dict({
			"check_box": 0,
			"workstation": row.name,
			"workstation_code": row.custom_machine_code,
			"make": row.custom_make,
			"department": row.custom_department
		}))

	return dict_list, defaut_workstation, default_workstation_code

@frappe.whitelist()
def update_workstation_in_operation(operation, default_workstation, other_workstations):
	op = frappe.get_doc("Operation", operation)
	op.workstation = default_workstation
	op.set("custom_other_workstations", [])

	if isinstance(other_workstations, str):
		other_workstations = json.loads(other_workstations)

	for workstation in other_workstations:
		op.append("custom_other_workstations", {
			"workstation": workstation,
		})

	op.save(ignore_permissions=True)
	return "Updated successfully"


@frappe.whitelist()
def get_tool_details(operation):
	dict_list = []
	checked_tools = []

	workstation_list = []
	default_workstation = frappe.db.get_value("Operation", operation, "workstation")
	if default_workstation:
		workstation_list.append(default_workstation)

	workstations = frappe.db.sql("""
		SELECT workstation
		FROM `tabWorkstation Child`
		WHERE parent = %s	  
	""", (operation,), as_dict=1)

	for row in workstations:
		workstation_list.append(row.workstation)

	tools = frappe.db.get_all(
		"Tool",
		filters={
			"operation": operation,
			"workstation": ["in", workstation_list],
			"status": "Active"
		},
		fields=["name", "tool_name", "tool_life", "total_processed_qty", "custom_balance_life"]
	)

	for row in tools:
		dict_list.append(frappe._dict({
			"check_box": 1,
			"tool": row.name,
			"tool_name": row.tool_name,
			"tool_life": row.tool_life,
			"total_processed_qty": row.total_processed_qty,
			"custom_balance_life": row.custom_balance_life,
		}))

	return dict_list


@frappe.whitelist()
def update_ot_details_table_in_dept():
	departments = frappe.db.get_all('Department', {"disabled": 0}, ['name'])
	for dept in departments:
		dept_doc = frappe.get_doc('Department', dept.name)
		updated = False
		for row in dept_doc.custom_ot_details:
			if row.allowed_ot != 0:
				row.allowed_ot = 0
				updated = True
		if updated:
			dept_doc.save(ignore_permissions=True)
			frappe.db.commit()


@frappe.whitelist()
def complete_last_job_card(doc, method):
	"""If all the quantities are accepted from the Quality Inspection, 
	close the Job Card"""

	manufactured_qty = frappe.db.get_value("Job Card", doc.reference_name, "total_completed_qty")
	accepted_qty_qc = frappe.db.sql("""
		SELECT sum(custom_accepted_qty)
		FROM `tabQuality Inspection`
		WHERE reference_name = %s 
			AND docstatus = 1
			AND status = 'Accepted'
		""", (doc.reference_name))[0][0] or 0
	
	total_accepted_qty = flt(accepted_qty_qc) + flt(doc.custom_accepted_qty)
	if flt(manufactured_qty) == flt(total_accepted_qty):
		if frappe.db.exists("Job Card", doc.reference_name):
			job_card = frappe.get_doc("Job Card", doc.reference_name)
			job_card.docstatus = 1
			job_card.status = "Completed"
			wo_sequence = frappe.db.count("Work Order Operation", {"parent": job_card.work_order})
			if job_card.sequence_id == wo_sequence:
				production_plan = frappe.db.get_value("Work Order", {"name": job_card.work_order}, "production_plan")
				next_wo = frappe.db.get_all(
					"Work Order",
					filters={
						"creation": [">", job_card.creation],
						"production_item": job_card.production_item,
						"production_plan": ["!=", production_plan] 
					},
					fields=["name", "production_item", "creation"],
					order_by="creation asc",
					limit=1
				)
				if next_wo:
					next_wo_name = next_wo[0].name
					frappe.db.set_value("Work Order", next_wo_name, "custom_is_queued", 0)
					frappe.db.set_value("Job Card", {"work_order": next_wo_name}, "custom_is_queued", 0)
			else:
				next_jc = frappe.db.get_all(
					"Job Card",
					filters={
						"creation": [">", job_card.creation],
						"production_item": job_card.production_item,
						"work_order": ["!=", job_card.work_order] 
					},
					fields=["name"],
					order_by="creation asc",
					limit=1
				)
				if next_jc:
					next_jc_name = next_jc[0].name
					frappe.db.set_value("Job Card", {"name": next_jc_name}, "custom_is_queued", 0)
			job_card.save(ignore_permissions=True)

@frappe.whitelist()
def get_rack_query(doctype, txt, searchfield, start, page_len, filters):
	item_code = filters.get("item_code")
	conditions = ""
	if item_code:
		conditions += f" AND parent = {item_code}"
	return frappe.db.sql(f"""
		SELECT rack
		FROM `tabItem Rack`
		WHERE {searchfield} LIKE %(txt)s
		AND parent = %(parent)s
		ORDER BY name ASC
		LIMIT %(start)s, %(page_len)s
	""", {
		"txt": f"%{txt}%",
		"parent": item_code,
		"start": start,
		"page_len": page_len
	})

# @frappe.whitelist()
# def get_rack_location_query(doctype, txt, searchfield, start, page_len, filters):
# 	item_code = filters.get("item_code")
# 	conditions = ""
# 	if item_code:
# 		conditions += f" AND parent = {item_code}"
# 	return frappe.db.sql(f"""
# 		SELECT location
# 		FROM `tabItem Rack Location`
# 		WHERE {searchfield} LIKE %(txt)s
# 		AND parent = %(parent)s
# 		ORDER BY name ASC
# 		LIMIT %(start)s, %(page_len)s
# 	""", {
# 		"txt": f"%{txt}%",
# 		"parent": item_code,
# 		"start": start,
# 		"page_len": page_len
# 	})
 

@frappe.whitelist()
def get_rack_location_query(doctype, txt, searchfield, start, page_len, filters):
	item_code = filters.get("item_code")
	rack = filters.get("rack")
	conditions = ""
	if item_code:
		conditions += f" AND parent = {item_code}"
	return frappe.db.sql(f"""
		SELECT location
		FROM `tabItem Rack`
		WHERE {searchfield} LIKE %(txt)s
		AND parent = %(parent)s
		AND rack = %(rack)s
		ORDER BY name ASC
		LIMIT %(start)s, %(page_len)s
	""", {
		"txt": f"%{txt}%",
		"parent": item_code,
		"rack":rack,
		"start": start,
		"page_len": page_len
	}) 

# @frappe.whitelist()
# def check_lr_status(doc,method):
#     if doc.custom_invoice_type=='Export Invoice':
#         if frappe.db.exists('Logistics Request',{'order_no':doc.name}):
#             lr=frappe.get_doc('Logistics Request',{'order_no':doc.name})
#             if doc.custom_logistics_status=='Request Approved':
#                     lr.status='Scheduled'
#                     lr.save(ignore_permissions=True)
#             else:
#                 frappe.throw('Logistcs Request is not Approved.To submit invoice, approval need in LR.')
#         else:
#             frappe.throw('Logistcs Request is not created.For export invoice logistics request need.')
	   

@frappe.whitelist()
def update_the_changes_to_lr(doc, method=None):
	"""If changes made in Sales Invoice,
	update the changes made in the Logistics Request
	if the LR Status (Sales Invoice) is None or Pending Export"""
	
	if isinstance(doc, str):
		doc = frappe.parse_json(doc)
	if isinstance(doc, dict):
		doc = frappe.get_doc(doc)

	lr_name = frappe.db.exists("Logistics Request", {"order_no": doc.name})
	if not lr_name:
		return

	if not doc.custom_lr_status or doc.custom_lr_status == "Pending Export":
		lr = frappe.get_doc("Logistics Request", {"order_no": doc.name})

		lr.set("product_description_so", [])

		for si_item in doc.items:
			row = {}
			for field in si_item.as_dict():
				if field in [d.fieldname for d in lr.meta.get_field("product_description_so").options and frappe.get_meta("Sales Invoice Item").fields]:
					row[field] = si_item.get(field)
			lr.append("product_description_so", row)

		result = frappe.db.sql(
			"""
			SELECT SUM(total_weight) as total 
			FROM `tabSales Invoice Item` 
			WHERE parent = %s
			""",
			(doc.name,),
			as_dict=True,
		)
		net_weight = result[0].total or 0

		lr.grand_total = doc.grand_total
		lr.cbm = doc.custom_total_cbm
		lr.gross_wt = doc.custom_total_gross_weight
		lr.seacontainertype = doc.seacontainertype
		lr.net_wt = net_weight
		lr.save(ignore_permissions=True)
		frappe.db.commit()

def validate_item_rate(doc, method):
	"""Do not allow to save Purchase Order with rate 0"""
	
	for row in doc.items:
		row_idx = "Row " + str(row.idx)
		if not row.rate:
			frappe.throw(
				_("{0}: Rate cannot be set to zero for the Item {1}")
				.format(frappe.bold(row_idx), frappe.bold(row.item_code))
			)
			frappe.throw(
				_("{0}: Item {1} க்கான Rate 0 ஆக இருக்க கூடாது")
				.format(frappe.bold(row_idx), frappe.bold(row.item_code))
			)


@frappe.whitelist()
def update_quality_inspection_age():
	quality_inspections = frappe.get_all("Quality Inspection",{'docstatus':0},['name'])

	for qi in quality_inspections:
		quality_ins = frappe.get_doc("Quality Inspection",qi.name)
		quality_ins.custom_age = date_diff(today(),quality_ins.report_date)
		quality_ins.save(ignore_permissions=True)



@frappe.whitelist()
def get_rejection_reason(doctype, txt, searchfield, start, page_len, filters): 
	if isinstance(filters, str):
		try:
			filters = json.loads(filters)
		except Exception:
			filters = {}

	rejection_category = filters.get("rejection_category") 
	
	if not rejection_category:
		return []
	
	
	return frappe.db.sql("""
		SELECT DISTINCT rejres.name
		FROM `tabRejection Reason` rejres
		INNER JOIN `tabRejection Reason Applicable` app
			ON app.parent = rejres.name
		WHERE app.rejection_category = %s
		  AND rejres.{sf} LIKE %s
		ORDER BY rejres.name
		LIMIT %s, %s
	""".format(sf=searchfield), (rejection_category, "%" + txt + "%", start, page_len))


@frappe.whitelist()
def get_rework_reason(doctype, txt, searchfield, start, page_len, filters): 
	if isinstance(filters, str):
		try:
			filters = json.loads(filters)
		except Exception:
			filters = {}

	rework_category = filters.get("rework_category") 
	
	if not rework_category:
		return []
	
	
	return frappe.db.sql("""
		SELECT DISTINCT rewres.name
		FROM `tabRework Reason` rewres
		INNER JOIN `tabRework Reason Applicable` app
			ON app.parent = rewres.name
		WHERE app.rework_category = %s
		  AND rewres.{sf} LIKE %s
		ORDER BY rewres.name
		LIMIT %s, %s
	""".format(sf=searchfield), (rework_category, "%" + txt + "%", start, page_len))

@frappe.whitelist()
def change_workflow_state(doc, method):
	if doc.order_no:
		sales_invoice = frappe.get_doc("Sales Invoice", doc.order_no)
		if sales_invoice.workflow_state == "Draft":
			frappe.db.set_value("Sales Invoice", sales_invoice.name, "workflow_state", "Pending for LR")
			frappe.db.commit()
			sales_invoice.reload()
@frappe.whitelist()
def update_workflow(doctype,name,transporter):
	sales_invoice =  frappe.get_doc("Sales Invoice", name)
	if sales_invoice.workflow_state == "Pending for LR":
		frappe.db.set_value("Sales Invoice", sales_invoice.name, "workflow_state", "LR Approved")
		frappe.db.set_value("Sales Invoice", sales_invoice.name, "transporter", transporter)
		frappe.db.set_value("Sales Invoice", sales_invoice.name,"custom_logistics_status", 'Request Approved')
		frappe.db.commit()
		sales_invoice.reload()
	
	


@frappe.whitelist()
def check_approved_lr(doc,method):
	if doc.workflow_state == "Pending for LR":
		if doc.custom_logistics_status!='Request Created':
			frappe.throw('Logistic Request not yet created.Kindy create request for further process.')
	if doc.workflow_state == "LR Approved":
		if frappe.db.exists('Logistics Request',{"order_no":doc.name}):
			if doc.custom_logistics_status!='Request Approved':
				frappe.throw('Logistic Request not yet Approved.Kindy get approval for further process.')
		else:
			frappe.throw('Logistic Request not yet created.Kindy create request for further process.')
	

def set_auto_name_for_time_logs(doc, method):
	for row in doc.time_logs:
		if not row.name.startswith("JC-ENT-"):
			row.name = frappe.model.naming.make_autoname("JC-ENT-.######")
			row.custom_docname = row.name

def set_docname_for_time_logs(doc, method):
	"""On every validation, the docname will be
	updated by the child's name"""
	for row in doc.time_logs:
		if not row.custom_docname:
			row.custom_docname = row.name

def create_quality_pending(args, item_code):
	"""Create Quality Pending on every Job Card entry"""
	
	if not frappe.db.exists("Item", {"item_billing_type": "Billing", "item_code": item_code}):
		return
	if args.accepted_qty > 0:
		quality_pending = frappe.db.get_value("Job Card", args.job_card_id, "custom_quality_pending")
		if frappe.db.exists("Quality Pending", quality_pending):
			qp = frappe.get_doc("Quality Pending", quality_pending)
			qp.possible_inspection_qty += args.accepted_qty
			qp.inspection_pending_qty += args.accepted_qty
			qp.save(ignore_permissions=True)
		else:
			qp = frappe.new_doc("Quality Pending")
			qp.job_card = args.job_card_id
			qp.datetime = now()
			qp.reference_type = "Job Card"
			qp.reference_name = args.job_card_id
			qp.possible_inspection_qty = args.accepted_qty
			qp.inspection_pending_qty = args.accepted_qty
			qp.inspection_completed_qty = 0
			qp.inspection_pending_type = "In Process"
			qp.save(ignore_permissions=True)

@frappe.whitelist()
def quality_inspection_state_change(quality_inspection):
	
	doc = frappe.get_doc("Quality Inspection", quality_inspection)
	if doc.is_new():
		return
	if doc.workflow_state != "Draft":
		return  
	
	if  flt(doc.custom_inspected_qty) > 0:
		for reading in doc.readings or []:
			if reading.reading_1 is None or reading.reading_2 is None or reading.reading_3 is None or reading.reading_4 is None or reading.reading_5 is None :
				return
			
		if doc.status=="Accepted":
			apply_workflow(doc,"Submit")
		else:
			if flt(doc.custom_accepted_qty) == 0:
				apply_workflow(doc,"Submit")
			else:
				apply_workflow(doc,"Send for Approval")

		return "ok"


@frappe.whitelist()
def get_possible_inspection_qty(job_card_entry):
	"""
	Return the possible inspection quantity for a given Job Card Entry (custom_docname).
	"""
	qty = frappe.db.get_value("Job Card Time Log", {"custom_docname": job_card_entry}, "completed_qty")
	return qty or 0


# @frappe.whitelist()
# def create_supplier_warehouse_for_existing_documents():
#     supplier_docs = frappe.db.get_all('Supplier', filters={'disabled': 0}, fields=['name'])
#     for doc in supplier_docs:
#         supplier_name = doc.get('name')
#         if supplier_name and not frappe.db.exists('Warehouse', {'name': supplier_name}):
#             warehouse_doc = frappe.new_doc('Warehouse')
#             warehouse_doc.warehouse_name = supplier_name
#             warehouse_doc.company = 'WONJIN AUTOPARTS INDIA PVT.LTD.'
#             warehouse_doc.parent_warehouse = 'Supplier Warehouse - WAIP'

#             warehouse_doc.save(ignore_permissions=True)
#             frappe.db.commit()

			
from frappe.utils.jinja import render_template

@frappe.whitelist()
def create_html_jobcard(doc):
	doc = frappe.parse_json(doc)
	template = """
		<table class="mb-5">
		<tr>
			<td rowspan=5>
				<img width="100px" src="https://api.qrserver.com/v1/create-qr-code/?size=100x100&amp;data={{ doc.custom_scan_barcode }}" alt="QR Code" />
			</td>
			<td><b>&nbsp;Date: {{ doc.posting_date }}</b></td>
		</tr>
		<tr>
			<td><b>&nbsp;Job Card: {{ doc.name }}</b></td>
		</tr>
		<tr>
			<td><b>&nbsp;Item: {{ doc.production_item }} - {{ doc.item_name }}</b></td>
		</tr>
		<tr>
			<td><b>&nbsp;Accepted Quantity: {{ doc.custom_accepted_qty | int }}</b></td>
		</tr>
	</table>

	{% if doc.custom_rejected_qty > 0 %}
	<table class="mb-5">
		<tr>
			<td rowspan=5>
				<img width="100px" src="https://api.qrserver.com/v1/create-qr-code/?size=100x100&amp;data={{ frappe.utils.get_url('/app/job-card/' + doc.name) }}" alt="QR Code" />
			</td>
			<td><b>&nbsp;Date: {{ doc.posting_date }}</b></td>
		</tr>
		<tr>
			<td><b>&nbsp;Job Card: {{ doc.name }}</b></td>
		</tr>
		<tr>
			<td><b>&nbsp;Item: {{ doc.production_item }} - {{ doc.item_name }}</b></td>
		</tr>
		<tr>
			<td><b>&nbsp;Rejected Quantity: {{ doc.custom_rejected_qty | int }}</b></td>
		</tr>
	</table>
	{% endif %}

	{% if doc.custom_rework_waiting_qty_details > 0 %}
	<table class="mb-5">
		<tr>
			<td rowspan=5>
				<img width="100px" src="https://api.qrserver.com/v1/create-qr-code/?size=100x100&amp;data={{ frappe.utils.get_url('/app/job-card/' + doc.name) }}" alt="QR Code" />
			</td>
			<td><b>&nbsp;Date: {{ doc.posting_date }}</b></td>
		</tr>
		<tr>
			<td><b>&nbsp;Job Card: {{ doc.name }}</b></td>
		</tr>
		<tr>
			<td><b>&nbsp;Item: {{ doc.production_item }} - {{ doc.item_name }}</b></td>
		</tr>
		<tr>
			<td><b>&nbsp;Rework Quantity: {{ doc.custom_rework_waiting_qty_details | int }}</b></td>
		</tr>
	</table>
	{% endif %}
		
	"""
	html = render_template(template, {"doc": doc})
	return {"html": html}



import frappe
import json
from datetime import datetime
from frappe.utils.jinja import render_template
@frappe.whitelist()
def create_html_mr(doc):

	# Parse JSON safely
	doc = json.loads(doc)

	# Handle datetime parsing gracefully
	if "creation" in doc:
		try:
			doc["creation"] = datetime.strptime(doc["creation"], "%Y-%m-%d %H:%M:%S.%f")
		except ValueError:
			doc["creation"] = datetime.strptime(doc["creation"], "%Y-%m-%d %H:%M:%S")

	# Extract items list
	items = doc.get("items") or []

	# Jinja Template
	template = """

	<style>
		td { font-size: 13px; }
		.dialog-page-wrapper {
			border: 1px solid black;
			padding-top: 15px;
			padding-right: 15px;
			padding-left: 15px;
			padding-bottom: 8px;
			border-radius: 10px;
		}
		.content-wrapper { flex: 1; }
		.item-table {
			width: 100%;
			border-collapse: collapse;
			margin-top: -5px;
		}
		.item-table th, .item-table td {
			border: 1px solid black;
			padding: 4px;
			text-align: center;
		}
		.item-table td.right-align { text-align: right; }
		.signature-footer {
			margin-top: auto;
			width: 100%;
			border-collapse: collapse;
		}
		.signature-footer td {
			font-weight: 700;
			margin-bottom: -80px;
		}
	</style>

	{% set total_items = items|length %}
	{% for i in range(0, total_items, 25) %}
		{% set batch = items[i:i+25] %}
		{% set last_index = total_items - 1 %}
		{% set is_last_batch = (i + batch|length) == total_items %}

		<div class="dialog-page-wrapper">
			<div class="content-wrapper">
				{% if i == 0 %}
					<!-- Header (only on first page) -->
					<table width="100%" style="border-bottom-style: hidden; margin-bottom:5px; border: none;">
						<tr style="border: solid 1px black;">
							<td><img src="/files/ci_img.png" width="120px"></td>
							<td style="text-align: center; font-weight: 700; font-size: 18px;">MATERIAL REQUEST NOTE</td>
							<td style="text-align: right;">DOCUMENT NO : STR\\R\\02</td>
						</tr>
					</table>

					<!-- MR Info -->
					<table style="width:100%; border-collapse: collapse; margin-top:-6px; border-top-style: hidden;">
						<tr>
							<td style="width:65%; border:1px solid black; padding:5px; border-bottom:none; border-right:none; font-size:12px;">
								<p style="line-height:1">
									<span style="display:inline-block; width:90px;">Department</span>: {{ doc.custom_department or '' }}<br><br>
									<span style="display:inline-block; width:90px;">Requested By</span>: <span style="font-size:12px;">{{ doc.custom_requested_by or '' }} - {{ doc.custom_requester_name or '' }}</span><br><br>
									<span style="display:inline-block; width:90px;">Warehouse</span>: {{ doc.set_from_warehouse or '' }}
								</p>
							</td>
							<td style="width:35%; border:1px solid black; padding:5px; border-bottom:none; border-left:none; font-size:12px;">
								<p style="line-height:1;">
									<span style="display:inline-block; width:80px;">MR No</span>: {{ doc.name }}<br><br>
									<span style="display:inline-block; width:80px;">MR Date</span>: {{ frappe.utils.format_date(doc.transaction_date, "dd-MM-yyyy") or '' }}
								</p>
							</td>
						</tr>
					</table>
				{% else %}
					<div style="page-break-before: always;"></div>
				{% endif %}

				<!-- Items Table -->
				<table class="item-table">
					<tr style="background-color: #fec76f;">
						<th style="font-size: 12px;">S.No.</th>
						<th style="font-size: 12px;">Item Description</th>
						<th style="font-size: 12px;">MR Req Qty</th>
						<th style="font-size: 12px;">Shop Floor Qty</th>
						<th style="font-size: 12px;">Today Req Qty</th>
					</tr>

					{% for row in items %}
					<tr>
						<td style="font-size: 12px;">{{ loop.index }}</td>
						<td style="font-size: 12px; text-align: left;">{{ row.item_code or '' }} - {{ row.item_name or '' }}</td>
						<td class="right-align" style="font-size: 12px;">{{ row.custom_total_req_qty or 0 }}</td>
						<td class="right-align" style="font-size: 12px;">{{ row.custom_shop_floor_stock or 0 }}</td>
						<td class="right-align" style="font-size: 12px;">{{ row.custom_requesting_qty or 0 }}</td>
					</tr>
					{% endfor %}

					<!-- ✅ Total row INSIDE same table -->
					<tr>
						<td colspan="4" style="font-weight:bold; text-align:center; font-size:12px; border-top:1px solid black;">
							Total
						</td>
						<td style="font-weight:bold; text-align:right; font-size:12px; border-top:1px solid black;">
							{{ items | sum(attribute='custom_requesting_qty') }}
						</td>
					</tr>

				</table>

			</div>
		
		<div style="border: 1px solid black; border-top: none; padding-top: 30px;">
			{% set prepared_signature = frappe.db.get_value("Employee", {"user_id": doc.owner}, "custom_digital_signature") if doc.owner else None %}
			{% set prepared_by_name = frappe.db.get_value("User", doc.owner, "full_name") if doc.owner else None %}
			{% set hod_signature = frappe.db.get_value("Employee", {"user_id": doc.custom_hod_approved_by}, "custom_digital_signature") if doc.custom_hod_approved_by else None %}
			{% set erp_team_signature = frappe.db.get_value("Employee", {"user_id": doc.custom_erp_team}, "custom_digital_signature") if doc.custom_erp_team else None %}
			
			<div style="display: flex; flex-direction: row; justify-content: space-around; text-align: center; margin-bottom: 10px;">
				<div class="d-flex flex-column">
					{% if prepared_signature %}
						<img src="{{ prepared_signature }}" style="height:40px;">
					{% else %}
						<b style="font-size: 12px; height: 40px;">{{prepared_by_name}}</b>
					{% endif %}
					{% if doc.creation %}
					<b style="font-size: 10px;">{{ frappe.utils.format_datetime(doc.creation, "dd-MM-yyyy HH:mm:ss") or '' }}</b>
					{% else %}
					<b style="font-size: 10px;">&nbsp;</b>
					{% endif %}
					<b>Prepared</b>
				</div>

				<div class="d-flex flex-column ">
					{% if hod_signature %}
						<img src="{{ hod_signature }}" style="height:40px; 	">
					{% else %}
						<img src="/files/Screenshot 2025-09-22 141452.png" style=" height:40px;">
					{% endif %}
					{% if doc.custom_hod_approved_on %}
						<b style="font-size: 10px;">{{ frappe.utils.format_datetime(doc.custom_hod_approved_on, "dd-MM-yyyy HH:mm:ss") or '' }}</b>
					{% else %}
						<b style="font-size: 10px;">&nbsp;</b>
					{% endif %}
					<b>HOD</b>
				</div>

				<div class="d-flex flex-column">
					{% if erp_team_signature %}
						<img src="{{ erp_team_signature }}" style="height:40px;">
					{% else %}
						<img src="/files/Screenshot 2025-09-22 141452.png" style="height:40px;">
					{% endif %}
					{% if doc.custom_erp_team_approved_on %}
						<b style="font-size: 10px;">{{ frappe.utils.format_datetime(doc.custom_erp_team_approved_on, "dd-MM-yyyy HH:mm:ss") or '' }}</b>
					{% else %}
						<b style="font-size: 10px;">&nbsp;</b>
					{% endif %}
					<b>ERP Team</b>
				</div>

			</div>

		</div>
	<div style="text-align:left; font-size:10px;margin-top:5px;font-weight:normal;">
		Note: This document is Digitally Signed
	</div>
	</div>
	{% endfor %}

	
	"""

	html = render_template(template, {"doc": doc, "items": items})
	return {"html": html}


# @frappe.whitelist()
# def create_html_mr(doc):
# 	# Parse JSON safely
# 	doc = json.loads(doc)

# 	# Handle datetime parsing gracefully
# 	if "creation" in doc:
# 		try:
# 			doc["creation"] = datetime.strptime(doc["creation"], "%Y-%m-%d %H:%M:%S.%f")
# 		except ValueError:
# 			doc["creation"] = datetime.strptime(doc["creation"], "%Y-%m-%d %H:%M:%S")

# 	# Extract items list
# 	items = doc.get("items") or []

# 	# Jinja Template
# 	template = """

# 	<style>
# 		td { font-size: 13px; }
# 		.page-wrapper {
# 			border: 1px solid black;
# 			padding: 10px;
# 			border-radius: 10px;
# 		}
# 		.content-wrapper { flex: 1; }
# 		.item-table {
# 			width: 100%;
# 			border-collapse: collapse;
# 			margin-top: -5px;
# 		}
# 		.item-table th, .item-table td {
# 			border: 1px solid black;
# 			padding: 4px;
# 			text-align: center;
# 		}
# 		.item-table td.right-align { text-align: right; }
# 		.signature-footer {
# 			margin-top: auto;
# 			width: 100%;
# 			border-collapse: collapse;
# 		}
# 		.signature-footer td {
# 			font-weight: 700;
# 			margin-bottom: -80px;
# 			border-top: 1px solid black;
# 		}
# 	</style>

# 	{% set total_items = items|length %}
# 	{% for i in range(0, total_items, 25) %}
# 		{% set batch = items[i:i+25] %}
# 		{% set last_index = total_items - 1 %}
# 		{% set is_last_batch = (i + batch|length) == total_items %}

# 		<div class="page-wrapper">
# 			<div class="content-wrapper">
# 				{% if i == 0 %}
# 					<!-- Header (only on first page) -->
# 					<table width="100%" style="border-bottom-style: hidden; margin-bottom:5px; border: none;">
# 						<tr style="border: solid 1px black;">
# 							<td><img src="/files/ci_img.png" width="120px"></td>
# 							<td style="text-align: center; font-weight: 700; font-size: 18px;">MATERIAL REQUEST NOTE</td>
# 							<td style="text-align: right;">DOCUMENT NO : STR\\R\\02</td>
# 						</tr>
# 					</table>

# 					<!-- MR Info -->
# 					<table style="width:100%; border-collapse: collapse; margin-top:-4px;">
# 						<tr>
# 							<td style="width:65%; border:1px solid black; padding:5px; border-bottom:none; border-right:none; font-size:11px;">
# 								<p style="line-height:1">
# 									<span style="display:inline-block; width:90px;">Department</span>: {{ doc.custom_department or '' }}<br><br>
# 									<span style="display:inline-block; width:90px;">Requested By</span>: <span style="font-size:9px;">{{ doc.custom_requested_by or '' }} - {{ doc.custom_requester_name or '' }}</span><br><br>
# 									<span style="display:inline-block; width:90px;">Warehouse</span>: {{ doc.set_from_warehouse or '' }}
# 								</p>
# 							</td>
# 							<td style="width:35%; border:1px solid black; padding:5px; border-bottom:none; border-left:none; font-size:11px;">
# 								<p style="line-height:1;">
# 									<span style="display:inline-block; width:80px;">MR No</span>: {{ doc.name }}<br><br>
# 									<span style="display:inline-block; width:80px;">MR Date</span>: {{ frappe.utils.format_date(doc.transaction_date, "dd-MM-yyyy") or '' }}
# 								</p>
# 							</td>
# 						</tr>
# 					</table>
# 				{% else %}
# 					<div style="page-break-before: always;"></div>
# 				{% endif %}

# 				<!-- Items Table -->
# 				<table class="item-table">
# 					<tr>
# 						<th style="font-size: 9px;">S.No.</th>
# 						<th style="font-size: 9px;">Item Description</th>
# 						<th style="font-size: 9px;">MR Req Qty</th>
# 						<th style="font-size: 9px;">Shop Floor Qty</th>
# 						<th style="font-size: 9px;">Today Req Qty</th>
# 					</tr>
# 					{% for row in batch %}
# 					<tr>
# 						<td style="font-size: 9px;">{{ loop.index + i }}</td>
# 						<td style="font-size: 9px; text-align: left;">{{ row.item_code or '' }} - {{ row.item_name or '' }}</td>
# 						<td class="right-align" style="font-size: 9px;">{{ row.custom_requesting_qty or 0 }}</td>
# 						<td class="right-align" style="font-size: 9px;">{{ row.custom_shop_floor_stock or 0 }}</td>
# 						<td class="right-align" style="font-size: 9px;">{{ row.custom_today_req_qty or 0 }}</td>
# 					</tr>
# 					{% endfor %}
# 				</table>

# 				{% if is_last_batch %}
# 				<table style="width:100%">
# 					<tr>
# 						<td colspan="4" style="font-weight:bold; text-align:center; font-size:10px; border-top:1px solid black; border-bottom:1px solid black;">Total</td>
# 						<td style="font-weight:bold; text-align:right; font-size:9px; border-top:1px solid black; border-bottom:1px solid black;">
# 							{{ items | sum(attribute='custom_today_req_qty') }}
# 						</td>
# 					</tr>
# 				</table>
# 				{% endif %}
# 			</div><br><br><br><br>
# 		<table class="signature-footer content-wrapper">

# 		{% set prepared_signature = frappe.db.get_value("Employee", {"user_id": doc.owner}, "custom_digital_signature") if doc.owner else None %}
# 		{% set hod_signature = frappe.db.get_value("Employee", {"user_id": doc.custom_hod}, "custom_digital_signature") if doc.custom_hod else None %}
# 		{% set plant_signature = frappe.db.get_value("Employee", {"user_id": doc.custom_erp_team}, "custom_digital_signature") if doc.custom_erp_team else None %}

# 		<tr>
# 			<td style="text-align:center; border-right:none;font-size:9px;border-bottom:none;"></td>
# 			<td style="text-align:center; border-left:none; border-right:none;font-size:9px;border-bottom:none;"></td> 
# 			<td style="text-align:center; border-left:none;font-size:9px;border-bottom:none;"></td>
# 		</tr>
# 		<tr>
# 			<td style="text-align:center; border-right:none;border-bottom:hidden;padding:top:8px;border-top:none;">
# 			{% if prepared_signature %}
# 			<img src="{{ prepared_signature }}" style="height:30px;">
# 			{%else%}
# 			<img src="/files/Screenshot 2025-09-22 141452.png" style="height:40px;">
# 			{% endif %}
# 			</td>
# 			<td style="text-align:center; border-right:none;border-bottom:hidden;padding:top:20px;border-top:none;">
# 			{% if hod_signature %}
# 			<img src="{{ hod_signature }}" style="height:30px;">
# 			{%else%}
# 			<img src="/files/Screenshot 2025-09-22 141452.png" style="height:40px;">
# 			{% endif %}
# 			</td>
# 			<td style="text-align:center; border-right:none;border-bottom:hidden;padding:top:8px;border-top:none;">
# 			{% if plant_signature %}
# 			<img src="{{ plant_signature }}" style="height:30px;">
# 			{%else%}
# 			<img src="/files/Screenshot 2025-09-22 141452.png" style="height:40px;">
# 			{% endif %}
# 			</td>
# 		</tr>
# 		<tr>
# 			<td style="text-align:center; border-right:none;font-size:9px;border-bottom:none;">{{ frappe.utils.format_datetime(doc.creation, "dd-MM-yyyy HH:mm:ss") or '' }}</td>
# 			<td style="text-align:center; border-left:none; border-right:none;font-size:9px;border-bottom:none;">{{ frappe.utils.format_datetime(doc.custom_hod_approved_on, "dd-MM-yyyy HH:mm:ss") or '' }}</td> 
# 			<td style="text-align:center; border-left:none;font-size:9px;border-bottom:none;">{{ frappe.utils.format_datetime(doc.custom_erp_team_approved_on, "dd-MM-yyyy HH:mm:ss") or '' }}</td>
# 		</tr>
# 		<tr>
# 			<td style="text-align:center; border-right:none;border-top:none;">Prepared</td>
# 			<td style="text-align:center; border-left:none; border-right:none;border-top:none;">HOD</td> 
# 			<td style="text-align:center; border-left:none;border-top:none;">ERP Team</td>
# 		</tr>
# 		<tr>
# 			<td colspan="3" style="border-top:none;">
# 			<div style="text-align:left; font-size:12px;margin-top:5px;font-weight:normal;">
# 				Note: This document is Digitally Signed
# 			</div>
# 		</td>
# 		</tr>
# 	</table>
# 		</div>
# 	{% endfor %}

	
# 	"""

# 	html = render_template(template, {"doc": doc, "items": items})
# 	return {"html": html}


@frappe.whitelist()
def create_html_lr(doc):
	doc = frappe.parse_json(doc)
	template = """
	<style>
	td {
	border: 1px solid black;
	
}
 table {
	width: 100%;
	border-collapse: collapse;
	font-size: 12px;
	font-family: 'Book Antiqua','Palatino Linotype',Palatino,serif;
  }
  td {
	border: 1px solid #000;
  }
  .label-cell {
	font-weight: bold;
	width: 10%;
	white-space: nowrap;
	line-height:0.9;
	border-right: none; /* remove right border */
	border-top: none !important;
	border-bottom: none !important;
	font-size:14px;
  }
  .label-text {
	display: inline-block;
	line-height:0.9;
	width: 100px; /* adjust for colon alignment */
	
  }
  .value-cell {
	text-align: left;
	line-height:0.9;
	width: 30%;
	border-left: none; /* remove left border */
	border-top: none !important;
	border-bottom: none !important;
	font-size:14px;
  }
  .no-border {
	border: none !important;
  }
 .remove-top-border {
		border-top: none !important;
	}
	
.remove-bottom-border {
	border-bottom: none !important;
}
.no-top-bottom {
	border-top: none !important;
	border-bottom: none !important;
	border-left: 1px solid black !important;   /* keep left if you want */
	border-right: 1px solid black !important;  /* keep right if you want */
	padding-top: 0px !important;
  padding-bottom: 0px !important;
	
}
	</style>
		<meta name="pdfkit-orientation" content="landscape" />
<meta name="page-size" content ='A4'/>
<table border="1" cellspacing="0" cellpadding="5" class="remove-bottom-border;">
  <tr>
	<td colspan=1 style=" border-right: none !important;">&nbsp;<img src="/files/ci_img.png" width="120px" height=35%;>
				 
			</td>
	<td colspan=9 style="border-left: none !important;font-size: 14px; font-family: 'Book Antiqua', 'Palatino Linotype', Palatino, serif; ">
						<div style="padding-left:250px;"><b>LOGISTICS REQUEST APPROVAL</b></div>
	</td>
  </tr>
  <tr>
	<td class="label-cell"><span class="label-text">Department</span>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; :</td>
	<td colspan="4" style="border-right: none;" class="value-cell">LOGISTICS</td>
	<td class="label-cell" style="border-left: none;"><span class="label-text">Request Date</span>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; :</td>
	<td colspan="4" class="value-cell">{{ frappe.utils.format_date(doc.creation) }}</td>
  </tr>

  <tr>
	<td class="label-cell"><span class="label-text">Requester Name</span> &nbsp;&nbsp;&nbsp;&nbsp; :</td>
	{% set name = frappe.db.get_value("Employee", {"user_id": doc.owner}, ["employee_name"]) %}
	<td colspan="4" style="border-right: none;" class="value-cell">{{ name or '' }}</td>
	<td class="label-cell" style="border-left: none;"><span class="label-text">Doc. No</span> &nbsp;&nbsp;&nbsp;&nbsp; :</td>
	<td colspan="4" class="value-cell">{{ doc.name }}</td>
  </tr>

  <tr style="border-top: 1px solid black !important; border-bottom: 1px solid black !important; ">
	<td class="label-cell"><span class="label-text" >Customer Name</span>&nbsp;&nbsp;&nbsp;&nbsp; :</td>
	<td colspan="9" class="value-cell" style="border-top: 1px solid black !important; border-bottom: 1px solid black !important;font-weight:bold; ">{{ doc.customer or "" }}</td>
  </tr>

  <tr>
	<td class="label-cell"><span class="label-text">Logistic Mode</span> &nbsp;&nbsp;&nbsp;&nbsp; :</td>
	<td colspan="9" class="value-cell">{{ doc.logistic_type or "" }}</td>
  </tr>

  <tr>
	<td class="label-cell"><span class="label-text">Scope of Delivery</span> &nbsp;&nbsp;&nbsp;&nbsp; :</td>
	<td colspan="9" class="value-cell">{{ doc.scope_of_delivery or "" }}</td>
  </tr>

  <tr>
	<td class="label-cell"><span class="label-text">Shipment Mode</span> &nbsp;&nbsp;&nbsp;&nbsp; :</td>
	<td colspan="9" class="value-cell">{{ doc.cargo_type or "" }}</td>
  </tr>

  <tr>
	<td class="label-cell"><span class="label-text">Invoice No</span> &nbsp;&nbsp;&nbsp;&nbsp; :</td>
	<td colspan="9" class="value-cell">{{ doc.order_no or "" }}</td>
  </tr>

  <tr class="remove-bottom-border;" style="border-bottom: hidden;">
	<td class="label-cell remove-bottom-border;" style="border-bottom: none !important;"><span class="label-text">Remarks</span> &nbsp;&nbsp;&nbsp;&nbsp; :</td>
	{% set name = frappe.db.get_value("Employee", {"user_id": doc.owner}, ["employee_name"]) %}
	<td colspan="9" style="border-bottom: none !important;" class="value-cell remove-bottom-border;">{{ name or '' }}</td>
  </tr>
</table>
  <div class="container remove-top-border; " style=" border-bottom: 1px solid black !important;border-left: 1px solid black !important;border-right: 1px solid black !important;">
	<div class="row remove-top-border;" style="border-bottom: none !important;">
		<div class="col-xs-12 p-0 remove-top-border;" style="font-size:10px;border-top: none !important;">
		<table class="remove-top-border;" border="1" cellspacing="0" cellpadding="5" style="width:100%; border-collapse:collapse; text-align:center;border-top: none;border-right: none;border-left: none;border-bottom: none !important;border-top: none !important;">

		<tr class="remove-top-border;" style="font-size: 13px;font-weight: bold; text-align:center; font-family: 'Book Antiqua', 'Palatino Linotype', Palatino, serif;border-left: none;border-right: none;">
			<td colspan=1 class="remove-top-border;" style="width:4%;border-left: none;">S.No</td>
			<td class="remove-top-border;" style="width:40%;">Forwarder Name</td>
			<td class="remove-top-border;" style="width:7%;">Volume in <br>details(CBM)</td>
			<td class="remove-top-border;" style="width:7%;">Invoice Value <br>INR</td>
			<td class="remove-top-border;" style="width:7%;">Freight Charges</td>
			<td class="remove-top-border;" style="width:7%;">Port/Handling Charges</td>
			<td class="remove-top-border;" style="width:7%;">CHA Charges</td>
			<td class="remove-top-border;" style="width:7%;">Total Expenses</td>
			<td class="remove-top-border;" style="width:7%;">%</td>
			<td class="remove-top-border;" style="width:7%;border-right: none;">Remarks</td>
		</tr>

{%- set max_rows = 8 -%}
{%- set row_count = (doc.ffw_quotation|length)  -%}

{%- for item in doc.ffw_quotation -%}
<tr style="font-size:13px; font-family:'Book Antiqua', 'Palatino Linotype', Palatino, serif;">
	{% set ffw_name = frappe.db.get_value("Supplier", item.ffw_name, 'supplier_name') %}
	<td style="text-align:center; border:1px solid black;border-bottom: hidden;border-left: none !important;">{{ item.idx }}</td>
	<td style="text-align:center; border:1px solid black;border-bottom: hidden;">{{ ffw_name }}</td>
	<td style="text-align:center; border:1px solid black;border-bottom: hidden;">{{ doc.cbm or "" }}</td>
	{% set tot = frappe.db.get_value("Sales Invoice", {"name": doc.order_no}, "base_grand_total") %}
	<td style="text-align:right; border:1px solid black;border-bottom: hidden;">{{ frappe.utils.fmt_money(tot) }}</td>
	<td style="text-align:right; border:1px solid black;border-bottom: hidden;">{{ item.freight_charges }}</td>
	<td style="text-align:right; border:1px solid black;border-bottom: hidden;">{{ item.porthandling_charges }}</td>
	<td style="text-align:right; border:1px solid black;border-bottom: hidden;">{{ item.cha_charges }}</td>
	<td style="text-align:right; border:1px solid black;border-bottom: hidden;">{{ frappe.utils.fmt_money(item.total_shipment_cost) }}</td>
	{% set per = ((item.quoted_value / doc.grand_total) * 100) %}
	<td style="text-align:right; border:1px solid black;border-bottom: hidden;">{{ "{:.2f}%".format(per) }}</td>
	<td style="border-right:1px solid black; border-top:1px solid black;border-bottom: hidden;"></td>
</tr>
{%- endfor -%}
{% if row_count < 3 %}
		{% set row_height = 18 %}
	{% else %}
		{% set row_height = 13 %}
	{% endif %}

{%- for i in range(row_count, max_rows) -%}

<!--<tr class="remove-bottom-border;" style="border-bottom: none !important;">-->
	<tr class="remove-bottom-border"style="font-size:10px;border-left: none !important;border: 1px solid black;{% if loop.last %}border-bottom:1px solid black;{% else %}border-bottom:none !important;{% endif %} border-top:none;height:{{ row_height }}px;">
	<td class="no-top-bottom remove-bottom-border" style="text-align:center;font-size:10px;border: 1px solid black;border-bottom: none !important;border-left: none !important;"></td>
	<td class="no-top-bottom remove-bottom-border" style="text-align:center;font-size:10px;border: 1px solid black;border-bottom: none !important;border-left: none !important;"></td>
	<td class="no-top-bottom remove-bottom-border" style="text-align:center;font-size:10px;border: 1px solid black;border-bottom: none !important;border-left: none !important;"></td>
	<td class="no-top-bottom remove-bottom-border" style="text-align:center;font-size:10px;border: 1px solid black;border-bottom: none !important;border-left: none !important;"></td>
	<td class="no-top-bottom remove-bottom-border" style="text-align:center;font-size:10px;border: 1px solid black;border-bottom: none !important;border-left: none !important;"></td>
	<td class="no-top-bottom remove-bottom-border" style="text-align:center;font-size:10px;border: 1px solid black;border-bottom: none !important;border-left: none !important;"></td>
	<td class="no-top-bottom remove-bottom-border" style="text-align:center;font-size:10px;border: 1px solid black;border-bottom: none !important;border-left: none !important;"></td>
	<td class="no-top-bottom remove-bottom-border" style="text-align:center;font-size:10px;border: 1px solid black;border-bottom: none !important;border-left: none !important;"></td>
	<td class="no-top-bottom remove-bottom-border" style="text-align:center;font-size:10px;border: 1px solid black;border-bottom: none !important;border-left: none !important;"></td>
	<td class="remove-bottom-border" style="border: 1px solid black;border-bottom:none !important;border-top: none !important;border-right:none !important;"></td>
</tr>
{%- endfor -%}


	</table>
   
	<table border=1 width=100% style="border-collapse:collapse;page-break-inside: avoid; break-inside: avoid;">
		<!--<tr class='remove-top-border;' style="border-top: none;">-->
		<!--    <td class="remove-top-border;" colspan=10 style="border-top: 1px solid black !important;border-bottom: none;">&nbsp;</td>-->
		<!--</tr>-->
		<tr style="font-size: 13px;  font-family: 'Book Antiqua', 'Palatino Linotype', Palatino, serif; text-align:left;border-left:hidden;">
			<td style="border-right: none;border-top: none;border-bottom: none;border-left:none;"><b>ETD</b>&nbsp;:&nbsp;{{frappe.format(doc.etd,{'fieldtype': 'Date'}) or ""}}</td>
			<td style="border-right: none;border-left: none;border-bottom: none;border-top: none !important;"><b>ETA</b>&nbsp;:&nbsp;{{frappe.format(doc.eta,{'fieldtype': 'Date'}) or ""}}</td>
			<td style="border-left: none;border-bottom: none;border-top: none !important;border-right:none;"><b>Transit Time</b>&nbsp;:&nbsp;{{doc.transit_time or ""}}</td>
		</tr>
		<!--<tr style="border-top: none;">-->
		<!--    <td colspan=10 style="border-top: none;border-bottom: none;">&nbsp;</td>-->
		<!--</tr>-->
	<tr style="border-top: none;border-left:hidden;border-right:hidden;">
		{%set supplier_name = frappe.db.get_value("Supplier",doc.recommended_ffw,'supplier_name') %}
			<td colspan=10 style="border-top: none;border-bottom: 1px solid black !important;font-size:13px;border-left:none;border-right:none;">Conclusion : After a several round of discussion we have finalized the deal with &nbsp;<b>{{supplier_name}}</b></td>
		</tr>
		</table>
		


		<table class="remove-top-border" border="1" cellspacing="0" cellpadding="5"
	style="width:100%;border-collapse:collapse;text-align:center;page-break-inside:avoid;break-inside:avoid;">

	<!-- HEADER ROW -->
	{% if doc.cargo_type == "Air" and doc.scope_of_delivery=="Wonjin" %} 
	<td colspan=1 class ="remove-top-border" style="width:12.5%;border-right:hidden;border:bottom:hidden;border-left:hidden;text-align:center;">Prepared By</td>
	<td colspan=1 class ="remove-top-border" style="width:12.5%;border-right:none;border:bottom:hidden;border-left:none;text-align:center;">Requested By</td> 
	<td colspan=1 class ="remove-top-border" style="width:12.5%;border-right:none;border:bottom:hidden;border-left:none;text-align:center;">Finance</td> 
	<td colspan=1 class ="remove-top-border" style="width:12.5%;border-right:none;border:bottom:hidden;border-left:none;text-align:center;">BMD</td> 
	<td colspan=1 class ="remove-top-border" style="width:12.5%;border-right:none;border:bottom:hidden;border-left:none;text-align:center;">CMD</td>
	<td colspan=1 class ="remove-top-border" style="width:12.5%;border-right:hidden;border-bottom:hidden;border-left:none;text-align:center;">SMD</td> 
	{% elif doc.cargo_type == "Sea" and doc.scope_of_delivery=="Wonjin" %} 
	<td colspan="1" class ="remove-top-border" style="width:12.5%;border-bottom:hidden;border-right:none;border-left:hidden;text-align:center;">Prepared By</td> 
	<td colspan=1 class ="remove-top-border" style="width:12.5%;border-right:none;border:bottom:hidden;border-left:none;text-align:center;">Requested By</td>
	<td colspan="2" class ="remove-top-border" style="width:12.5%;border-bottom:hidden;border-right:hidden;border-left:none;text-align:center;">Finance</td> 
	<td colspan="2" class ="remove-top-border" style="width:12.5%;border-right:none;border-left:hidden;border-bottom:hidden;text-align:center;">CMD</td>
	{% elif doc.scope_of_delivery=="Customer" %} 
	<td colspan=2 class ="remove-top-border" style="width:12.5%;border-bottom:hidden;border-right:hidden;border-left:hidden;text-align:center;">Prepared By</td> 
	<td colspan=2 class ="remove-top-border" style="width:12.5%;border-right:none;border:bottom:hidden;border-left:none;text-align:center;">Requested By</td> 
	<td colspan=2 class ="remove-top-border" style="width:12.5%;border-right:none;border-left:hidden;border-bottom:hidden;text-align:center;">Finance</td>
	{% endif %}
	</tr>
	<!-- SIGNATURE ROW -->
	<tr style="font-size:10px;font-family:'Book Antiqua','Palatino Linotype',Palatino,serif;border-bottom:hidden;border-right:hidden;border-left:hidden;">
		{% if doc.cargo_type == "Air" and doc.scope_of_delivery=="Wonjin" %}

			{% set prepared_by = frappe.db.get_value("Employee",{"user_id": doc.prepared_by},['custom_digital_signature']) %}
			<td style="height:60px;vertical-align:middle;border-top:hidden;border-right:hidden;border-left:hidden;text-align:center;">
				{% if prepared_by %}<img src="{{prepared_by}}" style="height:60px;width:35%;object-fit:contain;">{% endif %}
			</td>
			
			{% set si_owner = frappe.db.get_value("Sales Invoice",{"name": doc.order_no},['owner']) %}
			{% set requested_by = frappe.db.get_value("Employee",{"user_id": doc.owner},['custom_digital_signature']) %}
			<td style="height:60px;vertical-align:middle;border-top:hidden;border-right:hidden;border-left:hidden;text-align:center;">
				{% if requested_by %}<img src="{{requested_by}}" style="height:60px;width:35%;object-fit:contain;">{% endif %}
			</td>

			
			
			{% set finance = frappe.db.get_value("Employee", {"name": 'S0189'}, ["custom_digital_signature"]) %}
			<td style="height:60px;vertical-align:middle;border-top:hidden;border-right:hidden;border-left:hidden;text-align:center;">
				{% if finance and doc.approved_by_finance %}
					<img src="{{finance}}" style="height:60px;width:35%;object-fit:contain;">
				{% endif %}
			</td>
			
			{% set bmd = frappe.db.get_value("Employee", {"name": "BMD01"}, ["custom_digital_signature"]) %}
			<td style="height:60px;vertical-align:middle;border-top:hidden;border-right:hidden;border-left:hidden;text-align:center;">
				{% if bmd and doc.approved_by_bmd %}
					<img src="{{bmd}}" style="height:60px;width:35%;object-fit:contain;">
				{% endif %}
			</td>
			
			{% set cmd = frappe.db.get_value("Employee", {"name": "CMD01"}, ["custom_digital_signature"]) %}
			<td style="height:60px;vertical-align:middle;border-top:hidden;border-right:hidden;border-left:hidden;text-align:center;">
				{% if cmd and doc.approved_by_cmd %}
					<img src="{{cmd}}" style="height:60px;width:35%;object-fit:contain;">
				{% endif %}
			</td>
			
			{% set smd = frappe.db.get_value("Employee", {"name": "SMD01"}, ["custom_digital_signature"]) %}
			<td style="height:60px;vertical-align:middle;border-top:hidden;border-right:hidden;border-left:hidden;text-align:center;">
				{% if smd and doc.approved_by_smd %}
					<img src="{{smd}}" style="height:60px;width:35%;object-fit:contain;">
				{% endif %}
			</td>

		{% elif doc.cargo_type == "Sea" and doc.scope_of_delivery=="Wonjin" %}
			{% set prepared_by = frappe.db.get_value("Employee",{"user_id": doc.prepared_by},"custom_digital_signature") %}
			<td style="height:60px;vertical-align:middle;border-top:hidden;border-right:hidden;border-left:hidden;text-align:center;">
				{% if prepared_by %}<img src="{{prepared_by}}" style="height:60px;width:35%;object-fit:contain;">{% endif %}
			</td>
			{% set si_owner = frappe.db.get_value("Sales Invoice",{"name": doc.order_no},['owner']) %}
			{% set requested_by = frappe.db.get_value("Employee",{"user_id": doc.owner},['custom_digital_signature']) %}
			<td style="height:60px;vertical-align:middle;border-top:hidden;border-right:hidden;border-left:hidden;text-align:center;">
				{% if requested_by %}<img src="{{requested_by}}" style="height:60px;width:35%;object-fit:contain;">{% endif %}
			</td>
			
			
			{% set finance = frappe.db.get_value("Employee",{"name": 'S0189'},["custom_digital_signature"]) %}
			<td colspan="2" style="height:60px;vertical-align:middle;border-top:hidden;border-right:hidden;border-left:hidden;text-align:center;">
				{% if finance and doc.approved_by_finance %}<img src="{{finance}}" style="height:60px;width:35%;object-fit:contain;">{% endif %}
			</td>
			{% set cmd = frappe.db.get_value("Employee",{"name": 'CMD01'},["custom_digital_signature"]) %}
			<td colspan="2" style="height:60px;vertical-align:middle;border-top:hidden;border-right:hidden;border-left:hidden;text-align:center;">
				{% if cmd and doc.approved_by_cmd %}<img src="{{cmd}}" style="height:60px;width:35%;object-fit:contain;">{% endif %}
			</td>

		{% elif doc.scope_of_delivery=="Customer" %}
			{% set prepared_by = frappe.db.get_value("Employee",{"user_id": doc.prepared_by},['custom_digital_signature']) %}
			<td colspan="2" style="height:60px;vertical-align:middle;border-top:hidden;border-right:hidden;border-left:hidden;text-align:center;">
				{% if prepared_by %}<img src="{{prepared_by}}" style="height:50px;width:25%;object-fit:contain;">{% endif %}
			</td>
			{% set si_owner = frappe.db.get_value("Sales Invoice",{"name": doc.order_no},['owner']) %}
			{% set requested_by = frappe.db.get_value("Employee",{"user_id": doc.owner},['custom_digital_signature']) %}
			<td colspan=2 style="height:60px;vertical-align:middle;border-top:hidden;border-right:hidden;border-left:hidden;text-align:center;">
				{% if requested_by %}<img src="{{requested_by}}" style="height:60px;width:35%;object-fit:contain;">{% endif %}
			</td>
			
			
			{% set finance = frappe.db.get_value("Employee",{"name": 'S0189'},["custom_digital_signature"]) %}
			<td colspan="2" style="height:60px;vertical-align:middle;border-top:hidden;border-right:hidden;border-left:hidden;text-align:center;">
				{% if finance and doc.approved_by_finance %}<img src="{{finance}}" style="height:50px;width:25%;object-fit:contain;">{% endif %}
			</td>
		{% endif %}
	</tr>

	<!-- DATETIME ROW (separate) -->
	<tr class="remove-top-border remove-bottom-border"
		style="font-size:9px;text-align:center;font-family:'Book Antiqua','Palatino Linotype',Palatino,serif;border-top:none;border-right:hidden;border-left:hidden;">
		{% if doc.cargo_type == "Air" and doc.scope_of_delivery=="Wonjin" %}
			<td class="remove-bottom-border" style="border-right:hidden;border-left:hidden;font-size:9px;text-align:center;">{{ frappe.utils.format_datetime(doc.prepared_by_time, "dd-MM-yyyy HH:mm:ss") or ''}}</td>
			<td class="remove-bottom-border" style="border-right:hidden;border-left:hidden;font-size:9px;text-align:center;">{{ frappe.utils.format_datetime(doc.creation, "dd-MM-yyyy HH:mm:ss") or '' }}</td>
			<td class="remove-bottom-border" style="border-right:hidden;border-left:hidden;font-size:9px;text-align:center;">{{ frappe.utils.format_datetime(doc.approved_by_finance_date_and_time, "dd-MM-yyyy HH:mm:ss")  or ''}}</td>
			<td class="remove-bottom-border" style="border-right:hidden;border-left:hidden;font-size:9px;text-align:center;">{{ frappe.utils.format_datetime(doc.approved_by_bmd_date_and_time, "dd-MM-yyyy HH:mm:ss") or '' }}</td>
			<td class="remove-bottom-border" style="border-right:hidden;border-left:hidden;font-size:9px;text-align:center;">{{ frappe.utils.format_datetime(doc.approved_by_cmd_date_and_time, "dd-MM-yyyy HH:mm:ss") or '' }}</td>
			<td class="remove-bottom-border" style="border-right:hidden;border-left:hidden;font-size:9px;text-align:center;">{{ frappe.utils.format_datetime(doc.approved_by_smd_date_and_time, "dd-MM-yyyy HH:mm:ss") or ''}}</td>
		{% elif doc.cargo_type == "Sea" and doc.scope_of_delivery=="Wonjin" %}
			<td class="remove-bottom-border" style="border-right:hidden;border-left:hidden;font-size:9px;text-align:center;">{{ frappe.utils.format_datetime(doc.prepared_by_time, "dd-MM-yyyy HH:mm:ss") or ''}}</td>
			<td class="remove-bottom-border" style="border-right:hidden;border-left:hidden;font-size:9px;text-align:center;">{{ frappe.utils.format_datetime(doc.creation, "dd-MM-yyyy HH:mm:ss") or '' }}</td>
			<td colspan="2" class="remove-bottom-border" style="border-right:hidden;border-left:hidden;font-size:9px;text-align:center;">{{ frappe.utils.format_datetime(doc.approved_by_finance_date_and_time, "dd-MM-yyyy HH:mm:ss") or '' }}</td>
			<td colspan="2" class="remove-bottom-border" style="border-right:hidden;border-left:hidden;font-size:9px;text-align:center;">{{ frappe.utils.format_datetime(doc.approved_by_cmd_date_and_time, "dd-MM-yyyy HH:mm:ss") or ''}}</td>
		{% elif doc.scope_of_delivery=="Customer" %}
			<td colspan=2 class="remove-bottom-border" style="border-right:hidden;border-left:hidden;font-size:9px;text-align:center;">{{ frappe.utils.format_datetime(doc.prepared_by_time, "dd-MM-yyyy HH:mm:ss") or ''}}</td>
			<td colspan="2" class="remove-bottom-border" style="border-right:hidden;border-left:hidden;font-size:9px;text-align:center;">{{ frappe.utils.format_datetime(doc.creation, "dd-MM-yyyy HH:mm:ss") or '' }}</td>
			<td colspan="2" class="remove-bottom-border" style="border-right:hidden;border-left:hidden;font-size:9px;text-align:center;">{{ frappe.utils.format_datetime(doc.approved_by_finance_date_and_time, "dd-MM-yyyy HH:mm:ss") or '' }}</td>
		{% endif %}
	</tr>
	<tr class="remove-top-border remove-bottom-border"
		style="font-size:12px;text-align:left;font-family:'Book Antiqua','Palatino Linotype',Palatino,serif;border-top:none;border-right:hidden;border-left:hidden;">
		<td class="remove-bottom-border" style="border-right:hidden;border-left:hidden;" colspan="6">Note: This document is Digitally Signed</td>
		</tr>
</table>
</div>
</div> 
</div>

		"""
	html = render_template(template, {"doc": doc})
	return {"html": html}


@frappe.whitelist()
def create_html_lr_export(doc):
	doc = frappe.parse_json(doc)
	template = """
	<style>
	.tab1,
.tab1 tr,
.tab1 td{
	border:1px solid black;
}

.tab2,
.tab2 tr,
.tab2 td{
	border:1px solid black;
	border-bottom:none;

}

.tab3,
.tab3 tr,
.tab3 td{
	border:1px solid black;
	border-bottom:none;
	border:none;
}

.no-border{
	border:none !important;
}

.no-border-1{
	border-top:none !important;
	border-bottom:none !important;
}

/*.tab2 tr{*/
/*    page-break-inside:avoid; */
/*    page-break-after:auto;*/
/*}*/


.page-break {
	page-break-before: always;
	break-before: page;
}

.in-words-row {
	page-break-inside: avoid;
	break-inside: avoid;
	border-bottom:none;
	border-top: 1px solid black;
}


	</style>

	<!--{% set data = packing_list(doc.name)%}-->
<!--{{data}}-->


<div class="container">
	<div class="row ">
		<div class="col-xs-12 p-0 " >
			
		   <table width=100% style="border-collapse:collapse;border-bottom: none;font-size:10px;" class="tab1 mr-0 ml-0" >
			   <tr>
					<td colspan="5" style="font-size:15px;">
					   <center><b>PACKING LIST</b></center> 
					</td>
				</tr>
			   <tr>
				   <td rowspan="8" 
					style="width:30%; font-size:10px; border-bottom:1px solid black !important;vertical-align: middle;">
					<div style="margin-top:0px;">
					{% set address_title = frappe.db.get_value("Address", {'address_title': doc.company}, "address_title") %}
						{% set address_line1 = frappe.db.get_value("Address", {'address_title': doc.company}, "address_line1") %}
						{% set address_line2 = frappe.db.get_value("Address", {'address_title': doc.company}, "address_line2") %}
						{% set city = frappe.db.get_value("Address", {'address_title': doc.company}, "city") %}
						{% set state = frappe.db.get_value("Address", {'address_title': doc.company}, "state") %}
						{% set country = frappe.db.get_value("Address", {'address_title': doc.company}, "country") %}
						{% set phone = frappe.db.get_value("Address", {'address_title': doc.company}, "phone") %}
						{% set fax = frappe.db.get_value("Address", {'address_title': doc.company}, "fax") %}
						{% set gstin = frappe.db.get_value("Address", {'address_title': doc.company}, "gstin") %}
						<b>Factory Address</b><br>
						{% if address_title %}
							{{ address_title }}<br>
						{% endif %}
						{% if address_line1 %}
							{{ address_line1 }}<br>
						{% endif %}
						{% if address_line2 %}
							{{ address_line2 }}<br>
						{% endif %}
						{% if city and state %}
							{{ city }}, {{ state }}<br>
						{% elif city %}
							{{ city }}<br>
						{% elif state %}
							{{ state }}<br>
						{% endif %}
						{% if phone and fax %}
							Ph: {{ phone }}, Fax: {{ fax }}<br>
						{% elif phone %}
							{{ phone }}<br>
						{% elif fax %}
							{{ fax }}<br>
						{% endif %}
						{% if gstin %}
							GSTIN: {{ gstin }}<br>
						{% endif %}
						{% if country %}
							{{ country }}
						{% endif %}
	</div>
				</td>
						
					<td rowspan="8"  style="width:30%; font-size:10px; border-bottom:1px solid black !important;">
						<div style="margin-top:0px;">
						{% set address_title = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_title") %}
						{% set address_line1 = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_line1") %}
						{% set address_line2 = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_line2") %}
						{% set city = frappe.db.get_value("Address", {'address_title': doc.customer}, "city") %}
						{% set state = frappe.db.get_value("Address", {'address_title': doc.customer}, "state") %}
						{% set country = frappe.db.get_value("Address", {'address_title': doc.customer}, "country") %}
						{% set gstin = frappe.db.get_value("Address", {'address_title': doc.customer}, "gstin") %}
						{% set gstin_1 = frappe.db.get_value("Customer", {'customer_name': doc.customer}, "gstin") %}
						<b>Exporter</b><br>
						{% if address_title %}
							{{ address_title }}<br>
						{% endif %}
						{% if address_line1 %}
							{{ address_line1 }}<br>
						{% endif %}
						{% if address_line2 %}
							{{ address_line2 }}<br>
						{% endif %}
						{% if city %}
							{{ city }}<br>
						{% endif %}
						{% if state %}
							{{ state }}<br>
						{% endif %}
						{% if gstin and gstin_1 %}
							GSTIN: {{ gstin }}<br>
						{% elif gstin %}
							GSTIN: {{gstin}}<br>
						{% elif gstin_1 %}
							GSTIN: {{gstin_1}}<br>
						{% endif %}
						{% if country %}
							{{ country }}
						{% endif %}
						
						</div>
					</td>
				</tr>
			<tr>
					<td style="border-bottom:hidden;border-right:hidden;"><b>Invoice No & Date:</b></td>
					<td style="border-top:hidden;"><b>{{doc.order_no}}&nbsp;&nbsp;{{ frappe.format(frappe.db.get_value("Sales Invoice",{"name":doc.order_no},"posting_date"),{'fieldtype':'Date'}) or ""  }}</b></td>
						
					
				</tr>
			   <tr>        
						<td style="border-bottom:hidden;"><b>Exporter IEC</b></td>
						<td style="border-top:hidden;border-left:hidden;"><b>{{frappe.db.get_value('Company',{'name':doc.company},"custom_exporter_iec")or ""}}</b>
					</td>
			   </tr>
				
				<tr>
					<td><b>CIN No</b></td>
					<td style="border-top:hidden;border-left:hidden;"><b>{{frappe.db.get_value('Customer',{'name':doc.customer},"custom_cin_number")or ""}}</b></td>
				</tr>
			
				
		</table> 
		<table width=100% style="border-collapse:collapse;margin-top:5px;padding:0px;border-bottom:hidden;" class="tab1 mr-0 ml-0">
			
				<tr class= "remove-top-border">
					<td class= "remove-top-border" colspan="2" style="border-bottom:hidden;width:50%;font-size:10px;border-top: 1px solid black !important;">
						<b>Consignee</b>
				</td>
				<td colspan="2" style="width:50%;font-size:10px;border-bottom:hidden;border-top: 1px solid black !important;" class= "remove-top-border" >
						<b>Buyer if Other  than consignee</b>
				</td>
				</tr>
				<tr class= "remove-top-border">
				<td class="remove-top-border" colspan="2" 
					style="width:50%; font-size:10px; border-bottom:1px solid black !important;">
					<div>
					{% set address_title = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_title") %}
						{% set address_line1 = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_line1") %}
						{% set address_line2 = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_line2") %}
						{% set city = frappe.db.get_value("Address", {'address_title': doc.customer}, "city") %}
						{% set state = frappe.db.get_value("Address", {'address_title': doc.customer}, "state") %}
						{% set country = frappe.db.get_value("Address", {'address_title': doc.customer}, "country") %}
						
						{% if address_title %}
							{{ address_title }}<br>
						{% endif %}
						{% if address_line1 %}
							{{ address_line1 }}<br>
						{% endif %}
						{% if address_line2 %}
							{{ address_line2 }}<br>
						{% endif %}
						{% if city %}
							{{ city }}<br>
						{% endif %}
						{% if state %}
							{{ state }}<br>
						{% endif %}
						{% if country %}
							{{ country }}
						{% endif %}
	</div>
				</td>
						


					
					<td class="remove-top-border" colspan="2" style="width:50%; font-size:10px; border-bottom:1px solid black !important;">
						<div>
						{% set address_title = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_title") %}
						{% set address_line1 = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_line1") %}
						{% set address_line2 = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_line2") %}
						{% set city = frappe.db.get_value("Address", {'address_title': doc.customer}, "city") %}
						{% set state = frappe.db.get_value("Address", {'address_title': doc.customer}, "state") %}
						{% set country = frappe.db.get_value("Address", {'address_title': doc.customer}, "country") %}
						
						{% if address_title %}
							{{ address_title }}<br>
						{% endif %}
						{% if address_line1 %}
							{{ address_line1 }}<br>
						{% endif %}
						{% if address_line2 %}
							{{ address_line2 }}<br>
						{% endif %}
						{% if city %}
							{{ city }}<br>
						{% endif %}
						{% if state %}
							{{ state }}<br>
						{% endif %}
						{% if country %}
							{{ country }}
						{% endif %}
						
						</div>
					</td>
				</tr>
				<tr>
					
					<td style="width:25%;font-size:10px;border-bottom:hidden;"><b>Pre Carriage by:</b></td>
					<td style="width:25%;font-size:10px;border-bottom:hidden;"><b>Place of receipt by pre-carrier</b></td>
					<td style="width:50%;font-size:10px;border-bottom:hidden;">
						<b>Country of Origin of Goods</b></td>
					<!--<td style="width:25%;font-size:10px;">-->
					<!--    <b>Country of Final Destination</b></td>-->
					<!--</tr>-->
					<!--<td style="width:20%;font-size:10px;border-bottom:hidden;">-->
					<!--    <b>Country of Final Destination</b></td>-->
				<tr>
					<td style="width:25%;font-size:10px;border-bottom: 1px solid black !important;">{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"recommended_ffw")or ""}}</td>
					<td style="width:25%;font-size:10px;border-bottom: 1px solid black !important;">{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"custom_place_of_receipt_by_precarrier") or ""}}</td>
					<td style="width:50%;font-size:10px;border-bottom: 1px solid black !important;">
						{% if doc.cargo_type == 'Air' %}
						{{doc.pol_country_airport or ""}}
						{% elif doc.cargo_type == 'Sea' %}
						{{doc.pol_country_seaport or ""}}
						{% else %}
						{{doc.pol_country_airport or ""}}
						{% endif %}
					</td>

					
				</tr>
				<tr>
					
					<td style="width:25%;font-size:10px;border-bottom:hidden;"><b>Vessels / Flight No.</b></td>
					<td style="width:25%;font-size:10px;border-bottom:hidden;"><b>Port of Loading</b></td>
					<td style="width:50%;font-size:10px;border-bottom:hidden;">
						<b>Terms of Delivery & Payment</b></td>
					
				</tr>
				<tr>
					<td style="width:25%;font-size:10px;">BY&nbsp;{{doc.cargo_type or ""}}</td>
					<td style="width:25%;font-size:10px;">
						{% if doc.cargo_type == 'Sea' %}
							{{doc.pol_city_seaport or ""}}
						{% else %}
							{{doc.pol_city_airport or ""}}
						{% endif %}
					</td>
					<td style="width:50%;font-size:10px;border-bottom:hidden;">
						By&nbsp;{{doc.cargo_type or ""}}
					</td>
					
					
				</tr>
				<tr>
					
					<td style="width:25%;font-size:10px;border-bottom:hidden;"><b>Port Of Discharge</b></td>
					<td style="width:25%;font-size:10px;border-bottom:hidden;"><b>Final Destination</b></td>
					<td style="width:50%;font-size:10px;border-bottom:hidden;">
						<b>Payment</b></td>
					
				</tr>
				<tr>
					<td style="width:25%;font-size:10px;border-bottom: 1px solid black !important;">
						
							{% if doc.cargo_type == 'Sea' %}
							{{doc.pod_seaport or ""}}
						{% else %}
							{{doc.pod_airport or ""}}
						{% endif %}
						
					</td>
					<td style="width:25%;font-size:10px;border-bottom: 1px solid black !important;">
						
							{{ doc.final_destination or "" }}
					</td>
					<td colspan="2" style="width:50%;font-size:10px;border-bottom:hidden;">
						{{frappe.db.get_value("Sales Invoice",{"name":doc.order_no},"terms") or ""}}
					</td>
					
					
				</tr>
				<tr>
					<td colspan="2" style="font-size:10px;border-bottom:hidden;">
						<b>General Description of the below mentioned Products is Automobile</b>
					</td>
					
					<td colspan="2" style="font-size:10px;border-bottom:hidden;">
						<!--<b>Incoterms</b>-->
					</td>
				</tr>
			   <tr>
					<td colspan="2" style="font-size:10px;">
						<b>Components</b>
					</td>
					
					<td colspan="2" style="font-size:10px;">
						<!--{{doc.incoterm or ""}}-->
					</td>
				</tr>
		</table>
		</div>
	</div>
</div>

<div class="container">
	<div class="row">
		<div class="col-xs-12 p-0">
			<table width=100% style="border-collapse:collapse;font-size:10px; " class="tab2 mr-0 ml-0">
			   <thead style="display: table-header-group;" >    
				<tr style="border-bottom:1px solid black;border-top:none;">
					<td rowspan="2" style="width:2%; vertical-align: middle; text-align: center;"><b>PD</b></td>
					<td rowspan="2" style="width:2%; vertical-align: middle; text-align: center;"><b>Boxes</b></td>
				
					<td rowspan="2" style="width:5%; vertical-align: middle; text-align: center;"><b>Item Code</b></td>
					<td rowspan="2" style=" vertical-align: middle; text-align: center;"><b>Description of Goods</b></td>
					<td rowspan="2" style="width:10%; vertical-align: middle; text-align: center;"><b>Order No. & PO</b></td>
					<td style="width:8%; vertical-align: middle; text-align: center;border-bottom:hidden;"><b>HSN</b></td>
					<td style="width:5%; vertical-align: middle; text-align: center;border-bottom:hidden;"><b>Qty</b></td>
					<td style="width:5%; vertical-align: middle; text-align: center;border-bottom:hidden;"><b>Rate</b></td>
					<td style="width:5%; vertical-align: middle; text-align: center;border-bottom:hidden;"><b>Amount</b></td>
				</tr>

				<tr style="border-bottom:1px solid black;">
					<td style="width:5%;vertical-align: middle;"><b><center>Code</center></b></td>
					<td style="width:5%;vertical-align: middle;"><b><center>Nos</center></b></td>
					<td style="width:5%;vertical-align: middle;"><b><center>{{doc.currency or ""}}</center></b></td>
					<td style="width:5%;vertical-align: middle;"><b><center>{{doc.currency or ""}}</center></b></td>
				</tr>
				
			</thead>   
				<!--<tr style="border-bottom:1px solid black;">-->
				<!--    <td rowspan="2"><b><center>Parts/BOX</center></b></td>-->
				<!--    <td rowspan="2"><b><center>Item Code</center></b></td>-->
				<!--    <td rowspan="2"><b><center>Description of Goods</center></b></td>-->
				<!--    <td rowspan="2"><b><center>HSN Code</center></b></td>-->
				<!--    <td rowspan="2"><b><center>Quality Nos</center></b></td>-->
				<!--    <td rowspan="2"><b><center>Remark</center></b></td>-->
				<!--</tr>-->
				<!--<tr style="border-bottom:1px solid black;">-->
				<!--    <td><b><center>Pallet Details</center></b></td>-->
				<!--    <td><b><center>Boxes</center></b></td>-->
				<!--</tr>-->
				
				
				
				{%- for i in doc.product_description_so -%}
				 <tr style="border-bottom:none !important; border-top:none;">
					<td class="no-border-1" style="text-align:center;">{{i.custom_no_of_pallets or ""}}</td>
					<td class="no-border-1" style="text-align:center;">{{i.custom_no_of_boxes or ""}}</td>
					<td class="no-border-1">{{i.item_code or ""}}</td>
					<td class="no-border-1" style="font-size:10px;">{{i.description or ""}}</td>
					<td class="no-border-1">{{i.custom_customers_po_number or ""}}</td>
					<td class="no-border-1" style="text-align:center;">{{i.gst_hsn_code or ""}}</td>
					<td class="no-border-1"  style="text-align:center;">{{ ("%.2f"|format(i.qty|float)).rstrip('0').rstrip('.') if i.qty else "" }}</td>
					<td class="no-border-1" style="text-align:right;">{{i.rate or ""}}</td>
					<td class="no-border-1" style="text-align:right;">{{i.amount or ""}}</td>
					
				</tr>
		  
				
				{%- endfor-%}
				
				{% set p_total = doc.product_description_so | map(attribute="custom_no_of_pallets") | sum %}
				{% set b_total = doc.product_description_so | map(attribute="custom_no_of_boxes") | sum %}
				{% set q_total = doc.product_description_so | map(attribute="qty") | sum %}
				
				
				<tr style="border-bottom:none !important;">
					<td colspan="1" class="no-border">
						<center><p><b>{{p_total}}</b></p></center>
					</td>
					<td colspan="1" class="no-border">
					   <center><p><b>{{b_total}}</b></p></center> 
					</td>
					<td colspan="4" class="no-border">
					</td>
					<td colspan="1" class="no-border" style="text-align:center;">
					   <b>{{q_total}}</b>
					</td>
					<td colspan="2" class="no-border">
						
					</td>
				</tr>
				
				<tr style="border-bottom:none !important; border-top:none !important" >
					<td colspan="9" class="no-border">
						<p>This Shipment under duty draw back </p>
					</td>
				</tr>
				
				<tr style="border-top:none !important; border-bottom:none;">
					<td colspan="3" style="text-align:right; " class="no-border">
						<p><b><u>COUNTRY OF ORIGIN :{{doc.pol_country_airport or ""}}</u></b></p>
					</td>
					<td colspan="6" class="no-border">
						
					</td>
				</tr>
				
				<tr >
					<td colspan="3" style="width:32%;border-left:none;border-right:none;">
						<b>Total Gross Weight&nbsp;&nbsp;{{doc.gross_wt}} KGS</b>
					</td>
					<td colspan="3"  style="width:32%;text-align:center;border-left:none;border-right:hidden;">
						<b style=" left:50px;" >Total Net Weight&nbsp;&nbsp;{{doc.net_wt}} KGS</b> 
					</td>
					<td colspan="3"  style="width:36%;text-align:center;border-left:hidden;">   
						<b style=" left:350px;">Total No.of Packaged:{{p_total}} Nos</b>
					</td>
				</tr>

				
			</table>
			
			
			
		  
			
		</div>
	</div>
	
   </div> 

<style>
	@media print {
		.declaration-container {
			page-break-inside: avoid;
			break-inside: avoid;
		}

		.declaration-row {
			border: 1px solid black;
			border-top: 1px solid black;
		}

		/* If using wkhtmltopdf, avoid relying on @page:first */
	}

	.qrcode {
		width: 120px;
		height: 120px;
	}

	.text-center {
		text-align: center;
	}
</style>

{% set is_first_page = false %} {# Or dynamically determine this in your logic #}

   <table style="border: 1px solid black; border-collapse: collapse; width: 100%;">
   <tr style="border:none;">
			<td colspan="4" style="border: 1px solid black; border-bottom: none; border-right: 1px solid transparent !important;width: 50%;">

				<table style="border-collapse: collapse; font-size: 9px; table-layout: fixed;">
					<colgroup>
						
						<col style="width: 70px;">
						<col style="width: 70px;">
						<col style="width: 70px;">
						<col style="width: 100px;">
					</colgroup>
					<tr>
						<td colspan =4 style="border: 1px solid black; text-align: center;"><b>Pallet Details</b></td>
					</tr>
					<tr>
						
						<td style="border: 1px solid black; text-align: center;"><b>Length</b></td>
						<td style="border: 1px solid black; text-align: center;"><b>Width</b></td>
						<td style="border: 1px solid black; text-align: center;"><b>Height</b></td>
						<td style="border: 1px solid black; text-align: center;"><b>No of Pallets</b></td>
					</tr>
				
					{%- set ns = namespace(total_boxes_1=0, total_length_1=0, total_height_1=0, total_breadth_1 =0) -%}
                
                    {%- for row in doc["product_description_so"] -%}
                        {% set ns.total_boxes_1 = ns.total_boxes_1 + (row.custom_no_of_pallets or 0) %}
                        {% set ns.total_length_1 = ns.total_length_1 + (row.custom_pallet_length or 0) %}
                        {% set ns.total_height_1 = ns.total_height_1 + (row.custom_calculated_height or 0) %}
                        {% set ns.total_breadth_1 = ns.total_breadth_1 + (row.custom_pallet_breadth or 0) %}
                    <tr>
                        <td style="border: 1px solid black; text-align: center;">{{ row.custom_pallet_length or "" }}</td>
                        <td style="border: 1px solid black; text-align: center;">{{ row.custom_pallet_breadth or "" }}</td>
                        <td style="border: 1px solid black; text-align: center;">{{ row.custom_calculated_height or "" }}</td>
                        <td style="border: 1px solid black; text-align: center;">{{ row.custom_no_of_pallets or 0 }}</td>
                    </tr>
                    {%- endfor -%}
                    
                    <tr style="background-color:#D3D3D3;">  
                        
                        <td colspan="3" style="border: 1px solid black; text-align: right;"><b>Total No of Pallets</b></td>
                         <td style="border: 1px solid black; text-align: center;"><b>{{ ns.total_boxes_1 }}</b></td>
                        
                    </tr>
				</table>
			</td>
			<td colspan="4" style="border: 1px solid black; border-bottom: none; border-left: 1px solid transparent !important;">
			<table style="border-collapse: collapse; font-size: 9px; table-layout: fixed;">
				<colgroup>
					<col style="width: 70px;">
					<col style="width: 70px;">
					<col style="width: 70px;">
					<col style="width: 100px;">
				</colgroup>
				 <tr>
					<td colspan =4 style="border: 1px solid black; text-align: center;"><b>Box Details</b></td>
				</tr>
				<tr>
					
					<td style="border: 1px solid black; text-align: center;"><b>Length</b></td>
					<td style="border: 1px solid black; text-align: center;"><b>Width</b></td>
					<td style="border: 1px solid black; text-align: center;"><b>Height</b></td>
					<td style="border: 1px solid black; text-align: center; vertical-align: middle;">
						<b>No of Boxes</b>
					</td>
				</tr>
				{%- set ns = namespace(total_boxes=0, total_length=0, total_height=0, total_breadth =0) -%}
                
                {%- for row in doc["product_description_so"] -%}
                    {% set ns.total_boxes = ns.total_boxes + (row.custom_no_of_boxes or 0) %}
                    {% set ns.total_length = ns.total_length + (row.custom_box_length or 0) %}
                    {% set ns.total_height = ns.total_height + (row.custom_box_height or 0) %}
                    {% set ns.total_breadth = ns.total_breadth + (row.custom_box_breadth or 0) %}
                <tr>
                    <td style="border: 1px solid black; text-align: center;">{{ row.custom_box_length or "" }}</td>
                    <td style="border: 1px solid black; text-align: center;">{{ row.custom_box_breadth or "" }}</td>
                    <td style="border: 1px solid black; text-align: center;">{{ row.custom_box_height or "" }}</td>
                    <td style="border: 1px solid black; text-align: center;">{{ row.custom_no_of_boxes or 0 }}</td>
                </tr>
                {%- endfor -%}
                
                <tr style="background-color:#D3D3D3;">  
                    <td colspan="3" style="border: 1px solid black; text-align: right;"><b>Total No of Boxes</b></td>
                    <td style="border: 1px solid black; text-align: center;"><b>{{ ns.total_boxes }}</b></td>
                </tr>
			</table>
		</td>
	</tr>

	<tr style="border-top:none;">
		<td colspan="8" style="border: 1px solid black;border-top:none; padding-top: 15px;font-size:10px;">
			<div class="col-12 remove-top-border remove-bottom-border p-0">
			<strong><u>DECLARATION</u></strong><br>
			We declare that this packing list shows the actual<br>
			Weight of the goods<br>
			Described and that all particulars are true and correct
		</div>
		</td>
	</tr>
</table>


<!--</div>-->





		
	"""
	html = render_template(template, {"doc": doc})
	return {"html": html}

@frappe.whitelist()
def packing_list(sales_invoice):
	data = frappe.db.sql("""
		SELECT 
			si.custom_exporter_iec, si.custom_gstin,
			sii.custom_pallet, sii.custom_pallet_length,
			sii.custom_pallet_breadth, sii.custom_pallet_height,
			SUM(sii.total_weight) as net_weight,
			SUM(sii.custom_gross_weight) as gross_weight,
			SUM(sii.custom_no_of_boxes) as total_boxes,
			SUM(sii.custom_no_of_pallets) as total_pallets,
			sii.item_code, sii.description, SUM(sii.qty) as qty
		FROM `tabSales Invoice` si
		INNER JOIN `tabSales Invoice Item` sii ON si.name = sii.parent
		WHERE si.name = %s
		GROUP BY sii.custom_pallet, sii.item_code
		ORDER BY sii.custom_pallet
	""", (sales_invoice,), as_dict=True)

	if not data:
		return "<p>No packing list data available.</p>"

	iec = data[0].custom_exporter_iec or ''
	gstin = data[0].custom_gstin or ''

	# Group by pallet
	pallet_map = {}
	for row in data:
		pallet_map.setdefault(row.custom_pallet, []).append(row)

	# Header and first table
	html = f"""
	<h3 class="text-center">Packing List</h3>
	<p style="margin-left: 70%;">Exporter IEC: <span style="font-weight: 100;">{iec}</span></p>
	<p style="margin-left: 70%;">GSTIN: <span style="font-weight: 100;">{gstin}</span></p>
	<style>
	table {{
			width: 100%;
		}}

		th, td {{
			border: 1px solid black;
			text-align: left;
			font-size: 10px;
		}}
		
		.background {{
			background-color: #f68b1f;
			color: white;
			font-weight: 700;
		}}
	</style>
	<table class="mt-2" border="1" cellspacing="0" cellpadding="4">
		<tr>
			<td class="background text-center" style="background-color: #f68b1f;color:white;">Pallet</td>
			<td class="background text-center" style="background-color: #f68b1f;color:white;">Item</td>
			<td class="background text-center" style="background-color: #f68b1f;color:white;">Description</td>
			<td class="background text-center" style="background-color: #f68b1f;color:white;">Qty (Nos)</td>
			<td class="background text-center" style="background-color: #f68b1f;color:white;">No. of Boxes (Nos)</td>
			<td class="background text-center" style="background-color: #f68b1f;color:white;">No. of Pallets (Nos)</td>
		</tr>
	"""
	total_qty = 0
	total_boxes = 0
	total_pallets = 0
	for pallet, items in pallet_map.items():
		rowspan = len(items)
		for idx, i in enumerate(items):
			html += "<tr>"
			if idx == 0:
				html += f'<td rowspan="{rowspan}">{pallet or ""}</td>'
			total_qty += int(round(i.qty or 0))
			total_boxes += int(round(i.total_boxes or 0))
			total_pallets += int(round(i.total_pallets or 0))
			html += f"""
				<td>{i.item_code}</td>
				<td>{i.description}</td>
				<td class="text-right">{int(round(i.qty or 0))}</td>
				<td class="text-right">{int(round(i.total_boxes or 0))}</td>
				<td class="text-right">{int(round(i.total_pallets or 0))}</td>
			</tr>
			"""
	html += f"""
			<tr>
				<td colspan=3 class="text-right"><b>Total</b></td>
				<td class="text-right"><b>{total_qty}</b></td>
				<td class="text-right"><b>{total_boxes}</b></td>
				<td class="text-right"><b>{total_pallets}</b></td>
			</tr>"""
	html += "</table>"

	# Second table: Pallet summary
	html += f"""
	<div style="width: 70%;">
	<table class="mt-5" border="1" cellspacing="0" cellpadding="4" width=60%>
		<tr>
			<td rowspan="2" class="background text-center" style="background-color: #f68b1f;color:white;">Pallet</td>
			<td colspan="3" class="background text-center" style="background-color: #f68b1f;color:white;">Dimensions (mm)</td>
			<td colspan="2" class="background text-center" style="background-color: #f68b1f;color:white;">Weight (kg)</td>
		</tr>
		<tr>
			<td class="background text-center" style="background-color: #f68b1f;color:white;">L</td>
			<td class="background text-center" style="background-color: #f68b1f;color:white;">B</td>
			<td class="background text-center" style="background-color: #f68b1f;color:white;">H</td>
			<td class="background text-center" style="background-color: #f68b1f;color:white;">Net</td>
			<td class="background text-center" style="background-color: #f68b1f;color:white;">Gross</td>
		</tr>
	"""

	total_net = 0
	total_gross = 0

	for pallet, items in pallet_map.items():
		first = items[0]
		length = int(round(first.custom_pallet_length or 0))
		breadth = int(round(first.custom_pallet_breadth or 0))
		height = int(round(first.custom_pallet_height or 0))
		net_weight = int(round(first.net_weight or 0))
		gross_weight = int(round(first.gross_weight or 0))

		html += f"""
		<tr>
			<td class="text-left">{pallet or ""}</td>
			<td class="text-right">{length}</td>
			<td class="text-right">{breadth}</td>
			<td class="text-right">{height}</td>
			<td class="text-right">{net_weight}</td>
			<td class="text-right">{gross_weight}</td>
		</tr>
		"""
		total_net += net_weight
		total_gross += gross_weight

	html += f"""
		<tr>
			<td colspan="4" class="text-right"><b>Total</b></td>
			<td class="text-right"><b>{total_net}</b></td>
			<td class="text-right"><b>{total_gross}</b></td>
		</tr>
		"""

	html += "</table></div>"
	return html


@frappe.whitelist()
def create_html_EI(doc):
	doc = frappe.parse_json(doc)
	template="""
		<style>
			.tab1,
			.tab1 tr,
			.tab1 td{
				border:1px solid black;
			}

			.tab2,
			.tab2 tr,
			.tab2 td{
				border:1px solid black;
				border-bottom:none;

			}

			.tab3,
			.tab3 tr,
			.tab3 td{
				border:1px solid black;
				border-bottom:none;
				border:none;
			}

			.no-border{
				border:none !important;
			}

			.no-border-1{
				border-top:none !important;
				border-bottom:none !important;
			}

			/*.tab2 tr{*/
			/*    page-break-inside:avoid; */
			/*    page-break-after:auto;*/
			/*}*/


			.page-break {
				page-break-before: always;
				break-before: page;
			}

			.in-words-row {
				page-break-inside: avoid;
				break-inside: avoid;
				border-bottom:none;
				border-top: 1px solid black;
			}


		</style>
		{% if doc.custom_copy_type == 'Original' %}
			{% set copy_list = ['Original'] %}
		{% elif doc.custom_copy_type == 'Duplicate' %}
			{% set copy_list = ['Original','Duplicate'] %}
		{% elif doc.custom_copy_type == 'Triplicate' %}
			{% set copy_list = ['Original','Duplicate','Triplicate'] %}
		{% elif doc.custom_copy_type == 'Extra Copy 1' %}
			{% set copy_list = ['Original','Duplicate','Triplicate','Extra Copy 1'] %}
		{% elif doc.custom_copy_type == 'Extra Copy 2' %}
			{% set copy_list = ['Original','Duplicate','Triplicate','Extra Copy 1','Extra Copy 2'] %}
		{% else %}
			{% set copy_list = ['Original'] %}
		{% endif %}

		{% for copy in copy_list %}
		<div class="print-copy">
			{% if doc.custom_copy_type != 'Original' or not doc.custom_copy_type%}
				{% if loop.index ==1 %} 
				<div class="col-xs-12" style="vertical-align:top;text-align:right">
						<b>Original</b>
						
					</div>
				{% else %}
			<div class="col-xs-12" style="vertical-align:top;text-align:right">
					<b>Duplicate</b>
					
				</div>
		{% endif %}{% endif %}
		<div class="container">
		<div class="row" style="border: 1px solid black; border-bottom: none;">

				
				<div class="col-xs-2" style="vertical-align:top;">
					<b><u>Exporter</u></b>
					
				</div>
				

			<center><div class="col-xs-12" style="font-size:12px;width:600px;">
					<b>{{doc.company}}</b><br>
					(Under Rule 7 of GST Rule 2017)<br>
					For removal of Excise Goods from a factory or Warehouse on Payment of duty<br>
					{{frappe.db.get_value('Address',{'address_title':doc.company},'address_line1') or ""}}<br>
					{{frappe.db.get_value('Address',{'address_title':doc.company},'address_line2') or ""}}&nbsp;{{frappe.db.get_value('Address',{'address_title':doc.company},'city') or ""}}&nbsp;-&nbsp;{{frappe.db.get_value('Address',{'address_title':doc.company},'pincode') or ""}}<br>
					{% set phone = frappe.db.get_value('Address',{'address_title':doc.company},'phone') %}
					
					{% if phone%}
						
						Phone :&nbsp;{{phone}}
					{% endif %} 
					
					<br>
					
					
						
					{% set email = frappe.db.get_value('Address', {'address_title': doc.company}, 'email_id') %}
					{% set web = frappe.db.get_value('Company', {'name': doc.company}, 'domain') %}
					
					{% if email %}
						E-Mail : {{ email }}&nbsp;&nbsp;
					{% endif %}
					{% if web %}
						Web : {{ web }}
					{% endif %}
					<br>

					GSTIN:&nbsp;{{frappe.db.get_value('Address',{'address_title':doc.company},'gstin') or ""}} 
				</div></center>
			<div class="col-xs-2" style="margin-top:10px;text-align:center;">
			<img src="https://api.qrserver.com/v1/create-qr-code/?size=150x150&data={{ doc.custom_scan_barcode }}" alt="QR Code" />

</div>


		{% set e_invoice_log = frappe.db.get_value(
			"e-Invoice Log", doc.irn, ("invoice_data", "signed_qr_code"), as_dict=True
			) %}
		{% if e_invoice_log %}
		
		<div class="col-xs-2">
		   <img src="data:image/png;base64,{{ get_qr_code(e_invoice_log.signed_qr_code, scale=2) }}" class="qrcode" >
			
		</div>
		{% else %}
		
		<div class="col-xs-2">
		</div>
		
		{% endif %}
				
				
				
			
			</div>
		</div>

		<div class="container">
			<div class="row ">
				<div class="col-xs-12 p-0 " >
					
				<table width=100% style="border-collapse:collapse;font-size:14px;" class="tab1 mr-0 ml-0">
						<tr>
							<td colspan="3"><b><center><b>INVOICE</b></center></b></td>
						</tr>
						<tr>
							<td colspan="2" style="width:60%">
								E-WAY No. &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:{{doc.ewaybill or ""}}<br>
								E_WAY Date &nbsp;&nbsp;:{{frappe.db.get_value("e-Waybill Log",{"name":doc.ewaybill},"created_on") or ""}}<br>
								ACK No. &nbsp;&nbsp;&nbsp; &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:{{frappe.db.get_value('e-Invoice Log', {'sales_invoice': doc.name}, 'acknowledgement_number') or ""}}<br>
								ACK Date &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:{{frappe.db.get_value('e-Invoice Log', {'sales_invoice': doc.name}, 'acknowledged_on') or ""}}<br>
								IRN No. &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:{{doc.irn or ""}}
								
							</td>
							<td style="width:40%">
								
								Invoice No  &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; : {{doc.name}}<br>
								Invoice Date &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {{frappe.format(doc.posting_date,{'fieldtype':'Date'}) or ""}}<br>
								Exporter IEC &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {{frappe.db.get_value('Company',{'name':doc.company},"custom_exporter_iec")or ""}}<br>

								Vendor Code &nbsp;&nbsp;&nbsp;&nbsp;:{{frappe.db.get_value('Customer',{'name':doc.customer},"custom_vendor_code")or ""}}<br>
								CIN No&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:{{frappe.db.get_value('Customer',{'name':doc.customer},"custom_cin_number")or ""}}<br>
							</td>
						</tr>
						
				</table> 
				<table width=100% style="border-collapse:collapse;margin-top:5px;padding:0px;border-bottom:none;font-size:13px;" class="tab1 mr-0 ml-0">
					
						<tr>
							<td colspan="2" style="width:50%">
								<b>Consignee</b>
								{% set address_title = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_title") %}
								{% set address_line1 = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_line1") %}
								{% set address_line2 = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_line2") %}
								{% set city = frappe.db.get_value("Address", {'address_title': doc.customer}, "city") %}
								{% set state = frappe.db.get_value("Address", {'address_title': doc.customer}, "state") %}
								{% set country = frappe.db.get_value("Address", {'address_title': doc.customer}, "country") %}
								
								{% if address_title %}
									<br>{{ address_title }}
								{% endif %}
								{% if address_line1 %}
									<br>{{ address_line1 }}
								{% endif %}
								{% if address_line2 %}
									<br>{{ address_line2 }}
								{% endif %}
								{% if city %}
									<br>{{ city }}
								{% endif %}
								{% if state %}
									<br>{{ state }}
								{% endif %}
								{% if country %}
									<br>{{ country }}
								{% endif %}
							</td>

							
							<td style="width:50%">
								<b>Buyer if Other  than consignee</b>
								{% set address_title = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_title") %}
								{% set address_line1 = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_line1") %}
								{% set address_line2 = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_line2") %}
								{% set city = frappe.db.get_value("Address", {'address_title': doc.customer}, "city") %}
								{% set state = frappe.db.get_value("Address", {'address_title': doc.customer}, "state") %}
								{% set country = frappe.db.get_value("Address", {'address_title': doc.customer}, "country") %}
								
								{% if address_title %}
									<br>{{ address_title }}
								{% endif %}
								{% if address_line1 %}
									<br>{{ address_line1 }}
								{% endif %}
								{% if address_line2 %}
									<br>{{ address_line2 }}
								{% endif %}
								{% if city %}
									<br>{{ city }}
								{% endif %}
								{% if state %}
									<br>{{ state }}
								{% endif %}
								{% if country %}
									<br>{{ country }}
								{% endif %}
								
								
							</td>
						</tr>
						
						<tr>
							<td><b>Pre Carriage by:</b><br>
							
							{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"recommended_ffw")or ""}}
							
							
							</td>
							<td><b>Place of receipt by pre-carrier</b><br>
							{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"custom_place_of_receipt_by_precarrier") or ""}}
							
							
							</td>
							<td><b>Country of Origin of Goods</b><br>
								{% if doc.custom_cargo_mode == 'Air' %}
								{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"pol_country_airport")or ""}}
								{% elif doc.custom_cargo_mode == 'Sea' %}
								{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"pol_country_seaport")or ""}}
								{% else %}
								{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"pol_country_airport")or ""}}
								{% endif %}
							</td>
						</tr>
						
						<tr>
							<td>
								<b>Vessels / Flight No.</b><br>
								BY&nbsp;{{doc.custom_cargo_mode or ""}}
							</td>
							{% if doc.custom_cargo_mode == 'Sea' %}
							<td>
								<b>Port of Loading</b><br>
								{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"pol_city_seaport") or ""}}
							</td>
							{% else %}
							<td>
								<b>Port of Loading</b><br>
								{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"pol_city_airport") or ""}}
							</td>
							{% endif %}
							<td rowspan="3" style="border-bottom:none;">
								<b>Terms of Delivery & Payment</b><br>
								By&nbsp;{{doc.custom_cargo_mode or ""}}<br>
								
								<b>Payment</b><br>
								
								{{frappe.db.get_value("Customer",{'name':doc.customer},"payment_terms") or " "}}<br>
								
								<b>Incoterms</b><br>
								{{doc.incoterm or ""}}
							</td>
						</tr>
						<tr>
			{% if doc.custom_cargo_mode == 'Air' %}
				<td>
					<b>Port Of Discharge</b><br>
					{{ frappe.db.get_value("Logistics Request", {'order_no': doc.name}, "pod_city_airport") or "" }}
				</td>
				<td>
					<b>Final Destination</b><br>
					{{ frappe.db.get_value("Logistics Request", {'order_no': doc.name}, "final_destination") or "" }}
				</td>
			{% elif doc.custom_cargo_mode == 'Sea' %}
				<td>
					<b>Port Of Discharge</b><br>
					{{ frappe.db.get_value("Logistics Request", {'order_no': doc.name}, "pod_city_seaport") or "" }}
				</td>
				<td>
					<b>Final Destination</b><br>
					{{ frappe.db.get_value("Logistics Request", {'order_no': doc.name}, "final_destination") or "" }}
				</td>
			{% else %}
			<td>
					<b>Port Of Discharge</b><br>
					{{ frappe.db.get_value("Logistics Request", {'order_no': doc.name}, "pod_city_airport") or "" }}
				</td>
				<td>
					<b>Final Destination</b><br>
					{{ frappe.db.get_value("Logistics Request", {'order_no': doc.name}, "final_destination") or "" }}
				</td>
			{% endif %}
				
			
		</tr>

						
						<tr style="border-bottom:none !important;">
							<td colspan="2" style="border-bottom:none !important;">
								<b>General Description of the below mentioned Products is Automobile Components</b>
							</td>
						</tr>
				</table>
				</div>
			</div>
		</div>


		<div class="container ">
			<div class="row">
				<div class="col-xs-12 p-0">
					<table width=100% style="border-collapse:collapse;  " class="tab2 mr-0 ml-0">
					<thead style="display: table-header-group;" >    
						<tr style="border-bottom:1px solid black;">
							<td style="width:5%; vertical-align: middle; text-align: center;"><b>Marks & Nos</b></td>
							<td style="width:5%; vertical-align: middle; text-align: center;"><b>No of packages</b></td>
						
							<td rowspan="2" style="width:10%; vertical-align: middle; text-align: center;font-size:12px;"><b>Item Code</b></td>
							<td rowspan="2" style="width:38%; vertical-align: middle; text-align: center;font-size:12px;"><b>Description of Goods</b></td>
							<td rowspan="2" style="width:14%; vertical-align: middle; text-align: center;font-size:12px;"><b>Order No. & PO</b></td>
							<td rowspan="2" style="width:10%; vertical-align: middle; text-align: center;font-size:12px;"><b>HSN Code</b></td>
							<td rowspan="2" style="width:5%; vertical-align: middle; text-align: center;font-size:12px;"><b>Quality Nos</b></td>
							<td rowspan="2" style="width:5%; vertical-align: middle; text-align: center;font-size:12px;"><b>Rate&nbsp; {{doc.currency or ""}}</b></td>
							<td rowspan="2" style="width:10%; vertical-align: middle; text-align: center;font-size:12px;"><b>Amount&nbsp; {{doc.currency or ""}}</b></td>
						</tr>

						<tr style="border-bottom:1px solid black;">
							<td style="width:5%;vertical-align: middle;"><b><center>PD</center></b></td>
							<td style="width:5%;vertical-align: middle;"><b><center>Boxes</center></b></td>
						</tr>
						
					</thead>    
						{%- for i in doc["items"] -%}
						<tr style="{% if loop.last %}border-bottom:1px solid black;{% else %}border-bottom:none !important;{% endif %} border-top:none;">
							<td class="no-border-1" style="text-align:center;font-size:12px;">{{ i.custom_no_of_pallets or "" }}</td>
							<td class="no-border-1" style="text-align:center;font-size:12px;">{{ i.custom_no_of_boxes or "" }}</td>
							<td class="no-border-1" style="font-size:12px;">{{ i.item_code or "" }}</td>
							<td class="no-border-1" style="font-size:12px;">{{ i.description or "" }}</td>
							<td class="no-border-1" style="font-size:12px;">{{ i.custom_customers_po_number or "" }}</td>
							<td class="no-border-1" style="text-align:center;font-size:12px;">{{ i.gst_hsn_code or "" }}</td>
							<td class="no-border-1" style="text-align:center;font-size:12px;">{{ i.qty or "" }}</td>
							<td class="no-border-1" style="text-align:right;font-size:12px;">{{ "%.4f"|format(i.rate or 0) }}</td>
							<td class="no-border-1" style="text-align:right;font-size:12px;">{{ "%.2f"|format(i.amount or 0) }}</td>
						</tr>
						{%- endfor -%}
						<tr style="border-bottom:none !important; border-left:none; border-right:none;font-size:13px; " >
							<td colspan="6" class="no-border">
								This Shipment under duty draw back 
							</td>
							<td colspan="1" style="text-align:right;" class="no-border" >
								<b>{{doc.total_qty or ""}}</b>
							</td>
							<td colspan="2" class="no-border">
								
							</td>
						</tr>
						
						<tr style="border-top:none !important; border-bottom:none;font-size:13px;">
							<td colspan="3" style="text-align:right; " class="no-border">
								<b><u>COUNTRY OF ORIGIN :{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"pol_country_airport")or ""}}</u></b>
							</td>
							<td colspan="6" class="no-border">
								
							</td>
						</tr>
						
						<tr style="border-top:none !important; border-bottom:none;font-size:13px;">
							
							<td colspan="6" class="no-border">
								<b>Amount Chargable (in Words)</b>
							</td>
							<td colspan="1" class="no-border" style="text-align:right;">
								<b>TOTAL</b>
							</td>
							<td colspan="1" class="no-border" style="text-align:right;">
								<b>{{doc.currency or ""}}</b>
							</td>
							<td colspan="1" class="no-border" style="text-align:right;">
								<b>{{doc.grand_total or ""}}</b>
							</td>
							
						</tr>
						
						<tr style="border-top:none !important;border-bottom:1px solid black;font-size:13px;">
							<td colspan="9" class="no-border">
								{{doc.in_words or ""}}
							</td>
						</tr>
					</table>
					
				</div>
			</div>
			
		</div>

		<style>
			@media print {
				.declaration-container {
					page-break-inside: avoid;
					break-inside: avoid;
				}

				.declaration-row {
					border: 1px solid black;
					border-top: 1px solid black;
				}

				/* If using wkhtmltopdf, avoid relying on @page:first */
			}

			.qrcode {
				width: 120px;
				height: 120px;
			}

			.text-center {
				text-align: center;
			}
		</style>

		{% set is_first_page = false %} {# Or dynamically determine this in your logic #}

		<div class="container declaration-container">
			<div class="row declaration-row" style="border: 1px solid black; {{ 'border-top:none;' if is_first_page else 'border-top:none;' }}">
				<div class="col-12 p-0">
					<strong><u>DECLARATION</u></strong><br>
					We declare that this packing list shows the actual<br>
					Weight of the goods<br>
					Described and that all particulars are true and correct
				</div>
			</div>


		</div>
		</div>
		<table style="width:100%; border:1px solid black; border-collapse:collapse; font-size:9px; page-break-inside:avoid;">
	<tr >
		<td colspan="3" style="padding:2px;">
			<strong>LUT NO: {{ frappe.db.get_value("Customer", {"name": doc.customer}, "custom_lut_number") or "" }}</strong>
			<span style="float:right; white-space:nowrap;">For <strong>{{ doc.company }}</strong></span>
		</td>
	</tr>

	{% set prepared_by = frappe.db.get_value("Employee", {"user_id": doc.owner}, ["custom_digital_signature"]) %}
	{% set hod = frappe.db.get_value("Employee", {"user_id": doc.custom_hod}, ["custom_digital_signature"]) %}
	{% set finance = frappe.db.get_value("Employee", {"name": 'S0189'}, ["custom_digital_signature"]) %}

	<style>
		.signature-img {
			width: 100px;
			height: 50px;
			object-fit: contain;
			border: none;
		}
	</style>

	<!-- Signatures Row -->
	<tr style="height:60px; text-align:center;">
		<td style="vertical-align:bottom;">
			{% if prepared_by %}
				<img src="{{ prepared_by }}" class="signature-img">
			{% endif %}
		</td>
		<td style="vertical-align:bottom;">
			{% if hod and doc.workflow_state not in ['Pending For HOD', 'Draft'] %}
				<img src="{{ hod }}" class="signature-img">
			{% endif %}
		</td>
		<td style="vertical-align:bottom;">
			{% if finance and doc.workflow_state in ['Approved', 'Dispatched'] %}
				<img src="{{ finance }}" class="signature-img">
			{% endif %}
		</td>
	</tr>

	<!-- Dates Row -->
	<tr style="height:10px; text-align:center;">
		<td>{{ frappe.utils.format_datetime(doc.creation, "dd-MM-yyyy HH:mm:ss") or '' }}</td>
		<td>{{ frappe.utils.format_datetime(doc.custom_hod_approved_on, "dd-MM-yyyy HH:mm:ss") or '' }}</td>
		<td>{{ frappe.utils.format_datetime(doc.custom_finance_approver_approved_on, "dd-MM-yyyy HH:mm:ss") or '' }}</td>
	</tr>

	<!-- Labels Row -->
	<tr style="font-weight:bold; text-align:center; height:20px;">
		<td>Prepared by</td>
		<td>Checked by</td>
		<td>Authorised Signatory</td>
	</tr>
	<tr>
		<td colspan="3" style="padding:2px;border-top:1px solid black;">
			<strong>Note:</strong>This document is Digitally Signed.
		</td>
	</tr>
</table>

		{% if not loop.last %}
		<div class="page-break"></div>
		{% endif %}
		{% endfor %}    
	"""
		
	html = render_template(template, {"doc": doc})
	return {"html": html}


@frappe.whitelist()
def create_html_annexure(doc):
	doc = frappe.parse_json(doc)
	template = """
		<div class="col-12-xs">
		<table border=1 class="tab1" width=100%>
			<tr>
				<td colspan=7 style="text-align:center;font-size:16px;"><b>ANNEXURE</b></td>
			</tr>
			
			<tr>
				<td colspan=7 style="text-align:center;font-size:16px;">
					<b>Annexure for our Packing List & Inv. No. {{ doc.name }} Dt : {{ frappe.format(doc.posting_date, {'fieldtype': 'Date'}) }}</b>
				</td>
			</tr>
			
			<tr style="font-weight:bold; text-align:center;">
				<td>PART NO</td>
				<td>PART NAME</td>
				<td>Pallet No</td>
				<td>BOXES/PALLET</td>
				<td>PARTS/PALLET</td>
				<td>PARTS/BOX</td>
				<td>NET WEIGHT</td>
				<td>GROSS WEIGHT</td>
			</tr>

			{% set ns = namespace(total_pallets=0, total_boxes=0, total_parts=0, total_net=0, total_gross=0, pallet_counter=1) %}

			{% for row in doc["items"] %}
				{% set pallets = row.custom_no_of_pallets or 0 %}
		{% set boxes = row.custom_no_of_boxes or 0 %}
		{% set parts = row.parts_per_box or 0 %}
		{% set part_per_box = parts / boxes or 0 %}
		{% set net = row.total_weight or 0 %}
		{% set gross = row.custom_gross_weight or 0 %}
	
		{% if pallets > 0 %}
			{% set base_boxes = boxes // pallets %}
			{% set remainder_boxes = boxes % pallets %}
	
			{% for i in range(pallets) %}
				{% if i < pallets - 1 %}
					{% set current_boxes = base_boxes %}
				{% else %}
					{% set current_boxes = base_boxes + remainder_boxes %}
				{% endif %}
	
				<tr style="text-align:center;">
					{% if i == 0 %}
						<td rowspan="{{ pallets }}" style="text-align:left;">{{ row.item_code }}</td>
						<td rowspan="{{ pallets }}" style="text-align:left;">{{ row.item_name }}</td>
					{% endif %}
					<td>{{ ns.pallet_counter }}</td>
					<td>{{ current_boxes }}</td>
					<td>
					  {{ ('%.2f' % (row.qty / row.custom_no_of_pallets|float)).rstrip('0').rstrip('.') if row.custom_no_of_pallets else '' }}
					</td>
					<td>
					  {{ ('%.2f' % (row.qty / row.custom_no_of_boxes|float)).rstrip('0').rstrip('.') if row.custom_no_of_boxes else '' }}
					</td>
					<td>{{ "%.2f"|format(net / pallets) }}</td>
					<td>{{ "%.2f"|format(gross / pallets) }}</td>
				</tr>
	
				{% set ns.total_boxes = ns.total_boxes + current_boxes %}
				{% set ns.total_parts = ns.total_parts + row.qty/row.custom_no_of_boxes %}
				{% set ns.pallet_counter = ns.pallet_counter + 1 %}
			{% endfor %}
			{% set ns.total_pallets = ns.total_pallets + pallets %}
			{% set ns.total_net = ns.total_net + net %}
			{% set ns.total_gross = ns.total_gross + gross %}
			
		{% else %}
			<tr style="text-align:center;">
				<td style="text-align:left;">{{ row.item_code }}</td>
				<td style="text-align:left;">{{ row.item_name }}</td>
				<td>0</td>
				<td>{{ boxes }}</td>
				<td>{{ ("%.2f"|format(parts|float)).rstrip('0').rstrip('.') if parts else "" }}</td>
				<td>{{ ("%.2f"|format(part_per_box|float)).rstrip('0').rstrip('.') if part_per_box else ""}}</td>
				<td>{{ "%.2f"|format(net) }}</td>
				<td>{{ "%.2f"|format(gross) }}</td>
			</tr>
			{% set ns.total_boxes = ns.total_boxes + boxes %}
			{% set ns.total_parts = ns.total_parts + (parts / boxes) %}
			{% set ns.total_net = ns.total_net + net %}
			{% set ns.total_gross = ns.total_gross + gross %}
		{% endif %}
	{% endfor %}


	<tr style="font-weight:bold; text-align:center;">
		<td colspan=2>GRAND TOTAL</td>
		<td>{{ ns.total_pallets }}</td>
		<td>{{ ns.total_boxes }}</td>
		<td>{{ ("%.2f"|format(doc.total_qty|float)).rstrip('0').rstrip('.') if doc.total_qty else ""}}</td>
		<td>{{ ("%.2f"|format(ns.total_parts|float)).rstrip('0').rstrip('.') if ns.total_parts else "" }}</td>
		<td>{{ "%.2f"|format(ns.total_net) }}</td>
		<td>{{ "%.2f"|format(ns.total_gross) }}</td>
	</tr>
</table>

		</div>
	"""
	html = render_template(template, {"doc": doc})
	return {"html": html}


@frappe.whitelist()
def create_html_pallet(doc):
	doc = frappe.parse_json(doc)
	if "posting_date" in doc:
		try:
			doc["posting_date"] = datetime.strptime(doc["posting_date"], "%Y-%m-%d").strftime("%d/%m/%Y")
		except Exception:
			pass
	template = """
		<meta name="pdfkit-orientation" content="Landscape" />
		<meta name="page-size" content="A4" />

		<style>
			@import url('https://fonts.googleapis.com/css2?family=Chonburi&display=swap');

			@media print {
				body {
					margin: 0 !important;
					padding: 0 !important;
					margin-left: 150px;
					margin-right: 50px;
				}
			}
			.tab1,
			.tab1 tr,
			.tab1 td{
				border:1px solid black;
			}

			.page-break {
				page-break-after: always;
			}

			.tab1 {
				border-collapse: collapse;
				width: 100%;
			}

			.tab1 td, .tab1 th {
				border: 1px solid black;
				padding: 4px;
			}
		</style>

		<div class="col-12-xs">

			{% set pallet_row = namespace(no=1) %}
			{% set box_row = namespace(no=1) %}
			{% for i in doc["items"] %}
				{% set row = namespace(no=1) %}
				{% set total_qty = i.qty or 0 %}
				{% set total_boxes = i.custom_no_of_boxes or 1 %}
				{% set total_pallets = i.custom_no_of_pallets or 1 %}

				{% set boxes_per_pallet = (total_boxes // total_pallets) %}
				{% set remaining_boxes = (total_boxes % total_pallets) %}

				{% set qty_per_box = total_qty / total_boxes %}
				{% set net_we_per_box = qty_per_box * (i.weight_per_unit or 0) %}
				{% set gross_we_per_box = net_we_per_box + (i.custom_weight_per_unit_b or 0) %}

				{% for pallet_index in range(1, total_pallets + 1) %}
					{% set pallet_box_count = boxes_per_pallet + (1 if pallet_index <= remaining_boxes else 0) %}
					{% set pallet_qty = pallet_box_count * qty_per_box %}
					{% set pallet_net_we = pallet_box_count * net_we_per_box %}
					{% set pallet_gross_we = pallet_box_count * gross_we_per_box %}

					{% if pallet_box_count <= 12 %}
						{% set font_size = 15 %}
					{% elif 12 < pallet_box_count <= 20 %}
						{% set font_size = 12 %}
					{% else %}
						{% set font_size = 18 - ((pallet_box_count - 12) * 0.5) %}
						{% if font_size < 10 %}
							{% set font_size = 10 %}
						{% endif %}
					{% endif %}

					<table class="tab1" style="font-size: {{ font_size }}px;">

						<tr style="text-align:center; font-size:14px; font-weight:bold;">
							<td colspan="10">
								<b>PALLET - {{ "%02d"|format(pallet_row.no) }}</b>
							</td>
						</tr>

						<tr style="font-weight:bold; text-align:center; font-size:14px;">
							<td colspan="10">
								<b>Annexure for our Packing List & Inv. No. {{ doc.name }} Dt: {{ doc["posting_date"]}}</b>
							</td>
						</tr>

						<tr style="text-align:center; font-size:14px; font-weight:bold;">
							<td colspan="10">
								{% set idx_no = "%02d"|format(row.no or 0) %}
								<b>
						  PALLET - {{ "%02d"|format(pallet_row.no) }}
						  &nbsp;({{ "%02d"|format(box_row.no) }}-{{ "%02d"|format(pallet_box_count + box_row.no - 1) }} C/N Boxes)
						  &nbsp;(L&nbsp;
						  {{ ("%.2f"|format(i.custom_pallet_length|float)).rstrip('0').rstrip('.') if i.custom_pallet_length else "" }}
						  &nbsp;mm X W&nbsp;
						  {{ ("%.2f"|format(i.custom_pallet_breadth|float)).rstrip('0').rstrip('.') if i.custom_pallet_breadth else "" }}
						  &nbsp;mm X H&nbsp;
						  {{ ("%.2f"|format(i.custom_pallet_height|float)).rstrip('0').rstrip('.') if i.custom_pallet_height else "" }}
						  &nbsp;mm)
						</b>
							</td>
						</tr>

						<tr style="text-align:center; font-size:14px; font-weight:bold;">
							<td colspan="10">
								<b>{{ i.item_code or "" }}</b>
							</td>
						</tr>

						<tr style="text-align:center; font-weight:bold;">
							<td style="width:10%;">Carton Nos.</td>
							<td colspan="2">Quantity</td>
							<td>Description of Goods</td>
							<td>Net Weight in <br>Kgs.</td>
							<td>Gross Weight in <br>Kgs.</td>
							<td>MEASUREMENT (in mm)</td>
						</tr>

						{% for b in range(1, pallet_box_count + 1) %}
							<tr style="font-weight:bold;font-size:10px;">
								<td>C/NO. {{ "%02d"|format(row.no) }}</td>
								<td>{{ qty_per_box }}</td>
								<td>{{ i.uom or "" }}</td>
								<td>{{ i.description or "" }}</td>
								<td style="text-align:center;">{{ '{0:.2f}'.format(net_we_per_box) }}</td>
								<td style="text-align:center;">{{ '{0:.2f}'.format(gross_we_per_box) }}</td>
								<td>{{ ("%.2f"|format(i.custom_box_length|float)).rstrip('0').rstrip('.') if i.custom_box_length else "" }}X{{ ("%.2f"|format(i.custom_box_breadth|float)).rstrip('0').rstrip('.') if i.custom_box_breadth else "" }}X{{ ("%.2f"|format(i.custom_box_height|float)).rstrip('0').rstrip('.') if i.custom_box_height else "" }}</td>
							</tr>
							{% set row.no = row.no + 1 %}
							{% set box_row.no = box_row.no + 1 %}
						{% endfor %}

						<tr>
							<td></td>
							<td></td>
							<td></td>
							<td colspan="2" style="font-weight:bold; text-align:center;"><b>PALLET & LAYER PAD WEIGHT</b></td>
							<td style="text-align:center; font-weight:bold;">
			{{ '{0:.2f}'.format(i.custom_weight_per_unit_p) or  0 }}
		</td>

							<td></td>
						</tr>

						<tr style="font-weight:bold;">
							<td><b>Total</b></td>
							<td>{{ pallet_qty }}</td>
							<td>{{ i.uom or "" }}</td>
							<td></td>
							<td style="text-align:center;"><b>{{'{0:.2f}'.format(pallet_net_we) or 0}}</b></td>
							<td style="text-align:center;">{{'{0:.2f}'.format(pallet_gross_we + i.custom_weight_per_unit_p) or 0}}</td>
							<td></td>
						</tr>

					</table>
					{% set pallet_row.no = pallet_row.no + 1 %}
					<div class="page-break"></div>
				{% endfor %}
			{% endfor %}
		</div>

	"""
	html = render_template(template, {"doc": doc})
	return {"html": html}


@frappe.whitelist()
def create_html_address(doc):
	doc = frappe.parse_json(doc)
	template = """
		<meta name="pdfkit-orientation" content="Landscape">
		<meta name="page-size" content="A4"/>

		{% set pallet_counter = {'val': 1} %}
		{% set box_counter = {'val': 1} %}
		{% for item in doc["items"] %}
			{% set total_boxes = item.custom_no_of_boxes %}
			{% set total_pallets = item.custom_no_of_pallets %}

			{% if total_pallets and total_pallets > 0 %}
				{% set base_boxes_per_pallet = (total_boxes // total_pallets) %}
				{% set remainder = total_boxes % total_pallets %}

				{% for i in range(total_pallets) %}
					{% set extra_box = 1 if i < remainder else 0 %}
					{% set boxes_this_pallet = base_boxes_per_pallet + extra_box %}

					{% set start_box = box_counter.val %}
					{% set end_box = start_box + boxes_this_pallet - 1 %}

					<table style="width: 100%; border-collapse: collapse; page-break-after: always;" border="1">
						<tr>
							<td style="vertical-align: top; padding: 10px; font-size: 21px;width:60%;" rowspan="3">
								<b>To,</b><br>
								{{doc.customer}},<br>{{ (frappe.db.get_value("Customer", {"name": doc.customer}, "primary_address") or "")}}
							</td>
							<td style="padding: 10px; font-size: 20px; text-align: center; font-weight: bold; text-transform: uppercase;">INVOICE</td>
							<td style="padding: 10px; font-size: 20px; text-align: center; font-weight: bold; text-transform: uppercase;">{{ doc.name }}</td>
						</tr>
						<tr>
							<td colspan="2" style="padding: 20px 10px; font-size: 18px; text-align: center; font-weight: bold; text-transform: uppercase;">
								PALLET - {{ pallet_counter.val }}
							</td>
						</tr>
						<tr>
							<td colspan="2" style="padding: 20px 10px; font-size: 18px; text-align: center; font-weight: bold; text-transform: uppercase;">
								{{ doc.custom_customer_code }}/{{ "%02d" % start_box }} - {{ "%02d" % end_box }}
							</td>
						</tr>
						{% set address_1 = frappe.db.get_value("Address",{'name':doc.company_address},'address_line1') %}
				{% set address_2 = frappe.db.get_value("Address",{'name':doc.company_address},'address_line2') %}
				{% set city = frappe.db.get_value("Address",{'name':doc.company_address},'city') %}
				{% set country = frappe.db.get_value("Address",{'name':doc.company_address},'country') %}
				{% set state = frappe.db.get_value("Address",{'name':doc.company_address},'state') %}
				{% set postalcode = frappe.db.get_value("Address",{'name':doc.company_address},'pincode') %}
				{% set email = frappe.db.get_value("Address",{'name':doc.company_address},'email_id') %}
				{% set phone = frappe.db.get_value("Address",{'name':doc.company_address},'phone') %}
				{% set fax = frappe.db.get_value("Address",{'name':doc.company_address},'fax') %}
				<tr>
					<td style="vertical-align: top; padding: 10px; font-size: 18px;">
						<b>From,</b><br>
						{{doc.company or ''}}<br>
						{{address_1 or ''}},<br>
						{{address_2}}, {{city or ''}}, <br>
						{{country or ''}}, {{state or ''}} - {{postalcode or ''}}<br>
						Phone: {{phone or ''}}<br>
						Fax: {{fax or ''}}<br>
						Email: {{email or ''}}<br>
					</td>
					<td colspan="2" style="padding: 40px 40px; font-size: 18px; text-align: center; font-weight: bold; text-transform: uppercase;">
						<br>MADE IN INDIA
					</td>
				</tr>
					</table>
					<table style="width: 100%; border-collapse: collapse;" border="1">
				<tr>
					<td style="vertical-align: center; padding: 10px; font-size: 21px; width: 40%; border-top: none;text-align:center;"><b>{{ item.item_code }} - {{ item.item_name }}</b></td>
					<!--<td colspan="2" style="vertical-align: top; padding: 10px; font-size: 28px; border-top: none;">-->
					<!--    <b>{{ item.item_name }}</b>-->
					<!--</td>-->
				</tr>
			</table><br>

					{% set _ = pallet_counter.update({'val': pallet_counter.val + 1}) %}
					{% set _ = box_counter.update({'val': end_box + 1}) %}
				{% endfor %}
			{% else %}
				<center>No Pallet Found</center>
			{% endif %}
		{% endfor %}

	"""
	html = render_template(template, {"doc": doc})
	return {"html": html}


@frappe.whitelist()
def create_html_packing(doc):
	doc = frappe.parse_json(doc)
	template = """
		<style>
		.tab1,
		.tab1 tr,
		.tab1 td{
			border:1px solid black;
		}

		.tab2,
		.tab2 tr,
		.tab2 td{
			border:1px solid black;
			border-bottom:none;

		}

		.tab3,
		.tab3 tr,
		.tab3 td{
			border:1px solid black;
			border-bottom:none;
			border:none;
		}

		.no-border{
			border:none !important;
		}

		.no-border-1{
			border-top:none !important;
			border-bottom:none !important;
		}

		/*.tab2 tr{*/
		/*    page-break-inside:avoid; */
		/*    page-break-after:auto;*/
		/*}*/


		.page-break {
			page-break-before: always;
			break-before: page;
		}

		.in-words-row {
			page-break-inside: avoid;
			break-inside: avoid;
			border-bottom:none;
			border-top: 1px solid black;
		}


		</style>
		<div class="container">
			<div class="row ">
				<div class="col-xs-12 p-0 " >
					
				<table width=100% style="border-collapse:collapse;border-bottom: none;" class="tab1 mr-0 ml-0" >
					<tr>
							<td colspan="3">
							<center><p style="font-size:20px;"><b>PACKING LIST</b></p></center> 
							</td>
						</tr>
					<tr>
						<td rowspan="4" style="width:30%" >
								<p><b>Factory Address</b></p>
								<p>{{frappe.db.get_value("Address",{'address_title':doc.company},"address_title")or ""}}</p>
								<p>{{frappe.db.get_value("Address",{'address_title':doc.company},"address_line1") or ""}}</p>
								<p>{{frappe.db.get_value("Address",{'address_title':doc.company},"address_line2") or ""}}</p>
								<p>{{frappe.db.get_value("Address",{'address_title':doc.company},"city")or ""}}</p>
								<p>{{frappe.db.get_value("Address",{'address_title':doc.company},"state")or ""}}</p>
								<p>{{frappe.db.get_value("Address",{'address_title':doc.company},"country")or ""}}</p>
								<p>{{frappe.db.get_value("Address",{'address_title':doc.company},"gstin")or ""}}</p>
							</td>
							<td rowspan="4"  style="width:30%"> 
								<p>
									<b>Exporter</b></p>
									<p>{{frappe.db.get_value("Address",{'address_title':doc.company},"address_title")or ""}}</p>
								<p>{{frappe.db.get_value("Address",{'address_title':doc.company},"address_line1") or ""}}</p>
								<p>{{frappe.db.get_value("Address",{'address_title':doc.company},"address_line2") or ""}}</p>
								<p>{{frappe.db.get_value("Address",{'address_title':doc.company},"city")or ""}}</p>
								<p>{{frappe.db.get_value("Address",{'address_title':doc.company},"state")or ""}}</p>
								<p>{{frappe.db.get_value("Address",{'address_title':doc.company},"gstin")or ""}}</p>
								
								
								
							</td>
							<td  class="col-xs-5" >
								<span><b>Invoice No & Date:</b></span>
								<span style="float:right;"><b>{{doc.name}}&nbsp;&nbsp;{{ frappe.format(doc.posting_date) or ""  }}</b></span>
								
								
								<span><b>Exporter IEC</b></span>
								<span style="float:right;"><b>{{frappe.db.get_value('Company',{'name':doc.company},"custom_exporter_iec")or ""}}</b></span>

								<br><span><b>CIN NO</b></span>
								<span stylr="float:right;">{{frappe.db.get_value('Customer',{'name':doc.customer},"custom_cin_number")or ""}}</span>
							</td>
					</tr>
					
				
						
						<!--<tr>-->
						<!--    <td colspan="3"><b><center><b>INVOICE</b></center></b></td>-->
						<!--</tr>-->
						<!--<tr>-->
						<!--    <td colspan="2" style="width:60%">-->
						<!--        E-WAY No. &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:{{doc.ewaybill or ""}}<br>-->
						<!--        E_WAY Date &nbsp;&nbsp;:{{frappe.db.get_value("e-Waybill Log",{"name":doc.ewaybill},"created_on") or ""}}<br>-->
						<!--        ACK No. &nbsp;&nbsp;&nbsp; &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:{{frappe.db.get_value('e-Invoice Log', {'sales_invoice': doc.name}, 'acknowledgement_number') or ""}}<br>-->
						<!--        ACK Date &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:{{frappe.db.get_value('e-Invoice Log', {'sales_invoice': doc.name}, 'acknowledged_on') or ""}}<br>-->
						<!--        IRN No. &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:{{doc.irn or ""}}-->
								
						<!--    </td>-->
						<!--    <td style="width:40%">-->
								
						<!--        Invoice No  &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; : {{doc.name}}<br>-->
						<!--        Invoice Date &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {{frappe.format(doc.posting_date,{'fieldtype':'Date'}) or ""}}<br>-->
						<!--        Exporter IEC &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {{frappe.db.get_value('Company',{'name':doc.company},"custom_exporter_iec")or ""}}<br>-->
							<!--Order No. & PO :-->
							<!--          {% for row in doc.custom_sales_orders %}-->
							<!--              <span>{{ row.sales_order or "" }} {{ row.po_no or "" }}</span>{% if not loop.last %}, {% endif %}-->
							<!--          {% endfor %}-->
							<!--      <br>-->

						<!--        Vendor Code &nbsp;&nbsp;&nbsp;&nbsp;:{{frappe.db.get_value('Company',{'name':doc.company},"custom_vendor_code")or ""}}<br>-->
						<!--        CIN No&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:{{frappe.db.get_value('Company',{'name':doc.company},"custom_cin_number")or ""}}<br>-->
						<!--        GSTIN No &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:{{frappe.db.get_value('Address',{'address_title':doc.company},'gstin') or ""}} -->
								
						<!--    </td>-->
						<!--</tr>-->
						
				</table> 
				<table width=100% style="border-collapse:collapse;margin-top:5px;padding:0px;border-bottom:none;" class="tab1 mr-0 ml-0">
					
						<tr>
							<td colspan="2" style="width:50%">
								<b>Consignee</b>
								{% set address_title = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_title") %}
								{% set address_line1 = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_line1") %}
								{% set address_line2 = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_line2") %}
								{% set city = frappe.db.get_value("Address", {'address_title': doc.customer}, "city") %}
								{% set state = frappe.db.get_value("Address", {'address_title': doc.customer}, "state") %}
								{% set country = frappe.db.get_value("Address", {'address_title': doc.customer}, "country") %}
								
								{% if address_title %}
									<br>{{ address_title }}
								{% endif %}
								{% if address_line1 %}
									<br>{{ address_line1 }}
								{% endif %}
								{% if address_line2 %}
									<br>{{ address_line2 }}
								{% endif %}
								{% if city %}
									<br>{{ city }}
								{% endif %}
								{% if state %}
									<br>{{ state }}
								{% endif %}
								{% if country %}
									<br>{{ country }}
								{% endif %}
							</td>

							
							<td style="width:50%">
								<b>Buyer if Other  than consignee</b>
								{% set address_title = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_title") %}
								{% set address_line1 = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_line1") %}
								{% set address_line2 = frappe.db.get_value("Address", {'address_title': doc.customer}, "address_line2") %}
								{% set city = frappe.db.get_value("Address", {'address_title': doc.customer}, "city") %}
								{% set state = frappe.db.get_value("Address", {'address_title': doc.customer}, "state") %}
								{% set country = frappe.db.get_value("Address", {'address_title': doc.customer}, "country") %}
								
								{% if address_title %}
									<br>{{ address_title }}
								{% endif %}
								{% if address_line1 %}
									<br>{{ address_line1 }}
								{% endif %}
								{% if address_line2 %}
									<br>{{ address_line2 }}
								{% endif %}
								{% if city %}
									<br>{{ city }}
								{% endif %}
								{% if state %}
									<br>{{ state }}
								{% endif %}
								{% if country %}
									<br>{{ country }}
								{% endif %}
								
								
							</td>
						</tr>
						
						<tr>
							<td><b>Pre Carriage by:</b><br>
							
							{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"recommended_ffw")or ""}}
							
							
							</td>
							<td><b>Place of receipt by pre-carrier</b><br>
							{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"custom_place_of_receipt_by_precarrier") or "NA"}}
							
							
							</td>
							<td><b>Country of Origin of Goods</b><br>
							{% if doc.custom_cargo_mode == 'Air' %}
								<span style="position:absolute;top:23px; left:7px; bottom:2px; ">{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"pol_country_airport")or ""}}</span>
								{% elif doc.custom_cargo_mode == 'Sea' %}
								<span style="position:absolute;top:23px; left:7px; bottom:2px; ">{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"pol_country_seaport")or ""}}</span>
								{% else %}
								<span style="position:absolute;top:23px; left:7px; bottom:2px; ">{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"pol_country_airport")or ""}}</span>
								{% endif %}</td>
							
							</td>
						</tr>
						
						<tr>
							<td>
								<b>Vessels / Flight No.</b><br>
								BY&nbsp;{{doc.custom_cargo_mode or ""}}
							</td>
							{% if doc.custom_cargo_mode == 'Sea' %}
							<td>
								<b>Port of Loading</b><br>
								{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"pol_city_seaport") or ""}}
							</td>
							{% else %}
							<td>
								<b>Port of Loading</b><br>
								{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"pol_city_airport") or ""}}
							</td>
							{% endif %}
							<td rowspan="3" style="border-bottom:none;">
								<b>Terms of Delivery & Payment</b><br>
								By&nbsp;{{doc.custom_cargo_mode or ""}}<br>
								
								<b>Payment</b><br>
								
								{{frappe.db.get_value("Customer",{'name':doc.customer},"payment_terms") or " "}}<br>
								
								<b>Incoterms</b><br>
								{{doc.incoterm or ""}}
							</td>
						</tr>
						<tr>
							{% if doc.custom_cargo_mode == 'Air' %}
								<td>
									<b>Port Of Discharge</b><br>
									{{ frappe.db.get_value("Logistics Request", {'order_no': doc.name}, "pod_city_airport") or "" }}
								</td>
								<td>
									<b>Final Destination</b><br>
									{{ frappe.db.get_value("Logistics Request", {'order_no': doc.name}, "final_destination") or "" }}
								</td>
							{% elif doc.custom_cargo_mode == 'Sea' %}
								<td>
									<b>Port Of Discharge</b><br>
									{{ frappe.db.get_value("Logistics Request", {'order_no': doc.name}, "pod_city_seaport") or "" }}
								</td>
								<td>
									<b>Final Destination</b><br>
									{{ frappe.db.get_value("Logistics Request", {'order_no': doc.name}, "final_destination") or "" }}
								</td>
							{% else %}
							<td>
									<b>Port Of Discharge</b><br>
									{{ frappe.db.get_value("Logistics Request", {'order_no': doc.name}, "pod_city_airport") or "" }}
								</td>
								<td>
									<b>Final Destination</b><br>
									{{ frappe.db.get_value("Logistics Request", {'order_no': doc.name}, "final_destination") or "" }}
								</td>
							{% endif %}
						</tr>
						
						<tr style="border-bottom:none !important;">
							<td colspan="2" style="border-bottom:none !important;">
								<b>General Description of the below mentioned Products is Automobile Components</b>
							</td>
						</tr>
				</table>
				</div>
			</div>
		</div>


		<div class="container  " style="border-top:none;">
			<div class="row" style="border-top:none;" >
				<div class="col-xs-12 p-0" style="border-top:none;">
					<table width=100% style="border-collapse:collapse;border-top:none;  " class="tab2 mr-0 ml-0">
					<thead style="display: table-header-group;" >    
						<tr style="border-bottom:1px solid black;border-top:none;">
							<td style="width:5%; vertical-align: middle; text-align: center;"><b>Marks & Nos</b></td>
							<td style="width:5%; vertical-align: middle; text-align: center;"><b>No of packages</b></td>
						
							<td rowspan="2" style="width:10%; vertical-align: middle; text-align: center;"><b>Item Code</b></td>
							<td rowspan="2" style="width:40%; vertical-align: middle; text-align: center;"><b>Description of Goods</b></td>
							<td rowspan="2" style="width:10%; vertical-align: middle; text-align: center;"><b>Order No. & PO</b></td>
							<td rowspan="2" style="width:10%; vertical-align: middle; text-align: center;"><b>HSN Code</b></td>
							<td rowspan="2" style="width:5%; vertical-align: middle; text-align: center;"><b>Quality Nos</b></td>
							<td rowspan="2" style="width:5%; vertical-align: middle; text-align: center;"><b>Rate&nbsp; {{doc.currency or ""}}</b></td>
							<td rowspan="2" style="width:10%; vertical-align: middle; text-align: center;"><b>Amount&nbsp; {{doc.currency or ""}}</b></td>
						</tr>

						<tr style="border-bottom:1px solid black;">
							<td style="width:5%;vertical-align: middle;"><b><center>PD</center></b></td>
							<td style="width:5%;vertical-align: middle;"><b><center>Boxes</center></b></td>
						</tr>
						
					</thead>    
					{% set total_weight = 0 %}
		{% set custom_no_of_pallets = 0 %}
		{% set custom_no_of_boxes = 0 %}

		{%- for i in doc["items"] -%}
		<tr style="{% if loop.last %}border-bottom:1px solid black;{% else %}border-bottom:none !important;{% endif %} border-top:none;">
			<td class="no-border-1" style="text-align:center;">{{ i.custom_no_of_pallets or "" }}</td>
			<td class="no-border-1" style="text-align:center;">{{ i.custom_no_of_boxes or "" }}</td>
			<td class="no-border-1">{{ i.item_code or "" }}</td>
			<td class="no-border-1">{{ i.description or "" }}</td>
			<td class="no-border-1">{{ i.custom_customers_po_number or "" }}</td>
			<td class="no-border-1" style="text-align:center;">{{ i.gst_hsn_code or "" }}</td>
			<td class="no-border-1" style="text-align:center;">{{ i.qty or "" }}</td>
			<td class="no-border-1" style="text-align:right;">{{ i.rate or "" }}</td>
			<td class="no-border-1" style="text-align:right;">{{ i.amount or "" }}</td>
		</tr>

		{% set total_weight = total_weight + (i.total_weight or 0) %}
		{% set custom_no_of_pallets = custom_no_of_pallets + (i.custom_no_of_pallets or 0) %}
		{% set custom_no_of_boxes = custom_no_of_boxes + (i.custom_no_of_boxes or 0) %}

		{# Show summary row after last iteration #}
		{% if loop.last %}
		<tr style="border:none !important;">
			<td colspan="1" class="no-border" style="text-align:center;"><b>{{ doc.custom_total_no_of_pallets }}</b></td>
			<td colspan="1" class="no-border" style="text-align:center;"><b>{{ doc.custom_total_no_of_boxes }}</b></td>
			<td class="no-border"></td>
			<td class="no-border"></td>
			<td class="no-border"></td>
			<td class="no-border"></td>
			<td colspan="1" class="no-border" style="text-align:center;"><b>{{ doc.total_qty or 0 }}</b></td>
			<td class="no-border"></td>
			<td class="no-border"></td>
		</tr>
		{% endif %}
		{%- endfor -%}

					
					

						<tr style="border-bottom:none !important; border-left:none; border-right:none; " >
							<td colspan="6" class="no-border">
								This Shipment under duty draw back 
							</td>
							
							<td colspan="3" class="no-border">
								
							</td>
						</tr>
						
						<tr style="border-top:none !important; border-bottom:none;">
							<td colspan="3" style="text-align:right; " class="no-border">
								<b><u>COUNTRY OF ORIGIN :{{frappe.db.get_value("Logistics Request",{'order_no':doc.name},"pol_country_airport")or ""}}</u></b>
							</td>
							<td colspan="6" class="no-border">
								
							</td>
						</tr>
						{% set ns = namespace(total_pallets=0) %}
				{% for row in doc["items"] %}
					{% set ns.total_pallets = ns.total_pallets + (row.custom_no_of_pallets or 0) %}
				{% endfor %}
						
						<tr style="border-top:none !important; border-bottom:none;">
							
							<td colspan="3" style="border-left:none;border-right:none;">
								<b>Total Gross Weight:{{'{0:.2f}'.format(doc.custom_total_gross_weight) or ""}}KGS</b>
							</td>
							<td colspan="3"  style="text-align:center;border-left:none;border-right:none;">
								<b>Total Net Weight:{{'{0:.2f}'.format(doc.total_net_weight) or ""}}KGS </b>
							</td>
							<td colspan="3"  style="text-align:center;border-left:none;border-right:none;">
								<b>Total No.of Packaged:{{'{0:.2f}'.format(ns.total_pallets) or ""}}Nos</b>
							</td>
						</tr>
						
					</table>
					
				</div>
			</div>
			
		</div>

		<style>
			@media print {
				.declaration-container {
					page-break-inside: avoid;
					break-inside: avoid;
				}

				.declaration-row {
					border: 1px solid black;
					border-top: 1px solid black;
				}

				/* If using wkhtmltopdf, avoid relying on @page:first */
			}

			.qrcode {
				width: 120px;
				height: 120px;
			}

			.text-center {
				text-align: center;
			}
		</style>

		{% set is_first_page = false %} {# Or dynamically determine this in your logic #}

		<!--<div class="container declaration-container" style="border: 1px solid black;">-->
			<!--<div class="row declaration-row" style="border: 1px solid black; {{ 'border-top:none;' if is_first_page else 'border-top:1px solid black;' }}">-->
			<!--    <div class="col-12 p-0">-->
			<!--        <strong><u>DECLARATION</u></strong><br>-->
			<!--        We declare that this packing list shows the actual<br>-->
			<!--        Weight of the goods<br>-->
			<!--        Described and that all particulars are true and correct-->
			<!--    </div>-->
			<!--</div>-->
		<table style="border: 1px solid black; border-collapse: collapse; width: 100%;">
			<td colspan="4" style="border: 1px solid black; border-bottom: none; border-right: 1px solid transparent !important;width: 50%;">

				<table style="border-collapse: collapse; font-size: 11px; table-layout: fixed;">
					<colgroup>
						
						<col style="width: 70px;">
						<col style="width: 70px;">
						<col style="width: 70px;">
						<col style="width: 100px;">
					</colgroup>
					<tr>
						<td colspan =4 style="border: 1px solid black; text-align: center;"><b>Pallet Details</b></td>
					</tr>
					<tr>
						
						<td style="border: 1px solid black; text-align: center;"><b>Length</b></td>
						<td style="border: 1px solid black; text-align: center;"><b>Width</b></td>
						<td style="border: 1px solid black; text-align: center;"><b>Height</b></td>
						<td style="border: 1px solid black; text-align: center;"><b>No of Pallets</b></td>
					</tr>
					{%- set ns = namespace(total_boxes_1=0, total_length_1=0, total_height_1=0, total_breadth_1 =0) -%}
                
                    {%- for row in doc["items"] -%}
                        {% set ns.total_boxes_1 = ns.total_boxes_1 + (row.custom_no_of_pallets or 0) %}
                        {% set ns.total_length_1 = ns.total_length_1 + (row.custom_pallet_length or 0) %}
                        {% set ns.total_height_1 = ns.total_height_1 + (row.custom_calculated_height or 0) %}
                        {% set ns.total_breadth_1 = ns.total_breadth_1 + (row.custom_pallet_breadth or 0) %}
                    <tr>
                        <td style="border: 1px solid black; text-align: center;">{{ row.custom_pallet_length or "" }}</td>
                        <td style="border: 1px solid black; text-align: center;">{{ row.custom_pallet_breadth or "" }}</td>
                        <td style="border: 1px solid black; text-align: center;">{{ row.custom_calculated_height or "" }}</td>
                        <td style="border: 1px solid black; text-align: center;">{{ row.custom_no_of_pallets or 0 }}</td>
                    </tr>
                    {%- endfor -%}
                    
                    <tr style="background-color:#D3D3D3;">  
                        <td colspan="3" style="border: 1px solid black; text-align: right;"><b>Total No of Pallets</b></td>
                        <td style="border: 1px solid black; text-align: center;"><b>{{ ns.total_boxes_1 }}</b></td>
                    </tr>
				
				</table>
			</td>
			<td colspan="4" style="border: 1px solid black; border-bottom: none; border-left: 1px solid transparent !important;">
			<table style="border-collapse: collapse; font-size: 9px; table-layout: fixed;">
				<colgroup>
					<col style="width: 70px;">
					<col style="width: 70px;">
					<col style="width: 70px;">
					<col style="width: 100px;">
				</colgroup>
				<tr>
						<td colspan =4 style="border: 1px solid black; text-align: center;"><b>Box Details</b></td>
					</tr>
				<tr>
				   
					<td style="border: 1px solid black; text-align: center;"><b>Length</b></td>
					<td style="border: 1px solid black; text-align: center;"><b>Width</b></td>
					<td style="border: 1px solid black; text-align: center;"><b>Height</b></td>
					 <td style="border: 1px solid black; text-align: center; vertical-align: middle;">
						<b>No of Boxes</b>
					</td>
				</tr>
				{%- set ns = namespace(total_boxes=0, total_length=0, total_height=0, total_breadth =0) -%}
                
                {%- for row in doc["items"] -%}
                    {% set ns.total_boxes = ns.total_boxes + (row.custom_no_of_boxes or 0) %}
                    {% set ns.total_length = ns.total_length + (row.custom_box_length or 0) %}
                    {% set ns.total_height = ns.total_height + (row.custom_box_height or 0) %}
                    {% set ns.total_breadth = ns.total_breadth + (row.custom_box_breadth or 0) %}
                <tr>
                    <td style="border: 1px solid black; text-align: center;">{{ row.custom_box_length or "" }}</td>
                    <td style="border: 1px solid black; text-align: center;">{{ row.custom_box_breadth or "" }}</td>
                    <td style="border: 1px solid black; text-align: center;">{{ row.custom_box_height or "" }}</td>
                    <td style="border: 1px solid black; text-align: center;">{{ row.custom_no_of_boxes or 0 }}</td>
                </tr>
                {%- endfor -%}
                
                <tr style="background-color:#D3D3D3;">  
                    
                    <td colspan="3" style="border: 1px solid black; text-align: right;"><b>Total No of Boxes</b></td>
                    <td style="border: 1px solid black; text-align: center;"><b>{{ ns.total_boxes }}</b></td>
                </tr>
			</table>
		</td>

			<tr style="border-top:none;">
				<td colspan="8" style="border: none;border-top:none; padding-top: 15px;">
					<p><b><u>DECLARATION</u></b></p>
					<p>We declare that this packing list shows the actual</p>
					<p>weight of the goods</p>
					<p>described and that all particulars are true and correct.</p>
				</td>
			</tr>
		</table>
<table style="width:100%; border:1px solid black; border-collapse:collapse; font-size:9px; page-break-inside:avoid;">
	<tr >
		<td colspan="3" style="padding:2px;">
			<strong>LUT NO: {{ frappe.db.get_value("Customer", {"name": doc.customer}, "custom_lut_number") or "" }}</strong>
			<span style="float:right; white-space:nowrap;">For <strong>{{ doc.company }}</strong></span>
		</td>
	</tr>

	{% set prepared_by = frappe.db.get_value("Employee", {"user_id": doc.owner}, ["custom_digital_signature"]) %}
	{% set hod = frappe.db.get_value("Employee", {"user_id": doc.custom_hod}, ["custom_digital_signature"]) %}
	{% set finance = frappe.db.get_value("Employee", {"name": 'S0189'}, ["custom_digital_signature"]) %}

	<style>
		.signature-img {
			width: 100px;
			height: 50px;
			object-fit: contain;
			border: none;
		}
	</style>

	<!-- Signatures Row -->
	<tr style="height:60px; text-align:center;">
		<td style="vertical-align:bottom;">
			{% if prepared_by %}
				<img src="{{ prepared_by }}" class="signature-img">
			{% endif %}
		</td>
		<td style="vertical-align:bottom;">
			{% if hod and doc.workflow_state not in ['Pending For HOD', 'Draft'] %}
				<img src="{{ hod }}" class="signature-img">
			{% endif %}
		</td>
		<td style="vertical-align:bottom;">
			{% if finance and doc.workflow_state not in ['Pending For HOD', 'Draft', 'Pending for Finance'] %}
				<img src="{{ finance }}" class="signature-img">
			{% endif %}
		</td>
	</tr>

	<!-- Dates Row -->
	<tr style="height:10px; text-align:center;">
		<td>{{ frappe.utils.format_datetime(doc.creation, "dd-MM-yyyy HH:mm:ss") or '' }}</td>
		<td>{{ frappe.utils.format_datetime(doc.custom_hod_approved_on, "dd-MM-yyyy HH:mm:ss") or '' }}</td>
		<td>{{ frappe.utils.format_datetime(doc.custom_finance_approver_approved_on, "dd-MM-yyyy HH:mm:ss") or '' }}</td>
	</tr>

	<!-- Labels Row -->
	<tr style="font-weight:bold; text-align:center; height:20px;">
		<td>Prepared by</td>
		<td>Checked by</td>
		<td>Authorised Signatory</td>
	</tr>
	<tr>
		<td colspan="3" style="padding:2px;border-top:1px solid black;">
			<strong>Note:</strong>This document is Digitally Signed.
		</td>
	</tr>
</table>

		<!--</div>-->


	"""
	html = render_template(template, {"doc": doc})
	return {"html": html}

@frappe.whitelist()
def create_html_QID(doc):
	doc = frappe.parse_json(doc)
	template ="""
		<img width="100px" src="https://api.qrserver.com/v1/create-qr-code/?size=100x100&amp;data={{ doc.name }}" alt="QR Code" />
	"""
	html = render_template(template, {"doc": doc})
	return {"html": html}


@frappe.whitelist()
def create_html_supplier_delivery(doc):
	doc = frappe.parse_json(doc)
	template = """
		<div class="print-heading">
		<div style="display: flex;">
		<img src="https://api.qrserver.com/v1/create-qr-code/?size=100x100&amp;data={{ doc.scan_barcode }}" alt="QR Code" />
		<div class="ml-auto">
		<h2>
			<b>SUPPLIER DELIVERY NOTE</b>
			<br></h2>
			<h2><b>{{ doc.name }}</b></h2>
			<p>Date & Time: {{ frappe.format(doc.datetime,{"fieldtype":'Datetime'})}}</p>

		</div>
		</div>
		</div>
		<div style="display: flex; justify-content: space-between; width:100%;">
		<div style="flex: 1; padding-right: 10px;">
			<p><b>Billing Address: </b></p>
			<p style="text-transform: uppercase; font-size: 13px;">
			<b>{{doc.supplier_name1 or ''}} ({{doc.supplier or ''}})</b>
			</p>
			<p>{{ (doc.address_display or '').replace('\n', '<br>') | safe }}</p>
		</div>
		<div style="flex: 1; padding-left: 10px;">
			<p><b>Shipping Address: </b></p>
			<p style="text-transform: uppercase; font-size: 13px;">
			<b>{{doc.company}}</b>
			</p>
			<p>{{ (doc.billing_address_display or '').replace('\n', '<br>') | safe }}</p>
			<p class="mt-4" style="text-align:left;"><b>Supplier DC Number: {{doc.supplier_delivery_note or ''}}</b></p>
		</div>

		</div>

		<table width = 100% border= 1px solid black;>
		   <thead>
			<tr>
			<td style=font-size:11px;text-align:center;background-color:#f68b1f;color:white;width:2px;font-weight:bold;>S#</td> 
			<td style= text-align:center;font-size:11px;background-color:#f68b1f;color:white;font-weight:bold;>Item Code</td>
			<td style= text-align:center;font-size:11px;background-color:#f68b1f;color:white;font-weight:bold;>Item Name</td>
			<td style=text-align:center;font-size:11px;background-color:#f68b1f;color:white;font-weight:bold;>HSN/SAC</td>
			<td style='text-align:center;font-size:11px;background-color:#f68b1f;color:white;font-weight:bold;'>Qty</td>
			<td style=text-align:center;font-size:11px;background-color:#f68b1f;color:white;font-weight:bold;>UOM</td>
			<td style=text-align:center;font-size:11px;background-color:#f68b1f;color:white;font-weight:bold;>Remarks</td>
			</tr>
					</thead>

			{% for i in doc.item_table %}
		<tr>
			<td style=font-size:10px>{{i.idx}}</td>
			<td style=text-align:left;font-size:10px>{{i.item_code or ""}}</td>
			<td style=text-align:left;font-size:10px>{{i.item_name or ""}}</td>
			<td style=text-align:center;font-size:10px>{{i.hsnsac  or ''}}</td>
				<td style=text-align:center;font-size:10px>{{'%0.0f'|format(i.dis_qty) or ''}}</td>

			<td style=text-align:center;font-size:10px>{{i.uom  or ''}}</td>
			<td style=text-align:left;font-size:10px>{{i.remarks or ""}}</td>
		</tr> 
			{% endfor %}
		</table>
		<br>
		<div style="display: flex; justify-content: space-between; width:100%; margin-top:20px;">
		<div style="flex: 1; padding-right: 20px;">
			<p style="white-space: nowrap;">Mode of Transport: {{doc.mode_of_transport or ''}}</p>
		</div>

		<div style="flex: 1; padding-left: 20px;">
			<p>Vehicle No: {{doc.vehicle_no or ''}}</p>
		</div>
		<div style="flex: 1; padding-left: 20px;">
			<p>Driver Name & Contact: {{doc.driver_name or ''}}</p>
		</div>

		</div>

		<div style="display: flex; justify-content: space-between; width:100%; margin-top:20px;">
		<div style="flex: 1; padding-right: 15px;">
			<p>Thank you for giving us an opportunity to be at your service</p>
			<p><b>For <span style="text-transform: uppercase; font-size: 13px">{{doc.supplier_name1}}</span></b></p>

			<p style="margin-top: 0px;"><span style="display: inline-block; width: 120px;">Name</span>:</p>
			<p style="margin-top: 0px;"><span style="display: inline-block; width: 120px;">Designation</span>:</p>
			<p style="margin-top: 0px;"><span style="display: inline-block; width: 120px;">Contact</span>:</p>
			<p style="margin-top: 0px;"><span style="display: inline-block; width: 120px;">Company Seal</span>:</p>
		</div>

		<div style="flex: 1; padding-left: 15px;">
			<p>Received By (With signature,Date & Time)</p><br>
			<p style="font-size: 13px;"><b>For {{doc.company.upper()}}</b></p>

			<p style="margin-top: 0px;"><span style="display: inline-block; width: 120px;">Name</span>:</p>
			<p style="margin-top: 0px;"><span style="display: inline-block; width: 120px;">Designation</span>:</p>
			<p style="margin-top: 0px;"><span style="display: inline-block; width: 120px;">Contact</span>:</p>
		</div>

		</div>


	"""
	html = render_template(template, {"doc": doc})
	return {"html": html}





@frappe.whitelist()
def set_qty_sales_invoice(item_code, sales_order, schedule_month, schedule_year, qty, delivery_note=None):
	import traceback
	from frappe.utils import flt

	if delivery_note:
		delivered_qty = frappe.db.get_value("Delivery Note Item", {"parent": delivery_note, "item_code": item_code}, "qty")
		if flt(qty) > delivered_qty:
			return f"Quantity Exceeds. Delivered quantity: {delivered_qty}"
	else:
		qty = flt(qty)
		SOS = frappe.db.get_all(
			"Sales Order Schedule",
			filters={
				"item_code": item_code,
				"schedule_month": schedule_month,
				"schedule_year": schedule_year,
				"sales_order_number": sales_order
			},
			fields=["qty"]
		)
		sch_qty = sum(flt(r.get("qty", 0)) for r in SOS)
		if qty > sch_qty:
			return f"Quantity Exceeds. Available scheduled qty: {sch_qty}"

@frappe.whitelist()
def update_lr_status(doc, method):
	if doc.custom_invoice_type=='Export Invoice':
		try:
			lr=frappe.get_doc('Logistics Request',{'order_no':doc.name})
			if lr.status=='LR Approved':
				lr.status='Ready to Ship'
				lr.save(ignore_permissions=True)
		except Exception as e:
				frappe.log_error(f"status update failed {doc.name}: {str(e)}")

@frappe.whitelist()
def get_item_ledger(item_code, from_date, to_date, warehouse=None):
	"""Fetch Stock Ledger Entries for an item and warehouse between dates"""
	filters = {
		"item_code": item_code,
		"posting_date": ["between", [from_date, to_date]]
	}
	if warehouse:
		filters["warehouse"] = warehouse

	ledgers = frappe.get_all(
		"Stock Ledger Entry",
		filters=filters,
		fields=[
			"posting_date",
			"voucher_type",
			"voucher_no",
			"warehouse",
			"actual_qty",
			"qty_after_transaction"
		],
		order_by="posting_date asc"
	)

	result = []
	for idx, l in enumerate(ledgers, 1):
		result.append({
			"s_no": idx,
			"posting_date": l.posting_date,
			"voucher_type": l.voucher_type,
			"voucher_no": l.voucher_no,
			"warehouse": l.warehouse,
			"actual_qty": l.actual_qty,
			"balance_qty": l.qty_after_transaction
		})
	return result



@frappe.whitelist()
def set_query_for_jc_scrap_item(doctype, txt, searchfield, start, page_len, filters):
	
	job_card_name = filters.get("name")
	sequence_id = filters.get("sequence_id")
	work_order = filters.get("work_order")
	if sequence_id:  
		seq_int = int(sequence_id)
	doc = frappe.get_doc("Job Card",job_card_name)
		
	if not doc.is_new():
		direct = frappe.db.sql("""
			SELECT item_code
			FROM `tabJob Card Item`
			WHERE parent = %s
			AND item_code LIKE %s
			LIMIT %s OFFSET %s
		""", (job_card_name, f"%{txt}%",  page_len, start ))
		
		if direct:
			return direct
		else:
			for i in range(seq_int - 1, 0, -1):
				parent_job_card = frappe.db.sql("""
					SELECT name
					FROM `tabJob Card`
					WHERE sequence_id = %s
					AND work_order = %s 
					LIMIT 1
				""", (i,work_order))

				if parent_job_card:
					parent_name = parent_job_card[0][0]
					seq = frappe.db.sql("""
						SELECT item_code
						FROM `tabJob Card Item`
						WHERE parent = %s
						AND item_code LIKE %s
						LIMIT %s OFFSET %s
					""", (parent_name, f"%{txt}%", page_len, start))

					if seq:
						return seq

			return []
				 
 
 

@frappe.whitelist()
def check_items_in_jc_scrap_item(parent,item_code,sequence_id,work_order):
	if sequence_id:
		seq_int = int(sequence_id)
	doc = frappe.get_doc("Job Card",parent)
	
	if not doc.is_new():
		direct = frappe.db.sql("""
			SELECT item_code
			FROM `tabJob Card Item`
			WHERE parent = %s
			AND item_code = %s
			
		""", (parent, item_code))
	
		if direct:
			return direct
		else:
			for i in range(seq_int - 1, 0, -1):
				parent_job_card = frappe.db.sql("""
					SELECT name
					FROM `tabJob Card`
					WHERE sequence_id = %s
					AND work_order = %s 
					LIMIT 1
				""", (i,work_order))

				if parent_job_card:
					parent_name = parent_job_card[0][0]
					seq = frappe.db.sql("""SELECT item_code
						FROM `tabJob Card Item`
						WHERE parent = %s
						AND item_code = %s
						
					""", (parent_name, item_code))

					if seq:
						return seq

			return []	


@frappe.whitelist()
def si_qty_warehouse_qty_validation(item_code,name):
	result = frappe.db.sql("""
		SELECT SUM(actual_qty) AS total_qty
		FROM `tabBin`
		WHERE item_code = %s
		  AND warehouse = "Finished Goods - WAIP"
	""", (item_code,), as_dict=True)
	
	tot_qty = result[0].total_qty or 0
	si_data = frappe.db.sql("""
		SELECT SUM(sii.qty) AS total_si_qty
		FROM `tabSales Invoice Item` AS sii
		INNER JOIN `tabSales Invoice` AS si ON sii.parent = si.name
		WHERE si.docstatus = 0
		  AND si.name != %s
		  AND sii.item_code = %s
	""", (name, item_code), as_dict=True)
	
	draft_si_qty = si_data[0].total_si_qty or 0
	frappe.errprint(draft_si_qty)
	if tot_qty:
		frappe.errprint(tot_qty)
		qty=tot_qty-draft_si_qty
		if qty < 0:
			qty=0
		frappe.errprint(qty)
		return qty

def get_amount_in_words():
	print("hi")

def sales_order_schedule_reference(doc, method):
	# Collect already existing schedules in child table
	existing_order_schedule = {row.sales_order_schedule for row in doc.custom_sales_order_schedule}

	# Collect from PO Items
	for row in doc.po_items:
		schedules = [
			s.strip()
			for s in (row.get("custom_sales_order_schedule") or "").split(",")
			if s.strip()
		]
		for sos in schedules:
			if sos not in existing_order_schedule:
				sales_order = frappe.db.get_value("Sales Order Schedule", sos, "sales_order_number")
				doc.append("custom_sales_order_schedule", {
					"sales_order_schedule": sos,
					"sales_order": sales_order
				})
				existing_order_schedule.add(sos)

	# Collect from Sub Assembly Items
	for row in doc.sub_assembly_items:
		schedules = [
			s.strip()
			for s in (row.get("custom_sales_order_schedule") or "").split(",")
			if s.strip()
		]
		for sos in schedules:
			if sos not in existing_order_schedule:
				sales_order = frappe.db.get_value("Sales Order Schedule", sos, "sales_order_number")
				doc.append("custom_sales_order_schedule", {
					"sales_order_schedule": sos,
					"sales_order": sales_order
				})
				existing_order_schedule.add(sos)

def create_user_for_supplier(doc, method):
	"""User will be created for Suppliers on creation"""
	if doc.email_id:
		if not frappe.db.exists("User", {"email": doc.email_id}):
			user = frappe.new_doc("User")
			user.email = doc.email_id
			user.first_name = doc.supplier_name
			user.role_profile_name =  "Supplier"
			user.send_welcome_email = 1
			user.flags.ignore_permissions = True
			user.insert()

@frappe.whitelist()
def get_supervisor_name(supervisor):
	return frappe.db.get_value("Employee", supervisor, "employee_name") or ""




import urllib.parse
from frappe.utils import now_datetime

@frappe.whitelist()
def make_url_for_si(doc, method):
	if not doc.shipping_address_name:
		if doc.customer_address:
			doc.shipping_address_name=doc.customer_address
	if not doc.custom_allowed_hod:
		if frappe.db.exists("Employee",{'user_id':doc.owner}):
			reports_to=frappe.db.get_value('Employee',{'user_id':doc.owner},['reports_to'])
			if reports_to:
				hod_mail=frappe.db.get_value('Employee',reports_to,'user_id')
				if hod_mail:
					doc.custom_allowed_hod=hod_mail
	entry_time = now_datetime()
	params = {
		"entry_time": entry_time.isoformat(),
		"document_id": doc.name or "",
		"entry_document": 'Sales Invoice',
		"entry_type": 'Outward',
		"vehicle_number":doc.custom_vehicle_no or "",
		"driver_name":doc.custom_driver_name or "",
		"party_type":'Customer',
		"party":doc.customer,
		
	}
	query_string = urllib.parse.urlencode(params)
	full_url = f"https://erp.onegeneindia.in/app/gate-entry-update?{query_string}"
	encoded_url = urllib.parse.quote(full_url, safe="")  
	doc.custom_scan_barcode = encoded_url

@frappe.whitelist()
def get_rack_for_items(doctype, txt, searchfield, start, page_len, filters):
	if isinstance(filters, str):
		filters = json.loads(filters)

	warehouse = filters.get("warehouse")
	if not warehouse:
		return []

	start = int(start)
	page_len = int(page_len)

	search_text = f"%{txt.strip()}%" if txt else "%"

	return frappe.db.sql("""
		SELECT DISTINCT c.name
		FROM `tabWarehouse` p
		LEFT JOIN `tabWarehouse Rack` c ON c.parent = p.name
		WHERE p.name = %s
		  ANd c.rack LIKE %s
		LIMIT %s OFFSET %s
	""", (warehouse, search_text, page_len, start))

# @frappe.whitelist()
# def get_query_for_bom(doctype, txt, searchfield, start, page_len, filters):
# 	if isinstance(filters, str):
# 		filters = json.loads(filters)
	
# 	last_updated_on = filters.get('last_updated_on')
# 	pmr = frappe.db.get_value("Production Material Request", {"last_updated_on"})
# 	return frappe.db.sql("""
# 		SELECT item_code
# 		FROM `tabProduction Ma`
# 		WHERE department = %s
# 		  AND {key} LIKE %s
# 		ORDER BY name
# 		LIMIT %s, %s
# 	""".format(key=searchfield), (filters.get("department"), "%%%s%%" % txt, start, page_len))



@frappe.whitelist()
def set_hod(doc, method):
	if doc.workflow_state == "Pending for ERP Team":
		doc.custom_hod_approved_by = frappe.session.user
		doc.custom_hod_approved_on = now_datetime()
	if doc.workflow_state == "Material Pending":
		doc.custom_erp_team = frappe.session.user
		doc.custom_erp_team_approved_on = now_datetime()

@frappe.whitelist()
def get_custom_bom_pmr(doctype, txt, searchfield, start, page_len, filters):
	warehouse = filters.get("warehouse")

	# Get the latest Production Material Request
	pmr = frappe.db.get_value(
		"Production Material Request",
		{},
		"name",
		order_by="creation desc"
	)
	if not pmr:
		return []

	# Collect unique parent BOMs from both Sub Assembly Item & Raw Materials
	parent_boms = set()
	for child_table in ("Sub Assembly Item", "Raw Materials"):
		boms = frappe.get_all(
			child_table,
			filters={"parent": pmr, "warehouse": warehouse},
			pluck="parent_bom"
		)
		# Filter out None/empty values and ensure uniqueness
		parent_boms.update(bom for bom in boms if bom)

	if not parent_boms:
		return []

	# Fetch BOMs that match the parent_boms, are default, and match search text
	bom_list = frappe.get_all(
		"BOM",
		filters={
			"name": ["in", list(parent_boms)],
			"is_default": 1
		},
		fields=["name", "item", "item_name"],
	)

	# Apply text filter manually in Python
	if txt:
		bom_list = [
			bom for bom in bom_list
			if txt.lower() in bom.name.lower()
		]

	# Ensure uniqueness by name
	unique_boms = {bom.name: (bom.name, bom.item, bom.item_name) for bom in bom_list}
	return list(unique_boms.values())
@frappe.whitelist()
def validate_bom_warehouse(bom, warehouse):
	"""Validate if the given BOM's item exists in the latest Production Material Request for the given warehouse."""
	
	if not bom or not warehouse:
		return {"is_valid": False, "message": "BOM or Warehouse not provided."}

	# Safely get BOM document
	try:
		bom_doc = frappe.get_doc("BOM", bom)
	except frappe.DoesNotExistError:
		return {"is_valid": False, "message": f"BOM '{bom}' not found."}

	# Get the latest Production Material Request
	pmr_name = frappe.db.get_value(
		"Production Material Request",
		{},
		"name",
		order_by="creation desc"
	)

	if not pmr_name:
		return {"is_valid": False, "message": "No Production Material Request found."}

	# Collect all item codes from both Sub Assembly Item and Raw Materials tables
	parent_boms = set()

	for child_table in ("Sub Assembly Item", "Raw Materials"):
		boms = frappe.get_all(
			child_table,
			filters={"parent": pmr_name, "warehouse": warehouse},
			pluck="parent_bom"
		)
		parent_boms.update(bom for bom in boms if bom)

	# Check if the BOM's main item exists in PMR items
	if bom_doc.name in parent_boms:
		return {
			"is_valid": True,
			"message": f"BOM '{bom}' is valid for warehouse '{warehouse}'."
		}

	return {
		"is_valid": False,
		"message": f"Item '{bom_doc.name}' not found in the latest PMR for warehouse '{warehouse}'."
	}



import frappe

@frappe.whitelist()
def get_user_permitted_department():
	user = frappe.session.user
	permitted_departments = frappe.get_all(
		"User Permission",
		filters={
			"user": user,
			"allow": "Department"
		},
		fields=["for_value"]
	)

	return [d.for_value for d in permitted_departments]







@frappe.whitelist()
def create_so_iom(name):
	self=frappe.get_doc("Inter Office Memo",name)
	po_group = {}
			
	for item in self.approval_business_po:
		po_no = item.po_no  

		
		if po_no not in po_group:
			po_group[po_no] = []
		
		
		po_group[po_no].append(item)


	for po_no, items in po_group.items():
		
		if self.order_type =="Fixed Order":
			
			new_doc = frappe.new_doc("Sales Order")
			new_doc.custom_docname = po_no  
			new_doc.po_no = po_no  
			new_doc.po_date = items[0].po_date  or datetime.now() 
			new_doc.customer = self.customer
			new_doc.customer_order_type = "Fixed"  
			new_doc.delivery_date = items[0].po_date
			new_doc.company = "WONJIN AUTOPARTS INDIA PVT.LTD."
			new_doc.taxes_and_charges=self.taxes_and_charges
   
			
					
			
			for item in items:
				
				current_date = datetime.now()
				current_month = current_date.strftime("%b").upper() 
				first_date_of_curr_month = current_date.replace(day=1)
			
				new_doc.append("items", {
					"item_code": item.part_no,
					"qty":item.qty,
					"open_qty":item.qty,  
					"rate":item.po_priceinr  
				})
				
				
				new_doc.append("custom_schedule_table",{
					"item_code": item.part_no,
					"schedule_qty": item.qty,
					"schedule_date": first_date_of_curr_month,
					"schedule_month" :current_month
				})
				
			for t in self.taxes:
				new_doc.append("taxes",{
					"charge_type":t.charge_type,
					"account_head":t.account_head,
					"description":t.description,
					"rate":t.rate
				})
				
				
							
			
			new_doc.insert(ignore_permissions=True)
			
			
			try:
				apply_workflow(new_doc, action="Send to HOD")
				new_doc.reload()
				new_doc.save()

				apply_workflow(new_doc, action="Approve")
				new_doc.reload()
				new_doc.save()

				new_doc.submit()
			except Exception as e:
				frappe.log_error(f"Workflow transition failed for {new_doc.name}: {str(e)}")
				
		else:
		   
			current_date = datetime.now()
			current_year = current_date.year
			current_month = current_date.month
   
			if current_month <= 3:
				date_str = f"31-03-{current_year}"  
			else:
				
				date_str = f"31-03-{current_year + 1}"
	
	
			formatted_date = datetime.strptime(date_str, "%d-%m-%Y")
   
			new_doc = frappe.new_doc("Sales Order")
			new_doc.custom_docname = po_no  
			new_doc.po_no = po_no  
			new_doc.po_date = items[0].po_date  
			new_doc.customer = self.customer
			new_doc.customer_order_type = "Open"
			new_doc.delivery_date = formatted_date    
			new_doc.company = "WONJIN AUTOPARTS INDIA PVT.LTD."
			new_doc.taxes_and_charges=self.taxes_and_charges  

			
			for item in items:
				
				current_date = datetime.now()
				current_month = current_date.strftime("%b").upper() 
				first_date_of_curr_month = current_date.replace(day=1)
			
				new_doc.append("items", {
					"item_code": item.part_no,
					"qty":item.qty,
					"open_qty":1, 
					"rate":item.po_priceinr  
				})
				
				new_doc.append("custom_schedule_table",{
					"item_code": item.part_no,
					"schedule_qty": item.qty,
					"schedule_date": first_date_of_curr_month,
					"schedule_month" :current_month
				})
				
			for t in self.taxes:
				new_doc.append("taxes",{
					"charge_type":t.charge_type,
					"account_head":t.account_head,
					"description":t.description,
					"rate":t.rate
				})
				
				
							
			
			new_doc.insert(ignore_permissions=True)
			
			
			
			try:
				
				apply_workflow(new_doc, action="Send to HOD")
				
				new_doc.reload()
				new_doc.save()

				apply_workflow(new_doc, action="Approve")
				new_doc.reload()
				new_doc.save()

				new_doc.submit()
				
			except Exception as e:
				
				title = f"Document modified error for {new_doc.name}"  
				message = f"Error during workflow transition: {str(e)}"
				frappe.log_error(message=message, title=title)

@frappe.whitelist()
def create_quality_pending_documents(doc, method):
	for row in doc.items:
		qp_doc = frappe.new_doc('Quality Pending')
		qp_doc.inspection_pending_type = 'Incoming'
		qp_doc.status = 'Inspection Pending'
		qp_doc.reference_type = 'Purchase Receipt'
		qp_doc.reference_name = doc.name
		qp_doc.item_code = row.item_code
		qp_doc.possible_inspection_qty = row.qty
		qp_doc.inspection_pending_qty = row.qty
		qp_doc.insert(ignore_permissions=True)
		frappe.db.commit()
		

@frappe.whitelist()
def set_query_for_custom_racks(doctype, txt, searchfield, start, page_len, filters):
	name = filters.get("name")
	item_warehouse = filters.get("warehouse")
	
	warehouse = frappe.get_doc("Warehouse",item_warehouse)
	
	
	
	item_rack = frappe.db.sql("""
			SELECT rack
			FROM `tabWarehouse Rack`
			WHERE parent = %s
			AND rack LIKE %s
			LIMIT %s OFFSET %s
		""", (item_warehouse, f"%{txt}%",  page_len, start ))
	
	return item_rack or []
	
import json

@frappe.whitelist()
def check_rack_in_custom_racks(filters):
	
	filters = json.loads(filters)  
	
	item_warehouse = filters.get("warehouse")
	rack_name = filters.get("rack") 

	item_rack = frappe.db.sql("""
		SELECT rack FROM `tabWarehouse Rack`
		WHERE parent = %s AND rack = %s
	""", (item_warehouse, rack_name), as_dict=True)
	
	return item_rack or []

@frappe.whitelist()
def set_query_for_store_receipt_racks(doctype, txt, searchfield, start, page_len, filters):
	item_code = filters.get("item_code")
	
	
	
	store_receipt_rack = frappe.db.sql("""
			SELECT rack
			FROM `tabItem Rack`
			WHERE parent = %s
			AND rack LIKE %s
			LIMIT %s OFFSET %s
		""", (item_code, f"%{txt}%",  page_len, start ))
	
	return store_receipt_rack or []    
	
	
	
	
@frappe.whitelist()
def set_query_for_custom_locations(doctype, txt, searchfield, start, page_len, filters):
	name = filters.get("name")
	item_warehouse = filters.get("warehouse")
	
	warehouse = frappe.get_doc("Warehouse",item_warehouse)
	
	
	
	item_location = frappe.db.sql("""
			SELECT location
			FROM `tabWarehouse Rack Location`
			WHERE parent = %s
			AND location LIKE %s
			LIMIT %s OFFSET %s
		""", (item_warehouse, f"%{txt}%",  page_len, start ))
	
	return item_location or []
	
	

import json  

@frappe.whitelist()
def check_location_in_custom_locations(filters):
	
	filters = json.loads(filters)  
	
	item_warehouse = filters.get("warehouse")
	location_name = filters.get("location") 

	item_location = frappe.db.sql("""
		SELECT location FROM `tabWarehouse Rack Location`
		WHERE parent = %s AND location = %s
	""", (item_warehouse, location_name), as_dict=True)
	
	return item_location or []



@frappe.whitelist()
def update_datetime_and_approver_details_in_Si(doc, method):
	if doc.custom_invoice_type == "Export Invoice":
		if doc.has_value_changed("workflow_state") and doc.workflow_state == "Waiting for LR Request":
			doc.custom_hod_approved_on=now_datetime()
			doc.custom_hod=frappe.session.user
		if doc.has_value_changed("workflow_state") and doc.workflow_state == "Approved":
			doc.custom_finance_approver_approved_on=now_datetime()
			doc.custom_finance_approver=frappe.session.user

@frappe.whitelist()
def update_datetime_and_approver_details_finance(doc, method):
	if doc.custom_invoice_type == "Export Invoice":
		if doc.has_value_changed("workflow_state") and doc.workflow_state == "Approved":
			doc.custom_finance_approver_approved_on=now_datetime()
			doc.custom_finance_approver=frappe.session.user

@frappe.whitelist()
def update_hod_time(doc, method):
	if doc.custom_invoice_type == "Tax Invoice":
		if doc.workflow_state == "Pending for Finance":
			doc.custom_hod_approved_on=now_datetime()
			doc.custom_hod=frappe.session.user
		if doc.workflow_state == "Approved":
			doc.custom_finance_approver_approved_on=now_datetime()
			doc.custom_finance_approver=frappe.session.user

def remove_supervisor_and_department(doc, method):
	frappe.db.set_value("Job Card", doc.name, "custom_supervisor", None)
	frappe.db.set_value("Job Card", doc.name, "custom_department", None)
 


import frappe
from frappe.utils.html_utils import escape_html
import json

@frappe.whitelist()
def sales_invoice_pick_list_all(item_codes_json):
	item_codes = json.loads(item_codes_json) if item_codes_json else []

	if not item_codes:
		return []

	result = []

	

		
	placeholders = ", ".join(["%s"] * len(item_codes))

	
	query = """
		SELECT
			sle.item_code,
			COALESCE(i.item_name, '') AS item_name,
			COALESCE(i.item_group, '') AS item_group,
			COALESCE(sle.warehouse, '') AS warehouse,
			ir.rack,
			ir.location AS location,
			SUM(sle.actual_qty) AS total_qty
		FROM
			`tabStock Ledger Entry` sle
		LEFT JOIN
			`tabItem` i ON sle.item_code = i.name
		LEFT JOIN
			`tabItem Rack` ir ON ir.parent = i.name    
		WHERE
			sle.item_code IN ({placeholders})
			AND sle.is_cancelled = 0
			AND sle.warehouse = i.custom_warehouse
			AND sle.rack = ir.rack
			AND sle.rack_location = ir.location
		GROUP BY
			ir.rack,ir.location
		
	""".format(placeholders=placeholders)

	
	data = frappe.db.sql(query, tuple(item_codes), as_dict=True)

	if data:
		
		existing_codes = {d["item_code"] for d in data}

		
		for code in item_codes:
			if code not in existing_codes:
				itm = frappe.get_cached_doc("Item", code)
				result.append({
					"item_code": code,
					"item_name": itm.item_name or "",
					"item_group": itm.item_group or "",
					"warehouse": "",  
					"rack": "",
					"location": "",
					"available_qty": ""
				})

		
		for d in data:
			qty = d.get("total_qty") or 0
			d["available_qty"] = qty if qty > 0 else ""
			result.append(d)


		result_sorted = []
		for code in item_codes:
			rows = [r for r in result if r["item_code"] == code]
			result_sorted.extend(rows)

		return result_sorted






import frappe
from frappe.utils.jinja import render_template

@frappe.whitelist()
def create_html_view_tax(doc):
	doc = frappe.parse_json(doc)

	template = """
	{% set address = frappe.db.get_value("Customer",{"name":doc.customer},"primary_address") %}
<style>
@media print {
  .invoice-box {
	page-break-inside: avoid;
  }

  .invoice-footer {
	position: relative;
	bottom: 0;
	page-break-inside: avoid;
  }

  table {
	border-collapse: collapse;
  }

  tr, td, th {
	page-break-inside: avoid !important;
  }

  .signature-block {
	margin-top: 50px;
  }
}
	.invoice-box {
		width: 100%;
		font-family: Arial, sans-serif;
		font-size: 10px;
		color: #000;
		border:1px solid black;
	}
	.invoice-box table {
		width: 100%;
		border-collapse: collapse;
	}
	.invoice-box td, .invoice-box th {
		border: 1px solid black;
	}
	.no-border {
		border: none !important;
	}
	.text-center {
		text-align: center;
	}
	.text-right {
		text-align: right;
	}
	.text-left {
		text-align: left;
	}
	.remove-top-border td, 
	.remove-top-border th {
		border-top: none !important;
	}
	
	.remove-bottom-border td, 
	.remove-bottom-border th {
		border-bottom: none !important;
	}
	.compact-table td {
		line-height: 1.0;  /* reduce line height */
	}
	.compact-table-new td {
		line-height: 1.1;  /* reduce line height */
	}
	table{
		padding:0;
		width:100%;
	}

</style>

<div class="invoice-box" style="position: relative;"> <!-- adjust min-height as needed -->

	<!-- Header -->
	<table class="remove-top-border">
		<tr style="border:1px solid black;border-left:none;border-top:none;border-right:none;">
			<td rowspan="4" class="no-border" style="width:20%;border-right:none;border-left:none;">
				<img src="/files/Tax invoice logo.png" alt="logo" width="90">
			</td>
			<td style="font-size:10px;border-left:none;border-right:none ! important;text-align:center; border-right:none;">
				<strong><p style="font-size:11px;">Wonjin Autoparts India Pvt Ltd</p></strong>
				(Under Rule 7 of GST Rules 2017)<br>
				For removal of Excisable Goods from a factory or Warehouse on Payment of duty<br>
				Plot No: A1K, CMDA, Industrial Complex, Maraimalai Nagar, Chennai - 603 209<br>
				Phone No: 044 - 4740 4415 / 4740 4436<br>
				Email: finance@onegeneindia.in | Web: www.onegeneindia.in<br>
				GSTIN: 33AADCP2334E1ZY
			</td>
			<td rowspan="4" style="border-left:none ! important;border-top:none;border-bottom:none;border-right:none;">
				 <img src="https://api.qrserver.com/v1/create-qr-code/?size=150x150&data={{ ('https://erp.onegeneindia.in/app/sales-invoice/' + doc.name) | urlencode }}" alt="QR Code" width="100" />
			</td>
		</tr>
	</table>
	<table class="remove-top-border">
		<tr>
			<td style="text-align:center;font-size:14px;font-weight:bold;margin-top:-10px;line-height:0.7;border-left:none;border-right:none;">Tax Invoice</td>
		</tr>
	</table>
	<!--<table class="remove-top-border compact-table">-->
	<!--    <tr>-->
	<!--        <td style="border-right:none;border-bottom:none;width:24%">E-Way No </td><td style="border-left:none;border-bottom:none;width:25%">: {{ doc.ewaybill or '' }}</td>-->
	<!--        <td style="border-right:none;border-bottom:none;width:24%">Invoice No </td><td style="border-left:none;border-bottom:none;width:25%">: {{ doc.name }}</td>-->
	<!--    </tr>-->
	<!--    <tr>-->
	<!--        <td style="border-right:none;border-bottom:none;border-top:none;">E-Way Date</td><td style="border-left:none;border-bottom:none;border-top:none;">: {{frappe.db.get_value("e-Waybill Log",{"name":doc.ewaybill},"created_on") or ""}}</td>-->
	<!--        <td style="border-right:none;border-bottom:none;border-top:none;">Invoice Date </td><td style="border-left:none;border-bottom:none;border-top:none;">: {{ frappe.utils.format_date(doc.posting_date) }}</td>-->
	<!--    </tr>-->
	<!--    <tr>-->
	<!--        <td style="border-right:none;border-bottom:none;border-top:none;">ACK No </td><td style="border-left:none;border-bottom:none;border-top:none;">: {{frappe.db.get_value('e-Invoice Log', {'sales_invoice': doc.name}, 'acknowledgement_number') or ""}}</td>-->
	<!--        <td style="border-right:none;border-bottom:none;border-top:none;">Vendor Code </td><td style="border-left:none;border-bottom:none;border-top:none;">: {{frappe.db.get_value('Customer',{'name':doc.customer},"custom_vendor_code")or ""}}</td>-->
	<!--    </tr>-->
	<!--    <tr>-->
	<!--        <td style="border-right:none;border-bottom:none;border-top:none;">ACK Date </td><td style="border-left:none;border-bottom:none;border-top:none;">: {{frappe.db.get_value('e-Invoice Log', {'sales_invoice': doc.name}, 'acknowledged_on') or ""}}</td>-->
			<!--<td style="border-right:none;border-bottom:none;border-top:none;">Order No & Date </td><td style="border-left:none;border-bottom:none;border-top:none;">: {{ doc.po_no or '' }}</td>-->
	<!--        <td style="border-right:none;border-bottom:none;border-top:none;">Order No & Date </td><td style="border-left:none;border-bottom:none;border-top:none;">: -->
	<!--        {% for row in doc.custom_sales_orders %}-->
	<!--            <span>{{ row.sales_order or "" }}</span>{% if not loop.last %}, {% endif %}-->
	<!--        {% endfor %}-->
	<!--        </td>-->
	<!--    </tr>-->
	<!--    <tr>-->
	<!--        <td style="border-top:none; border-right:none;">IRN No </td><td style="border-left:none;border-top:none;">: {{doc.irn or ''}}</td>-->
	<!--        <td style="border-bottom:1px solid black;border-right:none"></td><td style="border-bottom:1px solid black;border-left:none"></td>-->
	<!--    </tr>-->
	<!--</table>-->
	<table class="remove-top-border">
		 <tr>
			<td colspan="2" style="width:50%;border-left:none;">
						E WAY No. &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:&nbsp;&nbsp; {{doc.ewaybill or ""}}<br>
						E WAY Date &nbsp;&nbsp;&nbsp;&nbsp;: &nbsp;&nbsp;{{frappe.db.get_value("e-Waybill Log",{"name":doc.ewaybill},"created_on") or ""}}<br>
						ACK No. &nbsp;&nbsp;&nbsp; &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:&nbsp;&nbsp; {{frappe.db.get_value('e-Invoice Log', {'sales_invoice': doc.name}, 'acknowledgement_number') or ""}}<br>
						ACK Date &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: &nbsp;&nbsp;{{frappe.db.get_value('e-Invoice Log', {'sales_invoice': doc.name}, 'acknowledged_on') or ""}}<br>
						IRN No. &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:&nbsp;&nbsp; {{doc.irn or ""}}
						
			</td>
			<td colspan="2" style="width:50%;border-right:none;">
						Invoice No. &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:&nbsp;&nbsp;{{doc.name or ""}}<br>
						Invoice Date. &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:&nbsp;&nbsp;{{frappe.format(doc.posting_date,{'fieldtype':'Date'}) or ""}}<br>
						vendor code &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;:&nbsp;&nbsp;{{frappe.db.get_value('Customer',{'name':doc.customer},"custom_vendor_code")or ""}}<br>
						Order No & Date &nbsp;&nbsp;:&nbsp;&nbsp;{% for row in doc.custom_sales_orders %}
							<span>{{ row.sales_order or "" }}</span>{% if not loop.last %}, {% endif %}
						{% endfor %}<br>
						
			</td>
		</tr>
	</table>
	{% set state = frappe.db.get_value("Address",{'address_title':doc.customer},'state') %}
	{% set statecode = frappe.db.get_value("Address",{'address_title':doc.customer},'gst_state_number') %}
	{% set gstin = frappe.db.get_value("Address",{'address_title':doc.customer},'gstin') %}
	<!--<table class="remove-top-border compact-table-new">-->
	<!--    <tr>-->
	<!--        <td colspan="2" style="border-right:none;font-weight:bold;">Billed To</td>-->
	<!--        <td colspan="2" style="border-left:none;font-weight:bold;">&nbsp;&nbsp;&nbsp;&nbsp;Shipped To</td>-->
	<!--    </tr>-->
	<!--    <tr>-->
	<!--        <td style="border-right:none;border-bottom:none;">Name</td><td style="border-left:none;border-bottom:none;" nowrap>: {{doc.customer or ''}}</td>-->
	<!--        <td style="border-right:none;border-bottom:none;">Name</td><td style="border-left:none;border-bottom:none;" nowrap>: {{doc.customer or ''}}</td>-->
	<!--    </tr>-->
	<!--    <tr>-->
	<!--        <td style="border-right:none;border-bottom:none;border-top:none;">Address</td><td style="border-left:none;border-bottom:none;border-top:none;">: {{doc.address_display or ''}}</td>-->
	<!--        <td style="border-right:none;border-bottom:none;border-top:none;">Address</td><td style="border-left:none;border-bottom:none;border-top:none;">: {{doc.address_display or ''}}</td>-->
	<!--    </tr>-->
	<!--    <tr>-->
	<!--        <td style="border-right:none;border-bottom:none;border-top:none;">State</td><td style="border-left:none;border-bottom:none;border-top:none;">: {{state or ''}}</td>-->
	<!--        <td style="border-right:none;border-bottom:none;border-top:none;">State</td><td style="border-left:none;border-bottom:none;border-top:none;">: {{state or ''}}</td>-->
	<!--    </tr>-->
	<!--    <tr>-->
	<!--        <td style="border-right:none;border-bottom:none;border-top:none;" nowrap>State Code</td><td style="border-left:none;border-bottom:none;border-top:none;">: {{statecode or ''}}</td>-->
	<!--        <td style="border-right:none;border-bottom:none;border-top:none;" nowrap>State Code</td><td style="border-left:none;border-bottom:none;border-top:none;">: {{statecode or ''}}</td>-->
	<!--    </tr>-->
	<!--    <tr>-->
	<!--        <td style="border-top:none; border-right:none;">GSTIN</td><td style="border-top:none; border-left:none;">: {{gstin or ''}}</td>-->
	<!--        <td style="border-top:none; border-right:none;">GSTIN</td><td style="border-top:none; border-left:none;">: {{gstin or ''}}</td>-->
	<!--    </tr>-->
	<!--</table>-->
	
	<table class="remove-top-border compact-table-new" style="width:100%;">
		
		<tr>
			<td colspan="2" style="border-right:none;font-weight:bold;line-height:0.8;border-left:none;">Billed To</td>
			<td colspan="2" style="border-left:none;font-weight:bold;line-height:0.8;border-right:none;">&nbsp;&nbsp;&nbsp;&nbsp;Shipped To</td>
		</tr>
		<tr>
			<!-- Billed To -->
			{% set address_line = frappe.db.get_value("Address",{"name":doc.customer_address},"address_line1") %}
			{% set city = frappe.db.get_value("Address",{'name':doc.customer_address},"city") %}
			{% set country = frappe.db.get_value("Address",{'name':doc.customer_address},"country") %}
			{% set state = frappe.db.get_value("Address",{'name':doc.customer_address},"state") %}
			{% set pincode = frappe.db.get_value("Address",{'name':doc.customer_address},"pincode") %}
			<td colspan="2" style="vertical-align:top; width:50%;border-left:none;">
				Name&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {{ doc.customer or '' }}<br>
				Address &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {{address_line}}<br>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;{{city}}, {{country}} <br>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;{{state}}, {{pincode}} <br>
				State&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {{ state or '' }}<br>
				State Code&nbsp;&nbsp;&nbsp;&nbsp;: {{ statecode or '' }}<br>
				GSTIN&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {{ gstin or '' }}
			</td>
	
			<td colspan="2" style="vertical-align:top; width:50%;border-right:none;">
				Name&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {{ doc.customer or '' }}<br>
				Address &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {{address_line}}<br>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;{{city}}, {{country}} <br>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;{{state}}, {{pincode}} <br>
				State&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {{ state or '' }}<br>
				State Code&nbsp;&nbsp;&nbsp;&nbsp;: {{ statecode or '' }}<br>
				GSTIN&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;: {{ gstin or '' }}
			</td>
		</tr>
	</table>


	<table class="remove-top-border">
		<thead>
			<tr style="line-height:0.8;border-left:none;border-right:none;">
				<td style="width:2%;font-weight:bold;text-align:center;font-size:10px;line-height:1.5;border-left:none;">S.No</td>
				<td style="width:17%;font-weight:bold;text-align:center;font-size:10px;border-left:none;">Part No</td>
				<td style="width:35%;font-weight:bold;text-align:center;font-size:10px;border-left:none;">Part Name</td>
				<td style="width:8%;font-weight:bold;text-align:center;font-size:10px;border-left:none;">HSN/SAC</td>
				<td style="width:3%;font-weight:bold;text-align:center;font-size:10px;border-left:none;">Qty</td>
				<td style="width:5%;font-weight:bold;text-align:center;font-size:10px;border-left:none;">Box</td>
				<td style="width:5%;font-weight:bold;text-align:center;font-size:10px;border-left:none;">UOM</td>
				<td style="width:8%;font-weight:bold;text-align:center;font-size:10px;border-left:none;">Rate (INR)</td>
				<td style="width:11%;font-weight:bold;text-align:center;font-size:10px;border-left:none;border-right:none;">Amount (INR)</td>
			</tr>
		</thead>
		<tbody>
		 {%- set max_rows = 14-%}
{%- set row_count = doc["items"]|length -%}
	{% for row in doc["items"] %}
	<tr style="line-height:1.1;">
		<td class="text-center" style="font-size:10px; border-left:none; border-right:1px solid black;border-bottom:none;">{{ loop.index }}</td>
		<td style="font-size:10px; border-left:1px solid black; border-right:1px solid black;border-bottom:none;">{{ row.item_code or '' }}</td>
		<td style="font-size:10px; border-left:1px solid black; border-right:1px solid black;border-bottom:none;">{{ row.item_name or '' }}</td>
		<td style="font-size:10px; border-left:1px solid black; border-right:1px solid black;border-bottom:none;">{{ row.gst_hsn_code or '' }}</td>
		<td class="text-right" style="font-size:10px; border-left:1px solid black; border-right:1px solid black;border-bottom:none;">{{ (row.qty or 0) | int }}</td>
		<td class="text-center" style="font-size:10px; border-left:1px solid black; border-right:1px solid black;border-bottom:none;">{{ row.custom_no_of_boxes or '' }}</td>
		<td class="text-center" style="font-size:10px; border-left:1px solid black; border-right:1px solid black;border-bottom:none;">{{ row.uom or '' }}</td>
		<td class="text-right" style="font-size:10px; border-left:1px solid black; border-right:1px solid black;border-bottom:none;">{{ "%.4f"|format(row.rate or 0) }}</td>
		<td class="text-right" style="font-size:10px; border-left:1px solid black; border-right:1px solid black;border-bottom:none;">{{ "%.2f" | format(row.amount or 0) }}</td>
	</tr>
	{% endfor %}
	{# filler rows to preserve borders #}
{%- for i in range(row_count, max_rows) -%}
<tr >
	<td class="no-top-bottom remove-bottom-border" style="text-align:center;font-size:9px;border: 1px solid black;border-bottom: none !important;border-left: none !important;">&nbsp;</td>
	<td class="no-top-bottom" style="text-align:center;font-size:9px;border: 1px solid black;border-bottom: none !important;border-left: none !important;">&nbsp;</td>
	<td class="no-top-bottom" style="text-align:center;font-size:9px;border: 1px solid black;border-bottom: none !important;border-left: none !important;">&nbsp;</td>
	<td class="no-top-bottom" style="text-align:center;font-size:9px;border: 1px solid black;border-bottom: none !important;border-left: none !important;">&nbsp;</td>
	<td class="no-top-bottom" style="text-align:center;font-size:9px;border: 1px solid black;border-bottom: none !important;border-left: none !important;">&nbsp;</td>
	<td class="no-top-bottom" style="text-align:center;font-size:9px;border: 1px solid black;border-bottom: none !important;border-left: none !important;">&nbsp;</td>
	<td class="no-top-bottom" style="text-align:center;font-size:9px;border: 1px solid black;border-bottom: none !important;border-left: none !important;">&nbsp;</td>
	<td class="no-top-bottom" style="text-align:center;font-size:9px;border: 1px solid black;border-bottom: none !important;border-left: none !important;">&nbsp;</td>
	<td style="border: 1px solid black;border-bottom:none !important;border-top: none !important;border-right:none !important;">&nbsp;</td>
</tr>
{%- endfor -%}
	

		</tbody>
		
<!--        <style>-->
<!--    table tbody tr.last-row td {-->
<!--        border-bottom: 1px solid black !important;-->
<!--    }-->
<!--</style>-->
 </table>


<div class="invoice-footer" style="width: 100%;">
		<!-- CGST/SGST and Net Total section -->
		<table style="table-layout:fixed; width:100%;">
			<tr>
				<td colspan="5" style="border-right:none;border-left:none;border-right:none;">
					{% if doc.taxes %}
					CGST AMT in words: {{ frappe.utils.money_in_words(doc.taxes[0].tax_amount_after_discount_amount or '') }}<br>
					SGST AMT in words: {{ frappe.utils.money_in_words(doc.taxes[1].tax_amount_after_discount_amount or '') }}
					{% else %}
					CGST AMT in words: <br>
					SGST AMT in words:
					{% endif %}
					
				</td>
				<td colspan="4" style="line-height:1;margin-top:5px;border-right:none;border-bottom:hidden;">
					<div style="display:flex; justify-content:space-between; margin:0; padding:0;">
						<span>Amount (INR)</span>
						<span style="text-align:right; min-width:80px;margin-left:180px;">{{"%.2f"|format(doc.total or 0)}}</span>
					</div><br>
					<div style="display:flex; justify-content:space-between; margin:0; padding:0;">
						<span>CGST 9%</span>
						{% if doc.taxes %}
						<span style="text-align:right; min-width:80px;margin-left:200px;">{{"%.2f"|format(doc.taxes[0].tax_amount_after_discount_amount or 0)}}</span>
						{% else %}
						<span style="text-align:right; min-width:80px;margin-left:200px;">0.00</span>
						{% endif %}
					</div><br>
					<div style="display:flex; justify-content:space-between; margin:0; padding:0;">
						<span>SGST 9%</span>
						{% if doc.taxes %}
						<span style="text-align:right; min-width:80px;margin-left:200px;">{{"%.2f"|format(doc.taxes[1].tax_amount_after_discount_amount or 0)}}</span>
						{% else %}
						<span style="text-align:right; min-width:80px;margin-left:200px;">0.00</span>
						{% endif %}
	
					</div><br>
				</td>
			</tr>
		</table>

		<table class="remove-top-border" style="table-layout:fixed; width:100%; border-left:none;">
			<tr>
				<td colspan="3" style="border:none; line-height:1.2em;">Lot Code: {{doc.custom_lot_code or ''}}</td>
				<td colspan="2" style="border:none; line-height:1.2em;">No of Boxes: {{ doc["items"] | sum(attribute='custom_no_of_boxes') }}</td>
				<td colspan="2"  style="border-left:none;border-right:none;font-weight:bold;border-top:hidden;line-height:1px;border-left:1px solid black;border-right:none;border-bottom:none;">Net Total : </td>
				{% if doc.taxes %}
				<td colspan="2" style="border-left:none;font-weight:bold;text-align:right;border-top:1px solid black;line-height:1px;border-left:none;border-right:none;border-bottom:none;"><div style="padding-right:6px;">{{"%.2f"|format((doc.total + doc.taxes[0].tax_amount_after_discount_amount + doc.taxes[1].tax_amount_after_discount_amount) or '') }}</div></td>
				{% else %}
				<td colspan="2" style="border-left:none;font-weight:bold;text-align:right;border-top:1px solid black;line-height:1px;border-left:none;border-right:none;border-bottom:none;">{{"%.2f"|format((doc.total + 0 + 0) or '') }}</td>
				{% endif %}
			</tr>
			<tr>
				<td colspan="3" style="border:none; line-height:1.2em;">Remarks: {{doc.remarks or ''}}</td>
				<td colspan="2" style="border:none; line-height:1.2em;border-right:1px solid black;">Vehicle No: {{doc.mode_of_transport or ''}}</td>
			</tr>
		</table>



		<table class="remove-bottom-border " style="table-layout:fixed; width:100%;">
			<tr>
				<td colspan="5" style="border-right:none;border-left:none;"><b>Amount in words :</b> {{doc.in_words}}</td>
				<td colspan="4" class="text-left" style="font-size:9px;border-right:none;">Certified that the particulars given above are true and correct and the
amount indicated represents the price actually charged and that there is<br> no
additional consideration directly or indirectly from the buyer </td>
			</tr>
		</table>
		<table>
	{% set current_time = frappe.utils.format_datetime(frappe.utils.now_datetime(), "dd/MM/yyyy HH:mm") %}
	<tr>
		<td colspan="6" style="border-left:none;border-bottom:none;border-right:none;">
			Date & Time of Invoice : {{ frappe.utils.format_datetime(doc.creation, "dd/MM/yyyy HH:mm") }}<br>
			Date & Time of Removal : {{ current_time or '' }}
		</td>
		<td colspan="3" class="text-right" style="border-bottom:none;border-left:none;border-right:none;">
			For<b> Wonjin Autoparts India Pvt Ltd</b>
		</td>
	</tr>
</table>

<!--<table class="remove-top-border">-->
<!--    <tr>-->
<!--        <td class="text-center" style="border-top:none;border-right:none;border-left:none;"><div style="padding-top:30px;">Prepared by</div></td>-->
<!--        <td class="text-center" style="border-top:none;border-left:none;border-right:none;border-left:none;"><div style="padding-top:30px;">Checked by</div></td>-->
<!--        <td class="text-center" style="border-top:none;border-left:none;border-left:none;"><div style="padding-top:30px;">Authorised Signatory</div></td>-->
<!--    </tr>-->
<!--</table>-->

	{% set prepared_by = frappe.db.get_value("Employee", {"user_id": doc.owner}, ["custom_digital_signature"]) %}
	{% set hod = frappe.db.get_value("Employee", {"user_id": doc.custom_hod}, ["custom_digital_signature"]) %}
	{% set finance = frappe.db.get_value("Employee", {"user_id": doc.custom_finance_approver}, ["custom_digital_signature"]) %}

	<style>
		.signature-img {
			width: 100px;
			height: 50px;
		}
	</style>

<div class="invoice-footer" style="margin-top:30px; page-break-inside: avoid;">

  <div class="signature-block" style="width:100%; text-align:center; margin-top:40px;">
	  <table style="width:100%; border:none;">
		  <tr>
			  <td class="text-center" style="border:none;width:30%">
			{% if prepared_by %}
			<div>
				{% if prepared_by %}
				<img src="{{ prepared_by }}" class="signature-img">
				{% endif %}<br>{{ frappe.utils.format_datetime(doc.creation, "dd-MM-yyyy HH:mm:ss") or '' }}<br>Prepared by
			</div>
			{% else %}
			<div style="padding-top:52px;">{{ frappe.utils.format_datetime(doc.creation, "dd-MM-yyyy HH:mm:ss") or '' }}<br>Prepared by</div>
			{% endif %}
			
		</td>
		<td class="text-center" style="border:none;width:40%">
			{% if hod %}
			<div>
				{% if hod %}
				<img src="{{ hod }}" class="signature-img">
				{% endif %}<br>{{ frappe.utils.format_datetime(doc.custom_hod_approved_on, "dd-MM-yyyy HH:mm:ss") or '' }}<br>Checked by
			</div>
			{% else %}
			<div style="padding-top:52px;">{{ frappe.utils.format_datetime(doc.custom_hod_approved_on, "dd-MM-yyyy HH:mm:ss") or '' }}<br>Checked by</div>
			{% endif %}
			
		</td>
		<td class="text-center" style="border:none;width:30%">
			{% if finance %}
			<div>
				{% if finance %}
				<img src="{{ finance }}" class="signature-img">
				{% endif %}<br>{{ frappe.utils.format_datetime(doc.custom_finance_approver_approved_on, "dd-MM-yyyy HH:mm:ss") or '' }}<br>Authorized by
			</div>
			{% else %}
			<div style="padding-top:52px;">{{ frappe.utils.format_datetime(doc.custom_finance_approver_approved_on, "dd-MM-yyyy HH:mm:ss") or '' }}<br>Authorized by</div>
			{% endif %}
			
		</td>
		  </tr>
	  </table>
  </div>
</div>


	</div>


</div>

	"""

	html = render_template(template, {"doc": doc})
	return {"html": html}

def set_required_item_in_work_order(doc, method):
	for row in doc.required_items:
		if row.idx > 1 or row.operation:
			return
		row.operation == frappe.db.get_value("Work Order Operation", {"idx": 1, "parent": doc.name}, "operation")

@frappe.whitelist()
def update_port_in_si(doc,method):
	country=frappe.db.get_value('Customer',doc.customer,'custom_country')
	if country!='India' and not doc.port_address:
		doc.port_address=doc.company_address


@frappe.whitelist()
def get_workflow_states(doctype):
	workflows = frappe.get_all(
		"Workflow",
		filters={"document_type": doctype,"is_active":1},
		pluck="name"
	)
	if not workflows:
		return []

	states = frappe.get_all(
		"Workflow Document State",
		filters={"parent": ["in", workflows],'state':["not in",['Cancelled']]},
		pluck="state"
	)

	unique_states = sorted(set(states))
	return unique_states

@frappe.whitelist()
def get_workflow_roles(doctype):
	workflows = frappe.get_all(
		"Workflow",
		filters={"document_type": doctype, "is_active": 1},
		pluck="name"
	)
	if not workflows:
		return []

	roles = frappe.get_all(
		"Workflow Document State",
		filters={"parent": ["in", workflows]},
		pluck="allow_edit" 
	)

	unique_roles = sorted(set(roles))
	return unique_roles

@frappe.whitelist()
def trigger_notification_based_on_the_workflow_in_si(doc,method):
	if not doc.is_new():
		if doc.has_value_changed("workflow_state") and doc.workflow_state !='Pending For HOD':
			create_workflow_notification(doc.doctype, doc.name, doc.workflow_state)
		elif doc.has_value_changed("workflow_state") and doc.workflow_state == 'Pending For HOD':
			emp = frappe.db.get_value(
				"Employee",
				{"user_id": doc.owner},
				["reports_to"],
				as_dict=True
			)
			recipients = []
			if emp:
				recipients = frappe.db.get_value("Employee",emp.reports_to,'user_id')

				if recipients:
					subject = f"{doc.doctype} Notification"

					message = f"""
						<p>Dear Sir/Madam,</p>
						<p>
							This is to inform you that 
							<strong>
								<a href="{frappe.utils.get_link_to_form(doc.doctype, doc.name)}" target="_blank">
									{doc.name}
								</a>
							</strong>
							is pending for your approval.
						</p>
						<p>
							Kindly review the document at your earliest convenience. 
							Your timely action will help us proceed without delays.
						</p>
						<p>Thank you for your prompt attention.</p>
						<br>
						<p>
							Best regards,<br>
							{emp.employee_name if emp else doc.owner}<br>
							{emp.department if emp and emp.department else ""}
						</p>
					"""

					frappe.sendmail(
						recipients=recipients,
						subject=subject,
						message=message
					)

@frappe.whitelist()
def trigger_notification_based_on_the_workflow_in_si_for_draft(doc,method):
	if doc.workflow_state:
		create_workflow_notification(doc.doctype, doc.name, doc.workflow_state)




import frappe
from collections import defaultdict

@frappe.whitelist()
def customer_wise_plan(month):
	
	grouped_data = defaultdict(lambda: defaultdict(float))

	
	schedules = frappe.db.get_all(
		'Sales Order Schedule',
		filters={'schedule_month': month},
		fields=[
			'customer_code',
			'schedule_amount_inr'
		],
		order_by='customer_code'
	)

	for entry in schedules:
		customer = entry.customer_code
		amount = entry.schedule_amount_inr or 0.0

		
		customer_group = frappe.db.get_value('Customer',{"customer_code": customer}, 'customer_group') or ''

		
		grouped_data[customer_group][customer] += amount

	return grouped_data






@frappe.whitelist()
def department_wise_plan(month):
	grouped_data = defaultdict(float)

	schedules = frappe.db.get_all(
		'Sales Order Schedule',
		filters={'schedule_month': month},
		fields=[
			'item_group',
			'schedule_amount_inr',
			'order'
		],
		
	)

	schedules.sort(key=lambda x: (x.order == 0 or x.order is None, x.order or 0))

	for entry in schedules:
		item_group = entry.item_group
		amount = entry.schedule_amount_inr or 0.0

		grouped_data[item_group] += amount

	return dict(grouped_data)


@frappe.whitelist()
def customer_group_wise(month):
	grouped_data = defaultdict(float)

	schedules = frappe.db.get_all(
		'Sales Order Schedule',
		filters={'schedule_month': month},
		fields=[
			'customer_group',
			'schedule_amount_inr',
			'order'
		],
		
	)

	for entry in schedules:
		customer_group = entry.customer_group
		amount = entry.schedule_amount_inr or 0.0

		grouped_data[customer_group] += amount

	return dict(grouped_data)




@frappe.whitelist()
def supplier_wise_plan(month):
	
	grouped_data = defaultdict(lambda: defaultdict(float))

	
	schedules = frappe.db.get_all(
		'Purchase Order Schedule',
		filters={'schedule_month': month},
		fields=[
			'supplier_code',
			'schedule_amount_inr'
		],
		order_by='supplier_code'
	)

	for entry in schedules:
		supplier = entry.supplier_code
		amount = entry.schedule_amount_inr or 0.0

		
		supplier_group = frappe.db.get_value('Supplier',{"supplier_code": supplier}, 'supplier_group') or ''

		
		grouped_data[supplier_group][supplier] += amount

	return grouped_data



@frappe.whitelist()
def department_wise_plan_pos(month):
	grouped_data = defaultdict(float)

	schedules = frappe.db.get_all(
		'Purchase Order Schedule',
		filters={'schedule_month': month},
		fields=[
			'item_group',
			'schedule_amount_inr'
		],
		
	)

	schedules.sort(key=lambda x: x.get('item_group', ''))

	for entry in schedules:
		item_group = entry.item_group
		amount = entry.schedule_amount_inr or 0.0

		grouped_data[item_group] += amount

	return dict(grouped_data)



@frappe.whitelist()
def supplier_group_wise(month):
	grouped_data = defaultdict(float)

	schedules = frappe.db.get_all(
		'Purchase Order Schedule',
		filters={'schedule_month': month},
		fields=[
			'supplier_group',
			'schedule_amount_inr',
			
		],
		
	)

	for entry in schedules:
		supplier_group = entry.supplier_group
		amount = entry.schedule_amount_inr or 0.0

		grouped_data[supplier_group] += amount

	return dict(grouped_data)




def get_permission_query_conditions(user):
	roles = frappe.get_roles(user)
	if "Stock User" in roles or user == "rawmaterial@onegeneindia.in":
		return None  # no filters = see all
	return None  # default

def has_permission(doc, user):
	roles = frappe.get_roles(user)
	if "Stock User" in roles or user == "rawmaterial@onegeneindia.in":
		return True  # allow opening any Material Request or Employee
	return None  # default permission



@frappe.whitelist()
def get_balance_schedule(pdate, item_code=None, po_no=None, customer_code=None):
	if not (pdate and po_no and customer_code):
		return None

	date_obj = getdate(pdate)
	year = date_obj.year
	month_str = date_obj.strftime("%b").upper()  

	schedule = frappe.db.get_value(
		"Sales Order Schedule",
		{
			"sales_order_number": po_no,
			"schedule_month": month_str,
			"schedule_year": year,
			"item_code":item_code,
			"customer_code":customer_code,
			"docstatus": 1,
		},
		["pending_qty"],
		as_dict=True,
	)

	if not schedule:
		return None

	return {"pending_qty": schedule.pending_qty}





# @frappe.whitelist()
# def request_vs_actual_mt(from_date, to_date):
	
	

	
# 	# results = frappe.db.sql("""
# 	# 	SELECT 
# 	# 		item_code,
# 	# 		item_name,
# 	# 		SUM(custom_requesting_qty) AS requested_qty,
# 	# 		SUM(ordered_qty) AS completed_qty
# 	# 	FROM 
# 	# 		`tabMaterial Request Item`
# 	# 	WHERE 
# 	# 		parent IN (
# 	# 			SELECT name FROM `tabMaterial Request`
# 	# 			WHERE transaction_date BETWEEN %s AND %s
# 	# 		)
# 	# 		AND docstatus = 1
# 	# 	GROUP BY 
# 	# 		item_code, item_name
# 	# """, (from_date, to_date), as_dict=True)
 
# 	dep_per =[]
# 	results = []
	
	
	
 
# 	if frappe.db.exists("User Permission", {"user": frappe.session.user, "allow": "Department"}):
		
# 		dep = frappe.db.get_all("User Permission", filters={"user": frappe.session.user, "allow": "Department"},pluck="for_value")

# 		if dep:
# 			for i in dep:
				 
# 				dep_per.append(i) 
# 				dep_per_each = frappe.db.get_all("Department",filters={"parent_department":i},pluck="name")
# 				if dep_per_each:
					
# 					dep_per.extend(dep_per_each)
				
					
		
# 		if dep_per :
		

# 			mr= frappe.get_all("Material Request", 
# 						filters={"transaction_date": ["between", [from_date, to_date]],"docstatus": 1,"custom_department":["in",dep_per]},
# 						pluck="name")
		
		
# 	else:
# 		mr= frappe.get_all("Material Request", 
# 				filters={"transaction_date": ["between", [from_date, to_date]],"docstatus": 1},
# 				pluck="name")    
	
# 	if mr:
  
  
# 		data = frappe.get_all(
# 			"Material Request Item",
# 			filters={
# 				"parent": ["in", mr ]
				
# 			},
# 			fields=["item_code", "item_name", "custom_requesting_qty", "ordered_qty"]
# 		)

		
# 		grouped_data = {}
# 		for item in data:
# 			key = (item["item_code"], item["item_name"])
# 			if key not in grouped_data:
# 				grouped_data[key] = {"requested_qty": 0, "completed_qty": 0}
# 			grouped_data[key]["requested_qty"] += item.get("custom_requesting_qty", 0) or 0
# 			grouped_data[key]["completed_qty"] += item.get("ordered_qty", 0) or 0

		
		
# 		for (item_code, item_name), qtys in grouped_data.items():
# 			results.append({
# 				"item_code": item_code,
# 				"item_name": item_name,
# 				"requested_qty": qtys["requested_qty"],
# 				"completed_qty": qtys["completed_qty"]
# 			})


	
# 	return results

@frappe.whitelist()
def request_vs_actual_mt(from_date, to_date):
	dep_per = []
	results = []

	
	def get_all_sub_departments(dept):
		children = frappe.db.get_all(
			"Department",
			filters={"parent_department": dept},
			pluck="name"
		)
		all_children = list(children)
		for child in children:
			all_children.extend(get_all_sub_departments(child))
		return all_children

	
	if frappe.db.exists("User Permission", {"user": frappe.session.user, "allow": "Department"}):
		dep = frappe.db.get_all(
			"User Permission",
			filters={"user": frappe.session.user, "allow": "Department"},
			pluck="for_value"
		)

		if dep:
			for i in dep:
				dep_per.append(i)  
				dep_per.extend(get_all_sub_departments(i))  

		if dep_per:
			mr = frappe.get_all(
				"Material Request",
				filters={
					"transaction_date": ["between", [from_date, to_date]],
					"docstatus": 1,
					"custom_department": ["in", dep_per]
				},
				pluck="name"
			)
	else:
		mr = frappe.get_all(
			"Material Request",
			filters={"transaction_date": ["between", [from_date, to_date]], "docstatus": 1},
			pluck="name"
		)

	
	if mr:
		data = frappe.get_all(
			"Material Request Item",
			filters={"parent": ["in", mr]},
			fields=["item_code", "item_name", "custom_requesting_qty", "ordered_qty"]
		)

		grouped_data = {}
		for item in data:
			key = (item["item_code"], item["item_name"])
			if key not in grouped_data:
				grouped_data[key] = {"requested_qty": 0, "completed_qty": 0}
			grouped_data[key]["requested_qty"] += item.get("custom_requesting_qty", 0) or 0
			grouped_data[key]["completed_qty"] += item.get("ordered_qty", 0) or 0

		for (item_code, item_name), qtys in grouped_data.items():
			results.append({
				"item_code": item_code,
				"item_name": item_name,
				"requested_qty": qtys["requested_qty"],
				"completed_qty": qtys["completed_qty"]
			})

	return results











@frappe.whitelist()
def download_request_vs_actual_excel(from_date, to_date):
	
	records = get_request_vs_actual_data(from_date, to_date)
	
	formatted_from_date = frappe.format(from_date,{"fieldtype":"Date"})
	formatted_to_date = frappe.format(to_date,{"fieldtype":"Date"})

	filename = f"Requested_vs_Actual_{formatted_from_date}_to_{formatted_to_date}.xlsx"
	xlsx_file = make_xlsx_css("Requested vs Actual", records, from_date, to_date)

	frappe.response['filename'] = filename
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


def get_request_vs_actual_data(from_date, to_date):
	
	dep_per = []
	results = []

	def get_all_sub_departments(dept):
		children = frappe.db.get_all(
			"Department",
			filters={"parent_department": dept},
			pluck="name"
		)
		all_children = list(children)
		for child in children:
			all_children.extend(get_all_sub_departments(child))
		return all_children

	
	if frappe.db.exists("User Permission", {"user": frappe.session.user, "allow": "Department"}):
		dep = frappe.db.get_all(
			"User Permission",
			filters={"user": frappe.session.user, "allow": "Department"},
			pluck="for_value"
		)
		if dep:
			for i in dep:
				dep_per.append(i)
				dep_per.extend(get_all_sub_departments(i))
		if dep_per:
			mr = frappe.get_all(
				"Material Request",
				filters={
					"transaction_date": ["between", [from_date, to_date]],
					"docstatus": 1,
					"custom_department": ["in", dep_per]
				},
				pluck="name"
			)
	else:
		mr = frappe.get_all(
			"Material Request",
			filters={"transaction_date": ["between", [from_date, to_date]], "docstatus": 1},
			pluck="name"
		)

	
	if mr:
		data = frappe.get_all(
			"Material Request Item",
			filters={"parent": ["in", mr]},
			fields=["item_code", "item_name", "custom_requesting_qty", "ordered_qty"]
		)

		grouped_data = {}
		for item in data:
			key = (item["item_code"], item["item_name"])
			if key not in grouped_data:
				grouped_data[key] = {"requested_qty": 0, "completed_qty": 0}
			grouped_data[key]["requested_qty"] += item.get("custom_requesting_qty", 0) or 0
			grouped_data[key]["completed_qty"] += item.get("ordered_qty", 0) or 0

		for (item_code, item_name), qtys in grouped_data.items():
			results.append({
				"item_code": item_code,
				"item_name": item_name,
				"requested_qty": qtys["requested_qty"],
				"completed_qty": qtys["completed_qty"]
			})

	return results


# def make_xlsx_css(sheet_name, records, from_date, to_date, wb=None):    
	
#     if wb is None:
#         wb = openpyxl.Workbook()

#     ws = wb.active
#     ws.title = sheet_name

	
#     header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
#     total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#     bold_font = Font(bold=True)
#     center = Alignment(horizontal="center", vertical="center")
#     left_align = Alignment(horizontal="left", vertical="center")
#     thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
#                          top=Side(style='thin'), bottom=Side(style='thin'))
	
#     formatted_from_date = frappe.format(from_date,{"fieldtype":"Date"})
#     formatted_to_date = frappe.format(to_date,{"fieldtype":"Date"})
	
#     ws.merge_cells('A1:D1')
#     ws['A1'] = f"Requested vs Actual (From {formatted_from_date} To {formatted_to_date})"
#     ws['A1'].font = Font(bold=True, size=14)
#     ws['A1'].alignment = center

	
#     headers = ["Item Code", "Item Name", "Requested", "Issued"]
#     ws.append(headers)

#     for col_num, header in enumerate(headers, 1):
#         cell = ws.cell(row=2, column=col_num)
#         cell.value = header
#         cell.fill = header_fill
#         cell.font = bold_font
#         cell.alignment = center
#         cell.border = thin_border

	
#     total_requested = 0
#     total_completed = 0
#     for i, row in enumerate(records, start=3):
#         ws.cell(i, 1, row["item_code"])
#         ws.cell(i, 2, row["item_name"])
#         ws.cell(i, 3, row["requested_qty"])
#         ws.cell(i, 4, row["completed_qty"])

#         total_requested += row["requested_qty"] or 0
#         total_completed += row["completed_qty"] or 0

#         for col in range(1, 5):
#             ws.cell(i, col).border = thin_border

   
#     total_row = len(records) + 3
#     ws.cell(total_row, 1, "Total")
#     ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
#     ws.cell(total_row, 3, total_requested)
#     ws.cell(total_row, 4, total_completed)

#     for col in range(1, 5):
#         cell = ws.cell(total_row, col)
#         cell.font = bold_font
#         cell.fill = total_fill
#         cell.border = thin_border
#         cell.alignment = center

	
#     for col in ws.columns:
#         max_length = 0
#         col_letter = openpyxl.utils.get_column_letter(col[0].column)
#         for cell in col:
#             try:
#                 if cell.value:
#                     max_length = max(max_length, len(str(cell.value)))
#             except:
#                 pass
#         ws.column_dimensions[col_letter].width = max_length + 3

	
#     xlsx_file = BytesIO()
#     wb.save(xlsx_file)
#     xlsx_file.seek(0)
#     return xlsx_file

def make_xlsx_css(sheet_name, records, from_date, to_date, wb=None):
	

	if wb is None:
		wb = openpyxl.Workbook()

	ws = wb.active
	ws.title = sheet_name

	
	header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
	total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
	bold_font = Font(bold=True)
	center = Alignment(horizontal="center", vertical="center")
	left_align = Alignment(horizontal="left", vertical="center")
	right_align = Alignment(horizontal="right", vertical="center")
	thin_border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)

	
	formatted_from_date = frappe.format(from_date, {"fieldtype": "Date"})
	formatted_to_date = frappe.format(to_date, {"fieldtype": "Date"})

	ws.merge_cells('A1:D1')
	ws['A1'] = f"Requested vs Actual (From {formatted_from_date} To {formatted_to_date})"
	ws['A1'].font = Font(bold=True, size=14)
	ws['A1'].alignment = center

	
	headers = ["Item Code", "Item Name", "Requested", "Issued"]
	ws.append(headers)

	for col_num, header in enumerate(headers, 1):
		cell = ws.cell(row=2, column=col_num)
		cell.value = header
		cell.fill = header_fill
		cell.font = bold_font
		cell.alignment = center
		cell.border = thin_border

	
	total_requested = 0
	total_completed = 0
	row_index = 3
	
	for i, row in enumerate(records, start=3):
		requested_qty = row["requested_qty"] or 0
		completed_qty = row["completed_qty"] or 0
		
		if requested_qty <= 0:
			continue

		ws.cell(i, 1, row["item_code"])
		ws.cell(i, 2, row["item_name"])
		ws.cell(i, 3, "-" if requested_qty == 0 else requested_qty)
		ws.cell(i, 4, "-" if completed_qty == 0 else completed_qty)

		total_requested += requested_qty
		total_completed += completed_qty

		for col in range(1, 5):
			ws.cell(i, col).border = thin_border
			ws.cell(i, col).alignment = Alignment(horizontal="right" if col > 2 else "left", vertical="center")
		row_index += 1     

	
	# total_row = len(records) + 3
	total_row = row_index
	ws.cell(total_row, 1, "Total")
	ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
	ws.cell(total_row, 3, "-" if total_requested == 0 else total_requested)
	ws.cell(total_row, 4, "-" if total_completed == 0 else total_completed)

	
	for col in range(1, 5):
		cell = ws.cell(total_row, col)
		cell.font = bold_font
		cell.fill = total_fill
		cell.border = thin_border
		
		if col in (1, 2):
			cell.alignment = left_align
		else:
			cell.alignment = right_align

	
	for col in ws.columns:
		max_length = 0
		col_letter = openpyxl.utils.get_column_letter(col[0].column)
		for cell in col:
			try:
				if cell.value:
					max_length = max(max_length, len(str(cell.value)))
			except Exception:
				pass
		ws.column_dimensions[col_letter].width = max_length + 3

	
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	xlsx_file.seek(0)
	return xlsx_file

	
@frappe.whitelist()
def today_req_qty_update(item_code, date):
	from frappe.utils import get_datetime, format_datetime
	from datetime import timedelta

	start_dt = get_datetime(date + " 08:31:00")
	end_dt = start_dt + timedelta(days=1) - timedelta(seconds=1)

	result = frappe.db.sql("""
	SELECT SUM(mti.issued_qty)
	FROM `tabMaterial Transfer Items` mti
	JOIN `tabMaterial Transfer` mt ON mti.parent = mt.name
	WHERE
	mt.creation >= %s
	AND mt.creation <= %s
	AND mti.item_code = %s
""", (start_dt, end_dt, item_code))

	
	total_issued_qty = result[0][0] if result and result[0][0] is not None else 0

	return total_issued_qty



@frappe.whitelist()
def get_rate_from_sales_order(item_code, customer_po, customer_code):
	rate = frappe.db.sql("""
		SELECT soi.rate
		FROM `tabSales Order Item` AS soi
		JOIN `tabSales Order` AS so ON so.name = soi.parent
		WHERE soi.item_code = %s
		AND so.po_no = %s AND so.custom_customer_code = %s
		AND so.docstatus = 1
		LIMIT 1
	""", (item_code, customer_po, customer_code), as_dict=True)

	if rate:
		return rate[0].rate
	else:
		return None
	
def validate_customer_po(doc, method):
	"""Do not allow saving Sales Order with duplicate Customer PO Number for the same Customer"""

	if doc.po_no and doc.customer:
		existing_so = frappe.db.get_value(
			"Sales Order",
			{
				"po_no": doc.po_no,
				"customer": doc.customer,
				"docstatus": ["!=", 2],
				"name": ["!=", doc.name]
			},
			"name"
		)

		if existing_so:
			frappe.throw(
				f"""
				<b>Duplicate Customer Purchase Order</b><br><br>
				Customer: <b>{doc.customer}</b><br>
				PO Number: <b>{doc.po_no}</b><br><br>
				This PO is already used in Sales Order:
				<a href="/app/sales-order/{existing_so}" target="_blank"><b>{existing_so}</b></a>
				""",
				title="Duplicate Customer PO"
			)


@frappe.whitelist()
def get_inspection_ids(exclude_qids, customer):
	""" QID data will be fetched from the Quality Inspection
	"""
	options = []
	exclude_qids = frappe.parse_json(exclude_qids) if exclude_qids else []

	sales_order_items = frappe.db.sql("""
		SELECT DISTINCT soi.item_code
		FROM `tabSales Order` so
		INNER JOIN `tabSales Order Item` soi on so.name = soi.parent
		WHERE so.customer = %s
	""", (customer,), as_list=True)

	# nothing to offer
	if not sales_order_items:
		return options

	item_codes = tuple(d[0] for d in sales_order_items)

	if frappe.db.exists("User Permission", {"user": frappe.session.user, "allow": "Item Group"}):
		item_groups = [d[0] for d in frappe.db.get_all(
			"User Permission",
			{"user": frappe.session.user, "allow": "Item Group"},
			"for_value",
			as_list=1
		)]

		if item_groups:
			qids = frappe.db.sql("""
				SELECT p.name AS parent, c.qid_data, c.quantity
				FROM `tabQuality Inspection` p
				INNER JOIN `tabQuality Inspection QID` c ON c.parent = p.name
				WHERE c.store_receipt = 1
				AND c.sales_invoice = 0
				AND p.item_code IN %(item_codes)s
				AND p.custom_item_group IN %(groups)s
				AND p.docstatus = 1
				AND p.workflow_state != 'Rejected'
				{exclude_clause}
			""".format(
				exclude_clause="AND c.qid_data NOT IN %(exclude_qids)s" if exclude_qids else ""
			), {
				"item_codes": item_codes,
				"groups": tuple(item_groups),
				"exclude_qids": tuple(exclude_qids)
			}, as_dict=1)

			for q in qids:
				options.append({
					"label": f"""{q.qid_data}<br>
						<span style='font-size: 11px; color: #888;'>
							Qty: {q.quantity}, Quality Inspection: {q.parent}
						</span>""",
					"value": q.qid_data
				})

	else:
		qids = frappe.db.sql("""
			SELECT p.name AS parent, c.qid_data, c.quantity
			FROM `tabQuality Inspection` p
			INNER JOIN `tabQuality Inspection QID` c ON c.parent = p.name
			WHERE c.store_receipt = 1
			AND c.sales_invoice = 0
			AND p.docstatus = 1
			AND p.item_code IN %(item_codes)s
			AND p.workflow_state != 'Rejected'
			{exclude_clause}
		""".format(
			exclude_clause="AND c.qid_data NOT IN %(exclude_qids)s" if exclude_qids else ""
		), {
			"item_codes": item_codes,
			"exclude_qids": tuple(exclude_qids)
		}, as_dict=1)

		for q in qids:
			options.append({
				"label": f"""{q.qid_data}<br>
					<span style='font-size: 11px; color: #888;'>
						Qty: {q.quantity}, Quality Inspection: {q.parent}
					</span>""",
				"value": q.qid_data
			})
	return options

@frappe.whitelist()
def get_quality_inspection_data_for_si(qid, customer):
	if not frappe.db.exists("Quality Inspection QID", {"qid_data": qid}):
		return "not found"

	if frappe.db.exists("Quality Inspection QID", {"qid_data": qid, "store_receipt": 0}):
		frappe.throw("The materials in the Bin has not moved to the stores", 
			   title="QID Expired")
		frappe.throw("Bin-இல் உள்ள materials இன்னும் move செய்யப்படவிள்ளை.", title="Invalid QID")
  
	if frappe.db.exists("Quality Inspection QID", {"qid_data": qid, "sales_invoice": 1}):
		frappe.throw("The materials in the Bin has already been invoiced", 
			   title="QID Expired")
		frappe.throw("Bin-இல் உள்ள materials ஏற்கனவே invoice செய்யப்பட்டது.", title="QID Expired")

	quality_inspection = frappe.db.get_value("Quality Inspection QID", {"qid_data": qid}, "parent")
	accepted_qty = frappe.db.get_value("Quality Inspection QID", {"qid_data": qid}, "quantity") or 0
	item_code = frappe.db.get_value("Quality Inspection", quality_inspection, "item_code")
	check_item_exists_for_customer(item_code, customer, qid)
	rack = frappe.db.get_value("Item Rack", {"parent": item_code, "is_default": 1}, "rack") or ""
	location = frappe.db.get_value("Item Rack Location", {"parent": item_code, "is_default": 1}, "location") or ""
	uom = frappe.db.get_value("Item", item_code, "stock_uom")
	item_name = frappe.db.get_value("Quality Inspection", quality_inspection,['item_name'])
	return item_code, accepted_qty, rack, location, uom, item_name, quality_inspection

def check_item_exists_for_customer(item_code, customer, qid=None):
	sales_orders = frappe.get_all(
		"Sales Order",
		filters={"customer": customer, "docstatus": 1},
		pluck="name"
	)
	if not sales_orders:
		frappe.throw(
			f"QID <b>{qid}</b> is not applicable for the customer <b>{customer}</b>",
			title="Invalid QID"
		)
		
	if not frappe.db.exists("Sales Order Item",{"parent": ["in", sales_orders], "item_code": item_code}):
		frappe.throw(
			f"QID <b>{qid}</b> is not applicable for the customer <b>{customer}</b>",
			title="Invalid QID"
		)

def mark_qid_for_si(doc, method):
	qids = []
	for row in doc.items:
		if row.custom_qid:
			parts = row.custom_qid.split(',')
			qids.extend(parts)
	for row in qids:
		print(row)
		frappe.db.set_value("Quality Inspection QID", {"qid_data": row}, "sales_invoice", 1)
  
def unmark_qid_for_si(doc, method):
	qids = []
	for row in doc.items:
		if row.custom_qid:
			parts = row.custom_qid.split(',')
			qids.extend(parts)
	for row in qids:
		print(row)
		frappe.db.set_value("Quality Inspection QID", {"qid_data": row}, "sales_invoice", 0)
	
@frappe.whitelist()
def update_existing_asn():
	asn_list = frappe.db.get_all('Advance Shipping Note', 
								 filters={"workflow_state": ['in', ['Gate Received', 'Material Received']]}, 
								 fields=['name'])

	for a in asn_list:
		a_doc = frappe.get_doc('Advance Shipping Note', a.name)

		if not frappe.db.exists('Gate Entry', {"entry_against": 'Advance Shipping Note', "entry_id": a.name}):
			print(a.name)
			ge = frappe.new_doc('Gate Entry')
			ge.entry_type = 'Inward'
			ge.entry_against = 'Advance Shipping Note'
			ge.entry_id = a.name
			ge.party_type = 'Supplier'
			ge.party = a_doc.supplier
			ge.supplier_code = frappe.db.get_value("Supplier",{"name":a_doc.supplier},"supplier_code") or ""

			if a_doc.workflow_state == 'Material Received':
				ge.entry_date = a_doc.received_date_time.date()
				ge.entry_time = a_doc.received_date_time
			else:
				ge.entry_date = a_doc.modified.date()
				ge.entry_time = a_doc.modified

			ge.ref_no = a_doc.confirm_supplier_dn
			ge.security_name = a_doc.security_name
			ge.vehicle_number = a_doc.vehicle_no
			ge.driver_name = a_doc.driver_name
			tot_box=0
			for i in a_doc.item_table:
				tot_box+=i.no_of_bins
				ge.append("gate_entry_items", {
					"item_code": i.item_code,
					"item_name": i.item_name,
					"uom": i.uom,
					"qty": i.dis_qty,
					"box": i.no_of_bins
				})
			ge.no_of_box=tot_box
			ge.save(ignore_permissions=True)
			frappe.db.commit() 

@frappe.whitelist()
def get_today_value():
	return today()

@frappe.whitelist()
def custom_update_items(data, sales_order=None, purchase_order=None):
	date = frappe.utils.today()
	month_abbr = frappe.utils.getdate(date).strftime("%b").upper()
	data = json.loads(data)
	items = data.get("items_table", [])

	if sales_order:
		order_doctype = "Sales Order"
		order_name = sales_order
		open_order_doctype = "Open Order"
		order_number_field = "sales_order_number"
		order_revision_doctype = "Sales Order Revision"
		order_schedule = "Sales Order Schedule"
		delivered_or_received_qty = "delivered_qty"
	
	elif purchase_order:
		order_doctype = "Purchase Order"
		order_name = purchase_order
		open_order_doctype = "Purchase Open Order"
		order_number_field = "purchase_order"
		order_revision_doctype = "Purchase Order Revision"
		order_schedule = "Purchase Order Schedule"
		delivered_or_received_qty = "received_qty"
	
	else:
		return False

	open_order = frappe.get_doc(open_order_doctype, {order_number_field: order_name})

	unmatched_items = []

	for order_row in items:
		matched = False

		for op_row in open_order.open_order_table:
			if order_row["item_code"] == op_row.item_code:
				matched = True

				# Update docname
				op_row.docname = order_row.get("docname")

				# Rate changed?
				if op_row.rate != order_row["rate"]:
					old_value = op_row.rate
					op_row.rate = order_row["rate"]

					# Add Revision Log
					next_idx = frappe.db.count(order_revision_doctype, {"parent": order_name}) + 1
					
					log = frappe.new_doc(order_revision_doctype)
					log.parent = order_name
					log.parenttype = order_doctype
					log.parentfield = "custom_revision_logs"
					log.idx = next_idx
					log.item_code = order_row["item_code"]
					log.old_value = old_value
					log.new_value = order_row["rate"]
					log.revised_on = now()
					log.revised_by = frappe.session.user
					log.insert(ignore_permissions=True)
					if order_doctype == "Purchase Order" and order_number_field == "purchase_order":
						order_number_field = order_number_field + "_number"
					# Update Order Schedule
					if frappe.db.exists(order_schedule, {
						"item_code": order_row["item_code"],
						order_number_field: order_name,
						"schedule_month": month_abbr
					}):
						os_doc = frappe.db.get_value(
							order_schedule,
							{
								"item_code": order_row["item_code"],
								order_number_field: order_name,
								"schedule_month": month_abbr
							},
							["name", "exchange_rate", "qty", delivered_or_received_qty, "pending_qty"],
							as_dict=True
						)
						if os_doc:
							if order_doctype == "Sales Order":
								update_vals = {
									"order_rate": order_row["rate"],
									"order_rate_inr": os_doc.exchange_rate * order_row["rate"],
									"schedule_amount": os_doc.qty * order_row["rate"],
									"schedule_amount_inr": os_doc.qty * order_row["rate"] * os_doc.exchange_rate,
									"delivered_amount": os_doc.delivered_qty * order_row["rate"],
									"delivered_amount_inr": os_doc.delivered_qty * order_row["rate"] * os_doc.exchange_rate,
									"pending_amount": os_doc.pending_qty * order_row["rate"],
									"pending_amount_inr": os_doc.pending_qty * order_row["rate"] * os_doc.exchange_rate
								}
							if order_doctype == "Purchase Order":
								update_vals = {
									"order_rate": order_row["rate"],
									"order_rate_inr": os_doc.exchange_rate * order_row["rate"],
									"schedule_amount": os_doc.qty * order_row["rate"],
									"schedule_amount_inr": os_doc.qty * order_row["rate"] * os_doc.exchange_rate,
									"received_amount": os_doc.received_qty * order_row["rate"],
									"received_amount_inr": os_doc.received_qty * order_row["rate"] * os_doc.exchange_rate,
									"pending_amount": os_doc.pending_qty * order_row["rate"],
									"pending_amount_inr": os_doc.pending_qty * order_row["rate"] * os_doc.exchange_rate
								}
							frappe.db.set_value(order_schedule, os_doc.name, update_vals)
				break  # Matching row handled → stop looping

		# AFTER LOOP: If NOT matched → append new row
		if not matched:
			open_order.append("open_order_table", {
				"item_code": order_row["item_code"],
				"rate": order_row["rate"],
				"qty": order_row.get("qty") or 1,
			})
			unmatched_items.append(order_row["item_code"])

	open_order.disable_update_items = 0
	open_order.save(ignore_permissions=True)
	return True

@frappe.whitelist()
def update_tax_category_by_state(doc, method):
	if doc.state and doc.country == "India":
		if doc.state == "Tamil Nadu":
			tax_category = "In-State"
		else: 
			tax_category = "Out-State"
		doc.tax_category = tax_category
		for row in doc.links:
			frappe.db.set_value(row.link_doctype, row.link_name, "tax_category", tax_category)

def bom_connection_from_fg_to_child(doc, method):
	frappe.db.set_value("BOM Item", {"item_code": doc.item,  "bom_no": ["in", [None, ""]]}, "bom_no", doc.name)

@frappe.whitelist()
def supplier_wise_plan_summary(month, year):
	total_sales_plan = 0
	# Get distinct supplier groups
	supplier_groups = frappe.db.get_all(
		"Purchase Order Schedule",
		filters={"schedule_month": month},
		fields=["supplier_group"],
		distinct=True
	)

	html = f"""
	<div style="border: 1px solid gray; color: black;">
		<h4 style="font-weight: 600; text-align: center; margin-top: 10px;">
			Supplier Wise Purchase Plan - {month}'{year}
		</h4>
		<div style="white-space: nowrap; overflow-x: auto; display:flex; gap:10px; padding-top:10px;">
	"""

	# Loop supplier groups
	for sg in supplier_groups:

		html += """
			<div>
			<table border="1" cellspacing="0" cellpadding="6" style="min-width:400px;">
				<tr>
					<th style="background-color:#ffc000; text-align:center; width: 150px; font-size: 12px; font-weight: bold;">Type</th>
					<th style="background-color:#ffc000; text-align:center; width: 100px; font-size: 12px; font-weight: bold;">Supplier</th>
					<th style="background-color:#ffc000; text-align:center; width: 150px; font-size: 12px; font-weight: bold; color: #2206f6;">Schedule Value</th>
				</tr>
		"""

		schedules = frappe.db.get_all(
			"Purchase Order Schedule",
			filters={"schedule_month": month, "supplier_group": sg.supplier_group},
			fields=[
				"supplier_code",
				"SUM(schedule_amount_inr) as schedule_amount"
			],
			group_by="supplier_code",
			order_by="supplier_code"
		)

		rowspan = len(schedules) + 1
		total_amount = 0
		first_row = True

		for row in schedules:
			supplier_name = frappe.db.get_value("Supplier", {"supplier_code": row.supplier_code}, "supplier_name")
			rounded_amount = round(row.schedule_amount)
			if first_row:
				html += f"""
				<tr>
					<td rowspan="{rowspan}" style="color: #b00d0a; font-weight: bold; width: 150px; text-align: center; font-size: 13px;">{sg.supplier_group}</td>
					<td style="font-weight: bold; width: 150px; font-size: 13px;">{supplier_name}</td>
					<td style="text-align:right; font-weight: bold; width: 150px; font-size: 13px;">₹ {fmt_money(int(rounded_amount), precision=0)}</td>
				</tr>
				"""
				first_row = False
			else:
				html += f"""
				<tr>
					<td style="font-weight: bold; font-size: 13px;">{supplier_name}</td>
					<td style="text-align:right; font-weight: bold; font-size: 13px;">₹ {fmt_money(int(rounded_amount), precision=0)}</td>
				</tr>
				"""

			total_amount += rounded_amount
		rounded_total_amount = round(total_amount)
		# Correct subtotal row (4 <td> values)
		html += f"""
			<tr>
				<td style="background-color: #00ffff;">SUB TOTAL</td>
				<td style="background-color: #00ffff; text-align:right;">₹ {fmt_money(int(rounded_total_amount), precision=0)}</td>
			</tr>
		</table>
		</div>
		"""
		total_sales_plan += rounded_total_amount

	html += f"""
		</div>
		<div style="text-align: center; background-color: #ffff00; margin-bottom: 0px; min-height: 30px;">
			<b style="padding-top: 5px;">Total Sales Plan: ₹ {fmt_money(int(total_sales_plan), precision=0)}</b>
		</div>
	</div>
	"""

	return html
 
@frappe.whitelist()
def custom_get_children(parent=None, is_root=False, **filters):
	"""Override against bom's get_children method"""

	if not parent or parent == "BOM":
		# frappe.msgprint(_("Please select a BOM"))
		return

	if parent:
		frappe.form_dict.parent = parent

	if frappe.form_dict.parent:
		bom_doc = frappe.get_cached_doc("BOM", frappe.form_dict.parent)
		frappe.has_permission("BOM", doc=bom_doc, throw=True)

		bom_items = frappe.get_all(
			"BOM Item",
			fields=["item_code", "bom_no as value", "stock_qty","uom","qty", "rate"],
			filters=[["parent", "=", frappe.form_dict.parent]],
			order_by="idx",
		)

		item_names = tuple(d.get("item_code") for d in bom_items)

		items = frappe.get_list(
			"Item",
			fields=["image", "description", "name", "stock_uom", "item_name", "is_sub_contracted_item"],
			filters=[["name", "in", item_names]],
		)  # to get only required item dicts

		for bom_item in bom_items:
			# extend bom_item dict with respective item dict
			bom_item.update(
				# returns an item dict from items list which matches with item_code
				next(item for item in items if item.get("name") == bom_item.get("item_code"))
			)

			bom_item.parent_bom_qty = bom_doc.quantity
			bom_item.expandable = 0 if bom_item.value in ("", None) else 1
			bom_item.image = frappe.db.escape(bom_item.image)

		return bom_items

def supplier_naming_before(doc, method):
	groupWV = ["Import- Bought-Out", "Import- Raw Material", "Local - Bought-Out", "Local - Raw Material", "Outsourcing"]
	groupWG = ["Manufacturer", "Electrical", "General", "Hardware", "Pharmaceutical", "Services", "Tooling & Fixtures"]
	groupFFW = [ "Transporter" ]
	
	if doc.supplier_group:
		
		if doc.supplier_group in groupWV:
			doc.naming_series = "WV.####"
		
		elif doc.supplier_group in groupWG:
			doc.naming_series = f"WG.####"  
			
		elif doc.supplier_group in groupFFW:
			doc.naming_series = f"FFW.####"  
			
		else:
			doc.naming_series = "WG.####"  
	else:
		doc.naming_series = "WG.####"  

   
	if not doc.supplier_code:
		doc.supplier_code = doc.naming_series
 

def supplier_naming_after(doc,method):
	
	if doc.supplier_code != doc.name:
		doc.supplier_code = doc.name  
		doc.save()  


@frappe.whitelist()
def get_next_supplier_code(prefix):
	
	prefix = prefix.upper()
	last_code = frappe.db.sql("""
		SELECT supplier_code
		FROM `tabSupplier`
		WHERE supplier_code LIKE %s
		ORDER BY supplier_code DESC
		LIMIT 1
	""", (prefix + "%",), as_dict=True)

	if last_code:
		last_code_str = last_code[0].supplier_code
		match = re.search(r'(\d+)$', last_code_str)
		if match:
			next_num = int(match.group(1)) + 1
		else:
			next_num = 1
	else:
		next_num = 1

	next_code = f"{prefix}{next_num:04d}"  

	return next_code  

@frappe.whitelist()
def change_supp_grp_user(name,old_grp, new_grp):
	
	usr_doc = frappe.db.get_value("User",{"username":name , "user_category":"Supplier", "supplier_group":old_grp},"name")
	
	
	if usr_doc:
		frappe.db.set_value("User",usr_doc,"supplier_group",new_grp)
		frappe.msgprint(f"Supplier group updated successfully for user {name}.")

  

@frappe.whitelist()
def get_calculated_height(doc,method):
	for i in doc.items:
		final_height=0
		pallet=i.custom_pallet
		box=i.custom_box
		if pallet and box :
			pal_doc=frappe.get_doc('Pallet',pallet)
			box_doc=frappe.get_doc('Box',box)
			pcount=i.custom_no_of_pallets
			bcount=i.custom_no_of_boxes
			if pcount > 0 and bcount > 0:
				# calculate pallet per box
				pox_per_pallet=bcount/pcount

				# calculate L*B for both pallet and box
				pallet_l_b=pal_doc.length*pal_doc.breadth
				box_l_b=box_doc.length*box_doc.breadth

				# divide L*B of pallet by box
				p_b_l_b=pallet_l_b/box_l_b
				p_b_l_b=int(p_b_l_b)

				# multiply p_b_l_b with box height
				v4=p_b_l_b*box_doc.height

				# add v4 with pallet height
				v5=v4+pal_doc.height

				final_height= v5+pal_doc.extra_height
		i.custom_calculated_height=final_height
	  
def mark_order_created_in_iom(doc, method):
	"""after insert of the Order, update Inter Office Memo
	that order was created"""
	
	frappe.db.set_value("Inter Office Memo", doc.custom_inter_office_memo, "order_created", 1)
	
def unmark_order_created_in_iom(doc, method):
	"""on deleting the Order, update Inter Office Memo
	that order was created"""
	
	frappe.db.set_value("Inter Office Memo", doc.custom_inter_office_memo, "order_created", 0)



@frappe.whitelist()
def update_customer_tax_category(doc, method):
	if doc.customer_group != "Export":
		if doc.territory == "India":
			country_address = frappe.db.get_value("Address",{'name':doc.customer_primary_address},'country')
			state_address = frappe.db.get_value("Address",{'name':doc.customer_primary_address},'state')
			if country_address == "India":
				if state_address == "Tamil Nadu":
					doc.tax_category = "In-State"
				else:
					doc.tax_category = "Out-State"
def test_check():
	data = frappe.db.sql("""
		SELECT parent AS sales_order, item_code, COUNT(*) AS duplicate_count
		FROM `tabSales Order Item`
		GROUP BY parent, item_code
		HAVING COUNT(*) > 1
	""", as_dict=True)

	sales_order = []
	for row in data:
		sales_order.append(row.sales_order)
	print(data)
	print(sales_order)

# Method to delete duplicate items from POI
def delete_duplicate():
	data = frappe.db.sql("""
		DELETE poi1
        FROM `tabSales Order Item` poi1
        INNER JOIN `tabSales Order Item` poi2
            ON poi1.parent = poi2.parent
            AND poi1.item_code = poi2.item_code
            AND poi1.creation < poi2.creation
		""")
	print(data)

# Method to correct the OT Balance
def correct_ot_balance():
	from frappe.utils import get_first_day, get_last_day, nowdate
	current_month_start_date = get_first_day(nowdate())
	current_month_end_date = get_last_day(nowdate())
	ot_balances = frappe.db.get_all("OT Balance", {"from_date": current_month_start_date, "to_date": current_month_end_date}, pluck="name")
	for ot_balance in ot_balances:
		print(ot_balance)
		ot_balance_doc = frappe.get_doc("OT Balance", ot_balance)
		total_ot_hours = frappe.db.sql("""
						SELECT SUM(custom_overtime_hours) as total_ot_hours
						FROM `tabAttendance`
						WHERE employee = %s AND 
							attendance_date BETWEEN %s AND %s
						""", (ot_balance_doc.employee, current_month_start_date, current_month_end_date), as_dict=1)[0]['total_ot_hours'] or 0
		
		coff_pending = frappe.db.sql("""
						SELECT SUM(custom_total_leave_days) as total_leave_days
						FROM `tabLeave Application`
						WHERE employee = %s AND custom_select_leave_type = "Comp-off from OT" AND 
							posting_date BETWEEN %s AND %s AND
							docstatus = 0 AND workflow_state not in ("Cancelled", "Rejected", "Approved")
						""", (ot_balance_doc.employee, current_month_start_date, current_month_end_date), as_dict=1)[0]['total_leave_days'] or 0
		
		coff_approved = frappe.db.sql("""
						SELECT SUM(custom_total_leave_days) as total_leave_days
						FROM `tabLeave Application`
						WHERE employee = %s AND custom_select_leave_type = "Comp-off from OT" AND
							posting_date BETWEEN %s AND %s AND
							docstatus = 1 AND workflow_state = "Approved"
						""", (ot_balance_doc.employee, current_month_start_date, current_month_end_date), as_dict=1)[0]['total_leave_days'] or 0
		
		ot_balance = total_ot_hours - ((float(coff_approved) * 8) + (float(coff_pending) * 8))
		
		ot_balance_doc.total_ot_hours = total_ot_hours
		ot_balance_doc.comp_off_pending_for_approval = coff_pending
		ot_balance_doc.comp_off_used = coff_approved
		ot_balance_doc.ot_balance = ot_balance
		ot_balance_doc.save()

def carry_forward_compoff():
	employees = frappe.db.get_all("Employee", {"employee_category": ("in", ("Staff", "Sub Staff")), "status": "Active"}, pluck="name")
	for employee in employees:

		allocation = frappe.db.sql("""
			SELECT total_leaves_allocated
			FROM `tabLeave Allocation`
			WHERE leave_type = 'Compensatory Off'
				AND from_date BETWEEN '2025-01-01' AND '2025-12-31'
				AND to_date BETWEEN '2025-01-01' AND '2025-12-31'
				AND employee = %s
				AND docstatus = 1
		""", (employee,), as_dict=True)

		leaves_allocated = allocation[0]['total_leaves_allocated'] if allocation else 0

		leaves_used = frappe.db.sql("""
			SELECT SUM(total_leave_days) AS leaves_used
			FROM `tabLeave Application`
			WHERE from_date BETWEEN '2025-01-01' AND '2025-12-31'
				AND to_date BETWEEN '2025-01-01' AND '2025-12-31'
				AND docstatus = 1
				AND workflow_state = 'Approved'
				AND employee = %s
				AND leave_type = "Compensatory Off"
		""", (employee,), as_dict=True)[0]['leaves_used'] or 0
  
		if leaves_allocated - leaves_used > 0:
			employee_category = frappe.db.get_value("Employee", employee, "employee_category")
			la = frappe.new_doc("Leave Allocation")
			la.leave_type = "Compensatory Off"
			la.from_date = "2026-01-01"
			la.to_date = "2026-12-31"
			la.custom_employee_category = employee_category
			la.employee = employee
			la.new_leaves_allocated = leaves_allocated - leaves_used
			la.total_leaves_allocated = leaves_allocated - leaves_used
			la.save()

@frappe.whitelist()
def update_pallet_height(name):
	so = frappe.get_doc("Sales Invoice", name)

	for row in so.items:
		height = update_pallet_h(name, row.item_code)
		row.db_set('custom_calculated_height', height, update_modified=False)

	so.save(ignore_permissions=True)
	so.reload()

	return "OK"


@frappe.whitelist()
def update_pallet_h(name,item):
	doc=frappe.get_doc("Sales Invoice",name)
	for i in doc.items:
		if i.item_code==item:
			final_height=0
			pallet=i.custom_pallet
			box=i.custom_box
			if pallet and box :
				pal_doc=frappe.get_doc('Pallet',pallet)
				box_doc=frappe.get_doc('Box',box)
				pcount=i.custom_no_of_pallets
				bcount=i.custom_no_of_boxes
				if pcount > 0 and bcount > 0:
					# calculate pallet per box
					pox_per_pallet=bcount/pcount

					# calculate L*B for both pallet and box
					pallet_l_b=pal_doc.length*pal_doc.breadth
					box_l_b=box_doc.length*box_doc.breadth

					# divide L*B of pallet by box
					p_b_l_b=pallet_l_b/box_l_b
					# p_b_l_b=p_b_l_b/bcount
					p_b_l_b=int(p_b_l_b)

					# multiply p_b_l_b with box height
					v4=p_b_l_b*box_doc.height

					# add v4 with pallet height
					v5=v4+pal_doc.height

					final_height= v5+pal_doc.extra_height
			i.custom_calculated_height=final_height
	return final_height


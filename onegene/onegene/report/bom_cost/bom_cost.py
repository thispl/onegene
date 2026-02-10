# Copyright (c) 2013, Frappe Technologies Pvt. Ltd. and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.utils import flt, now_datetime
from openpyxl import Workbook
from frappe.utils.file_manager import save_file
import json

ITEM_CACHE = {}
PRICE_CACHE = {}
SALES_RATE_CACHE = {}

def execute(filters=None):
    preload_master_data()
    data = []
    columns = get_columns()
    get_data(data, filters)
    # Apply formatting AFTER calculations
    for row in data:
        row["last_purchase_rate"] = format_value(row.get("last_purchase_rate"), 2, "Currency")
        row["last_sales_rate"] = format_value(row.get("last_sales_rate"), 2, "Currency")
        row["item_rate"] = format_value(row.get("item_rate"), 2, "Currency")
        row["process_cost"] = format_value(row.get("process_cost"), 2, "Currency")
        row["rm_cost"] = format_value(row.get("rm_cost"), 2, "Currency")
        row["percentage"] = format_value(row.get("percentage"), 2, "Percentage")
        row["qty"] = format(row.get("qty"), f".6f")

        if row["indent"] == 0:
            row["item_code"] = f"<b>{row['item_code']}</b>"
            row["item_name"] = f"<b>{row['item_name']}</b>"
            row["item_type"] = f"<b>{row['item_type']}</b>"
            row["qty"] = f"<b>{row['qty']}</b>"
            row["uom"] = f"<b>{row['uom']}</b>"
            row["last_sales_rate"] = f"<b>{row['last_sales_rate']}</b>"
            row["rm_cost"] = f"<b>{row['rm_cost']}</b>"
            row["percentage"] = f"<b>{row['percentage']}</b>"
    return columns, data

def get_columns():
    return [
        {"label": _("Item Code"), "fieldtype": "Data", "fieldname": "item_code", "width": 300},
        {"label": _("Item Name"), "fieldtype": "data", "fieldname": "item_name", "width": 400},
        {"label": _("Item Type"), "fieldtype": "data", "fieldname": "item_type", "width": 180},
        {"label": _("Qty"), "fieldtype": "Data", "fieldname": "qty", "width": 100, "align": "right"},
        {"label": _("UOM"), "fieldtype": "data", "fieldname": "uom", "width": 100},
        {"label": _("Last Purchase Rate"), "fieldtype": "Data", "fieldname": "last_purchase_rate", "width": 160, "align": "right"},
        {"label": _("Last Sales Rate"), "fieldtype": "Data", "fieldname": "last_sales_rate", "width": 150, "align": "right"},
        # {"label": _("Item Rate"), "fieldtype": "Data", "fieldname": "item_rate", "width": 150, "align": "right"},
        {"label": _("RM Cost"), "fieldtype": "Data", "fieldname": "rm_cost", "width": 150, "align": "right"},
        {"label": _("Process Cost"), "fieldtype": "Data", "fieldname": "process_cost", "width": 150, "align": "right"},
        {"label": _("Sales Rate %"), "fieldtype": "Data", "fieldname": "percentage", "width": 150, "align": "right"},
    ]

def get_data(data, filters=None):
    bom_list = []
    if not filters:
        bom_list = frappe.db.get_all("BOM",{"custom_item_billing_type": "Billing", "is_active": 1, "is_default": 1, "docstatus": 1},pluck="name")
    else:
        bom_filter = filters.get('bom')
        # if not frappe.db.exists("BOM", {"name": bom_filter, "custom_item_billing_type": "Billing"}):
        # 	frappe.publish_realtime("report_error")
        # 	frappe.throw(f"The selected BOM <b>{bom_filter}</b> is not a billing item", title="Invalid BOM")
        bom_list.append(bom_filter)

    if bom_list:
        for bom_filter in bom_list:
            item = frappe.db.get_value("BOM", bom_filter,
                ["item", "item_name", "uom", "custom_item_billing_type"], as_dict=1)

            item_type = frappe.db.get_value("Item", item.item, "item_type")
            last_sales_rate = 0
            if item.custom_item_billing_type == "Billing":
                last_sales_rate = get_last_sales_rate(item.item)
            warehouse = frappe.db.get_value("Item", item.item, "custom_warehouse")
            data.append({
                "item_code": item.item,
                "item_name": item.item_name,
                "item_type": item_type,
                "indent": 0,
                "bom_level": 0,
                "bom": bom_filter,
                "qty": 1,
                "uom": item.uom,
                "last_sales_rate": last_sales_rate,
                "warehouse": warehouse
            })
            get_exploded_items(bom=bom_filter, data=data)
        calculate_totals(data)

def calculate_totals(data):
    for i, row in enumerate(data):
        current_indent = row.get("indent")
        total = 0

        item_type = row.get("item_type")
        warehouse = row.get("warehouse")
  
        # only_direct_child = (
        # 	item_type in ("Process Item")
        # 	and warehouse in ("Semi")
        # 	and current_indent != 0
        # )

        for j in range(i + 1, len(data)):
            next_row = data[j]

            # Stop when we move out of this parent tree
            if next_row.get("indent") <= current_indent:
                break

            # If only direct children are allowed, skip deeper levels
            # if only_direct_child and next_row.get("indent") != current_indent + 1:
            # 	continue

            item_rate = flt(next_row.get("item_rate") or 0)
            process_cost = flt(next_row.get("process_cost") or 0)
            total += item_rate + process_cost

        row["rm_cost"] = total if total > 0 else ""

        # For Purchase / Consumables, RM cost is just item_rate
        if row["item_type"] in ["Purchase Item", "Consumables"]:
            row["rm_cost"] = row.get("item_rate")

    # Sales percentage only on top row
    if data and data[0].get("last_sales_rate") and data[0].get("rm_cost"):
        data[0]["percentage"] = flt(data[0]["rm_cost"] * 100 / flt(data[0]["last_sales_rate"]))

  
# Build child structure
def get_exploded_items(bom, data, indent=0, qty=1):
    exploded_items = frappe.get_all(
        "BOM Item",
        filters={"parent": bom},
        fields=["qty", "bom_no", "item_code", "item_name", "uom"],
        order_by="idx"
    )

    for item in exploded_items:
        last_purchase_rate = get_last_purchase_rate(item.item_code) if not item.bom_no else 0
        warehouse = ITEM_CACHE.get(item.item_code, {}).get("custom_warehouse")
        item_type = ITEM_CACHE.get(item.item_code, {}).get("item_type")
        item_rate = flt(last_purchase_rate) * flt(item.qty) * flt(qty)
        process_cost = flt(get_process_cost(item.item_code, item.bom_no)) * flt(item.qty) * flt(qty)
        data.append({
            "item_code": item.item_code,
            "item_name": item.item_name,
            "item_type": item_type,
            "indent": indent + 1,
            "bom_level": indent + 1,
            "bom": item.bom_no,
            "qty": flt(item.qty) * flt(qty),
            "uom": item.uom,
            "last_purchase_rate": last_purchase_rate,
            "item_rate": item_rate,
            "process_cost": process_cost,
            "warehouse": warehouse,
            "parent": bom
        })

        if item.bom_no:
            get_exploded_items(bom=item.bom_no, data=data, indent=indent + 1, qty=flt(item.qty))

def get_last_purchase_rate(item_code):
    return ITEM_CACHE.get(item_code, {}).get("last_purchase_rate") or 0

def get_last_sales_rate(item_code):
    rate = frappe.db.sql("""
        SELECT soi.base_rate
        FROM `tabSales Order` so
        INNER JOIN `tabSales Order Item` soi ON soi.parent = so.name
        WHERE so.docstatus = 1 AND soi.item_code = %s
        ORDER BY so.transaction_date DESC
    """, (item_code,), as_dict=1)
    return rate[0].base_rate if rate else 0

def get_process_cost(item_code, bom=None):
    process_cost = 0
    # if bom:
    # 	operating_cost = frappe.db.sql("""
    # 		SELECT poi.base_rate
    # 		FROM `tabPurchase Order` po
    # 		INNER JOIN `tabPurchase Order Item` poi ON poi.parent = po.name
    # 		WHERE po.docstatus = 1 AND poi.item_code = %s AND custom_po_type = 'Job Order'
    # 		ORDER BY po.transaction_date DESC
    # 	""", (item_code,), as_dict=1)

    # 	operating_cost = operating_cost[0].base_rate if operating_cost else 0

    # 	if operating_cost == 0:
    # 		operating_cost = frappe.db.get_value("BOM", bom, "operating_cost") or 0

    # 	scrap_material_cost = frappe.db.get_value("BOM", bom, "scrap_material_cost") or 0
    # 	process_cost = operating_cost + scrap_material_cost

    return process_cost

def preload_master_data():
    global ITEM_CACHE, PRICE_CACHE

    items = frappe.db.get_all("Item", fields=["name", "item_type", "custom_warehouse", "last_purchase_rate"])
    ITEM_CACHE = {i.name: i for i in items}

    prices = frappe.db.get_all(
        "Item",
        fields=["name", "last_purchase_rate"]
    )
    PRICE_CACHE = {p.name: p.last_purchase_rate for p in prices}
 
def preload_sales_rates():
    global SALES_RATE_CACHE

    rates = frappe.db.sql("""
        SELECT t.item_code, t.base_rate
        FROM (
            SELECT soi.item_code, soi.base_rate, so.transaction_date
            FROM `tabSales Order` so
            INNER JOIN `tabSales Order Item` soi ON soi.parent = so.name
            WHERE so.docstatus = 1
            ORDER BY so.transaction_date DESC
        ) t
        GROUP BY t.item_code
    """, as_dict=1)

    SALES_RATE_CACHE = {r.item_code: r.base_rate for r in rates}
 
def format_value(value, precision, unit):
    from decimal import Decimal, ROUND_HALF_UP
    if not value or float(value) == 0:
        return ""
    rounded = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    result = format(rounded, f".{precision}f")
    if unit == "Currency":
        # result = "₹ " + str(result)
        result = str(result)
    if unit == "Percentage":
        result = str(result)
    if unit == "FLoat":
        result = float(result)
    return result

def strip_html(value):
    return value.replace("<b>", "").replace("</b>", "") if isinstance(value, str) else value

# Excel Report
@frappe.whitelist()
def get_excel_report(filters=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from frappe.utils import cint
 
    bold_font = Font(bold=True)
    sfg_fill = PatternFill(start_color="ffe699", end_color="ffe699", fill_type="solid")
    fg_fill = PatternFill(start_color="4a7085", end_color="4a7085", fill_type="solid")
    fg_font  = Font(bold=True, color="ffffff")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    
    if filters and isinstance(filters, str):
        filters = json.loads(filters)

    columns, data = execute(filters)
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Cost"
    ws.append(["BOM LIST"])
    ws.append(["Detail"])
    headers = [
        "S#", "Item Code", "Item Name", "Item Type", "Qty", "UOM", "Last Purchase Rate", "Last Sales Rate", "RM Cost", "Process Cost", "Sales Rate %"
    ]
    ws.append(headers)
    idx = 0
    for row in data:
        if strip_html(row.get('indent')) == 0:
            idx += 1
            serial_no = idx
        else:
            serial_no = ""
        values = [
            serial_no,
            (row.get('indent') * "   ") + strip_html(row.get("item_code")),
            strip_html(row.get("item_name")),
            strip_html(row.get("item_type")),
            flt(strip_html(row.get("qty"))),
            strip_html(row.get("uom")),
            flt(strip_html(row.get("last_purchase_rate"))),
            flt(strip_html(row.get("last_sales_rate"))),
            flt(strip_html(row.get("rm_cost"))),
            flt(strip_html(row.get("process_cost"))),
            flt(strip_html(row.get("percentage"))),
        ]
        ws.append(values)
        if strip_html(row.get("item_type")) == "Process Item":
            last_row = ws.max_row
            for col in range(1, len(values) + 1):
                cell = ws.cell(row=last_row, column=col)
                cell.font = bold_font
                if strip_html(row.get('indent')) == 0:
                    cell.fill = fg_fill
                    cell.font = fg_font
     
    # styles
    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column):
        for cell in row:
            cell.border = border

    for row in ws["A1:K3"]:
        for cell in row:
            cell.fill = fg_fill
            cell.font = Font(bold=True, color="ffffff")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells("A1:K1")
    ws["A1"] = "BOM LIST"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(bold=True, size=15, color="ffffff")
    
    ws.merge_cells("A2:K2")
    ws["A2"] = "Detail"
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].font = Font(bold=True, size=12, color="ffffff")
 
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["K"].width = 18

    file_name = f"bom_cost.xlsx"

    from io import BytesIO
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)

    file_doc = save_file(
        file_name,
        xlsx_file.read(),
        dt="Report",
        dn="BOM Cost",
        is_private=0
    )

    return file_doc.file_url

@frappe.whitelist()
def download_report_json(filters=None):
    if filters and isinstance(filters, str):
        filters = json.loads(filters)

    columns, data = execute(filters)

    return data

@frappe.whitelist()
def update_cost(data):
    if isinstance(data, str):
        data = json.loads(data)

    # ---------------- GROUP BY ITEM ----------------
    item_map = {}

    for row in data:
        item_code = strip_html(row.get("item_code"))
        item_type = strip_html(row.get("item_type"))
        parent = strip_html(row.get("parent"))

        qty = flt(strip_html(row.get("qty"))) or 0
        rate = flt(strip_html(row.get("last_purchase_rate"))) / qty if qty > 0 else flt(strip_html(row.get("rm_cost")))
        rm_cost = (
            flt(strip_html(row.get("rm_cost"))) / qty
            if qty > 0 else flt(strip_html(row.get("rm_cost")))
        )

        if item_code not in item_map:
            item_map[item_code] = {
                "total_qty": 0,
                "total_amount": 0,
                "item_type": item_type,
                "parent": parent,
                "rm_cost": rm_cost
            }

        item_map[item_code]["total_qty"] += qty
        item_map[item_code]["total_amount"] += qty * rate

    # ---------------- PROCESS EACH ITEM ONCE ----------------
    for item_code, d in item_map.items():

        total_qty = d["total_qty"] or 1
        avg_rate = d["total_amount"] / total_qty
        item_type = d["item_type"]
        parent = d["parent"]
        rm_cost = d["rm_cost"]

        # -------- RM COST UPDATE (once) --------
        if item_type == "Process Item" and flt(rm_cost) > 0:
            old_rm_cost = frappe.db.get_value("Item", item_code, "custom_rm_cost")
            precision = frappe.get_precision("Item", "custom_rm_cost")

            if round(flt(old_rm_cost), precision) != round(flt(rm_cost), precision):
                rm_rvision_len = frappe.db.count("Item RM Cost Revision", {"parent": item_code})
                frappe.db.set_value("Item", item_code, "custom_rm_cost", rm_cost)

                frappe.get_doc({
                    "doctype": "Item RM Cost Revision",
                    "parent": item_code,
                    "parenttype": "Item",
                    "parentfield": "custom_rm_cost_revisions",
                    "idx": rm_rvision_len + 1,
                    "revised_on": now_datetime(),
                    "revised_by": frappe.session.user,
                    "rm_cost": old_rm_cost,
                    "revised_rm_cost": rm_cost
                }).insert(ignore_permissions=True)

        # -------- BOM RATE UPDATE (once) --------
        bom_item = frappe.db.get_value(
            "BOM Item",
            {"parent": parent, "item_code": item_code},
            ["rate", "qty"],
            as_dict=True
        )

        if not bom_item:
            continue

        current_rate = flt(bom_item.rate)
        bom_qty = flt(bom_item.qty) or 1

        new_rate = rm_cost if item_type == "Process Item" else avg_rate
        new_amount = bom_qty * new_rate

        rate_precision = frappe.get_precision("BOM Item", "rate")

        if round(current_rate, rate_precision) != round(new_rate, rate_precision):
            frappe.db.set_value(
                "BOM Item",
                {"parent": parent, "item_code": item_code},
                {
                    "base_rate": new_rate,
                    "rate": new_rate,
                    "amount": new_amount,
                    "base_amount": new_amount,
                }
            )

@frappe.whitelist() 
def update_bom_cost_scheduler():
    data = download_report_json(filters=None)
    if not data:
        return
    update_cost(data)
    frappe.log_error(title="BOM Cost Scheduler", message=f"Auto Update Cost executed rows: {len(data)}")

@frappe.whitelist()
def update_cost_from_bom(filters):
    if isinstance(filters, str):
        filters = json.loads(filters)

    columns, data = execute(filters)
  
    for row in data:
        item_code = strip_html(row.get("item_code"))
        item_type = strip_html(row.get("item_type"))
        rm_cost = flt(strip_html(row.get("rm_cost"))) / flt(strip_html(row.get("qty")), 9) if flt(strip_html(row.get("qty")), 9) > 0 else flt(strip_html(row.get("rm_cost")))
        parent = strip_html(row.get("parent"))
        rate = flt(strip_html(row.get("last_purchase_rate")))
        amount = flt(strip_html(row.get("qty"))) * flt(strip_html(row.get("last_purchase_rate")))
  
        # Update RM Cost in Item
        if item_type == "Process Item" and flt(rm_cost) > 0:
            old_rm_cost = frappe.db.get_value("Item", item_code, "custom_rm_cost")
            if flt(old_rm_cost) != flt(rm_cost):
                rm_rvision_len = frappe.db.count("Item RM Cost Revision", {"parent": item_code})
                frappe.db.set_value("Item", item_code, "custom_rm_cost", rm_cost)
                
                # Revision Logs
                frappe.get_doc({
                    "doctype": "Item RM Cost Revision",
                    "parent": item_code,
                    "parenttype": "Item",
                    "parentfield": "custom_rm_cost_revisions",

                    "idx": rm_rvision_len + 1,
                    "revised_on": now_datetime(),
                    "revised_by": frappe.session.user,
                    "rm_cost": old_rm_cost,
                    "revised_rm_cost": rm_cost
                }).insert(ignore_permissions=True)
    
        # Update Rate in BOM
        if item_type != "Process Item":
            frappe.db.set_value("BOM Item", {"parent": parent, "item_code": item_code}, 
                        {"base_rate": rate, "rate": rate, 
                        "amount": amount, "base_amount": amount})
        else:
            qty = frappe.db.get_value("BOM Item", {"parent": parent, "item_code": item_code}, "qty") or 1
            frappe.db.set_value("BOM Item", {"parent": parent, "item_code": item_code}, 
                        {"base_rate": rm_cost, "rate": rm_cost, 
                        "amount": qty * rm_cost, "base_amount": qty * rm_cost})
   
def test_check():
    doc = frappe.new_doc("Scheduled Job Type")
    doc.method = "onegene.onegene.report.bom_cost.bom_cost.update_bom_cost_scheduler"
    doc.frequency = "Cron"
    doc.cron_format = "0 */2 * * *"
    doc.insert(ignore_permissions=True)
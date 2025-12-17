# Copyright (c) 2023, TEAMPRO and contributors
# For license information, please see license.txt

import calendar
from datetime import timedelta
import csv
from datetime import datetime
import math, json
import re
from frappe import _dict
import frappe
from frappe import _
from frappe.utils import flt
from frappe import _dict
from frappe.utils import strip_html
from frappe.utils import (
	add_days,
	ceil,
	cint,
	comma_and,
	flt,
	get_link_to_form,
	getdate,
	datetime,
	now_datetime,
	nowdate,
	today,
	cstr
)
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
import erpnext
import datetime
from six import BytesIO
import openpyxl
from openpyxl import Workbook
from frappe.utils.csvutils import UnicodeWriter,read_csv_content
from frappe.utils.file_manager import get_file
import openpyxl
from openpyxl import Workbook
from frappe.utils.csvutils import UnicodeWriter,read_csv_content
from frappe.utils.file_manager import get_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
import frappe, io
import os
from frappe.utils.file_manager import get_file_path
import base64



def execute(filters=None):
	columns = get_columns(filters)
	data = get_data(filters)
	planned_data = plan_data(data,filters)
	return columns, planned_data

def get_data(filters):
	data = []
	year = datetime.date.today().year
	month_str = filters.get("month")
	month_map = {
		'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
		'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
	}
	month = month_map.get(month_str)
	start_date = datetime.date(year, month, 1)
	last_day = calendar.monthrange(year, month)[1]
	to_date_obj =  datetime.date(year, month, last_day)
	to_date =str(to_date_obj)

	query = """
		SELECT 
			item_code , item_name , item_group , SUM(qty) as qty
		FROM 
			`tabSales Order Schedule` 
		WHERE 
			docstatus = 1 AND schedule_date BETWEEN %(from_date)s AND %(to_date)s
			{customer_condition}
			{item_group_condition}
			{item_code_condition}
		GROUP BY
			item_code
	"""

	conditions = {
		"from_date": start_date,
		"to_date": to_date,
	}

	customer_condition = ""
	item_group_condition = ""
	item_code_condition = ""

	if filters.get("customer"):
		customer_condition = "AND customer_name = %(customer)s"
		conditions["customer"] = filters["customer"]
	if filters.get("item_group"):
		item_group_condition = "AND item_group = %(item_group)s"
		conditions["item_group"] = filters["item_group"]

	if filters.get("item_code"):
		item_code_condition = "AND item_code = %(item_code)s"
		conditions["item_code"] = filters["item_code"]
	
	query = query.format(
		customer_condition=customer_condition,
		item_group_condition=item_group_condition,
		item_code_condition=item_code_condition,

	)
	
	sales_order_schedule = frappe.db.sql(query, conditions, as_dict=True)
	for row in sales_order_schedule:
		
		rej_allowance = frappe.get_value("Item",row.item_code,['rejection_allowance'])
		pack_size = frappe.get_value("Item",row.item_code,['pack_size'])
		work_days = frappe.db.get_single_value("Production Plan Settings", "working_days")
		with_rej = (row.qty * (rej_allowance/100)) + row.qty
		per_day = with_rej / int(work_days)
		if pack_size > 0:
			cal = per_day/ pack_size
			total = ceil(cal) * pack_size
		else:
			total = ceil(per_day)
		data.append({
			"item_code": row.item_code,
			"item_name": row.item_name,
			"item_group": row.item_group,
			"with_rej": with_rej,
			"pack_size": pack_size,
			"total": total
		})

	return data


def get_columns(filters):
	
	year = datetime.date.today().year
	month_str = filters.get("month")
	month_map = {
		'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
		'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
	}
	month_lower=month_str.lower()
	month = month_map.get(month_str)
	start_date = datetime.date(year, month, 1)
	last_day = calendar.monthrange(year, month)[1]
 
	if filters.view_rm==1:
	 
		cols=[
			{"label": _("Item Code"),"fieldtype": "Data","fieldname": "item_code","width": 200,"options": "Item"},
			{"label": _("Item Name"), "fieldtype": "Data", "fieldname": "item_name", "width": 200},
			{"label": _("Item Group"), "fieldtype": "Data", "fieldname": "item_group", "width": 150},
			{"label": _("Item Type"), "fieldtype": "Data", "fieldname": "item_type", "width": 150},
			{"label": _("Monthly Schedule"), "fieldtype": "Data", "fieldname": "with_rej", "width": 150},
			{"label": _("Bin Qty"), "fieldtype": "Data", "fieldname": "pack_size", "width": 100},
			{"label": _("Per Day Plan"), "fieldtype": "Data", "fieldname": "total", "width": 150,},
		]
  
		for day in range(1, last_day + 1):
			current_date = datetime.date(year, month, day).strftime('%d')
			cols.extend([
						{
							"label": _(f"{month_str}-{current_date} Plan"),
							"fieldtype": "HTML",  
							"fieldname": f"{month_lower}_{current_date}_plan",
							"width": 120
						},
						{
							"label": _(f"{month_str}-{current_date} Actual"),
							"fieldtype": "HTML",  
							"fieldname": f"{month_lower}_{current_date}_actual",
							"width": 120
						},
						
	  
	  
					])
   
		cols.extend([
			
			{
						"label":_("Total Plan"),
						"fieldtype": "HTML",  
						"fieldname": "total_plan_sum",
						"width": 120
					},
					{
						"label":_("Total Actual"),
						"fieldtype": "HTML",  
						"fieldname": "total_actual_sum",
						"width": 120
					}
			
		])
	
		return cols
   
   
   
	else:
	 
		cols=[
		
			{"label": _("Item Code"),"fieldtype": "Data","fieldname": "item_code","width": 200,"options": "Item"},
			{"label": _("Item Name"), "fieldtype": "Data", "fieldname": "item_name", "width": 200},
			{"label": _("Item Group"), "fieldtype": "Data", "fieldname": "item_group", "width": 150},
			{"label": _("Monthly Schedule"), "fieldtype": "Data", "fieldname": "with_rej", "width": 150},
			{"label": _("Bin Qty"), "fieldtype": "Data", "fieldname": "pack_size", "width": 100},
			{"label": _("Per Day Plan"), "fieldtype": "Data", "fieldname": "total", "width": 150,},
			
   
		]
  
		for day in range(1, last_day + 1):
			current_date = datetime.date(year, month, day).strftime('%d')

			cols.extend([
						{
							"label": _(f"{month_str}-{current_date} Plan"),
							"fieldtype": "HTML",  
							"fieldname": f"{month_lower}_{current_date}_plan",
							"width": 120
						},
						{
							"label": _(f"{month_str}-{current_date} Actual"),
							"fieldtype": "HTML",  
							"fieldname": f"{month_lower}_{current_date}_actual",
							"width": 120
						},
	  
	  
					])
   
		cols.extend([
			
			{
						"label":_("Total Plan"),
						"fieldtype": "HTML",  
						"fieldname": "total_plan_sum",
						"width": 120
					},
					{
						"label":_("Total Actual"),
						"fieldtype": "HTML",  
						"fieldname": "total_actual_sum",
						"width": 120
					}
			
		])

		return cols

def plan_data(data, filters):
	from collections import defaultdict
	from datetime import date, timedelta

	year = date.today().year
	month_str = filters.get("month")
	month_lower = month_str.lower()
	month_map = {
		'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
		'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
	}
	month = month_map.get(month_str)
	start_date = date(year, month, 1)
	last_day = calendar.monthrange(year, month)[1]
	to_date = date(year, month, last_day)

	# --- 1. Bulk fetch Plan Qty ---
	plan_qty_rows = frappe.get_all(
		"Daily Production Plan Item",
		filters={"date": ["between", [start_date, to_date]]},
		fields=["item_code", "date", "plan"],
		as_list=True
	)
	plan_qty_map = {(item_code, d): plan or 0 for item_code, d, plan in plan_qty_rows}

	# --- 2. Bulk fetch Quality Inspection ---
	qi_rows = frappe.db.sql("""
		SELECT item_code, report_date, custom_shift, custom_shift_time,
			   SUM(custom_accepted_qty) AS qty
		FROM `tabQuality Inspection`
		WHERE docstatus = 1
		  AND workflow_state != 'Rejected'
		  AND custom_inspection_type = 'In Process'
		  AND report_date BETWEEN %s AND %s
		GROUP BY item_code, report_date, custom_shift, custom_shift_time
	""", (start_date, to_date), as_dict=True)

	qi_map = defaultdict(list)
	for r in qi_rows:
		key = (r.item_code, r.report_date, r.custom_shift)
		qi_map[key].append(r)

	# --- 3. Get shift times once ---
	check_out_end_time = frappe.db.get_value("Shift Type", '3', "custom_checkout_end_time")
	check_in_start_time = frappe.db.get_value("Shift Type", '3', "custom_checkin_start_time")

	# --- 4. Build planned data ---
	planned_data = []
	total_plan = defaultdict(float)
	total_actual = defaultdict(float)

	for row in data:
		row_data = {
			'item_code': f"<span style='text-align: left:'>{row['item_code']}</span>",
			'item_name': row['item_name'],
			'item_group': row['item_group'],
			'with_rej': int(row['with_rej']),
			'pack_size': row['pack_size'],
			'total': row['total']
		}
  
		total_plan_val = 0
		total_actual_val = 0

		current_date = start_date
		while current_date <= to_date:
	  
			
			# --- Plan Qty ---
			plan_qty = plan_qty_map.get((row['item_code'], current_date), 0)

			# --- Actual Qty ---
			actual_qty = 0
			for shift in ["1", "2", "3"]:
				shift_data = qi_map.get((row['item_code'], current_date, shift), [])
				for rec in shift_data:
					if shift in ["1", "2"]:
						actual_qty += rec.qty
					elif shift == "3":
						if rec.custom_shift_time >= 24:
							actual_qty += rec.qty
						else:
							# Shift 3 that spills to next day
							next_day = current_date + timedelta(days=1)
							shift_next = qi_map.get((row['item_code'], next_day, '3'), [])
							for r2 in shift_next:
								if r2.custom_shift_time <= check_out_end_time:
									actual_qty += r2.qty

			field_plan = f"{month_lower}_{current_date.strftime('%d')}_plan"
			field_actual = f"{month_lower}_{current_date.strftime('%d')}_actual"

			row_data[field_plan] = f"<div style='background-color:#D3D3D3; width:110px; height:24px; border-radius:4px;  padding:4px; font-weight:bold; text-align:right; '>{int(plan_qty) if plan_qty > 0 else ''}</div>"
			row_data[field_actual] = f"<div style='background-color:#FFD580; width:110px; height:24px; border-radius:4px;  padding:4px; font-weight:bold; text-align:right; '>{int(actual_qty) if actual_qty > 0 else ''}</div>"
			total_plan_val += plan_qty
			total_actual_val += actual_qty

			total_plan[field_plan] += plan_qty
			total_actual[field_actual] += actual_qty

			current_date += timedelta(days=1)
		row_data["total_plan_sum"] = f"<div style='font-weight:bold; text-align:right;'>{int(total_plan_val) if total_plan_val else ''}</div>"
		row_data["total_actual_sum"] = f"<div style='font-weight:bold; text-align:right;'>{int(total_actual_val) if total_actual_val else ''}</div>"
		planned_data.append(row_data)
	# ---- ADD TOTAL ROW ----
	total_row = {
		"item_code": "<b>Total</b>",
		"item_name": "",
		"item_group": "",
		"with_rej": "",
		"pack_size": "",
		"total": "",
	}
   
	total_plan_val_2 = 0
	total_actual_val_2 = 0

	for day in range(1, last_day + 1):
		plan_field = f"{month_lower}_{str(day).zfill(2)}_plan"
		actual_field = f"{month_lower}_{str(day).zfill(2)}_actual"
		total_row[plan_field] = f"<div style='font-weight:bold; text-align:right; '>{int(total_plan[plan_field]) if total_plan[plan_field] else ''}</div>"
  
		total_plan_val_2 += total_plan[plan_field]
		total_row[actual_field] = f"<div style='font-weight:bold; text-align:right;'>{int(total_actual[actual_field]) if total_actual[actual_field] else ''}</div>"
		total_actual_val_2 += total_actual[actual_field]
	total_row["total_plan_sum"] = f"<div style='font-weight:bold; text-align:right;'>{int(total_plan_val_2) if total_plan_val_2 else ''}</div>"
	total_row["total_actual_sum"] = f"<div style='font-weight:bold; text-align:right;'>{int(total_actual_val_2) if total_actual_val_2 else ''}</div>"
		
	
	planned_data.append(total_row)
	return planned_data


import openpyxl
from openpyxl.styles import Alignment
@frappe.whitelist(allow_guest=False)
def download_kqty_template():
	import json, re, base64, io, datetime, calendar
	from openpyxl import Workbook

	data = frappe.form_dict.get('data')
	month_str = frappe.form_dict.get('month')
	
	if isinstance(data, str):
		data = json.loads(data)

	wb = Workbook()
	ws = wb.active
	ws.title = "Production Kanban Qty"
	ws.freeze_panes = 'C2'

	month_map = {
		'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
		'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
	}
	month = month_map.get(month_str)

	today = datetime.date.today()
	last_day = calendar.monthrange(today.year, month)[1]
	
	if month == today.month:
		start_day = today.day
	else:
		start_day = 1
	if start_day > last_day:
		start_day = last_day

	header = ['Item Code', 'Item Name']
	day_headers = [f"{month_str}-{str(day).zfill(2)}" for day in range(start_day, last_day + 1)]
	header.extend(day_headers)

	ws.append(header)
	ws.column_dimensions['A'].width = 20  
	ws.column_dimensions['B'].width = 40  

	for idx, day_header in enumerate(day_headers, start=3):  
		cell = ws.cell(row=1, column=idx)
		cell.value = str(day_header)
		cell.number_format = '@'

	# 🚫 Skip the last row (Total row)
	data_to_export = data[:-1] if data and isinstance(data, list) and len(data) > 1 else data

	for row in data_to_export:
		item_code = re.sub(r'<.*?>', '', row.get("item_code", ""))  # remove HTML
		row_data = [item_code, row.get("item_name", "")]
		row_data.extend([""] * len(day_headers))
		ws.append(row_data)

	file_data = io.BytesIO()
	wb.save(file_data)
	file_data.seek(0)

	encoded = base64.b64encode(file_data.getvalue()).decode()
	filename = "kanban_qty_template.xlsx"
	return {"filename": filename, "data": encoded}

import frappe
from frappe.utils.file_manager import get_file
import pandas as pd
import datetime
import math

@frappe.whitelist()
def enqueue_upload(month, to_date, file):
	year = datetime.date.today().year
	month_map = {
		'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
		'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
	}
	month_num = month_map.get(month)

	# Get file path
	file_doc = get_file(file)
	file_path = file_doc[1]

	# Read Excel file
	df = pd.read_excel(file_path)

	for _, row in df.iterrows():
		item_code = row.get("Item Code")
		item_name = row.get("Item Name")

		for col in df.columns:
			if col in ["Item Code", "Item Name"]:
				continue

			value = row.get(col)

			# Skip empty cells
			if value is None or (isinstance(value, float) and math.isnan(value)):
				continue

			try:
				qty = float(value)
			except ValueError:
				frappe.errprint(f"Invalid quantity {value} for item {item_code}")
				continue

			# Construct plan date
			try:
				plan_date = datetime.datetime.strptime(f"{col}-{year}", "%b-%d-%Y").date()
    
				today = datetime.date.today()
				if plan_date.month < today.month:
					plan_date = plan_date.replace(year=year + 1)
				
			except Exception:
				frappe.errprint(f"Invalid date format in column {col}")
				continue

			# --- Ensure parent document exists ---
			parent_doc_name = frappe.db.get_value("Daily Production Plan", {"date": plan_date}, "name")

			if not parent_doc_name:
				parent_doc = frappe.get_doc({
					"doctype": "Daily Production Plan",
					"date": plan_date,
					"items": []  # initialize child table
				}).insert(ignore_permissions=True)
				parent_doc_name = parent_doc.name

			# --- Check if child exists ---
			existing_item = frappe.db.exists(
				"Daily Production Plan Item",
				{"parent": parent_doc_name, "parentfield": "items", "item_code": item_code}
			)

			if existing_item:
				frappe.db.set_value("Daily Production Plan Item", existing_item, "plan", qty)
				frappe.db.set_value("Daily Production Plan Item", existing_item, "date", plan_date)
			else:
				frappe.get_doc({
					"doctype": "Daily Production Plan Item",
					"parent": parent_doc_name,
					"parenttype": "Daily Production Plan",
					"parentfield": "items",  # MUST match child table fieldname
					"item_code": item_code,
					"plan": qty,
					"date": plan_date
				}).insert(ignore_permissions=True)


	return "OK"


import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import frappe

@frappe.whitelist()
def download():
	args = frappe.local.form_dict
	filters = json.loads(args.get("filters"))
	filters = frappe._dict(filters)
	
	return production_plan_template(filters)

def production_plan_template(filters):
	columns, data, *_ = execute(filters)  

	wb = Workbook()
	ws = wb.active
	ws.title = "Production Plan"

	# Header formatting
	bold_font = Font(bold=True)
	align_center = Alignment(horizontal='center')
	thin_border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)

	# Color mapping for plan/actual cells
	color_map = {
		'#D3D3D3': 'D3D3D3',  # Light gray
		'#FFD580': 'FFD580',  # Light orange
	}

	# Write Serial Number header
	cell = ws.cell(row=1, column=1, value="S.No")
	cell.font = bold_font
	cell.alignment = align_center
	cell.border = thin_border

	# Write other headers starting from column 2
	for col_idx, col in enumerate(columns, start=2):
		header_name = col.get('label') if isinstance(col, dict) else col
		if not header_name:
			header_name = col.get('fieldname') if isinstance(col, dict) else ''
		cell = ws.cell(row=1, column=col_idx, value=header_name)
		cell.font = bold_font
		cell.alignment = align_center
		cell.border = thin_border

	ws.freeze_panes = 'E2'

	# Write data rows
	for row_idx, row in enumerate(data, start=2):
		# Serial Number (first column)
		sn_cell = ws.cell(row=row_idx, column=1, value=row_idx-1)
		sn_cell.border = thin_border
		sn_cell.alignment = Alignment(horizontal='right')

		for col_idx, col in enumerate(columns, start=2):
			key = col.get('fieldname') if isinstance(col, dict) else col
			value = row.get(key, '') if isinstance(row, dict) else row
			cell = ws.cell(row=row_idx, column=col_idx)

			# Remove HTML tags only for first data column (col_idx == 2)
			if col_idx == 2 and isinstance(value, str):
				value = re.sub(r'<.*?>', '', value)

			# Handle colored divs for other columns
			elif isinstance(value, str) and '<div' in value:
				match = re.search(r'background-color:\s*(#[0-9A-Fa-f]{6})', value)
				if match:
					hex_color = match.group(1)
					fill_color = color_map.get(hex_color)
					if fill_color:
						cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
				inner_text_match = re.search(r'>(.*?)</div>', value)
				value = inner_text_match.group(1) if inner_text_match else ''

			cell.value = value
			cell.border = thin_border

			# Alignment: first three columns right, rest left
			if col_idx in [2, 3, 4]:
				cell.alignment = Alignment(horizontal='left')
			else:
				cell.alignment = Alignment(horizontal='right')

	# Adjust column widths
	for col in ws.columns:
		max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
		adjusted_width = max_length + 2
		ws.column_dimensions[col[0].column_letter].width = adjusted_width

	# Save workbook to BytesIO and send as response
	output = BytesIO()
	wb.save(output)
	output.seek(0)

	frappe.local.response.filename = "Production Plan.xlsx"
	frappe.local.response.filecontent = output.read()
	frappe.local.response.type = "download"

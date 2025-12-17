from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.utils import get_column_letter
from io import BytesIO
import frappe
from openpyxl import Workbook

@frappe.whitelist()
def download(date):
    filename = 'Attendance_Report.xlsx'
    xlsx_content = build_xlsx_content(date)
    frappe.response['filename'] = filename
    frappe.response['filecontent'] = xlsx_content.getvalue()
    frappe.response['type'] = 'binary'

def build_xlsx_content(date):
    wb = Workbook()
    sheet = wb.active
    
    # Insert a row above headers and merge cells for 'Shift' and 'Category'
    sheet.insert_rows(1)
    sheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    sheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=9)
    sheet.merge_cells(start_row=1, start_column=10, end_row=1, end_column=13)
    sheet.merge_cells(start_row=1, start_column=14, end_row=2, end_column=14)
    sheet.cell(row=1, column=1).value = 'Parent Department'
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    sheet.cell(row=1, column=2).value = 'Department'
    sheet.cell(row=1, column=2).alignment = Alignment(horizontal='center')
    sheet.cell(row=1, column=3).value = 'Shift'
    sheet.cell(row=1, column=3).alignment = Alignment(horizontal='center')
    sheet.cell(row=1, column=10).value = 'Category'
    sheet.cell(row=1, column=10).alignment = Alignment(horizontal='center')
    sheet.cell(row=1, column=14).value = 'Total Present'
    sheet.cell(row=1, column=14).alignment = Alignment(horizontal='center') 
    columns_to_adjust = [10, 11, 12, 13, 14]
    for col_idx in columns_to_adjust:
        sheet.column_dimensions[get_column_letter(col_idx)].width = 15 
    columns_for_dep=[1,2] 
    for col_indx in columns_for_dep:
        sheet.column_dimensions[get_column_letter(col_indx)].width = 27  
    
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    black_font = Font(color='000000')  # Black color font
    
    border = Border(left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'))

    # Apply fill and font to all cells
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            
            cell.fill = orange_fill
            cell.font = black_font
            cell.border=border
    shifts = frappe.get_all("Shift Type", ['name'], order_by='name ASC')
    num_shifts = len(shifts)
    shift_col_start = 3  # Start shifts after three empty headers
    
    # Insert shift headers
    for idx, shift in enumerate(shifts, start=0):
        sheet.cell(row=2, column=shift_col_start + idx).value = shift['name']
        sheet.cell(row=2, column=shift_col_start + idx).alignment = Alignment(horizontal='center')
    
    # Determine starting column for categories
    category_col_start = shift_col_start + num_shifts

    # Insert category headers
    categories = frappe.get_all("Employee Category", {'name': ('not in', ['Sub Staff', 'Director'])}, ['name'])
    for idx, category in enumerate(categories, start=0):
        sheet.cell(row=2, column=category_col_start + idx).value = category['name']
        sheet.cell(row=2, column=category_col_start + idx).alignment = Alignment(horizontal='center')
    sheet.cell(row=2, column=category_col_start + len(categories) + 1).alignment = Alignment(horizontal='center')

    parent_departments = frappe.get_all("Department", {'disabled': ('!=', 1), "parent_department": "All Departments"}, ['name'])
    row_num = 3  # Start populating data from row 3
    for parent_dept in parent_departments:
        parent_dept_name = parent_dept['name']
        sheet.cell(row=row_num, column=1).value = parent_dept_name
        
        parent_categories = frappe.get_all("Employee Category", {'name': ('not in', ['Sub Staff', 'Director'])}, ['name'])
        
        populate_department_data(sheet, date, parent_dept_name, row_num, shift_col_start, category_col_start, shifts, parent_categories)
        linked_depts = frappe.get_all("Department", {'disabled': ('!=', 1), "parent_department": parent_dept_name}, ['name'])
        if linked_depts:
            for linked_idx, linked_dept in enumerate(linked_depts, start=1):
                linked_dept_name = linked_dept['name']
                linked_row_num = row_num + linked_idx
                sheet.cell(row=linked_row_num, column=2).value = linked_dept_name
                linked_categories = frappe.get_all("Employee Category", {'name': ('not in', ['Sub Staff', 'Director'])}, ['name'])
                populate_department_data(sheet, date, linked_dept_name, linked_row_num, shift_col_start, category_col_start, shifts, linked_categories)          
            row_num += len(linked_depts) + 1  
        else:
            row_num += 1
    
    # Insert row for totals
    total_row = sheet.max_row + 1
    # sheet.cell(row=total_row, column=1).value = 'Total'
    sheet.cell(row=total_row, column=1).alignment = Alignment(horizontal='center')
    
    # Calculate and insert totals for shifts only
    for s_idx in range(num_shifts):
        total_shifts_count = sum(sheet.cell(row=row, column=shift_col_start + s_idx).value or 0 for row in range(3, total_row))
        sheet.cell(row=total_row, column=shift_col_start + s_idx).value = total_shifts_count
        sheet.cell(row=total_row, column=shift_col_start + s_idx).alignment = Alignment(horizontal='center')

    # Calculate and insert total count for all shifts
    total_all_count = sum(sheet.cell(row=total_row, column=col).value or 0 for col in range(3, category_col_start + len(categories) + 1))
    sheet.cell(row=total_row, column=category_col_start + len(categories)).value = total_all_count
    sheet.cell(row=total_row, column=category_col_start + len(categories)).alignment = Alignment(horizontal='center')
    
    # Set row height
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        sheet.row_dimensions[row[0].row].height = 20  # Adjust height as needed

    output = BytesIO()
    wb.save(output)
    return output


def populate_department_data(sheet, date, department_name, row_num, shift_col_start, category_col_start, shifts, categories):
    for s_idx, shift in enumerate(shifts, start=0):
        shift_attendance_count = frappe.db.sql("""
            SELECT COUNT(*) AS count
            FROM `tabAttendance`
            WHERE attendance_date = %s
            AND shift = %s
            AND department = %s
            AND in_time IS NOT NULL
        """, (date, shift['name'], department_name), as_dict=True)
        shift_attendance = shift_attendance_count[0].count if shift_attendance_count else 0
        sheet.cell(row=row_num, column=shift_col_start + s_idx).value = shift_attendance
        sheet.cell(row=row_num, column=shift_col_start + s_idx).alignment = Alignment(horizontal='center')

    for c_idx, category in enumerate(categories, start=0):
        category_attendance_count = frappe.db.sql("""
            SELECT COUNT(*) AS count
            FROM `tabAttendance`
            WHERE attendance_date = %s
            AND custom_employee_category = %s
            AND department = %s
            AND in_time IS NOT NULL
        """, (date, category['name'], department_name), as_dict=True)
        category_attendance = category_attendance_count[0].count if category_attendance_count else 0
        sheet.cell(row=row_num, column=category_col_start + c_idx).value = category_attendance
        sheet.cell(row=row_num, column=category_col_start + c_idx).alignment = Alignment(horizontal='center')

    # Calculate and insert total count for shifts in the "Total" column
    total_shifts_count = sum(sheet.cell(row=row_num, column=shift_col_start + s_idx).value or 0 for s_idx in range(len(shifts)))
    sheet.cell(row=row_num, column=category_col_start + len(categories)).value = total_shifts_count
    sheet.cell(row=row_num, column=category_col_start + len(categories)).alignment = Alignment(horizontal='center')

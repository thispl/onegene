import frappe
from frappe.utils import add_days,date_diff
from frappe.utils.csvutils import UnicodeWriter
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from six import BytesIO
from datetime import date, timedelta, datetime,time
from calendar import monthrange
from frappe.utils import getdate, nowdate


@frappe.whitelist()
def fetch_attendance_summary(from_date, to_date, dept, categories, shift, mode):
	condition_shift = "AND att.shift = %s" if shift else ""
	condition_date = {
		"greater_3": "emp.date_of_joining < DATE_SUB(CURDATE(), INTERVAL 3 MONTH)",
		"bw_1_3": "emp.date_of_joining BETWEEN DATE_SUB(CURDATE(), INTERVAL 3 MONTH) AND DATE_SUB(CURDATE(), INTERVAL 1 MONTH)",
		"less_1": "emp.date_of_joining > DATE_SUB(CURDATE(), INTERVAL 1 MONTH)"
	}

	query = f"""
		SELECT COUNT(*) AS count
		FROM `tabAttendance` att
		INNER JOIN `tabEmployee` emp ON att.employee = emp.name
		WHERE att.attendance_date BETWEEN %s AND %s
		AND att.department = %s
		AND att.custom_employee_category IN %s
		AND att.docstatus != 2
		AND att.status = 'Present'
		{condition_shift}
		AND {condition_date[mode]}
	"""

	params = [from_date, to_date, dept, categories]
	if shift:
		params.append(shift)

	return frappe.db.sql(query, tuple(params), as_dict=True)[0]["count"] or 0
@frappe.whitelist()
def fetch_ot_count(from_date, to_date, dept, shift=None):
	shift_clause = "AND shift = %s" if shift else ""
	query = f"""
		SELECT COUNT(*) AS count
		FROM `tabAttendance`
		WHERE attendance_date BETWEEN %s AND %s
		AND department = %s
		AND docstatus != 2
		AND custom_overtime_hours > 0
		{shift_clause}
	"""
	params = [from_date, to_date, dept]
	if shift:
		params.append(shift)
	return frappe.db.sql(query, tuple(params), as_dict=True)[0]["count"] or 0


@frappe.whitelist()
def download(from_date,to_date,department=None,shift=None):
	filename = 'Manpower Usage Service Wise'
	args = {'from_date':from_date,'to_date':to_date,'department':department,'shift':shift}
	# args = frappe.local.form_dict
	frappe.msgprint("Report is generating in the background,kindly check after few mins in the same page.")
	enqueue(build_xlsx_response, queue='long', timeout=36000, event='build_xlsx_response',filename=filename,args=args)
	

def make_xlsx(data, args, sheet_name=None, wb=None, column_widths=None):
	
	
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	#table1
	header_date = title(args)
	ws.append(header_date)

	header_date = title1(args)
	ws.append(header_date)

	header_date = title2(args)
	ws.append(header_date)

	table_count =0
	data_rows = title3(args,table_count)
	for row in data_rows:
		ws.append(row)
	if not args.get('department'):
		ws.append([])
		ws.append([])
		ws.append([])
		#table2	
		header_date = title(args)
		ws.append(header_date)

		header_date = title1(args)
		ws.append(header_date)

		header_date = title2(args)
		ws.append(header_date)

		table_count = 1
		data_rows = title3(args,table_count)
		for row in data_rows:
			ws.append(row)

	border = Border(
		left=Side(border_style='thin'),
		right=Side(border_style='thin'),
		top=Side(border_style='thin'),
		bottom=Side(border_style='thin'))
	align_center = Alignment(horizontal='center',vertical='center')
	align_right = Alignment(horizontal='right', vertical='top', wrap_text=True)
	align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
	if args.get('department'):
		ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= 25)
		ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column= 2)
		ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column= 3)
		ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column= 6)
		ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column= 9)
		ws.merge_cells(start_row=2, start_column=10, end_row=2, end_column= 12)
		ws.merge_cells(start_row=2, start_column=13, end_row=2, end_column= 15)
		ws.merge_cells(start_row=2, start_column=16, end_row=3, end_column= 16)
		ws.merge_cells(start_row=2, start_column=17, end_row=2, end_column= 28)
	
		for rows in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=28):
			for cell in rows:
			
				cell.font = Font(name="Times New Roman", size=12,bold=True)
				cell.alignment = align_center
		for rows in ws.iter_rows(min_row=4, max_row=3+len(title3(args,0)), min_col=1, max_col=1):
			for cell in rows:
				ws.merge_cells(start_row=cell.row, start_column=1, end_row=cell.row, end_column= 2)
				cell.alignment = align_center
		for rows in ws.iter_rows(min_row=3+len(title3(args,0)), max_row=3+len(title3(args,0)), min_col=1, max_col=2):
			for cell in rows:
				cell.alignment = align_center
				cell.font = Font(bold=True,name="Times New Roman", size=12)
		for rows in ws.iter_rows(min_row=3+len(title3(args,0)), max_row=3+len(title3(args,0)), min_col=3, max_col=28):
			for cell in rows:
				cell.alignment = align_right
				cell.font = Font(bold=True,name="Times New Roman", size=12)
		for rows in ws.iter_rows(min_row=4, max_row=2+len(title3(args,0)), min_col=1, max_col=28):
			for cell in rows:
				cell.font = Font(name="Times New Roman", size=12)
		for rows in ws.iter_rows(min_row=1, max_row=3+len(title3(args,0)), min_col=1, max_col=28):
			for cell in rows:
				cell.border = border
	else:
		ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= 28)
		ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column= 2)
		ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column= 3) 
		ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column= 6) # Operators
		ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column= 9) # trainees
		ws.merge_cells(start_row=2, start_column=10, end_row=2, end_column= 12) # Apperentice
		ws.merge_cells(start_row=2, start_column=13, end_row=2, end_column= 15) # Contractors
		ws.merge_cells(start_row=2, start_column=16, end_row=3, end_column= 16) #Oneday(OD)
		ws.merge_cells(start_row=2, start_column=17, end_row=2, end_column= 28)
		ws.merge_cells(start_row=15, start_column=1, end_row=15, end_column= 2)
		ws.merge_cells(start_row=4, start_column=1, end_row=8, end_column= 1)
		ws.merge_cells(start_row=9, start_column=1, end_row=14, end_column= 1)
		ws.merge_cells(start_row=16, start_column=1, end_row=18, end_column= 28)

		# table 2
		ws.merge_cells(start_row=19, start_column=1, end_row=19, end_column= 28)
		ws.merge_cells(start_row=20, start_column=1, end_row=21, end_column= 2)
		ws.merge_cells(start_row=20, start_column=3, end_row=21, end_column= 3)
		ws.merge_cells(start_row=20, start_column=4, end_row=20, end_column= 6)
		ws.merge_cells(start_row=20, start_column=7, end_row=20, end_column= 9)
		ws.merge_cells(start_row=20, start_column=10, end_row=20, end_column= 12)
		ws.merge_cells(start_row=20, start_column=13, end_row=20, end_column= 15)
		ws.merge_cells(start_row=20, start_column=16, end_row=21, end_column= 16) 
		ws.merge_cells(start_row=20, start_column=17, end_row=20, end_column= 28)

		ws.merge_cells(start_row=23, start_column=1, end_row=32, end_column= 1)
		ws.merge_cells(start_row=33, start_column=1, end_row=40, end_column= 1)
		ws.merge_cells(start_row=41, start_column=1, end_row=41, end_column= 2)
		for rows in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=28):
			for cell in rows:
				cell.font = Font(name="Times New Roman", size=12,bold=True)
				cell.alignment = align_center
		for rows in ws.iter_rows(min_row=4, max_row=3+len(title3(args,0)), min_col=1, max_col=1):
			for cell in rows:
				# ws.merge_cells(start_row=cell.row, start_column=1, end_row=cell.row, end_column= 2)
				cell.alignment = align_center
		for rows in ws.iter_rows(min_row=3+len(title3(args,0)), max_row=3+len(title3(args,0)), min_col=1, max_col=2):
			for cell in rows:
				cell.alignment = align_center
				cell.font = Font(bold=True,name="Times New Roman", size=12)
		for rows in ws.iter_rows(min_row=3+len(title3(args,0)), max_row=3+len(title3(args,0)), min_col=3, max_col=28):
			for cell in rows:
				cell.alignment = align_right
				cell.font = Font(bold=True,name="Times New Roman", size=12)
		for rows in ws.iter_rows(min_row=4, max_row=2+len(title3(args,0)), min_col=1, max_col=28):
			for cell in rows:
				cell.font = Font(name="Times New Roman", size=12)
		for rows in ws.iter_rows(min_row=1, max_row=3+len(title3(args,0)), min_col=1, max_col=28):
			for cell in rows:
				cell.border = border
		
		for rows in ws.iter_rows(min_row=16, max_row=18, min_col=1, max_col=28):
			for cell in rows:
				cell.font = Font(name="Times New Roman", size=12,bold=True)
				cell.alignment = align_center
		#table 2
		for rows in ws.iter_rows(min_row=19, max_row=21, min_col=1, max_col=28):
			for cell in rows:
				cell.font = Font(name="Times New Roman", size=12,bold=True)
				cell.alignment = align_center
		for rows in ws.iter_rows(min_row=22, max_row=21+len(title3(args,1)), min_col=1, max_col=1):
			for cell in rows:
				# ws.merge_cells(start_row=cell.row, start_column=1, end_row=cell.row, end_column= 2)
				cell.alignment = align_center
		for rows in ws.iter_rows(min_row=9+len(title3(args,1))+len(title3(args,0)), max_row=9+len(title3(args,1))+len(title3(args,0)), min_col=1, max_col=2):
			for cell in rows:
				cell.alignment = align_center
				cell.font = Font(bold=True,name="Times New Roman", size=12)
		for rows in ws.iter_rows(min_row=9+len(title3(args,1))+len(title3(args,0)), max_row=9+len(title3(args,1))+len(title3(args,0)), min_col=3, max_col=28):
			for cell in rows:
				cell.alignment = align_right
				cell.font = Font(bold=True,name="Times New Roman", size=12)
		for rows in ws.iter_rows(min_row=22, max_row=20+len(title3(args,1)), min_col=1, max_col=28):
			for cell in rows:
				cell.font = Font(name="Times New Roman", size=12)
		for rows in ws.iter_rows(min_row=19, max_row=21+len(title3(args,1)), min_col=1, max_col=28):
			for cell in rows:
				cell.border = border
	ws.column_dimensions['A'].width = 25
	ws.column_dimensions['B'].width = 25
	ws.column_dimensions['C'].width = 15
	ws.column_dimensions['P'].width = 15
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename,args):
	frappe.log_error('manpower',filename)
	xlsx_file = make_xlsx(filename,args)
	ret = frappe.get_doc({
			"doctype": "File",
			"attached_to_name": ' ',
			"attached_to_doctype": 'Reports Dashboard',
			"attached_to_field": 'attach',
			"file_name": filename + '.xlsx',
			"is_private": 0,
			"content": xlsx_file.getvalue(),
			"decode": False
		})
	ret.save(ignore_permissions=True)
	frappe.db.commit()
	attached_file = frappe.get_doc("File", ret.name)
	frappe.db.set_value('Reports Dashboard',None,'downloaded_report',attached_file.file_url)
	doc = frappe.get_single("Reports Dashboard")
	doc.reload()
	doc.downloaded_report = attached_file.file_url
	doc.save()

@frappe.whitelist()
def title(args):
	start_date = datetime.strptime(str(args.get('from_date')),'%Y-%m-%d')
	start_dat = str(start_date.strftime('%d') +' - '+ str(start_date.strftime('%b')))
	end_date = datetime.strptime(str(args.get('to_date')),'%Y-%m-%d')
	end_dat = str(end_date.strftime('%d') +' - '+ str(end_date.strftime('%b')))
	data = ["Manpower Usage Service Wise Report for " + str(start_dat) + " to " +str(end_dat)]
	return data

@frappe.whitelist()
def title1(args):
	data = ["Category","","Allotted MP","Operators","","","Trainees","","","APP","","","Cont","","","Oneday(OD)","Total(Mandays)"]
	return data

@frappe.whitelist()
def title2(args):
	data = ["","","","R","1 to 3M","1M","R","1 to 3M","1M","R","1 to 3M","1M","R","1 to 3M","1M","","R","1 to 3M","1M","OD","OT","Total","R","1 to 3M","1M","OD","OT","Total"]
	return data






# @frappe.whitelist()
# def get_attendance_breakdown(from_date, to_date, dept, shift=None):
# 	CATEGORY_GROUPS = {
# 		"trainee": ("Trainee",),
# 		"comp": ("Staff", "Sub Staff", "Operator"),
# 		"app": ("Apprentice",),
# 		"cont": ("Contractor",)
# 	}
# 	EXPERIENCE_BANDS = ["greater_3", "bw_1_3", "less_1"]
# 	result = {}
# 	for cat_key, categories in CATEGORY_GROUPS.items():
# 		for exp_key in EXPERIENCE_BANDS:
# 			result[f"{cat_key}_{exp_key}"] = fetch_attendance_summary(from_date, to_date, dept, categories, shift, exp_key)
# 			result[f"{cat_key}_{exp_key}_half"] = 0  # Optional: add real half-day support here
# 	return result

# @frappe.whitelist()
# def build_row(args, dept, parent_label, plan, breakdown, ot_count):
# 	def calc_total(*keys):
# 		return sum(breakdown[k] + (breakdown.get(f"{k}_half", 0) / 2) for k in keys)

# 	days_count = date_diff(add_days(args.get('to_date'), 1), args.get('from_date'))
# 	total_plan = plan * days_count if plan else 0

# 	tg3 = calc_total("trainee_greater_3", "comp_greater_3", "app_greater_3", "cont_greater_3")
# 	t13 = calc_total("trainee_bw_1_3", "comp_bw_1_3", "app_bw_1_3", "cont_bw_1_3")
# 	t1 = calc_total("trainee_less_1", "comp_less_1", "app_less_1", "cont_less_1")
# 	total = tg3 + t13 + t1
# 	plan_value = total_plan[0]["plan"] if total_plan and total_plan[0].get("plan") else 0

# 	percent = lambda val: round((val / plan_value) * 100, 2) if plan_value else 0
# 	# percent = lambda val: round((val / total_plan[0]) * 100, 2) if total_plan else 0
# 	# percent = lambda val: round((val[0] / total_plan) * 100, 2) if total_plan and isinstance(val, list) else round((val / total_plan) * 100, 2) if total_plan else 0


# 	row = [
# 		parent_label, dept, total_plan,
# 		calc_total("comp_greater_3"), calc_total("comp_bw_1_3"), calc_total("comp_less_1"),
# 		calc_total("trainee_greater_3"), calc_total("trainee_bw_1_3"), calc_total("trainee_less_1"),
# 		calc_total("app_greater_3"), calc_total("app_bw_1_3"), calc_total("app_less_1"),
# 		calc_total("cont_greater_3"), calc_total("cont_bw_1_3"), calc_total("cont_less_1"),
# 		"", tg3, t13, t1, "", ot_count, total,
# 		percent(tg3), percent(t13), percent(t1), "", percent(ot_count), percent(total)
# 	]
# 	return row, {
# 		"tg3": tg3, "t13": t13, "t1": t1, "total": total, "ot": ot_count,"plan":plan or 0,
# 		"percent_tg3": percent(tg3), "percent_t13": percent(t13),
# 		"percent_t1": percent(t1), "percent_ot": percent(ot_count),
# 		"percent_total": percent(total), "plan_days": total_plan
# 	}

@frappe.whitelist()
def title3(args,table_count):
	data = []
	row = []
	departments = ''
	ttrainee_greater_3 = 0
	ttrainee_1_3 =0
	ttrainee_less_1 =0
	tcomp_greater_3 = 0
	tapp_greater_3 = 0
	tcont_greater_3 = 0
	tcomp_1_3 = 0
	tapp_1_3 = 0
	tcont_1_3 = 0
	tcomp_less_1 = 0
	tapp_less_1 = 0
	tcont_less_1 = 0
	ttotal_greater_3 = 0
	ttotal_1_3 = 0
	ttotal_less_1 = 0
	ttotal=0
	tot_count=0
	tmanpower_plan=0
	tper_greater_3=0
	tper_1_3=0
	tper_less_1=0
	tot_per=0
	ttotal_per=0
	avag_1=0
	avag_13=0
	avag_3=0
	avag_ot=0
	avag_tot=0
	i = 1
	if args.get('department'):
		frappe.log_error('Shift', args.get('shift'))
		start_date = datetime.strptime(args.get('from_date'), "%Y-%m-%d")
		end_date = datetime.strptime(args.get('to_date'), "%Y-%m-%d")
		month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
		start_month_name = month_names[start_date.month - 1]
		end_month_name = month_names[end_date.month - 1]
		start_month_number = start_date.month
		end_month_number = end_date.month
		if start_month_name==end_month_name:
			manpower_plan = manpower_plan_method(args.get('department'),start_month_name,start_date.year,end_month_name)
		else:
			manpower_plan = manpower_plan_method(args.get('department'),start_month_name,start_date.year,end_month_name)
		if args.get('shift') !='':
			trainee_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), args.get('department'),("Trainee",),args.get('shift'))
			trainee_greater_3_half = greater_3_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Trainee",),args.get('shift'))
			trainee_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), args.get('department'),("Trainee",),args.get('shift'))
			trainee_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Trainee",),args.get('shift'))
			trainee_less_1 = less_1(args.get('from_date'), args.get('to_date'), args.get('department'),("Trainee",),args.get('shift'))
			trainee_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Trainee",),args.get('shift'))
			comp_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), args.get('department'),('Staff', 'Sub Staff', 'Operator',),args.get('shift'))
			comp_greater_3_half = greater_3(args.get('from_date'), args.get('to_date'), args.get('department'),('Staff', 'Sub Staff', 'Operator',),args.get('shift'))
			comp_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), args.get('department'),('Staff', 'Sub Staff', 'Operator',),args.get('shift'))
			comp_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), args.get('department'),('Staff', 'Sub Staff', 'Operator',),args.get('shift'))
			comp_less_1 = less_1(args.get('from_date'), args.get('to_date'), args.get('department'),('Staff', 'Sub Staff', 'Operator',),args.get('shift'))
			comp_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), args.get('department'),('Staff', 'Sub Staff', 'Operator',),args.get('shift'))
			app_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), args.get('department'),("Apprentice",),args.get('shift'))
			app_greater_3_half = greater_3_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Apprentice",),args.get('shift'))
			app_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), args.get('department'),("Apprentice",),args.get('shift'))
			app_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Apprentice",),args.get('shift'))
			app_less_1 = less_1(args.get('from_date'), args.get('to_date'), args.get('department'),("Apprentice",),args.get('shift'))
			app_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Apprentice",),args.get('shift'))	
			cont_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), args.get('department'),("Contractor",),args.get('shift'))
			cont_greater_3_half = greater_3_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Contractor",),args.get('shift'))
			cont_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), args.get('department'),("Contractor",),args.get('shift'))
			cont_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Contractor",),args.get('shift'))
			cont_less_1 = less_1(args.get('from_date'), args.get('to_date'), args.get('department'),("Contractor",),args.get('shift'))
			cont_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Contractor",),args.get('shift'))			
			ot_count = fetch_ot_count(args.get('from_date'), args.get('to_date'), args.get('department'),args.get('shift'))
		else:
			trainee_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), args.get('department'),("Trainee",))
			trainee_greater_3_half = greater_3_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Trainee",))
			trainee_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), args.get('department'),("Trainee",))
			trainee_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Trainee",))
			trainee_less_1 = less_1(args.get('from_date'), args.get('to_date'), args.get('department'),("Trainee",))
			trainee_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Trainee",))			
			comp_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), args.get('department'),("Staff", "Sub Staff", "Operator",))
			comp_greater_3_half = greater_3_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Staff", "Sub Staff", "Operator",))
			comp_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), args.get('department'),("Staff", "Sub Staff", "Operator",))
			comp_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Staff", "Sub Staff", "Operator",))
			comp_less_1 = less_1(args.get('from_date'), args.get('to_date'), args.get('department'),("Staff", "Sub Staff", "Operator",))
			comp_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Staff", "Sub Staff", "Operator",))
			app_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), args.get('department'),("Apprentice",))
			app_greater_3_half = greater_3_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Apprentice",))
			app_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), args.get('department'),("Apprentice",))
			app_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Apprentice",))
			app_less_1 = less_1(args.get('from_date'), args.get('to_date'), args.get('department'),("Apprentice",))
			app_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Apprentice",))			
			cont_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), args.get('department'),("Contractor",))
			cont_greater_3_half = greater_3_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Contractor",))
			cont_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), args.get('department'),("Contractor",))
			cont_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Contractor",))
			cont_less_1 = less_1(args.get('from_date'), args.get('to_date'), args.get('department'),("Contractor",))
			cont_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), args.get('department'),("Contractor",))			
			ot_count = fetch_ot_count(args.get('from_date'), args.get('to_date'), args.get('department'))
		total_greater_3=trainee_greater_3[0]['count']+comp_greater_3[0]['count']+app_greater_3[0]['count']+cont_greater_3[0]['count']+((trainee_greater_3_half[0]['count']+ comp_greater_3_half[0]['count']+app_greater_3_half[0]['count']+cont_greater_3_half[0]['count'])/2)
		total_1_3=trainee_1_3[0]['count']+comp_1_3[0]['count']+app_1_3[0]['count']+cont_1_3[0]['count']+((trainee_1_3_half[0]['count']+comp_1_3_half[0]['count']+app_1_3_half[0]['count']+cont_1_3_half[0]['count'])/2)
		total_less_1=trainee_less_1[0]['count']+comp_less_1[0]['count']+app_less_1[0]['count']+cont_less_1[0]['count']+((trainee_less_1_half[0]['count'] + comp_less_1_half[0]['count']+app_less_1_half[0]['count']+cont_less_1_half[0]['count'])/2)
		total=total_greater_3+total_1_3+total_less_1		
		per_greater_3=total_greater_3/(manpower_plan[0]['plan']*(date_diff(add_days(args.get('to_date'),1),args.get('from_date')))) * 100 if manpower_plan[0]['plan'] else 0
		per_1_3=total_1_3/(manpower_plan[0]['plan']*(date_diff(add_days(args.get('to_date'),1),args.get('from_date')))) * 100 if manpower_plan[0]['plan'] else 0
		per_less_1=total_less_1/(manpower_plan[0]['plan']*(date_diff(add_days(args.get('to_date'),1),args.get('from_date')))) * 100 if manpower_plan[0]['plan'] else 0
		ot_per=ot_count [0]['count']/(manpower_plan[0]['plan']*(date_diff(add_days(args.get('to_date'),1),args.get('from_date')))) * 100 if manpower_plan[0]['plan'] else 0
		total_per=total/(manpower_plan[0]['plan']*(date_diff(add_days(args.get('to_date'),1),args.get('from_date')))) * 100 if manpower_plan[0]['plan'] else 0
		row = [args.get('department'),"",(manpower_plan[0]['plan']*(date_diff(add_days(args.get('to_date'),1),args.get('from_date')))) if manpower_plan[0]['plan'] else 0,
		comp_greater_3[0]['count']+(comp_greater_3_half[0]['count']/2),comp_1_3[0]['count']+(comp_1_3_half[0]['count']/2),comp_less_1[0]['count']+(comp_less_1_half[0]['count']/2),
		trainee_greater_3[0]['count']+(trainee_greater_3_half[0]['count']/2),trainee_1_3[0]['count']+(trainee_1_3_half[0]['count']/2),trainee_less_1[0]['count']+(trainee_less_1_half[0]['count']/2),
		app_greater_3[0]['count']+(app_greater_3_half[0]['count']/2),app_1_3[0]['count']+(app_1_3_half[0]['count']/2),app_less_1[0]['count']+(app_less_1_half[0]['count']/2),
		cont_greater_3[0]['count']+(cont_greater_3_half[0]['count']/2),cont_1_3[0]['count']+(cont_1_3_half[0]['count']/2),cont_less_1[0]['count']+(cont_less_1_half[0]['count']/2),'',total_greater_3,total_1_3,total_less_1,'',ot_count[0]['count'],total,round(per_greater_3,2),round(per_1_3,2),round(per_less_1,2),'',round(ot_per,2),round(total_per,2)]
		if per_greater_3>0:
			avag_3+=1
		if per_1_3>0:
			avag_13+=1
		if per_less_1>0:
			avag_1+=1
		if ot_per>0:
			avag_ot+=1
		if total_per>0:
			avag_tot+=1
		ttrainee_greater_3+=trainee_greater_3[0]['count']+(trainee_greater_3_half[0]['count']/2)
		ttrainee_1_3+=trainee_1_3[0]['count']+(trainee_1_3_half[0]['count']/2)
		ttrainee_less_1+=trainee_less_1[0]['count']+(trainee_less_1_half[0]['count']/2)
		tcomp_greater_3+=comp_greater_3[0]['count']+(comp_greater_3_half[0]['count']/2)
		tapp_greater_3+=app_greater_3[0]['count']+(app_greater_3_half[0]['count']/2)
		tcont_greater_3+=cont_greater_3[0]['count']+(cont_greater_3_half[0]['count']/2)
		tcomp_1_3+=comp_1_3[0]['count']+(comp_1_3_half[0]['count']/2)
		tapp_1_3+=app_1_3[0]['count']+(app_1_3_half[0]['count']/2)
		tcont_1_3+=cont_1_3[0]['count']+(cont_1_3_half[0]['count']/2)
		tcomp_less_1+=comp_less_1[0]['count']+(comp_less_1_half[0]['count']/2)
		tapp_less_1+=app_less_1[0]['count']+(app_less_1_half[0]['count']/2)
		tcont_less_1+=cont_less_1[0]['count']+(cont_less_1_half[0]['count']/2)
		ttotal_greater_3+=total_greater_3
		ttotal_1_3+=total_1_3
		ttotal_less_1+=total_less_1
		ttotal+=total
		tot_count+=ot_count [0]['count']
		tper_greater_3+=round(per_greater_3,2)
		tper_1_3 +=round(per_1_3,2)
		tper_less_1 +=round(per_less_1,2)
		tot_per +=round(ot_per,2)
		ttotal_per +=round(total_per,2)
		tmanpower_plan+=(manpower_plan[0]['plan']*(date_diff(add_days(args.get('to_date'),1),args.get('from_date')))) if manpower_plan[0]['plan'] else 0
		data.append(row)
	else:
		parent_dept_list1 =['Manufacturing - WAIP','Other QMFG']
		parent_dept_list2 =['Quality - WAIP','Supporting-Variable','Supporting-Fixed']
		parent_dept_list = parent_dept_list1 +parent_dept_list2
		manufacturing_dept_list =['ASSEMBLY - WAIP','Evap & AC-Tubing - WAIP','IHX-Tubing - WAIP','PSA - WAIP','HTR-Tubing - WAIP']
		other_dept_list =['Cutting - WAIP','Ex-Cutting & Forging - WAIP','Forging - WAIP','Machining - WAIP','Washing - WAIP','Sensor - WAIP']
		quality_dept = ['Quality - WAIP']
		supporting_variable_dept_list =['Delivery - WAIP','Innovation - WAIP','Maintenance - WAIP','Packing - WAIP','NPD - WAIP','ME -Regular - WAIP']
		supporting_fixed_dept_list =['AUDITOR - WAIP','Driver - WAIP','ETP - WAIP','House Keeping - WAIP','IT - WAIP','LOGISTICS - WAIP',
		'Outsourcing - WAIP','PPC - WAIP','Raw Material - WAIP','Security - WAIP','Supplier Debit - WAIP','Data Entry']
		if table_count == 0:
			# table_count +=1
			departments = manufacturing_dept_list + other_dept_list
		if table_count == 1:
			departments = quality_dept + supporting_variable_dept_list + supporting_fixed_dept_list
		for dept in departments:
			# row = [i]
			if dept in manufacturing_dept_list:
				index =0
			if dept in other_dept_list:
				index =1
			if dept in quality_dept:
				index =2
			if dept in supporting_variable_dept_list:
				index =3
			if dept in supporting_fixed_dept_list:
				index =4
			start_date = datetime.strptime(args.get('from_date'), "%Y-%m-%d")
			end_date = datetime.strptime(args.get('to_date'), "%Y-%m-%d")
			month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
			start_month_name = month_names[start_date.month - 1]
			end_month_name = month_names[end_date.month - 1]
			start_month_number = start_date.month
			end_month_number = end_date.month
			if start_month_name==end_month_name:
				manpower_plan = manpower_plan_method(dept,start_month_name,start_date.year,end_month_name)
			else:
				manpower_plan = manpower_plan_method(dept,start_month_name,start_date.year,end_month_name)
			if args.get('shift') !='':
				# trainee
				trainee_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), dept,("Trainee",),args.get('shift'))
				trainee_greater_3_half = greater_3_half(args.get('from_date'), args.get('to_date'), dept,("Trainee",),args.get('shift'))
				trainee_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), dept,("Trainee",),args.get('shift'))
				trainee_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), dept,("Trainee",),args.get('shift'))
				trainee_less_1 = less_1(args.get('from_date'), args.get('to_date'), dept,("Trainee",),args.get('shift'))
				trainee_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), dept,("Trainee",),args.get('shift'))				
				comp_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), dept,("Staff", "Sub Staff", "Operator",),args.get('shift'))
				comp_greater_3_half = greater_3_half(args.get('from_date'), args.get('to_date'), dept,("Staff", "Sub Staff", "Operator",),args.get('shift'))
				comp_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), dept,("Staff", "Sub Staff", "Operator",),args.get('shift'))
				comp_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), dept,("Staff", "Sub Staff", "Operator",),args.get('shift'))
				comp_less_1 = less_1(args.get('from_date'), args.get('to_date'), dept,("Staff", "Sub Staff", "Operator",),args.get('shift'))
				comp_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), dept,("Staff", "Sub Staff", "Operator",),args.get('shift'))				
				app_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), dept,("Apprentice",),args.get('shift'))
				app_greater_3_half = greater_3_half(args.get('from_date'), args.get('to_date'), dept,("Apprentice",),args.get('shift'))
				app_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), dept,("Apprentice",),args.get('shift'))
				app_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), dept,("Apprentice",),args.get('shift'))
				app_less_1 = less_1(args.get('from_date'), args.get('to_date'), dept,("Apprentice",),args.get('shift'))
				app_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), dept,("Apprentice",),args.get('shift'))				
				cont_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), dept,("Contractor",),args.get('shift'))
				cont_greater_3_half = greater_3_half(args.get('from_date'), args.get('to_date'), dept,("Contractor",),args.get('shift'))
				cont_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), dept,("Contractor",),args.get('shift'))
				cont_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), dept,("Contractor",),args.get('shift'))
				cont_less_1 = less_1(args.get('from_date'), args.get('to_date'), dept,("Contractor",),args.get('shift'))
				cont_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), dept,("Contractor",),args.get('shift'))				
				ot_count = fetch_ot_count(args.get('from_date'), args.get('to_date'), dept,args.get('shift'))
			else:
				trainee_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), dept,("Trainee",))
				trainee_greater_3_half = greater_3_half(args.get('from_date'), args.get('to_date'), dept,("Trainee",))
				trainee_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), dept,("Trainee",))
				trainee_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), dept,("Trainee",))
				trainee_less_1 = less_1(args.get('from_date'), args.get('to_date'), dept,("Trainee",))
				trainee_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), dept,("Trainee",))			
				comp_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), dept,("Staff", "Sub Staff", "Operator",))
				comp_greater_3_half = greater_3_half(args.get('from_date'), args.get('to_date'), dept,("Staff", "Sub Staff", "Operator",))
				comp_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), dept,("Staff", "Sub Staff", "Operator",))
				comp_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), dept,("Staff", "Sub Staff", "Operator",))
				comp_less_1 = less_1(args.get('from_date'), args.get('to_date'), dept,("Staff", "Sub Staff", "Operator",))
				comp_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), dept,("Staff", "Sub Staff", "Operator",))				
				app_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), dept,("Apprentice",))
				app_greater_3_half = greater_3_half(args.get('from_date'), args.get('to_date'), dept,("Apprentice",))
				app_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), dept,("Apprentice",))
				app_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), dept,("Apprentice",))
				app_less_1 = less_1(args.get('from_date'), args.get('to_date'), dept,("Apprentice",))
				app_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), dept,("Apprentice",))
				cont_greater_3 = greater_3(args.get('from_date'), args.get('to_date'), dept,("Contractor",))
				cont_greater_3_half = greater_3_half(args.get('from_date'), args.get('to_date'), dept,("Contractor",))
				cont_1_3 = bw_1_3(args.get('from_date'), args.get('to_date'), dept,("Contractor",))
				cont_1_3_half = bw_1_3_half(args.get('from_date'), args.get('to_date'), dept,("Contractor",))
				cont_less_1 = less_1(args.get('from_date'), args.get('to_date'), dept,("Contractor",))
				cont_less_1_half = less_1_half(args.get('from_date'), args.get('to_date'), dept,("Contractor",))			
				ot_count = fetch_ot_count(args.get('from_date'), args.get('to_date'), dept)	
			total_greater_3= trainee_greater_3[0]['count'] +comp_greater_3[0]['count']+app_greater_3[0]['count']+cont_greater_3[0]['count']+((trainee_greater_3_half[0]['count'] + comp_greater_3_half[0]['count']+app_greater_3_half[0]['count']+cont_greater_3_half[0]['count'])/2)
			total_1_3= trainee_1_3[0]['count']+comp_1_3[0]['count']+app_1_3[0]['count']+cont_1_3[0]['count']+((trainee_1_3_half[0]['count']+comp_1_3_half[0]['count']+app_1_3_half[0]['count']+cont_1_3_half[0]['count'])/2)
			total_less_1= trainee_less_1[0]['count']+comp_less_1[0]['count']+app_less_1[0]['count']+cont_less_1[0]['count']+((trainee_less_1_half[0]['count']+comp_less_1_half[0]['count']+app_less_1_half[0]['count']+cont_less_1_half[0]['count'])/2)
			total=total_greater_3+total_1_3+total_less_1
			per_greater_3=total_greater_3/(manpower_plan[0]['plan']*(date_diff(add_days(args.get('to_date'),1),args.get('from_date')))) * 100 if manpower_plan[0]['plan'] else 0			
			per_1_3=total_1_3/(manpower_plan[0]['plan']*(date_diff(add_days(args.get('to_date'),1),args.get('from_date')))) * 100 if manpower_plan[0]['plan'] else 0
			per_less_1=total_less_1/(manpower_plan[0]['plan']*(date_diff(add_days(args.get('to_date'),1),args.get('from_date')))) * 100 if manpower_plan[0]['plan'] else 0
			ot_per=ot_count [0]['count']/(manpower_plan[0]['plan']*(date_diff(add_days(args.get('to_date'),1),args.get('from_date')))) * 100 if manpower_plan[0]['plan'] else 0
			total_per=total/(manpower_plan[0]['plan']*(date_diff(add_days(args.get('to_date'),1),args.get('from_date')))) * 100 if manpower_plan[0]['plan'] else 0
			row = [parent_dept_list[index],dept,(manpower_plan[0]['plan'] *(date_diff(add_days(args.get('to_date'),1),args.get('from_date')))) if manpower_plan[0]['plan'] else 0,
			  comp_greater_3[0]['count']+(comp_greater_3_half[0]['count']/2),comp_1_3[0]['count']+(comp_1_3_half[0]['count']/2),comp_less_1[0]['count']+(comp_less_1_half[0]['count']/2),
			  trainee_greater_3[0]['count']+(trainee_greater_3_half[0]['count']/2),trainee_1_3[0]['count']+(trainee_1_3_half[0]['count']/2),trainee_less_1[0]['count']+(trainee_less_1_half[0]['count']/2),
			app_greater_3[0]['count']+(app_greater_3_half[0]['count']/2),app_1_3[0]['count']+(app_1_3_half[0]['count']/2),app_less_1[0]['count']+(app_less_1_half[0]['count']/2),
			cont_greater_3[0]['count']+(cont_greater_3_half[0]['count']/2),cont_1_3[0]['count']+(cont_1_3_half[0]['count']/2),cont_less_1[0]['count']+(cont_less_1_half[0]['count']/2),
			'',total_greater_3,total_1_3,total_less_1,'',ot_count[0]['count'],total,round(per_greater_3,2),round(per_1_3,2),round(per_less_1,2),'',round(ot_per,2),round(total_per,2)]
			if per_greater_3>0:
				avag_3+=1
			if per_1_3>0:
				avag_13+=1
			if per_less_1>0:
				avag_1+=1
			if ot_per>0:
				avag_ot+=1
			if total_per>0:
				avag_tot+=1
			ttrainee_greater_3+=trainee_greater_3[0]['count']+(trainee_greater_3_half[0]['count']/2)
			ttrainee_1_3+=trainee_1_3[0]['count']+(trainee_1_3_half[0]['count']/2)
			ttrainee_less_1+=trainee_less_1[0]['count']+(trainee_less_1_half[0]['count']/2)
			tcomp_greater_3+=comp_greater_3[0]['count']+(comp_greater_3_half[0]['count']/2)
			tapp_greater_3+=app_greater_3[0]['count']+(app_greater_3_half[0]['count']/2)
			tcont_greater_3+=cont_greater_3[0]['count']+(cont_greater_3_half[0]['count']/2)
			tcomp_1_3+=comp_1_3[0]['count']+(comp_1_3_half[0]['count']/2)
			tapp_1_3+=app_1_3[0]['count']+(app_1_3_half[0]['count']/2)
			tcont_1_3+=cont_1_3[0]['count']+(cont_1_3_half[0]['count']/2)
			tcomp_less_1+=comp_less_1[0]['count']+(comp_less_1_half[0]['count']/2)
			tapp_less_1+=app_less_1[0]['count']+(app_less_1_half[0]['count']/2)
			tcont_less_1+=cont_less_1[0]['count']+(cont_less_1_half[0]['count']/2)
			ttotal_greater_3+=total_greater_3
			ttotal_1_3+=total_1_3
			ttotal_less_1+=total_less_1
			ttotal+=total
			tot_count+=ot_count [0]['count']
			tper_greater_3+=round(per_greater_3,2)
			tper_1_3 +=round(per_1_3,2)
			tper_less_1 +=round(per_less_1,2)
			tot_per +=round(ot_per,2)
			ttotal_per +=round(total_per,2)
			tmanpower_plan+=(manpower_plan[0]['plan']*(date_diff(add_days(args.get('to_date'),1),args.get('from_date')))) if manpower_plan[0]['plan'] else 0
			data.append(row)		
	row1 = ["Total"," ",tmanpower_plan,tcomp_greater_3,tcomp_1_3,tcomp_less_1,
			ttrainee_greater_3,ttrainee_1_3,ttrainee_less_1,
			tapp_greater_3,tapp_1_3,tapp_less_1,
			tcont_greater_3,tcont_1_3,tcont_less_1,
			'',ttotal_greater_3,ttotal_1_3,ttotal_less_1,'',tot_count,ttotal,tper_greater_3/avag_3 if avag_3 else 0,tper_1_3/avag_13 if avag_13 else 0,tper_less_1/avag_1 if avag_1 else 0,'',tot_per/avag_ot if avag_ot else 0,ttotal_per/avag_tot if avag_tot else 0]
	data.append(row1)
	return data



@frappe.whitelist()
def manpower_plan_method(dept,month,year,end_month):
	if month==end_month:
		return frappe.db.sql("""
			SELECT 
				sum(business_plan) as plan
			FROM 
				`tabManpower Plan`
			WHERE 
				department = '%s' 
				AND month = '%s'
				AND year = '%s' 
		"""% (dept,month,year), as_dict=True)
	else:
		return frappe.db.sql("""
			SELECT 
				sum(business_plan) as plan
			FROM 
				`tabManpower Plan`
			WHERE 
				department = '%s' 
				AND FIELD(month, 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')
					between '%s' and '%s' 
				AND year = '%s' 
		"""% (dept,month,end_month,year), as_dict=True)

@frappe.whitelist()
def greater_3(from_date, to_date, dept, employee_category, shift=None):
	# Ensure employee_category is a tuple or list
	if isinstance(employee_category, str):
		employee_category = (employee_category,)  # convert to tuple

	conditions = """
		att.attendance_date BETWEEN %s AND %s
		AND att.department = %s
		AND att.custom_employee_category IN %s
		AND att.docstatus != 2
		AND att.status = 'Present'
		AND emp.date_of_joining < DATE_SUB(CURDATE(), INTERVAL 3 MONTH)
	"""
	params = [from_date, to_date, dept, employee_category]

	if shift:
		conditions += " AND att.shift = %s"
		params.append(shift)

	query = f"""
		SELECT COUNT(*) AS count
		FROM `tabAttendance` att
		INNER JOIN `tabEmployee` emp ON att.employee = emp.name
		WHERE {conditions}
	"""

	return frappe.db.sql(query, tuple(params), as_dict=True)

@frappe.whitelist()
def greater_3_half(from_date, to_date, dept, employee_category, shift=None):
	# Ensure employee_category is a tuple or list
	if isinstance(employee_category, str):
		employee_category = (employee_category,)  # convert to tuple

	conditions = """
		att.attendance_date BETWEEN %s AND %s
		AND att.department = %s
		AND att.custom_employee_category IN %s
		AND att.docstatus != 2
		AND att.status = 'Half Day'
		AND emp.date_of_joining < DATE_SUB(CURDATE(), INTERVAL 3 MONTH)
	"""
	params = [from_date, to_date, dept, employee_category]

	if shift:
		conditions += " AND att.shift = %s"
		params.append(shift)

	query = f"""
		SELECT COUNT(*) AS count
		FROM `tabAttendance` att
		INNER JOIN `tabEmployee` emp ON att.employee = emp.name
		WHERE {conditions}
	"""

	return frappe.db.sql(query, tuple(params), as_dict=True)

@frappe.whitelist()
def bw_1_3(from_date, to_date, dept, employee_category, shift=None):
	# Ensure employee_category is a tuple or list
	if isinstance(employee_category, str):
		employee_category = (employee_category,)  # convert to tuple

	conditions = """
		att.attendance_date BETWEEN %s AND %s
		AND att.department = %s
		AND att.custom_employee_category IN %s
		AND att.docstatus != 2
		AND att.status = 'Present'
		AND DATE_SUB(CURDATE(), INTERVAL 3 MONTH) <= emp.date_of_joining 
		AND emp.date_of_joining <= DATE_SUB(CURDATE(), INTERVAL 1 MONTH)
	"""
	params = [from_date, to_date, dept, employee_category]

	if shift:
		conditions += " AND att.shift = %s"
		params.append(shift)

	query = f"""
		SELECT COUNT(*) AS count
		FROM `tabAttendance` att
		INNER JOIN `tabEmployee` emp ON att.employee = emp.name
		WHERE {conditions}
	"""

	return frappe.db.sql(query, tuple(params), as_dict=True)

@frappe.whitelist()
def bw_1_3_half(from_date, to_date, dept, employee_category, shift=None):
	# Ensure employee_category is a tuple or list
	if isinstance(employee_category, str):
		employee_category = (employee_category,)  # convert to tuple

	conditions = """
		att.attendance_date BETWEEN %s AND %s
		AND att.department = %s
		AND att.custom_employee_category IN %s
		AND att.docstatus != 2
		AND att.status = 'Half Day'
		AND DATE_SUB(CURDATE(), INTERVAL 3 MONTH) <= emp.date_of_joining 
		AND emp.date_of_joining <= DATE_SUB(CURDATE(), INTERVAL 1 MONTH)
	"""
	params = [from_date, to_date, dept, employee_category]

	if shift:
		conditions += " AND att.shift = %s"
		params.append(shift)

	query = f"""
		SELECT COUNT(*) AS count
		FROM `tabAttendance` att
		INNER JOIN `tabEmployee` emp ON att.employee = emp.name
		WHERE {conditions}
	"""

	return frappe.db.sql(query, tuple(params), as_dict=True)


@frappe.whitelist()
def less_1(from_date, to_date, dept, employee_category, shift=None):
	# Ensure employee_category is a tuple or list
	if isinstance(employee_category, str):
		employee_category = (employee_category,)  # convert to tuple

	conditions = """
		att.attendance_date BETWEEN %s AND %s
		AND att.department = %s
		AND att.custom_employee_category IN %s
		AND att.docstatus != 2
		AND att.status = 'Present'
		AND emp.date_of_joining > DATE_SUB(CURDATE(), INTERVAL 1 MONTH)
	"""
	params = [from_date, to_date, dept, employee_category]

	if shift:
		conditions += " AND att.shift = %s"
		params.append(shift)

	query = f"""
		SELECT COUNT(*) AS count
		FROM `tabAttendance` att
		INNER JOIN `tabEmployee` emp ON att.employee = emp.name
		WHERE {conditions}
	"""

	return frappe.db.sql(query, tuple(params), as_dict=True)

@frappe.whitelist()
def less_1_half(from_date, to_date, dept, employee_category, shift=None):
	# Ensure employee_category is a tuple or list
	if isinstance(employee_category, str):
		employee_category = (employee_category,)  # convert to tuple

	conditions = """
		att.attendance_date BETWEEN %s AND %s
		AND att.department = %s
		AND att.custom_employee_category IN %s
		AND att.docstatus != 2
		AND att.status = 'Half Day'
		AND emp.date_of_joining > DATE_SUB(CURDATE(), INTERVAL 1 MONTH)
	"""
	params = [from_date, to_date, dept, employee_category]

	if shift:
		conditions += " AND att.shift = %s"
		params.append(shift)

	query = f"""
		SELECT COUNT(*) AS count
		FROM `tabAttendance` att
		INNER JOIN `tabEmployee` emp ON att.employee = emp.name
		WHERE {conditions}
	"""

	return frappe.db.sql(query, tuple(params), as_dict=True)

@frappe.whitelist()
def fetch_ot_count(from_date, to_date, dept, shift=None):
	shift_clause = "AND shift = %s" if shift else ""
	query = f"""
		SELECT COUNT(*) AS count
		FROM `tabAttendance`
		WHERE attendance_date BETWEEN %s AND %s
		AND department = %s
		AND docstatus != 2
		AND custom_overtime_hours > 0
		{shift_clause}
	"""
	params = [from_date, to_date, dept]
	if shift:
		params.append(shift)
	return frappe.db.sql(query, tuple(params), as_dict=True)
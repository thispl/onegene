from __future__ import unicode_literals
import frappe
# from frappe.utils import _, bold
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, GradientFill, PatternFill
from six import BytesIO
from frappe.utils import cstr, getdate, format_date, add_days
import datetime
from datetime import date, timedelta, datetime,time

@frappe.whitelist()
def download_hr_to_accounts():
	filename = 'Salary_HR_to_Accounts.xlsx'
	build_xlsx_response(filename)

def make_xlsx(sheet_name=None):
	args = frappe.local.form_dict
	wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	data = get_data(args)
	for row in data:
		ws.append(row)

	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= 8)
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column= 8)
	ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column= 8)
	ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column= 8)
	ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column= 8)
	
	align_center = Alignment(horizontal='center',vertical='center')
	border = Border(
		left=Side(border_style='thin'),
		right=Side(border_style='thin'),
		top=Side(border_style='thin'),
		bottom=Side(border_style='thin'))
	for rows in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=8):
		for cell in rows:
			cell.font = Font(bold=True)
			cell.alignment = align_center
			cell.border = border
	for rows in ws.iter_rows(min_row=3, max_row=5, min_col=1, max_col=8):
		for cell in rows:
			cell.font = Font(bold=True)
			# cell.alignment = align_center
			cell.border = border
	for rows in ws.iter_rows(min_row=6, max_row=6, min_col=1, max_col=8):
		for cell in rows:
			cell.font = Font(bold=True)
			cell.alignment = align_center
			cell.border = border
	for rows in ws.iter_rows(min_row=7, max_row=len(get_data(args)), min_col=1, max_col=8):
		for cell in rows:
			# cell.font = Font(bold=True)
			# cell.alignment = align_center
			cell.border = border
	ws.sheet_view.zoomScale = 100 
	ws.column_dimensions['A'].width = 10
	ws.column_dimensions['B'].width = 10
	ws.column_dimensions['C'].width = 30
	ws.column_dimensions['D'].width = 20
	ws.column_dimensions['E'].width = 30
	ws.column_dimensions['F'].width = 20
	ws.column_dimensions['G'].width = 25
	ws.column_dimensions['H'].width = 25
	
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

@frappe.whitelist()
def get_data(args):
	month = datetime.strptime(str(args.from_date), '%Y-%m-%d')
	month_str = month.strftime('%B %Y')
	employee_category  = args.employee_category or ' '
	bank = args.bank or ' '
	branch = frappe.get_value("Bank",{'name':args.bank},['custom_branch']) or ' '
	branch_code = frappe.get_value("Bank",{'name':args.bank},['custom_branch_code']) or ' '
	data = [
		["WONJIN AUTOPARTS INDIA PVT.LTD."],
		["Salary Report - " + month_str],
		["Category", employee_category, "", "", "", ""],
		["Bank",bank, "", "", "", ""],
		["Branch",branch, "", "", "", ""],
		["S No", "Emp.Code", "Emp Name","Bank Name","Branch Code","IFSC Code", "S/B Ac No", "Net Salary"]
	]
	employees = get_employees(args)
	for i, employee in enumerate(employees, start=0 + 1):
		employeecode= employee.get("employee")
		frappe.errprint(employeecode)
		bankname = frappe.get_value("Employee",{'name':employeecode},['bank_name'])
		branchname =frappe.get_value("Employee",{'name':employeecode},['branch_name'])
		ifsc =frappe.get_value("Employee",{'name':employeecode},['ifsc_code'])
		frappe.errprint(ifsc)
		data.append([i, employee.get("employee"), employee.get("employee_name"), bankname, branchname, ifsc ,employee.get("bank_account_no"), employee.get("rounded_total")])  # Adjust as per your employee data structure
	return data


@frappe.whitelist()
def get_employees(args):
	conditions = []
	if args.from_date:
		conditions.append("start_date = '%s'" % args.from_date)
	if args.to_date:
		conditions.append("end_date = '%s'" % args.to_date)
	if args.bank:
		conditions.append("bank_name = '%s'" % args.bank)
	if args.employee_category:
		conditions.append("custom_employee_category = '%s'" % args.employee_category)

	where_clause = " AND ".join(conditions)
	sql_query = """SELECT * FROM `tabSalary Slip` WHERE {0}""".format(where_clause)
    
	try:
		employees = frappe.db.sql(sql_query, as_dict=True)
	except Exception as e:
		frappe.log_error(f"Error in get_employees: {e}")
		employees = []

	return employees
import frappe
from frappe.utils import add_days
from frappe.utils.csvutils import UnicodeWriter
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue
import openpyxl
from frappe.utils import (getdate, cint, add_months, date_diff, add_days)
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from six import BytesIO
from datetime import date, timedelta, datetime,time
from frappe.utils.background_jobs import enqueue
from frappe.utils import time_diff_in_hours, formatdate, get_first_day,get_last_day, nowdate, now_datetime

class ReportsDashboard(Document):
	pass

@frappe.whitelist()
def download(from_date,to_date,department=None,shift=None):
	filename = 'Manpower_Plan_vs_Actual_Report.xlsx'
	# args = frappe.local.form_dict
	# xlsx_content = build_xlsx_content()
	args = {'from_date':from_date,'to_date':to_date,'department':department,'shift':shift}
	frappe.msgprint("Report is generating in the background,kindly check after few mins in the same page.")
	enqueue(build_xlsx_response, queue='long', timeout=36000, event='build_xlsx_response',filename=filename,args=args)
	# frappe.response['filename'] = filename
	# frappe.response['filecontent'] = xlsx_content
	# frappe.response['type'] = 'binary'

def make_xlsx(args, sheet_name=None, wb=None, column_widths=None):
	# args = frappe.local.form_dict
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)

	header_date = title(args)
	ws.append(header_date)

	header_date = title1(args)
	ws.append(header_date)

	header_date = title2(args)
	ws.append(header_date)

	header_date = title3(args)
	ws.append(header_date)

	data_rows = title4(args)
	for row in data_rows:
		ws.append(row)

	header_date = title5(args)
	ws.append(header_date)

	data_rows = title6(args)
	for row in data_rows:
		ws.append(row)

	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= 18)
	ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column= 18)
	ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column= 1)
	ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column= 2)
	ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column= 3)
	ws.merge_cells(start_row=2, start_column=4, end_row=3, end_column= 4)
	# ws.merge_cells(start_row=2, start_column=5, end_row=3, end_column= 5)
	ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=10)
	ws.merge_cells(start_row=2, start_column=11, end_row=2, end_column= 16)
	ws.merge_cells(start_row=2, start_column=18, end_row=3, end_column= 18)
	ws.merge_cells(start_row=2, start_column=17, end_row=3, end_column= 17)
	# ws.merge_cells(start_row=2, start_column=16, end_row=2, end_column=16)
	ws.merge_cells(start_row=4+len(title4(args))+1, start_column=1, end_row=4+len(title4(args))+1, end_column=18)
	ws.merge_cells(start_row=4+len(title4(args)), start_column=1, end_row=4+len(title4(args)), end_column=3)
	ws.merge_cells(start_row=4+len(title4(args))-1, start_column=1, end_row=4+len(title4(args))-1, end_column=3)
	ws.merge_cells(start_row=(len(title4(args))+ len(title6(args)) + 2), start_column=1, end_row=(len(title4(args))+ len(title6(args)) + 2), end_column=3)
	ws.merge_cells(start_row=(len(title4(args))+ len(title6(args)) + 5), start_column=1, end_row=(len(title4(args))+ len(title6(args)) + 5), end_column=3)
	ws.merge_cells(start_row=(len(title4(args))+ len(title6(args)) + 4), start_column=1, end_row=(len(title4(args))+ len(title6(args)) + 4), end_column=3)
	ws.merge_cells(start_row=(len(title4(args))+ len(title6(args)) + 3), start_column=1, end_row=(len(title4(args))+ len(title6(args)) + 3), end_column=3)
	align_center = Alignment(horizontal='center',vertical='center')
	border = Border(
		left=Side(border_style='thin'),
		right=Side(border_style='thin'),
		top=Side(border_style='thin'),
		bottom=Side(border_style='thin'))
	for rows in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=18):
		for cell in rows:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for rows in ws.iter_rows(min_row=4+len(title4(args))-1, max_row=4+len(title4(args))+1, min_col=1, max_col=18):
		for cell in rows:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for rows in ws.iter_rows(min_row=(len(title4(args))+ len(title6(args)) + 2), max_row=(len(title4(args))+ len(title6(args)) + 5), min_col=1, max_col=17):
		for cell in rows:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for rows in ws.iter_rows(min_row=5, max_row=(len(title4(args))+ len(title6(args)) + 5), min_col=4, max_col=18):
		for cell in rows:
			cell.alignment = align_center
	for rows in ws.iter_rows(min_row=1, max_row=(len(title4(args))+ len(title6(args)) + 5), min_col=1, max_col=18):
		for cell in rows:
			cell.border = border
	ws.freeze_panes = 'D4'
	ws.column_dimensions['O'].width = 20
	ws.column_dimensions['B'].width = 25
	ws.column_dimensions['C'].width = 20
	ws.column_dimensions['D'].width = 20
	# ws.column_dimensions['E'].width = 20
	ws.column_dimensions['Q'].width = 15
	ws.column_dimensions['R'].width = 25
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename,args):
	frappe.log_error('manpower',filename)
	xlsx_file = make_xlsx(args)
	ret = frappe.get_doc({
			"doctype": "File",
			"attached_to_name": ' ',
			"attached_to_doctype": 'Reports Dashboard',
			"attached_to_field": 'attach',
			"file_name": filename + '.xlsx',
			"is_private": 0,
			"content": xlsx_file.getvalue(),
			"decode": False
		})
	ret.save(ignore_permissions=True)
	frappe.db.commit()
	attached_file = frappe.get_doc("File", ret.name)
	# frappe.db.set_value('Reports Dashboard',None,'manpower_vs_actual_report',attached_file.file_url)
	doc = frappe.get_single("Reports Dashboard")
	doc.reload()
	doc.manpower_vs_actual_report = attached_file.file_url
	doc.save()


@frappe.whitelist()
def title(args):
	start_month = datetime.strptime(str(args.get('from_date')), '%Y-%m-%d')
	start_mon = start_month.strftime('%B %Y')
	end_month = datetime.strptime(str(args.get('to_date')), '%Y-%m-%d')
	end_mon = end_month.strftime('%B %Y')
	data = [
		"Manpower Plan vs Actual Report for " + str(start_month.day) + " " + start_mon +
		" to " + str(end_month.day) + " " + end_mon
	]
	return data

@frappe.whitelist()
def title1(args):
	data = ["S NO" ,"Particulars","Department","Business Plan","Man Days(A)","","","","","","Overtime(B)","","","","","","Total(A+B)","%"]
	return data

@frappe.whitelist()
def title2(args):
	data = ["","","","","Operators","Trainee","App","Cont","Total","%","Operators","Trainee","App","Cont","Total","%","",""]
	return data

@frappe.whitelist()
def title3(args):
	data = ["Production Team"]
	return data

@frappe.whitelist()
def title4(args):
	# frappe.errprint("STARTEQUSLEND")
	data = []
	row = []
	departments = frappe.get_all('Department', {'parent_department': 'Manufacturing - WAIP'}, order_by='name ASC')
	department_names = [dept['name'] for dept in departments]
	bp_value = 0
	tbp = 0
	tacp = 0
	tco = 0
	tap = 0
	tcon = 0
	tco_ot = 0
	tap_ot = 0
	tcon_ot = 0
	tot = 0
	overall =0
	ttrainee =0
	ttrainee_ot =0
	
	i = 1
	for dept in department_names:
		row = [i]
		start_date = datetime.strptime(args.get('from_date'), "%Y-%m-%d")
		end_date = datetime.strptime(args.get('to_date'), "%Y-%m-%d")
		month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
		start_month_name = month_names[start_date.month - 1]
		end_month_name = month_names[end_date.month - 1]
		start_month_number = start_date.month
		end_month_number = end_date.month
		if start_month_name==end_month_name:
			bp = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = '%s' 
						AND month = '%s'
						AND year = '%s' 
				"""% (dept,start_month_name, start_date.year), as_dict=True)
		else:
			
			bp=frappe.db.sql("""
				SELECT 
					sum(business_plan) as plan
				FROM 
					`tabManpower Plan`
				WHERE 
					department = '%s' 
					AND FIELD(month, 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')
						between '%s' and '%s' 
					AND year = '%s' 
			"""% (dept,start_month_number,end_month_number, end_date.year), as_dict=True)
		
		if start_month_name == end_month_name:
			no_of_days = date_diff(add_days(args.get('to_date'), 1), args.get('from_date'))
			bp_value = bp[0]['plan'] if bp and bp[0]['plan'] else 0
			if bp_value>0:
				bp_value=bp_value*no_of_days
		else:
			bp1 = frappe.db.sql("""
				SELECT 
					sum(business_plan) as plan
				FROM 
					`tabManpower Plan`
				WHERE 
					department = %s 
					AND month = %s
					AND year = %s
			""", (str(dept), str(start_month_name), int(start_date.year)), as_dict=True)
			bp_value1 = bp1[0]['plan'] if bp1 and bp1[0]['plan'] else 0
			bp2 = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = '%s' 
						AND month = '%s'
						AND year = '%s' 
				"""% (str(dept),str(end_month_number), int(end_date.year)), as_dict=True)
			bp_value2 = bp2[0]['plan'] if bp2 and bp2[0]['plan'] else 0
			first_date=get_first_day(end_date)
			last_date=get_last_day(start_date)
			start_days = date_diff(add_days(last_date, 1), start_date)
			end_days = date_diff(add_days(end_date, 1), first_date)
			if bp_value1>0:
				bp_value1=bp_value1*start_days
			else:
				bp_value1=0
			if bp_value2>0:
				bp_value1=bp_value2*end_days
			else:
				bp_value2=0 
			bp_value=bp_value1+bp_value2
		if start_month_name==end_month_name:
			acp = frappe.get_value("Manpower Plan",{'department':dept,'month':end_month_number},['plan']) or 0
		else:
			acp = frappe.get_value("Manpower Plan",{'department':dept,'month':start_month_number,'year':start_date.year},['plan']) or 0
			acp_2=frappe.get_value("Manpower Plan",{'department':dept,'month':end_month_number,'year':start_date.year},['plan']) or 0
			acp=acp+acp_2
		
		trainee_ma_p = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s'  AND shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Trainee") AND docstatus != 2 AND status = 'Present' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
		trainee_ma_h = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s' AND shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Trainee") AND docstatus != 2 AND status = 'Half day' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
		count_trainee_ma_p1 = trainee_ma_p[0].get('count', 0)
		count_trainee_ma_h1 = trainee_ma_h[0].get('count', 0)
		trainee_ma = count_trainee_ma_p1 + (count_trainee_ma_h1 / 2)

		co_ma_p = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s'  AND shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 AND status = 'Present' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
		co_ma_h = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s' AND shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 AND status = 'Half day' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
		count_co_ma_p1 = co_ma_p[0].get('count', 0)
		count_co_ma_h1 = co_ma_h[0].get('count', 0)
		co_ma = count_co_ma_p1 + (count_co_ma_h1 / 2)

		ap_ma_p = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s' AND shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Apprentice") AND docstatus != 2 AND status = 'Present' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
		ap_ma_h = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s' AND shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Apprentice") AND docstatus != 2 AND status = 'Half day' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
		count_ap_ma_p1 = ap_ma_p[0].get('count', 0)
		count_ap_ma_h1 = ap_ma_h[0].get('count', 0)
		ap_ma = count_ap_ma_p1 + (count_ap_ma_h1 / 2)

		con_ma_p = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s' AND shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Contractor") AND docstatus != 2 AND status = 'Present' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
		con_ma_h = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s' AND  shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Contractor") AND docstatus != 2 AND status = 'Half day' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
		count_con_ma_p1 = con_ma_p[0].get('count', 0)
		count_con_ma_h1 = con_ma_h[0].get('count', 0)
		con_ma = count_con_ma_p1 + (count_con_ma_h1 / 2)

		trainee_ma_ot = frappe.db.sql("""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s' AND  shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Trainee")AND docstatus != 2 AND status = 'Present'AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
		co_ma_ot = frappe.db.sql("""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s' AND  shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 AND status = 'Present'AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
		ap_ma_ot = frappe.db.sql("""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s' AND  shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Apprentice")AND docstatus != 2 AND status = 'Present'AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
		con_ma_ot = frappe.db.sql("""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s' AND  shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Contractor")AND docstatus != 2 AND status = 'Present'AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
		
		trainee_ot_value = trainee_ma_ot[0].get('ot', 0)
		if trainee_ot_value is not None:
			count_trainee_ma_ot = round(trainee_ot_value / 8, 2)
		else:
			count_trainee_ma_ot = 0.0

		co_ot_value = co_ma_ot[0].get('ot', 0)
		if co_ot_value is not None:
			count_co_ma_ot = round(co_ot_value / 8, 2)
		else:
			count_co_ma_ot = 0.0

		ap_ot_value = ap_ma_ot[0].get('ot', 0)
		if ap_ot_value is not None:
			count_ap_ma_ot = round(ap_ot_value / 8, 2)
		else:
			count_ap_ma_ot = 0.0

		con_ot_value = con_ma_ot[0].get('ot', 0)
		if con_ot_value is not None:
			count_con_ma_ot = round(con_ot_value / 8, 2)
		else:
			count_con_ma_ot = 0.0
		
		if args.get('shift'):
			start_date = datetime.strptime(args.get('from_date'), "%Y-%m-%d")
			end_date = datetime.strptime(args.get('to_date'), "%Y-%m-%d")
			month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
			start_month_name = month_names[start_date.month - 1]
			end_month_name = month_names[end_date.month - 1]
			start_month_number = start_date.month
			end_month_number = end_date.month
			if start_month_name==end_month_name:
				bp = frappe.db.sql("""
						SELECT 
							sum(business_plan) as plan
						FROM 
							`tabManpower Plan`
						WHERE 
							department = '%s' 
							AND month = '%s'
							AND year = '%s' 
					"""% (dept,start_month_name, start_date.year), as_dict=True)
			else:
				
				bp=frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = '%s' 
						AND FIELD(month, 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')
							between '%s' and '%s' 
						AND year = '%s' 
				"""% (dept,start_month_number,end_month_number, start_date.year), as_dict=True)
			if start_month_name == end_month_name:
				no_of_days = date_diff(add_days(args.get('to_date'), 1), args.get('from_date'))
				bp_value = bp[0]['plan'] if bp and bp[0]['plan'] else 0
				if bp_value>0:
					bp_value=bp_value*no_of_days
			else:
				bp1 = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = %s 
						AND month = %s
						AND year = %s
				""", (str(dept), str(start_month_name), int(start_date.year)), as_dict=True)
				bp_value1 = bp1[0]['plan'] if bp1 and bp1[0]['plan'] else 0
				bp2 = frappe.db.sql("""
						SELECT 
							sum(business_plan) as plan
						FROM 
							`tabManpower Plan`
						WHERE 
							department = '%s' 
							AND month = '%s'
							AND year = '%s' 
					"""% (str(dept),str(end_month_number), int(end_date.year)), as_dict=True)
				bp_value2 = bp2[0]['plan'] if bp2 and bp2[0]['plan'] else 0
				first_date=get_first_day(end_date)
				last_date=get_last_day(start_date)
				start_days = date_diff(add_days(last_date, 1), start_date)
				end_days = date_diff(add_days(end_date, 1), first_date)
				if bp_value1>0:
					bp_value1=bp_value1*start_days
				else:
					bp_value1=0
				if bp_value2>0:
					bp_value1=bp_value2*end_days
				else:
					bp_value2=0 
				bp_value=bp_value1+bp_value2
			if start_month_name==end_month_name:
				acp = frappe.get_value("Manpower Plan",{'department':dept,'month':end_month_number},['plan']) or 0
			else:
				acp = frappe.get_value("Manpower Plan",{'department':dept,'month':start_month_number,'year':start_date.year},['plan']) or 0
				acp_2=frappe.get_value("Manpower Plan",{'department':dept,'month':end_month_number,'year':start_date.year},['plan']) or 0
				acp=acp+acp_2

			trainee_ma_p = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s'  AND shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Trainee") AND docstatus != 2 AND status = 'Present' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
			trainee_ma_h = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s' AND shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Trainee") AND docstatus != 2 AND status = 'Half day' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
			count_trainee_ma_p1 = trainee_ma_p[0].get('count', 0)
			count_trainee_ma_h1 = trainee_ma_h[0].get('count', 0)
			trainee_ma = count_trainee_ma_p1 + (count_trainee_ma_h1 / 2)

			co_ma_p = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s'  AND shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 AND status = 'Present' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
			co_ma_h = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s' AND shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 AND status = 'Half day' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
			count_co_ma_p1 = co_ma_p[0].get('count', 0)
			count_co_ma_h1 = co_ma_h[0].get('count', 0)
			co_ma = count_co_ma_p1 + (count_co_ma_h1 / 2)

			ap_ma_p = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s' AND shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Apprentice") AND docstatus != 2 AND status = 'Present' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
			ap_ma_h = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s' AND shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Apprentice") AND docstatus != 2 AND status = 'Half day' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
			count_ap_ma_p1 = ap_ma_p[0].get('count', 0)
			count_ap_ma_h1 = ap_ma_h[0].get('count', 0)
			ap_ma = count_ap_ma_p1 + (count_ap_ma_h1 / 2)
			con_ma_p = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s' AND shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Contractor") AND docstatus != 2 AND status = 'Present' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
			con_ma_h = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s' AND  shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Contractor") AND docstatus != 2 AND status = 'Half day' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
			count_con_ma_p1 = con_ma_p[0].get('count', 0)
			count_con_ma_h1 = con_ma_h[0].get('count', 0)
			con_ma = count_con_ma_p1 + (count_con_ma_h1 / 2)
			trainee_ma_ot = frappe.db.sql("""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s' AND  shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Trainee") AND docstatus != 2 AND status = 'Present'AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
			co_ma_ot = frappe.db.sql("""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s' AND  shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 AND status = 'Present'AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
			ap_ma_ot = frappe.db.sql("""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s' AND  shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Apprentice")AND docstatus != 2 AND status = 'Present'AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
			con_ma_ot = frappe.db.sql("""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s' AND  shift ='%s' AND department = '%s'  AND custom_employee_category IN ("Contractor")AND docstatus != 2 AND status = 'Present'AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'),args.get('shift'), dept), as_dict=True)
			
			
			trainee_ot_value = trainee_ma_ot[0].get('ot', 0)
			if trainee_ot_value is not None:
				count_trainee_ma_ot = round(trainee_ot_value / 8, 2)
			else:
				count_trainee_ma_ot = 0.0
			co_ot_value = co_ma_ot[0].get('ot', 0)
			if co_ot_value is not None:
				count_co_ma_ot = round(co_ot_value / 8, 2)
			else:
				count_co_ma_ot = 0.0
			ap_ot_value = ap_ma_ot[0].get('ot', 0)
			if ap_ot_value is not None:
				count_ap_ma_ot = round(ap_ot_value / 8, 2)
			else:
				count_ap_ma_ot = 0.0
			con_ot_value = con_ma_ot[0].get('ot', 0)
			if con_ot_value is not None:
				count_con_ma_ot = round(con_ot_value / 8, 2)
			else:
				count_con_ma_ot = 0.0
		else:
			start_date = datetime.strptime(args.get('from_date'), "%Y-%m-%d")
			end_date = datetime.strptime(args.get('to_date'), "%Y-%m-%d")
			month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
			start_month_name = month_names[start_date.month - 1]
			end_month_name = month_names[end_date.month - 1]
			start_month_number = start_date.month
			end_month_number = end_date.month

			if start_month_name == end_month_name:
				bp = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = '%s' 
						AND month = '%s'
						AND year = '%s' 
				""" % (dept, start_month_name, start_date.year), as_dict=True)
			else:
				bp = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = '%s' 
						AND FIELD(month, 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')
						BETWEEN '%s' AND '%s' 
						AND year = '%s' 
				""" % (dept, start_month_number, end_month_number, start_date.year), as_dict=True)

			if start_month_name == end_month_name:
				no_of_days = date_diff(add_days(args.get('to_date'), 1), args.get('from_date'))
				bp_value = bp[0]['plan'] if bp and bp[0]['plan'] else 0
				if bp_value>0:
					bp_value=bp_value*no_of_days
			else:
				bp1 = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = %s 
						AND month = %s
						AND year = %s
				""", (str(dept), str(start_month_name), int(start_date.year)), as_dict=True)
				bp_value1 = bp1[0]['plan'] if bp1 and bp1[0]['plan'] else 0
				bp2 = frappe.db.sql("""
						SELECT 
							sum(business_plan) as plan
						FROM 
							`tabManpower Plan`
						WHERE 
							department = '%s' 
							AND month = '%s'
							AND year = '%s' 
					"""% (str(dept),str(end_month_number), int(end_date.year)), as_dict=True)
				bp_value2 = bp2[0]['plan'] if bp2 and bp2[0]['plan'] else 0
				first_date=get_first_day(end_date)
				last_date=get_last_day(start_date)
				start_days = date_diff(add_days(last_date, 1), start_date)
				end_days = date_diff(add_days(end_date, 1), first_date)
				if bp_value1>0:
					bp_value1=bp_value1*start_days
				else:
					bp_value1=0
				if bp_value2>0:
					bp_value1=bp_value2*end_days
				else:
					bp_value2=0 
				bp_value=bp_value1+bp_value2
			if start_month_name == end_month_name:
				acp = frappe.get_value("Manpower Plan", {'department': dept, 'month': end_month_name, 'year': start_date.year}, ['plan']) or 0
			else:
				acp = frappe.get_value("Manpower Plan", {'department': dept, 'month': start_month_number, 'year': start_date.year}, ['plan']) or 0
				acp_2 = frappe.get_value("Manpower Plan", {'department': dept, 'month': end_month_number, 'year': start_date.year}, ['plan']) or 0
				acp = acp + acp_2

			# Ensure you handle acp properly when assigning values
			
			trainee_ma_p = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s'   AND department = '%s'  AND custom_employee_category IN ("Trainee") AND docstatus != 2 AND status = 'Present' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'), dept), as_dict=True)
			trainee_ma_h = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s'  AND department = '%s'  AND custom_employee_category IN ("Trainee") AND docstatus != 2 AND status = 'Half day' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'), dept), as_dict=True)
			count_trainee_ma_p1 = trainee_ma_p[0].get('count',0)
			count_trainee_ma_h1 = trainee_ma_h[0].get('count',0)
			trainee_ma = count_trainee_ma_p1 + (count_trainee_ma_h1 / 2)

			co_ma_p = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s'   AND department = '%s'  AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 AND status = 'Present' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'), dept), as_dict=True)
			co_ma_h = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s'  AND department = '%s'  AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 AND status = 'Half day' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'), dept), as_dict=True)
			count_co_ma_p1 = co_ma_p[0].get('count',0)
			count_co_ma_h1 = co_ma_h[0].get('count',0)
			co_ma = count_co_ma_p1 + (count_co_ma_h1 / 2)
			ap_ma_p = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s'  AND department = '%s'  AND custom_employee_category IN ("Apprentice") AND docstatus != 2 AND status = 'Present' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'), dept), as_dict=True)
			ap_ma_h = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date  BETWEEN '%s' AND '%s'  AND department = '%s'  AND custom_employee_category IN ("Apprentice") AND docstatus != 2 AND status = 'Half day' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'), dept), as_dict=True)
			count_ap_ma_p1 = ap_ma_p[0].get('count',0)
			count_ap_ma_h1 = ap_ma_h[0].get('count',0)
			ap_ma = count_ap_ma_p1 + (count_ap_ma_h1 / 2)
			con_ma_p = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s'  AND department = '%s'  AND custom_employee_category IN ("Contractor") AND docstatus != 2 AND status = 'Present' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'), dept), as_dict=True)
			con_ma_h = frappe.db.sql("""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s'  AND department = '%s'  AND custom_employee_category IN ("Contractor") AND docstatus != 2 AND status = 'Half day' AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'), dept), as_dict=True)
			count_con_ma_p1 = con_ma_p[0].get('count', 0)
			count_con_ma_h1 = con_ma_h[0].get('count', 0)
			con_ma = count_con_ma_p1 + (count_con_ma_h1 / 2)

			
			trainee_ma_ot = frappe.db.sql("""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s'  AND department = '%s'  AND custom_employee_category IN ("Trainee") AND docstatus != 2 AND status = 'Present'AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'), dept), as_dict=True)
			co_ma_ot = frappe.db.sql("""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s'  AND department = '%s'  AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 AND status = 'Present'AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'), dept), as_dict=True)
			ap_ma_ot = frappe.db.sql("""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s'  AND department = '%s'  AND custom_employee_category IN ("Apprentice")AND docstatus != 2 AND status = 'Present'AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'), dept), as_dict=True)
			con_ma_ot = frappe.db.sql("""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN '%s' AND '%s'  AND department = '%s'  AND custom_employee_category IN ("Contractor")AND docstatus != 2 AND status = 'Present'AND attendance_request is NULL """ % (args.get('from_date'),args.get('to_date'), dept), as_dict=True)

			trainee_ot_value = trainee_ma_ot[0].get('ot', 0)
			if trainee_ot_value is not None:
				count_trainee_ma_ot = round(trainee_ot_value / 8, 2)
			else:
				count_trainee_ma_ot = 0.0
			
			co_ot_value = co_ma_ot[0].get('ot', 0)
			if co_ot_value is not None:
				count_co_ma_ot = round(co_ot_value / 8, 2)
			else:
				count_co_ma_ot = 0.0
			ap_ot_value = ap_ma_ot[0].get('ot', 0)
			if ap_ot_value is not None:
				count_ap_ma_ot = round(ap_ot_value / 8, 2)
			else:
				count_ap_ma_ot = 0.0
			con_ot_value = con_ma_ot[0].get('ot', 0)
			if con_ot_value is not None:
				count_con_ma_ot = round(con_ot_value / 8, 2)
			else:
				count_con_ma_ot = 0.0

				
		if bp_value>0:
			pp1=((trainee_ma+co_ma+ap_ma+con_ma/bp_value)*100)
			pp_2=(count_trainee_ma_ot+count_co_ma_ot+count_ap_ma_ot+count_con_ma_ot/bp_value)*100
			pp_3=(count_trainee_ma_ot+count_co_ma_ot+count_ap_ma_ot+count_con_ma_ot+trainee_ma+co_ma+ap_ma+con_ma/bp_value)*100
		else:
			pp1=0
			pp_2=0
			pp_3=0
		overall = (trainee_ma+co_ma+ap_ma+con_ma) + (count_trainee_ma_ot+count_co_ma_ot + count_ap_ma_ot+count_con_ma_ot)		
		row += [dept, 'Production', bp_value or 0, co_ma,trainee_ma,ap_ma,con_ma,(co_ma+ap_ma+con_ma+trainee_ma),f"{pp1:.2f}",count_co_ma_ot,count_trainee_ma_ot,count_ap_ma_ot,count_con_ma_ot,(count_trainee_ma_ot+count_co_ma_ot+count_ap_ma_ot+count_con_ma_ot),f"{pp_2:.2f}",((trainee_ma+co_ma+ap_ma+con_ma)+(count_trainee_ma_ot+count_co_ma_ot+count_ap_ma_ot+count_con_ma_ot)),f"{pp_3:.2f}"]
		data.append(row)

		tbp += bp_value
		tacp += acp
		ttrainee +=trainee_ma
		tco += co_ma
		tap += ap_ma
		tcon += con_ma 
		ttrainee_ot += count_trainee_ma_ot
		tco_ot += count_co_ma_ot
		tap_ot += count_ap_ma_ot
		tcon_ot += count_con_ma_ot
		tot += (trainee_ma+co_ma+ap_ma+con_ma)+(count_trainee_ma_ot+count_co_ma_ot+count_ap_ma_ot+count_con_ma_ot)
		i += 1
	if tbp>0:
		per_1=(ttrainee+tco + tap + tcon/tbp)*100
		per2=(ttrainee_ot+tco_ot + tap_ot + tcon_ot/tbp)*100
		per_tot=(tot/tbp)*100
	else:
		per_1=0
		per2=0
		per_tot=0
	
	subtotal_row = ["Sub Total (A)", "", "", tbp, tco,ttrainee, tap, tcon, tco + tap + tcon,per_1, tco_ot,ttrainee_ot, tap_ot, tcon_ot,
					tco_ot + tap_ot + tcon_ot,per2,tot,per_tot]
	data.append(subtotal_row)

	if tbp != 0:
		percentage_row = ["Percentage", "", "", "", "", "", "",
						f"{((ttrainee+tco + tap + tcon) / tbp) * 100:.2f} %",
						"", "", "",
						f"{((ttrainee_ot+tco_ot + tap_ot + tcon_ot) / tbp) * 100:.2f} %",
						"", "", "",
						f"{((ttrainee+tco + tap + tcon + ttrainee_ot + tco_ot + tap_ot + tcon_ot) / tbp) * 100:.2f} %"]
	else:
		percentage_row = ["Percentage", "",  "", "", "", "", "", "", "", "", "", "","",""]

	data.append(percentage_row)

	return data


@frappe.whitelist()
def title5(args):
	data = ["Supporting Team"]
	return data

@frappe.whitelist()
def title6(args):
	tot_bp=0
	department_data = []
	department = ['ASSEMBLY - WAIP','Packing - WAIP','Quality - WAIP', 'NPD - WAIP', 'M P L & Purchase - WAIP', 'ME -Regular - WAIP', 'Maintenance - WAIP','Management - WAIP', 'Delivery - WAIP', 'Finance - WAIP', 'HR - WAIP']
	for d in department:
		d_group = frappe.get_value("Department", {'name': d}, ['is_group'])
		if d_group == 1:
			dept = frappe.get_all("Department", {'parent_department': d}, ['*'])
			for i in dept:
				department_data.append(i.name)
		else:
			department_data.append(d)

	data = []
	row = []
	j = 1
	tbp = tacp = tco = tap = tcon = tco_ot = tap_ot = tcon_ot = tot = 0
	ttrainee =0
	ttrainee_ot =0
	for i in department_data:
		start_date = datetime.strptime(args.get('from_date'), "%Y-%m-%d")
		end_date = datetime.strptime(args.get('to_date'), "%Y-%m-%d")
		month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
		start_month_name = month_names[start_date.month - 1]
		end_month_name = month_names[end_date.month - 1]
		start_month_number = start_date.month
		end_month_number = end_date.month
		if start_month_name==end_month_name:
			bp = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = '%s' 
						AND month = '%s'
						AND year = '%s' 
				"""% (i,start_month_name, start_date.year), as_dict=True)
		else:
			bp=frappe.db.sql("""
				SELECT 
					sum(business_plan) as plan
				FROM 
					`tabManpower Plan`
				WHERE 
					department = '%s' 
					AND FIELD(month, 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')
						between '%s' and '%s' 
					AND year = '%s' 
			"""% (i,start_month_number,end_month_number, start_date.year), as_dict=True)
		if start_month_name == end_month_name:
			no_of_days = date_diff(add_days(args.get('to_date'), 1), args.get('from_date'))
			bp_value = bp[0]['plan'] if bp and bp[0]['plan'] else 0
			if bp_value>0:
				bp_value=bp_value*no_of_days
		else:
				bp1 = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = %s 
						AND month = %s
						AND year = %s
				""", (i, str(start_month_name), int(start_date.year)), as_dict=True)
				bp_value1 = bp1[0]['plan'] if bp1 and bp1[0]['plan'] else 0
				bp2 = frappe.db.sql("""
						SELECT 
							sum(business_plan) as plan
						FROM 
							`tabManpower Plan`
						WHERE 
							department = '%s' 
							AND month = '%s'
							AND year = '%s' 
					"""% (i,str(end_month_number), int(end_date.year)), as_dict=True)
				bp_value2 = bp2[0]['plan'] if bp2 and bp2[0]['plan'] else 0
				first_date=get_first_day(end_date)
				last_date=get_last_day(start_date)
				start_days = date_diff(add_days(last_date, 1), start_date)
				end_days = date_diff(add_days(end_date, 1), first_date)
				if bp_value1>0:
					bp_value1=bp_value1*start_days
				else:
					bp_value1=0
				if bp_value2>0:
					bp_value1=bp_value2*end_days
				else:
					bp_value2=0 
				bp_value=bp_value1+bp_value2
		# tot_bp+=bp_value
		if start_month_name == end_month_name:
			acp = frappe.db.get_value("Manpower Plan", {'department': i, 'month':end_month_name, 'year':start_date.year}, ['plan']) or 0
		else:
			acp = frappe.get_value("Manpower Plan", {'department': i, 'month': start_month_number, 'year': start_date.year}, ['plan']) or 0
			acp_2 = frappe.get_value("Manpower Plan", {'department': i, 'month': end_month_number, 'year': start_date.year}, ['plan']) or 0
			acp = acp + acp_2
		
		trainee_ma_p = frappe.db.sql("""
			SELECT count(*) as count 
			FROM `tabAttendance` 
			WHERE 
				attendance_date BETWEEN '%s' AND '%s' AND 
				department = '%s' AND 
				shift ='%s' AND
				custom_employee_category IN ("Trainee") AND 
				docstatus != 2 AND 
				status = 'Present' AND 
				attendance_request is NULL 
		""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
		
		trainee_ma_h = frappe.db.sql("""
			SELECT count(*) as count 
			FROM `tabAttendance` 
			WHERE 
				attendance_date  BETWEEN '%s' AND '%s' AND 
				department = '%s' AND 
				shift ='%s' AND		  
				custom_employee_category IN ("Trainee") AND 
				docstatus != 2 AND 
				status = 'Half day' AND 
				attendance_request is NULL 
		""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
		
		count_trainee_ma_p1 = trainee_ma_p[0].get('count', 0)
		count_trainee_ma_h1 = trainee_ma_h[0].get('count', 0)
		trainee_ma = count_trainee_ma_p1 + (count_trainee_ma_h1 / 2)
		
		co_ma_p = frappe.db.sql("""
			SELECT count(*) as count 
			FROM `tabAttendance` 
			WHERE 
				attendance_date BETWEEN '%s' AND '%s' AND 
				department = '%s' AND 
				shift ='%s' AND
				custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND 
				docstatus != 2 AND 
				status = 'Present' AND 
				attendance_request is NULL 
		""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
		
		co_ma_h = frappe.db.sql("""
			SELECT count(*) as count 
			FROM `tabAttendance` 
			WHERE 
				attendance_date  BETWEEN '%s' AND '%s' AND 
				department = '%s' AND 
				shift ='%s' AND		  
				custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND 
				docstatus != 2 AND 
				status = 'Half day' AND 
				attendance_request is NULL 
		""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
		
		count_co_ma_p1 = co_ma_p[0].get('count', 0)
		count_co_ma_h1 = co_ma_h[0].get('count', 0)
		co_ma = count_co_ma_p1 + (count_co_ma_h1 / 2)
		
		ap_ma_p = frappe.db.sql("""
			SELECT count(*) as count 
			FROM `tabAttendance` 
			WHERE 
				attendance_date  BETWEEN '%s' AND '%s' AND 
				department = '%s' AND 
				shift ='%s' AND
				custom_employee_category IN ("Apprentice")AND 
				docstatus != 2 AND 
				status = 'Present' AND 
				attendance_request is NULL 
		""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
		
		ap_ma_h = frappe.db.sql("""
			SELECT count(*) as count 
			FROM `tabAttendance` 
			WHERE 
				attendance_date  BETWEEN '%s' AND '%s' AND 
				department = '%s' AND 
				shift ='%s' AND
				custom_employee_category IN ("Apprentice")AND 
				docstatus != 2 AND 
				status = 'Half day' AND 
				attendance_request is NULL 
		""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
		
		count_ap_ma_p1 = ap_ma_p[0].get('count', 0)
		count_ap_ma_h1 = ap_ma_h[0].get('count', 0)
		ap_ma = count_ap_ma_p1 + (count_ap_ma_h1 / 2)
		
		con_ma_p = frappe.db.sql("""
			SELECT count(*) as count 
			FROM `tabAttendance` 
			WHERE 
				attendance_date  BETWEEN '%s' AND '%s' AND 
				department = '%s' AND
				shift ='%s' AND		    
				custom_employee_category IN ("Contractor") AND 
				docstatus != 2 AND 
				status = 'Present' AND 
				attendance_request is NULL 
		""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
		
		con_ma_h = frappe.db.sql("""
			SELECT count(*) as count 
			FROM `tabAttendance` 
			WHERE 
				attendance_date  BETWEEN '%s' AND '%s' AND 
				department = '%s' AND
				shift ='%s' AND		    
				custom_employee_category IN ("Contractor") AND 
				docstatus != 2 AND 
				status = 'Half day' AND 
				attendance_request is NULL 
		""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
		
		count_con_ma_p1 = con_ma_p[0].get('count', 0)
		count_con_ma_h1 = con_ma_h[0].get('count', 0)
		con_ma = count_con_ma_p1 + (count_con_ma_h1 / 2)
		
		trainee_ma_ot = frappe.db.sql("""
			SELECT sum(custom_overtime_hours) as ot 
			FROM `tabAttendance` 
			WHERE 
				attendance_date  BETWEEN '%s' AND '%s' AND 
				department = '%s' AND 
				shift ='%s' AND		   
				custom_employee_category IN ("Trainee") AND 
				docstatus != 2 AND 
				status = 'Present' AND 
				attendance_request is NULL 
		""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
		
		co_ma_ot = frappe.db.sql("""
			SELECT sum(custom_overtime_hours) as ot 
			FROM `tabAttendance` 
			WHERE 
				attendance_date  BETWEEN '%s' AND '%s' AND 
				department = '%s' AND 
				shift ='%s' AND		   
				custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND 
				docstatus != 2 AND 
				status = 'Present' AND 
				attendance_request is NULL 
		""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
		
		ap_ma_ot = frappe.db.sql("""
			SELECT sum(custom_overtime_hours) as ot 
			FROM `tabAttendance` 
			WHERE 
				attendance_date  BETWEEN '%s' AND '%s' AND 
				department = '%s' AND
				shift ='%s' AND		    
				custom_employee_category IN ("Apprentice") AND 
				docstatus != 2 AND 
				status = 'Present' AND 
				attendance_request is NULL 
		""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
		
		con_ma_ot = frappe.db.sql("""
			SELECT sum(custom_overtime_hours) as ot 
			FROM `tabAttendance` 
			WHERE 
				attendance_date  BETWEEN '%s' AND '%s' AND 
				department = '%s' AND 
				shift ='%s' AND			
				custom_employee_category IN ("Contractor") AND 
				docstatus != 2 AND 
				status = 'Present' AND 
				attendance_request is NULL 
		""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)

		
		trainee_ot_value = trainee_ma_ot[0].get('ot', 0)
		count_trainee_ma_ot = round(trainee_ot_value / 8, 2) if trainee_ot_value is not None else 0.0

		co_ot_value = co_ma_ot[0].get('ot', 0)
		count_co_ma_ot = round(co_ot_value / 8, 2) if co_ot_value is not None else 0.0
		
		ap_ot_value = ap_ma_ot[0].get('ot', 0)
		count_ap_ma_ot = round(ap_ot_value / 8, 2) if ap_ot_value is not None else 0.0
		
		con_ot_value = con_ma_ot[0].get('ot', 0)
		count_con_ma_ot = round(con_ot_value / 8, 2) if con_ot_value is not None else 0.0

		if args.get('shift'):
			start_date = datetime.strptime(args.get('from_date'), "%Y-%m-%d")
			end_date = datetime.strptime(args.get('to_date'), "%Y-%m-%d")
			month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
			start_month_name = month_names[start_date.month - 1]
			end_month_name = month_names[end_date.month - 1]
			start_month_number = start_date.month
			end_month_number = end_date.month
			if start_month_name==end_month_name:
				bp = frappe.db.sql("""
						SELECT 
							sum(business_plan) as plan
						FROM 
							`tabManpower Plan`
						WHERE 
							department = '%s' 
							AND month = '%s'
							AND year = '%s' 
					"""% (i,start_month_name, start_date.year), as_dict=True)
			else:
				bp=frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = '%s' 
						AND FIELD(month, 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')
							between '%s' and '%s' 
						AND year = '%s' 
				"""% (i,start_month_number,end_month_number, start_date.year), as_dict=True)
			if start_month_name == end_month_name:
				no_of_days = date_diff(add_days(args.get('to_date'), 1), args.get('from_date'))
				bp_value = bp[0]['plan'] if bp and bp[0]['plan'] else 0
				if bp_value>0:
					bp_value=bp_value*no_of_days
			else:
				bp1 = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = %s 
						AND month = %s
						AND year = %s
				""", (i, str(start_month_name), int(start_date.year)), as_dict=True)
				bp_value1 = bp1[0]['plan'] if bp1 and bp1[0]['plan'] else 0
				bp2 = frappe.db.sql("""
						SELECT 
							sum(business_plan) as plan
						FROM 
							`tabManpower Plan`
						WHERE 
							department = '%s' 
							AND month = '%s'
							AND year = '%s' 
					"""% (i,str(end_month_number), int(end_date.year)), as_dict=True)
				bp_value2 = bp2[0]['plan'] if bp2 and bp2[0]['plan'] else 0
				first_date=get_first_day(end_date)
				last_date=get_last_day(start_date)
				start_days = date_diff(add_days(last_date, 1), start_date)
				end_days = date_diff(add_days(end_date, 1), first_date)
				if bp_value1>0:
					bp_value1=bp_value1*start_days
				else:
					bp_value1=0
				if bp_value2>0:
					bp_value1=bp_value2*end_days
				else:
					bp_value2=0 
				bp_value=bp_value1+bp_value2
			tot_bp+=bp_value
			if start_month_name==end_month_name:
				acp = frappe.get_value("Manpower Plan",{'department':i,'month':end_month_number},['plan']) or 0
			else:
				acp = frappe.get_value("Manpower Plan",{'department':i,'month':start_month_number,'year':start_date.year},['plan']) or 0
				acp_2=frappe.get_value("Manpower Plan",{'department':i,'month':end_month_number,'year':start_date.year},['plan']) or 0
				acp=acp+acp_2
			
			trainee_ma_p = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					shift ='%s' AND
					custom_employee_category IN ("Trainee") AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
			
			trainee_ma_h = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					shift ='%s' AND		  
					custom_employee_category IN ("Trainee") AND 
					docstatus != 2 AND 
					status = 'Half day' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
			
			count_trainee_ma_p1 = trainee_ma_p[0].get('count', 0)
			count_trainee_ma_h1 = trainee_ma_h[0].get('count', 0)
			trainee_ma = count_trainee_ma_p1 + (count_trainee_ma_h1 / 2)

			co_ma_p = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					shift ='%s' AND
					custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
			
			co_ma_h = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					shift ='%s' AND		  
					custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND 
					docstatus != 2 AND 
					status = 'Half day' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
			
			count_co_ma_p1 = co_ma_p[0].get('count', 0)
			count_co_ma_h1 = co_ma_h[0].get('count', 0)
			co_ma = count_co_ma_p1 + (count_co_ma_h1 / 2)
			
			ap_ma_p = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					shift ='%s' AND
					custom_employee_category IN ("Apprentice")AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
			
			ap_ma_h = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					shift ='%s' AND
					custom_employee_category IN ("Apprentice")AND 
					docstatus != 2 AND 
					status = 'Half day' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
			
			count_ap_ma_p1 = ap_ma_p[0].get('count', 0)
			count_ap_ma_h1 = ap_ma_h[0].get('count', 0)
			ap_ma = count_ap_ma_p1 + (count_ap_ma_h1 / 2)
			
			con_ma_p = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND
					shift ='%s' AND		    
					custom_employee_category IN ("Contractor") AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
			
			con_ma_h = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND
					shift ='%s' AND		    
					custom_employee_category IN ("Contractor") AND 
					docstatus != 2 AND 
					status = 'Half day' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
			
			count_con_ma_p1 = con_ma_p[0].get('count', 0)
			count_con_ma_h1 = con_ma_h[0].get('count', 0)
			con_ma = count_con_ma_p1 + (count_con_ma_h1 / 2)
			
			trainee_ma_ot = frappe.db.sql("""
				SELECT sum(custom_overtime_hours) as ot 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					shift ='%s' AND		   
					custom_employee_category IN ("Trainee") AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
			co_ma_ot = frappe.db.sql("""
				SELECT sum(custom_overtime_hours) as ot 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					shift ='%s' AND		   
					custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
			
			ap_ma_ot = frappe.db.sql("""
				SELECT sum(custom_overtime_hours) as ot 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND
					shift ='%s' AND		    
					custom_employee_category IN ("Apprentice") AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)
			
			con_ma_ot = frappe.db.sql("""
				SELECT sum(custom_overtime_hours) as ot 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					shift ='%s' AND			
					custom_employee_category IN ("Contractor") AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i,args.get('shift')), as_dict=True)

			trainee_ot_value = trainee_ma_ot[0].get('ot', 0)
			count_trainee_ma_ot = round(trainee_ot_value / 8, 2) if trainee_ot_value is not None else 0.0
			
			co_ot_value = co_ma_ot[0].get('ot', 0)
			count_co_ma_ot = round(co_ot_value / 8, 2) if co_ot_value is not None else 0.0
			
			ap_ot_value = ap_ma_ot[0].get('ot', 0)
			count_ap_ma_ot = round(ap_ot_value / 8, 2) if ap_ot_value is not None else 0.0
			
			con_ot_value = con_ma_ot[0].get('ot', 0)
			count_con_ma_ot = round(con_ot_value / 8, 2) if con_ot_value is not None else 0.0
				
		else:
			start_date = datetime.strptime(args.get('from_date'), "%Y-%m-%d")
			end_date = datetime.strptime(args.get('to_date'), "%Y-%m-%d")
			month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
			start_month_name = month_names[start_date.month - 1]
			end_month_name = month_names[end_date.month - 1]
			start_month_number = start_date.month
			end_month_number = end_date.month
			if start_month_name==end_month_name:
				bp = frappe.db.sql("""
						SELECT 
							sum(business_plan) as plan
						FROM 
							`tabManpower Plan`
						WHERE 
							department = '%s' 
							AND month = '%s'
							AND year = '%s' 
					"""% (i,start_month_name, start_date.year), as_dict=True)
			else:
				bp=frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = '%s' 
						AND FIELD(month, 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')
							between '%s' and '%s' 
						AND year = '%s' 
				"""% (i,start_month_number,end_month_number, start_date.year), as_dict=True)
			if start_month_name == end_month_name:
				no_of_days = date_diff(add_days(args.get('to_date'), 1), args.get('from_date'))
				bp_value = bp[0]['plan'] if bp and bp[0]['plan'] else 0
				if bp_value>0:
					bp_value=bp_value*no_of_days
			else:
				bp1 = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = %s 
						AND month = %s
						AND year = %s
				""", (i, str(start_month_name), int(start_date.year)), as_dict=True)
				bp_value1 = bp1[0]['plan'] if bp1 and bp1[0]['plan'] else 0
				bp2 = frappe.db.sql("""
						SELECT 
							sum(business_plan) as plan
						FROM 
							`tabManpower Plan`
						WHERE 
							department = '%s' 
							AND month = '%s'
							AND year = '%s' 
					"""% (i,str(end_month_number), int(end_date.year)), as_dict=True)
				bp_value2 = bp2[0]['plan'] if bp2 and bp2[0]['plan'] else 0
				first_date=get_first_day(end_date)
				last_date=get_last_day(start_date)
				start_days = date_diff(add_days(last_date, 1), start_date)
				end_days = date_diff(add_days(end_date, 1), first_date)
				if bp_value1>0:
					bp_value1=bp_value1*start_days
				else:
					bp_value1=0
				if bp_value2>0:
					bp_value1=bp_value2*end_days
				else:
					bp_value2=0 
				bp_value=bp_value1+bp_value2
			tot_bp+=bp_value
			if start_month_name==end_month_name:
				acp = frappe.get_value("Manpower Plan",{'department':i,'month':end_month_number},['plan']) or 0
			else:
				acp = frappe.get_value("Manpower Plan",{'department':i,'month':start_month_number,'year':start_date.year},['plan']) or 0
				acp_2=frappe.get_value("Manpower Plan",{'department':i,'month':end_month_number,'year':start_date.year},['plan']) or 0
				acp=acp+acp_2

			trainee_ma_p = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					custom_employee_category IN ("Trainee") AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i), as_dict=True)
			
			trainee_ma_h = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					custom_employee_category IN ("Trainee") AND 
					docstatus != 2 AND 
					status = 'Half day' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i), as_dict=True)
			
			count_trainee_ma_p1 = trainee_ma_p[0].get('count', 0)
			count_trainee_ma_h1 = trainee_ma_h[0].get('count', 0)
			trainee_ma = count_trainee_ma_p1 + (count_trainee_ma_h1 / 2)
		
			co_ma_p = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i), as_dict=True)
			
			co_ma_h = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND 
					docstatus != 2 AND 
					status = 'Half day' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i), as_dict=True)
			
			count_co_ma_p1 = co_ma_p[0].get('count', 0)
			count_co_ma_h1 = co_ma_h[0].get('count', 0)
			co_ma = count_co_ma_p1 + (count_co_ma_h1 / 2)
			
			ap_ma_p = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					custom_employee_category IN ("Apprentice")AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i), as_dict=True)
			
			ap_ma_h = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					custom_employee_category IN ("Apprentice")AND 
					docstatus != 2 AND 
					status = 'Half day' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i), as_dict=True)
			
			count_ap_ma_p1 = ap_ma_p[0].get('count', 0)
			count_ap_ma_h1 = ap_ma_h[0].get('count', 0)
			ap_ma = count_ap_ma_p1 + (count_ap_ma_h1 / 2)
			
			con_ma_p = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND
					custom_employee_category IN ("Contractor") AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i), as_dict=True)
			
			con_ma_h = frappe.db.sql("""
				SELECT count(*) as count 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND
					custom_employee_category IN ("Contractor") AND 
					docstatus != 2 AND 
					status = 'Half day' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i), as_dict=True)
			
			count_con_ma_p1 = con_ma_p[0].get('count', 0)
			count_con_ma_h1 = con_ma_h[0].get('count', 0)
			con_ma = count_con_ma_p1 + (count_con_ma_h1 / 2)
			
			
			trainee_ma_ot = frappe.db.sql("""
				SELECT sum(custom_overtime_hours) as ot 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					custom_employee_category IN ("Trainee") AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i), as_dict=True)
			co_ma_ot = frappe.db.sql("""
				SELECT sum(custom_overtime_hours) as ot 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i), as_dict=True)
			
			ap_ma_ot = frappe.db.sql("""
				SELECT sum(custom_overtime_hours) as ot 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND
					custom_employee_category IN ("Apprentice") AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i), as_dict=True)
			
			con_ma_ot = frappe.db.sql("""
				SELECT sum(custom_overtime_hours) as ot 
				FROM `tabAttendance` 
				WHERE 
					attendance_date  BETWEEN '%s' AND '%s' AND 
					department = '%s' AND 
					custom_employee_category IN ("Contractor") AND 
					docstatus != 2 AND 
					status = 'Present' AND 
					attendance_request is NULL 
			""" % (args.get('from_date'),args.get('to_date'), i), as_dict=True)

			
			trainee_ot_value = trainee_ma_ot[0].get('ot', 0)
			count_trainee_ma_ot = round(trainee_ot_value / 8, 2) if trainee_ot_value is not None else 0.0
			
			co_ot_value = co_ma_ot[0].get('ot', 0)
			count_co_ma_ot = round(co_ot_value / 8, 2) if co_ot_value is not None else 0.0
			
			ap_ot_value = ap_ma_ot[0].get('ot', 0)
			count_ap_ma_ot = round(ap_ot_value / 8, 2) if ap_ot_value is not None else 0.0
			
			con_ot_value = con_ma_ot[0].get('ot', 0)
			count_con_ma_ot = round(con_ot_value / 8, 2) if con_ot_value is not None else 0.0

		overall = (trainee_ma+co_ma+ap_ma+con_ma) + (count_trainee_ma_ot+count_co_ma_ot + count_ap_ma_ot+count_con_ma_ot)	
		if bp_value>0:
			per=(((trainee_ma+co_ma+ap_ma+con_ma)/bp_value)*100)
		else:
			per=0
		if bp_value>0:
			per1=((count_co_ma_ot+count_ap_ma_ot+count_con_ma_ot)/bp_value)*100
		else:
			per1=0
		if bp_value >0:
			per3=((count_co_ma_ot+count_ap_ma_ot+count_con_ma_ot+trainee_ma+co_ma+ap_ma+con_ma)/bp_value)*100
		else:
			per3=0
		row = [j, i, i, bp_value or 0, co_ma,trainee_ma,ap_ma,con_ma,(trainee_ma+co_ma+ap_ma+con_ma),f"{per:.2f}",count_co_ma_ot,count_trainee_ma_ot,count_ap_ma_ot,count_con_ma_ot,(count_trainee_ma_ot+count_co_ma_ot+count_ap_ma_ot+count_con_ma_ot),f"{per1:.2f}",((trainee_ma+co_ma+ap_ma+con_ma)+(count_trainee_ma_ot+count_co_ma_ot+count_ap_ma_ot+count_con_ma_ot)),f"{per3:.2f}"]
		data.append(row)
		# trainee_ma_p
		# trainee_ma_h
		# count_trainee_ma_p1
		# count_trainee_ma_h1
		# trainee_ma
		# trainee_ma_ot
		# trainee_ot_value
		# count_trainee_ma_ot
		# ttrainee
		# ttrainee_ot	
		j += 1
		tbp += bp_value
		tacp += acp
		ttrainee += trainee_ma
		tco += co_ma
		tap += ap_ma
		tcon += con_ma
		ttrainee_ot += count_trainee_ma_ot
		tco_ot += count_co_ma_ot
		tap_ot += count_ap_ma_ot
		tcon_ot += count_con_ma_ot
		tot += ((trainee_ma+co_ma + ap_ma + con_ma)) + ((count_trainee_ma_ot+count_co_ma_ot + count_ap_ma_ot + count_con_ma_ot))
	if tbp > 0:
		pp1=((ttrainee+tco + tap + tcon)/tbp)*100
		# pp1=f"{pp1:.2f} %"
	else:
		pp1=0
	if tbp > 0:
		pp2=((ttrainee_ot+tco_ot + tap_ot + tcon_ot)/tbp)*100
		# pp2=f"{pp2:.2f} %"
	else:
		pp2=0
	if tbp > 0:
		pp3=((ttrainee_ot +tco_ot + tap_ot + tcon_ot+ttrainee+tco + tap + tcon)/tbp)*100
		# pp3=f"{pp3:.2f}"
	else:
		pp3=0
	row1 = ["Sub Total (B)", " ", " ", tbp, tco,ttrainee, tap, tcon, (ttrainee+tco + tap + tcon), f"{pp1:.2f}",tco_ot,ttrainee_ot, tap_ot, tcon_ot, (ttrainee_ot+tco_ot + tap_ot + tcon_ot),f"{pp2:.2f}",tot,f"{pp3:.2f}"]

	if tbp != 0:
		if tacp > 0:
			p1=(((ttrainee+tco + tap + tcon) / tbp) * 100)
			p1=f"{p1:.2f} %"
		else:
			p1=0
		if tacp > 0:
			p2=(((ttrainee_ot +tco_ot + tap_ot + tcon_ot) / tbp) * 100)
			p2=f"{p2:.2f} %"
		else:
			p2=0
		if tacp > 0:
			p3=(((ttrainee_ot + tco_ot + tap_ot + tcon_ot+ttrainee+tco + tap + tcon) / tbp) * 100)
			p3=f"{p3:.2f} %"
		else:
			p3=0
		row2 = ["Percentage", " ",  " ", " ", " ", " ", p1, '', " ", " ", " ", p2,'',p3]
	else:
		row2 = ["Percentage", " ",  " ", " ", " ", " ", " ", " ", " ", " ", " ", " ", " ",'',"",""]

	data.append(row1)
	data.append(row2)

	departments = frappe.get_all('Department', {'parent_department': 'Manufacturing - WAIP'}, order_by='name ASC')
	department_names = [dept['name'] for dept in departments]
	total_business_plan = tbp
	total_actual_plan = tacp
	total_staff_count = tco
	total_apprentice_count = tap
	total_contractor_count = tcon
	total_staff_overtime_hours = tco_ot
	total_apprentice_overtime_hours = tap_ot
	total_contractor_overtime_hours = tcon_ot

	for dept in department_names:
		start_date = datetime.strptime(args.get('from_date'), "%Y-%m-%d")
		end_date = datetime.strptime(args.get('to_date'), "%Y-%m-%d")
		month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
		start_month_name = month_names[start_date.month - 1]
		end_month_name = month_names[end_date.month - 1]
		start_month_number = start_date.month
		end_month_number = end_date.month
		if start_month_name==end_month_name:
			bp = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = '%s' 
						AND month = '%s'
						AND year = '%s' 
				"""% (dept,start_month_name, start_date.year), as_dict=True)
		else:
			bp=frappe.db.sql("""
				SELECT 
					sum(business_plan) as plan
				FROM 
					`tabManpower Plan`
				WHERE 
					department = '%s' 
					AND FIELD(month, 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')
						between '%s' and '%s' 
					AND year = '%s' 
			"""% (dept,start_month_number,end_month_number, start_date.year), as_dict=True)
		if start_month_name == end_month_name:
			no_of_days = date_diff(add_days(args.get('to_date'), 1), args.get('from_date'))
			bp_value = bp[0]['plan'] if bp and bp[0]['plan'] else 0
			if bp_value>0:
				bp_value=bp_value*no_of_days
		else:
			bp1 = frappe.db.sql("""
				SELECT 
					sum(business_plan) as plan
				FROM 
					`tabManpower Plan`
				WHERE 
					department = %s 
					AND month = %s
					AND year = %s
			""", (str(dept), str(start_month_name), int(start_date.year)), as_dict=True)
			bp_value1 = bp1[0]['plan'] if bp1 and bp1[0]['plan'] else 0
			bp2 = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = '%s' 
						AND month = '%s'
						AND year = '%s' 
				"""% (str(dept),str(end_month_number), int(end_date.year)), as_dict=True)
			bp_value2 = bp2[0]['plan'] if bp2 and bp2[0]['plan'] else 0
			first_date=get_first_day(end_date)
			last_date=get_last_day(start_date)
			start_days = date_diff(add_days(last_date, 1), start_date)
			end_days = date_diff(add_days(end_date, 1), first_date)
			if bp_value1>0:
				bp_value1=bp_value1*start_days
			else:
				bp_value1=0
			if bp_value2>0:
				bp_value1=bp_value2*end_days
			else:
				bp_value2=0 
			bp_value=bp_value1+bp_value2
		# tot_bp+=bp_value
		if start_month_name==end_month_name:
			actual_plan = frappe.get_value("Manpower Plan", {'department': dept, 'month': start_month_number,'year':start_date.year}, ['plan']) or 0
		else:
			actual_plan = frappe.get_value("Manpower Plan", {'department': dept, 'month': start_month_number,'year':start_date.year}, ['plan']) or 0
			actual_plan_next = frappe.get_value("Manpower Plan", {'department': dept, 'month': end_month_number,'year':start_date.year}, ['plan']) or 0
			actual_plan=actual_plan+actual_plan_next
		staff_present_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance`  WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s AND department = %(department)s  
												AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 
												AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift'),'department': dept}, as_dict=True)
		staff_half_day_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s AND department = %(department)s  
												AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 
												AND status = 'Half day' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift'),'department': dept}, as_dict=True)
		staff_present_count = staff_present_query[0].get('count', 0)
		staff_half_day_count = staff_half_day_query[0].get('count', 0)
		staff_count = staff_present_count + (staff_half_day_count / 2)

		apprentice_present_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
													AND custom_employee_category IN ("Apprentice") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		apprentice_half_day_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
													AND custom_employee_category IN ("Apprentice") AND docstatus != 2 
													AND status = 'Half day' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		apprentice_present_count = apprentice_present_query[0].get('count', 0)
		apprentice_half_day_count = apprentice_half_day_query[0].get('count', 0)
		apprentice_count = apprentice_present_count + (apprentice_half_day_count / 2)

		contractor_present_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance`  WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
													AND custom_employee_category IN ("Contractor") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		contractor_half_day_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
													AND custom_employee_category IN ("Contractor") AND docstatus != 2 
													AND status = 'Half day' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		contractor_present_count = contractor_present_query[0].get('count', 0)
		contractor_half_day_count = contractor_half_day_query[0].get('count', 0)
		contractor_count = contractor_present_count + (contractor_half_day_count / 2)

		staff_overtime_query = frappe.db.sql(f"""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
												AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 
												AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		apprentice_overtime_query = frappe.db.sql(f"""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
													AND custom_employee_category IN ("Apprentice") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		contractor_overtime_query = frappe.db.sql(f"""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
													AND custom_employee_category IN ("Contractor") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		staff_overtime_hours = staff_overtime_query[0].get('ot', 0) / 8 if staff_overtime_query[0].get('ot', 0) else 0
		apprentice_overtime_hours = apprentice_overtime_query[0].get('ot', 0) / 8 if apprentice_overtime_query[0].get('ot', 0) else 0
		contractor_overtime_hours = contractor_overtime_query[0].get('ot', 0) / 8 if contractor_overtime_query[0].get('ot', 0) else 0

		total_business_plan += bp_value
		total_actual_plan += actual_plan
		total_staff_count += staff_count
		total_apprentice_count += apprentice_count
		total_contractor_count += contractor_count
		total_staff_overtime_hours += staff_overtime_hours
		total_apprentice_overtime_hours += apprentice_overtime_hours
		total_contractor_overtime_hours += contractor_overtime_hours

		if total_actual_plan != 0:
			staff_percentage = ((total_staff_count) / total_actual_plan) * 100
			apprentice_percentage = ((total_apprentice_count) / total_actual_plan) * 100
			contractor_percentage = ((total_contractor_count) / total_actual_plan) * 100
		else:
			staff_percentage = 0
			apprentice_percentage = 0
			contractor_percentage = 0

	if args.get('shift'): 

		start_date = datetime.strptime(args.get('from_date'), "%Y-%m-%d")
		end_date = datetime.strptime(args.get('to_date'), "%Y-%m-%d")
		month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
		start_month_name = month_names[start_date.month - 1]
		end_month_name = month_names[end_date.month - 1]
		start_month_number = start_date.month
		end_month_number = end_date.month
		if start_month_name==end_month_name:
			bp = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = '%s' 
						AND month = '%s'
						AND year = '%s' 
				"""% (dept,start_month_name, start_date.year), as_dict=True)
		else:
			bp=frappe.db.sql("""
				SELECT 
					sum(business_plan) as plan
				FROM 
					`tabManpower Plan`
				WHERE 
					department = '%s' 
					AND FIELD(month, 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')
						between '%s' and '%s' 
					AND year = '%s' 
			"""% (dept,start_month_number,end_month_number, start_date.year), as_dict=True)
		if start_month_name == end_month_name:
			no_of_days = date_diff(add_days(args.get('to_date'), 1), args.get('from_date'))
			bp_value = bp[0]['plan'] if bp and bp[0]['plan'] else 0
			if bp_value>0:
				bp_value=bp_value*no_of_days
		else:
			bp1 = frappe.db.sql("""
				SELECT 
					sum(business_plan) as plan
				FROM 
					`tabManpower Plan`
				WHERE 
					department = %s 
					AND month = %s
					AND year = %s
			""", (str(dept), str(start_month_name), int(start_date.year)), as_dict=True)
			bp_value1 = bp1[0]['plan'] if bp1 and bp1[0]['plan'] else 0
			bp2 = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = '%s' 
						AND month = '%s'
						AND year = '%s' 
				"""% (str(dept),str(end_month_number), int(end_date.year)), as_dict=True)
			bp_value2 = bp2[0]['plan'] if bp2 and bp2[0]['plan'] else 0
			first_date=get_first_day(end_date)
			last_date=get_last_day(start_date)
			start_days = date_diff(add_days(last_date, 1), start_date)
			end_days = date_diff(add_days(end_date, 1), first_date)
			if bp_value1>0:
				bp_value1=bp_value1*start_days
			else:
				bp_value1=0
			if bp_value2>0:
				bp_value1=bp_value2*end_days
			else:
				bp_value2=0 
			bp_value=bp_value1+bp_value2
		tot_bp+=bp_value
		if start_month_name==end_month_name:
			actual_plan = frappe.get_value("Manpower Plan", {'department': dept, 'month': start_month_number,'year':start_date.year}, ['plan']) or 0
		else:
			actual_plan = frappe.get_value("Manpower Plan", {'department': dept, 'month': start_month_number,'year':start_date.year}, ['plan']) or 0
			actual_plan_next = frappe.get_value("Manpower Plan", {'department': dept, 'month': end_month_number,'year':start_date.year}, ['plan']) or 0
			actual_plan=actual_plan+actual_plan_next
			
		staff_present_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance`  WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
												AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 
												AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		staff_half_day_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
												AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 
												AND status = 'Half day' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		staff_present_count = staff_present_query[0].get('count', 0)
		staff_half_day_count = staff_half_day_query[0].get('count', 0)
		staff_count = staff_present_count + (staff_half_day_count / 2)

		apprentice_present_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
													AND custom_employee_category IN ("Apprentice") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		apprentice_half_day_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
													AND custom_employee_category IN ("Apprentice") AND docstatus != 2 
													AND status = 'Half day' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		apprentice_present_count = apprentice_present_query[0].get('count', 0)
		apprentice_half_day_count = apprentice_half_day_query[0].get('count', 0)
		apprentice_count = apprentice_present_count + (apprentice_half_day_count / 2)

		contractor_present_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance`  WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
													AND custom_employee_category IN ("Contractor") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		contractor_half_day_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
													AND custom_employee_category IN ("Contractor") AND docstatus != 2 
													AND status = 'Half day' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		contractor_present_count = contractor_present_query[0].get('count', 0)
		contractor_half_day_count = contractor_half_day_query[0].get('count', 0)
		contractor_count = contractor_present_count + (contractor_half_day_count / 2)

		staff_overtime_query = frappe.db.sql(f"""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
												AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 
												AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		apprentice_overtime_query = frappe.db.sql(f"""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
													AND custom_employee_category IN ("Apprentice") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		contractor_overtime_query = frappe.db.sql(f"""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s  AND shift = %(shift)s  
													AND custom_employee_category IN ("Contractor") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date'),'shift': args.get('shift')}, as_dict=True)
		staff_overtime_hours = staff_overtime_query[0].get('ot', 0) / 8 if staff_overtime_query[0].get('ot', 0) else 0
		apprentice_overtime_hours = apprentice_overtime_query[0].get('ot', 0) / 8 if apprentice_overtime_query[0].get('ot', 0) else 0
		contractor_overtime_hours = contractor_overtime_query[0].get('ot', 0) / 8 if contractor_overtime_query[0].get('ot', 0) else 0

		total_business_plan = bp_value
		total_actual_plan = actual_plan
		total_staff_count = staff_count
		total_apprentice_count = apprentice_count
		total_contractor_count = contractor_count
		total_staff_overtime_hours = staff_overtime_hours
		total_apprentice_overtime_hours = apprentice_overtime_hours
		total_contractor_overtime_hours = contractor_overtime_hours

		if total_actual_plan != 0:
			staff_percentage = ((total_staff_count) / total_actual_plan) * 100
			apprentice_percentage = ((total_apprentice_count) / total_actual_plan) * 100
			contractor_percentage = ((total_contractor_count) / total_actual_plan) * 100
		else:
			staff_percentage = 0
			apprentice_percentage = 0
			contractor_percentage = 0
	else:
		start_date = datetime.strptime(args.get('from_date'), "%Y-%m-%d")
		end_date = datetime.strptime(args.get('to_date'), "%Y-%m-%d")
		month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
		start_month_name = month_names[start_date.month - 1]
		end_month_name = month_names[end_date.month - 1]
		start_month_number = start_date.month
		end_month_number = end_date.month
		if start_month_name==end_month_name:
			bp = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = '%s' 
						AND month = '%s'
						AND year = '%s' 
				"""% (dept,start_month_name, start_date.year), as_dict=True)
		else:
			bp=frappe.db.sql("""
				SELECT 
					sum(business_plan) as plan
				FROM 
					`tabManpower Plan`
				WHERE 
					department = '%s' 
					AND FIELD(month, 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')
						between '%s' and '%s' 
					AND year = '%s' 
			"""% (dept,start_month_number,end_month_number, end_date.year), as_dict=True)
		if start_month_name == end_month_name:
			no_of_days = date_diff(add_days(args.get('to_date'), 1), args.get('from_date'))
			bp_value = bp[0]['plan'] if bp and bp[0]['plan'] else 0
			if bp_value>0:
				bp_value=bp_value*no_of_days
		else:
			bp1 = frappe.db.sql("""
				SELECT 
					sum(business_plan) as plan
				FROM 
					`tabManpower Plan`
				WHERE 
					department = %s 
					AND month = %s
					AND year = %s
			""", (str(dept), str(start_month_name), int(start_date.year)), as_dict=True)
			bp_value1 = bp1[0]['plan'] if bp1 and bp1[0]['plan'] else 0
			bp2 = frappe.db.sql("""
					SELECT 
						sum(business_plan) as plan
					FROM 
						`tabManpower Plan`
					WHERE 
						department = '%s' 
						AND month = '%s'
						AND year = '%s' 
				"""% (str(dept),str(end_month_number), int(end_date.year)), as_dict=True)
			bp_value2 = bp2[0]['plan'] if bp2 and bp2[0]['plan'] else 0
			first_date=get_first_day(end_date)
			last_date=get_last_day(start_date)
			start_days = date_diff(add_days(last_date, 1), start_date)
			end_days = date_diff(add_days(end_date, 1), first_date)
			if bp_value1>0:
				bp_value1=bp_value1*start_days
			else:
				bp_value1=0
			if bp_value2>0:
				bp_value1=bp_value2*end_days
			else:
				bp_value2=0 
			bp_value=bp_value1+bp_value2
		tot_bp+=bp_value
		if start_month_name==end_month_name:
			actual_plan = frappe.get_value("Manpower Plan", {'department': dept, 'month': start_month_number,'year':start_date.year}, ['plan']) or 0
		else:
			actual_plan = frappe.get_value("Manpower Plan", {'department': dept, 'month': start_month_number,'year':start_date.year}, ['plan']) or 0
			actual_plan_next = frappe.get_value("Manpower Plan", {'department': dept, 'month': end_month_number,'year':start_date.year}, ['plan']) or 0
			actual_plan=actual_plan+actual_plan_next
			
		staff_present_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance`  WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s   
												AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 
												AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date')}, as_dict=True)
		staff_half_day_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s   
												AND custom_employee_category IN ("Staff", "Sub Staff", "Operator") AND docstatus != 2 
												AND status = 'Half day' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date')}, as_dict=True)
		staff_present_count = staff_present_query[0].get('count', 0)
		staff_half_day_count = staff_half_day_query[0].get('count', 0)
		staff_count = staff_present_count + (staff_half_day_count / 2)

		apprentice_present_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s   
													AND custom_employee_category IN ("Apprentice") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date')}, as_dict=True)
		apprentice_half_day_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s   
													AND custom_employee_category IN ("Apprentice") AND docstatus != 2 
													AND status = 'Half day' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date')}, as_dict=True)
		apprentice_present_count = apprentice_present_query[0].get('count', 0)
		apprentice_half_day_count = apprentice_half_day_query[0].get('count', 0)
		apprentice_count = apprentice_present_count + (apprentice_half_day_count / 2)

		contractor_present_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance`  WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s   
													AND custom_employee_category IN ("Contractor") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date')}, as_dict=True)
		contractor_half_day_query = frappe.db.sql(f"""SELECT count(*) as count FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s   
													AND custom_employee_category IN ("Contractor") AND docstatus != 2 
													AND status = 'Half day' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date')}, as_dict=True)
		contractor_present_count = contractor_present_query[0].get('count', 0)
		contractor_half_day_count = contractor_half_day_query[0].get('count', 0)
		contractor_count = contractor_present_count + (contractor_half_day_count / 2)

		staff_overtime_query = frappe.db.sql(f"""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s   
												AND custom_employee_category IN ("Staff", "SUB STAFF", "Operators") AND docstatus != 2 
												AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date')}, as_dict=True)
		apprentice_overtime_query = frappe.db.sql(f"""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s   
													AND custom_employee_category IN ("Apprentice") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date')}, as_dict=True)
		contractor_overtime_query = frappe.db.sql(f"""SELECT sum(custom_overtime_hours) as ot FROM `tabAttendance` WHERE attendance_date BETWEEN %(from_date)s AND %(to_date)s   
													AND custom_employee_category IN ("Contractor") AND docstatus != 2 
													AND status = 'Present' AND attendance_request is NULL """,{'from_date': args.get('from_date'),'to_date': args.get('to_date')}, as_dict=True)
		staff_overtime_hours = staff_overtime_query[0].get('ot', 0) / 8 if staff_overtime_query[0].get('ot', 0) else 0
		apprentice_overtime_hours = apprentice_overtime_query[0].get('ot', 0) / 8 if apprentice_overtime_query[0].get('ot', 0) else 0
		contractor_overtime_hours = contractor_overtime_query[0].get('ot', 0) / 8 if contractor_overtime_query[0].get('ot', 0) else 0

		total_business_plan = tot_bp
		total_actual_plan = actual_plan
		total_staff_count = staff_count
		total_apprentice_count = apprentice_count
		total_contractor_count = contractor_count
		total_staff_overtime_hours = staff_overtime_hours
		total_apprentice_overtime_hours = apprentice_overtime_hours
		total_contractor_overtime_hours = contractor_overtime_hours

		if total_actual_plan != 0:
			staff_percentage = ((total_staff_count) / total_actual_plan) * 100
			apprentice_percentage = ((total_apprentice_count) / total_actual_plan) * 100
			contractor_percentage = ((total_contractor_count) / total_actual_plan) * 100
		else:
			staff_percentage = 0
			apprentice_percentage = 0
			contractor_percentage = 0
	
	if total_business_plan>0:
		act=total_staff_count + total_apprentice_count + total_contractor_count
		ot=total_staff_overtime_hours + total_apprentice_overtime_hours + total_contractor_overtime_hours
		tot_1=((total_staff_count + total_apprentice_count + total_contractor_count)/total_business_plan)*100
		tot_2=((total_staff_overtime_hours + total_apprentice_overtime_hours + total_contractor_overtime_hours)/total_business_plan)*100
		tot_3=((act+ot)/total_business_plan)*100
	else:
		tot_1=0
		tot_2=0
		tot_3=0
	data.append(["Total (A + B)", "", "", total_business_plan, total_staff_count, total_apprentice_count, total_contractor_count, 
				(total_staff_count + total_apprentice_count + total_contractor_count),f"{tot_1:.2f}", total_staff_overtime_hours, total_apprentice_overtime_hours, 
				total_contractor_overtime_hours, (total_staff_overtime_hours + total_apprentice_overtime_hours + total_contractor_overtime_hours), f"{tot_2:.2f}",
				((total_staff_count + total_apprentice_count + total_contractor_count) + 
				(total_staff_overtime_hours + total_apprentice_overtime_hours + total_contractor_overtime_hours)),f"{tot_3:.2f}"])
	
	tot_act=0
	act_tot=((total_staff_count + total_apprentice_count + total_contractor_count) + 
				(total_staff_overtime_hours + total_apprentice_overtime_hours + total_contractor_overtime_hours))
	tot_act_ot=0
	act_tot_per=0
	if total_business_plan>0:
		tot_act=((total_staff_count+total_apprentice_count+total_contractor_count)/total_business_plan)*100
		tot_act_ot=((total_staff_overtime_hours + total_apprentice_overtime_hours + total_contractor_overtime_hours)/total_business_plan)*100
		act_tot_per=(act_tot/total_business_plan)*100
	data.append(["Percentage", "", "", "",  '', '','','',f"{tot_act:.2f} %",'' ,'','','',  f"{tot_act_ot:.2f} %",'',f"{act_tot_per:.2f} %"])

	return data
 
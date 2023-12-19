# Copyright (c) 2023, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file
from frappe.utils import cstr, add_days, date_diff, getdate
from frappe.utils.background_jobs import enqueue
import datetime
from datetime import datetime


from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

class OrderSchedule(Document):
    def after_insert(self):
        if self.customer_code:
            self.customer_name = frappe.db.get_value("Customer",{"customer_code":self.customer_code},["name"])
    def on_update(self):
        if self.sales_order_number:
            order_type = frappe.db.get_value("Sales Order",self.sales_order_number,'customer_order_type')
            if order_type == "Open":
                order_open = frappe.get_doc("Open Order",{"sales_order_number":self.sales_order_number})
                for ord in order_open.open_order_table:
                    item_qty = 0
                    order_schedules = frappe.get_all("Order Schedule", {"sales_order_number": self.sales_order_number,"item_code": self.item_code}, ["qty"])
                    for sche in order_schedules:
                        item_qty += sche.qty
                    if ord.item_code == self.item_code:
                        ord.qty = item_qty
                    order_open.save(ignore_permissions=True)
            else:
                s_qty = 0
                so_exist = frappe.db.exists('Order Schedule',{"sales_order_number":self.sales_order_number,"customer_code":self.customer_code,"item_code":self.item_code})
                if so_exist:
                    exist_so = frappe.get_all("Order Schedule",{"sales_order_number":self.sales_order_number,"customer_code":self.customer_code,"item_code":self.item_code},["*"])
                    for i in exist_so:
                        old_qty = i.qty
                        s_qty += old_qty
                    if order_type == "Fixed":
                        sales_order = frappe.get_all("Sales Order Item",{"parent": self.sales_order_number, "item_code": self.item_code},["qty","item_code"])
                        if sales_order and len(sales_order) > 0:
                            sales_order_qty = sales_order[0].get("qty")
                            item_cde = sales_order[0].get("item_code")
                            idx = sales_order[0].get("idx")
                            if sales_order_qty < s_qty:
                                frappe.throw(f"Validation failed: Quantity <b>{s_qty}</b> exceeds Sales Order quantity <b>{sales_order_qty}</b> in Line Item <b>{item_cde}</b>.")
                        else:
                            frappe.throw(f"Item <b>{self.item_code}</b> not found in the Sales Order - <b>{self.sales_order_number}</b>")


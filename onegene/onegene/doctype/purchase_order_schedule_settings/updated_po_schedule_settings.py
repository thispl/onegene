# Copyright (c) 2025, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file
from frappe.utils import cstr, add_days, date_diff, getdate
from frappe.utils.background_jobs import enqueue
import datetime
from datetime import datetime
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import json
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

class SalesOrderScheduleSettings(Document):
    pass

import frappe
import uuid

import uuid
from datetime import datetime
import frappe
from frappe.utils.file_manager import get_file
from frappe.exceptions import PermissionError
from frappe import _

@frappe.whitelist()
def enqueue_upload(file):
    import uuid

    job_id = f"upload_job_{uuid.uuid4().hex}"

    file = get_file(file)
    pps = read_xlsx_file_from_attached_file(fcontent=file[1])
    pps = [pp for pp in pps if pp and pp[0] != "Supplier Code"]

    if len(pps) <=500:
        validate_attached_file(file[0],pps,job_id )    
    else:
        frappe.throw(_("Upload supports only up to 500 rows"), title=_("Too Many Rows"))


    
def validate_attached_file(file, records, job_id):
    from frappe.utils.background_jobs import enqueue
    from collections import defaultdict
    from datetime import datetime
    import frappe
    from frappe.utils.file_manager import remove_file

    error_logs = []
    valid_records = []
    allowed_months = {"JAN", "FEB", "MAR", "APR", "MAY", "JUN",
                      "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"}
    distinct_po_numbers = list({r[1] for r in records if r[1]})

    for idx, pp in enumerate(records, start=1):
        try:
            sup_code, po_number, item, schedule_month = pp[0], pp[1], pp[2], (pp[3] or "").strip()
            s_qty = pp[4]

            if not sup_code or not frappe.db.exists("Supplier", {"supplier_code": sup_code}):
                error_logs.append(f"Row {idx}: Supplier '{sup_code}' does not exist.")
                continue

            if not po_number or not frappe.db.exists("Purchase Order", po_number):
                error_logs.append(f"Row {idx}: Purchase Order '{po_number}' does not exist.")
                continue

            purchase = frappe.get_doc("Purchase Order", po_number)

            if purchase.custom_order_type != "Open":
                error_logs.append(f"Row {idx}: Purchase Order '{po_number}' is not of type 'Open'.")
                continue

            so_item = next((i for i in purchase.items if i.item_code == str(item)), None)
            if not so_item:
                error_logs.append(f"Row {idx}: Item '{item}' is not part of Purchase Order '{po_number}'.")
                continue

            try:
                if isinstance(s_qty, float) or ('.' in str(s_qty)):
                    raise ValueError
                s_qty = int(s_qty)
                if s_qty <= 0:
                    raise ValueError
            except:
                error_logs.append(f"Row {idx}: Schedule Qty '{s_qty}' must be a positive integer (no decimals).")
                continue

            schedule_month = schedule_month.upper()
            if schedule_month not in allowed_months:
                error_logs.append(f"Row {idx}: Invalid Schedule Month '{schedule_month}'. Use Jan/Feb/... format.")
                continue

            current_year = datetime.now().year
            schedule_date = datetime.strptime(f"01-{schedule_month}-{current_year}", "%d-%b-%Y")

            exists = frappe.db.exists("Purchase Order Schedule", {
                "purchase_order_number": po_number,
                "schedule_month": schedule_month,
                "item_code": item,
                "docstatus": 1
            })
            if exists:
                error_logs.append(f"Row {idx}: Schedule already exists for {po_number}, {item}, {schedule_month}.")
                continue

            valid_records.append([sup_code, po_number, item, schedule_month, s_qty])

        except Exception as e:
            error_logs.append(f"Row {idx}: Unexpected error - {str(e)}")
            frappe.log_error(f"Row {idx}: {str(e)}", "Purchase Order Upload Error")

    # If any validation errors occur, remove the uploaded file
    if error_logs:
        if file:
            try:
                remove_file(file)
            except Exception as e:
                frappe.log_error(f"Failed to remove file after validation errors: {str(e)}", "File Remove Error")
        frappe.throw("<b>Validation Errors:</b><br>" + "<br>".join(error_logs))

    # If all records are valid, enqueue the background job
    frappe.msgprint(_("All rows validated successfully."), alert=True)
    
@frappe.whitelist()
def enqueue_upload_validate(file):
    from frappe.utils.file_manager import get_file
    file = get_file(file)
    pps = read_xlsx_file_from_attached_file(fcontent=file[1])
    pps = [pp for pp in pps if pp and pp[0] != "Supplier Code"]

    if len(pps) > 500:
        return "<b>Upload supports only up to 500 rows</b>"
    error_logs = validate_attached_file_return_errors(file[0], pps)

    if error_logs:
        html = '''
        <div style="max-height: 300px; overflow-y: auto; border:1px solid #ddd; padding:10px; border-radius:5px; background:#f9f9f9;">
            <table class="table table-striped table-hover table-bordered">
                <thead style="background:#fd7e14; color:white;">
                    <tr>
                        <th style="width:60px;text-align:center">Row</th>
                        <th style="text-align:center">Error</th>
                    </tr>
                </thead>
                <tbody>
        '''

        for idx, err in enumerate(error_logs, start=1):
            html += f'''
            <tr>
                <td style="font-weight:bold; color:#dc3545;">{idx}</td>
                <td><i class="fa fa-exclamation-triangle" style="color:#dc3545; margin-right:5px;"></i>{err}</td>
            </tr>
            '''

        html += '''
                </tbody>
            </table>
        </div>
        '''

        return html
   
# @frappe.whitelist()
# def enqueue_upload_validate(file):
#     from frappe.utils.file_manager import get_file
#     file = get_file(file)
#     pps = read_xlsx_file_from_attached_file(fcontent=file[1])
#     pps = [pp for pp in pps if pp and pp[0] != "Supplier Code"]

#     if len(pps) > 500:
#         return {"errors_html": "<b>Upload supports only up to 500 rows</b>", "warnings_html": ""}

#     error_logs, warnings = validate_attached_file_return_errors(file[0], pps)

#     def build_html(messages, title):
#         html = f'''
#         <div style="max-height: 300px; overflow-y: auto; border:1px solid #ddd; padding:10px; border-radius:5px; background:#f9f9f9;">
#             <table class="table table-striped table-hover table-bordered">
#                 <thead style="background:#fd7e14; color:white;">
#                     <tr>
#                         <th style="width:60px;text-align:center">Row</th>
#                         <th style="text-align:center">{title}</th>
#                     </tr>
#                 </thead>
#                 <tbody>
#         '''
#         for idx, msg in enumerate(messages, start=1):
#             html += f'''
#             <tr>
#                 <td style="font-weight:bold; color:#dc3545;">{idx}</td>
#                 <td><i class="fa fa-exclamation-triangle" style="color:#dc3545; margin-right:5px;"></i>{msg}</td>
#             </tr>
#             '''
#         html += '''
#                 </tbody>
#             </table>
#         </div>
#         '''
#         return html

#     errors_html = build_html(error_logs, "Error") if error_logs else ""
#     warnings_html = build_html(warnings, "Warning") if warnings else ""

#     return {
#         "errors_html": errors_html,
#         "warnings_html": warnings_html
#     }




def validate_attached_file_return_errors(file, records):
    import frappe
    from frappe.utils.file_manager import remove_file
    error_logs = []
    warning =[]
    valid_records = []
    allowed_months = {"JAN", "FEB", "MAR", "APR", "MAY", "JUN",
                      "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"}

    for idx, pp in enumerate(records, start=1):
        row_errors = []  # collect all errors for this row

        try:
            sup_code, po_number, item, schedule_month = pp[0], pp[1], pp[2], (pp[3] or "").strip()
            s_qty = pp[4]

            # ✅ Validate Supplier
            if not sup_code or not frappe.db.exists("Supplier", {"supplier_code": sup_code}):
                row_errors.append(f"Supplier '{sup_code}' does not exist.")

            # ✅ Validate Purchase Order
            if not po_number or not frappe.db.exists("Purchase Order", po_number):
                row_errors.append(f"Purchase Order '{po_number}' does not exist.")
            else:
                purchase = frappe.get_doc("Purchase Order", po_number)

                if purchase.custom_order_type != "Open":
                    row_errors.append(f"Purchase Order '{po_number}' is not of type 'Open'.")
                
                so_item = next((i for i in purchase.items if i.item_code == str(item)), None)
                if not so_item:
                    row_errors.append(f"Item '{item}' is not part of Purchase Order '{po_number}'.")

            # ✅ Validate Schedule Qty
            try:
                if isinstance(s_qty, float) or ('.' in str(s_qty)):
                    raise ValueError
                s_qty = int(s_qty)
                if s_qty < 0:
                    raise ValueError
            except:
                row_errors.append(f"Schedule Qty '{s_qty}' must be a positive integer (no decimals).")

            # ✅ Validate Schedule Month
            schedule_month = schedule_month.upper()
            if schedule_month not in allowed_months:
                row_errors.append(f"Invalid Schedule Month '{schedule_month}'. Use Jan/Feb/... format.")

            # ✅ Check if Schedule already exists
            exists = frappe.db.exists("Purchase Order Schedule", {
                "purchase_order_number": po_number,
                "schedule_month": schedule_month,
                "item_code": item,
                "docstatus": 1
            })
            if exists:
                # warning.append(f"Schedule already exists for {po_number}, {item}, {schedule_month}.")
                exists_val = frappe.db.get_value("Purchase Order Schedule",{"purchase_order_number": po_number,"schedule_month": schedule_month,"item_code": item, "docstatus": 1},"received_qty")
                if exists_val and exists_val > s_qty:
                    row_errors.append(f"While Revising the Schedule for {po_number}, {item}, {schedule_month}, the schedule qty{s_qty} is less than received qty{exists_val}.")

            # ✅ Append result
            if row_errors:
                # Combine multiple errors into a single string (with <br> for display)
                error_logs.append(f"<b>Row {idx}</b>: " + "<br>".join(row_errors))
            else:
                valid_records.append([sup_code, po_number, item, schedule_month, s_qty])

        except Exception as e:
            error_logs.append(f"<b>Row {idx}</b>: Unexpected error - {str(e)}")
            frappe.log_error(f"Row {idx}: {str(e)}", "Purchase Order Upload Error")

    return error_logs 

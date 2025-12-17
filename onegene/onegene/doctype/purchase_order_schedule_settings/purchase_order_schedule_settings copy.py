# Copyright (c) 2025, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file
from frappe.utils import cstr, add_days, date_diff, getdate
from frappe.utils.background_jobs import enqueue
import datetime
from datetime import datetime
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import json
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

class PurchaseOrderScheduleSettings(Document):
	pass

import frappe
import uuid

import uuid
from datetime import datetime
import frappe
from frappe.utils.file_manager import get_file
from frappe.exceptions import PermissionError
from frappe import _
from onegene.onegene.doctype.purchase_open_order.purchase_open_order import update_child_qty_rate
from onegene.onegene.doctype.purchase_order_schedule.purchase_order_schedule import revise_schedule_qty

@frappe.whitelist()
def enqueue_upload(file):
	import uuid

	job_id = f"upload_job_{uuid.uuid4().hex}"

	file = get_file(file)
	pps = read_xlsx_file_from_attached_file(fcontent=file[1])
	pps = [pp for pp in pps if pp and pp[0] != "Supplier Code"]

	if len(pps) < 20:
		return _process_upload(file=file[0], records=pps)
	elif len(pps) <= 500:
		frappe.msgprint(_("Upload is being processed in background..."), alert=True)
		frappe.enqueue(
			method="onegene.onegene.doctype.purchase_order_schedule_settings.purchase_order_schedule_settings._process_upload",
			queue="long",
			timeout=1800,
			is_async=True,
			file=file[0],
			records=pps,
			job_name=job_id,
		)
	else:
		frappe.throw(_("Upload supports only up to 500 rows"), title=_("Too Many Rows"))

def _process_upload(file, records):
	from collections import defaultdict
	from datetime import datetime
	import frappe

	docs_to_submit = []
	grouped_docs = defaultdict(list)  # PO Number => List of Schedule Doc Names
	distinct_po_numbers = list({r[1] for r in records if r[1]})
	total_records = len(records)
	error_log = []

	for idx, pp in enumerate(records, start=1):
		try:
			sup_code, po_number, item, schedule_month = pp[0], pp[1], pp[2], pp[3]
			s_qty = pp[4]

			if not po_number:
				continue

			purchase = frappe.get_doc("Purchase Order", po_number)

			if purchase.custom_order_type == "Open":
				current_year = datetime.now().year
				schedule_month = schedule_month.upper() if schedule_month else ""
				schedule_date = datetime.strptime(f"01-{schedule_month}-{current_year}", "%d-%b-%Y")
				if not frappe.db.exists("Purchase Order Schedule",{"item_code":item,"purchase_order_number":po_number,"schedule_month":schedule_month}):
					doc = frappe.new_doc('Purchase Order Schedule')
     
					doc.supplier_code = sup_code
					doc.supplier_name = frappe.db.get_value("Supplier",{"supplier_code":sup_code},"name")
					doc.purchase_order_number = po_number
					doc.item_code = item
					doc.schedule_date = schedule_date
					doc.qty = s_qty
					doc.schedule_month = schedule_month
					doc.pending_qty = s_qty
					doc.disable_update_items = 1
					doc.save(ignore_permissions=True)

					docs_to_submit.append(doc.name)
					grouped_docs[po_number].append(doc.name)
				else:
					if not frappe.db.exists("Purchase Order Schedule",{"item_code":item,"purchase_order_number":po_number,"schedule_month":schedule_month,"qty":s_qty}):
						existing_pos = frappe.db.get_value("Purchase Order Schedule",{"item_code":item,"purchase_order_number":po_number,"schedule_month":schedule_month},['name'])
						revise_schedule_qty(existing_pos,s_qty,'Revised from Purchase Order Schedule Settings')
				# doc = frappe.new_doc('Purchase Order Schedule')
				# doc.supplier_code = sup_code
				# doc.purchase_order_number = po_number
				# doc.item_code = item
				# doc.schedule_date = schedule_date
				# doc.qty = s_qty
				# doc.schedule_month = schedule_month
				# doc.pending_qty = s_qty
				# doc.disable_update_items = 1
				# doc.save(ignore_permissions=True)

				# docs_to_submit.append(doc.name)
				# grouped_docs[po_number].append(doc.name)

			frappe.publish_realtime(
				event='purchase_order_upload_progress',
				message={
					"stage": "Creating Purchase Order Schedule",
					"progress": round(float(idx) * 100 / total_records, 2),
					"description": po_number
				},
				user=frappe.session.user
			)

		except Exception as e:
			error_msg = f"Row {idx} failed: {e}"
			frappe.log_error(error_msg, "Purchase Order Upload Error")
			error_log.append(error_msg)

	# Second pass: submit docs
	total_to_submit = len(docs_to_submit)
	for idx, name in enumerate(docs_to_submit, start=1):
		try:
			doc = frappe.get_doc("Purchase Order Schedule", name)
			if doc.docstatus == 0:
				frappe.flags.ignore_permissions = True
				doc.submit()
			frappe.publish_realtime(
				event='purchase_order_upload_progress',
				message={
					"stage": "Submitting Schedule",
					"progress": round(float(idx) * 100 / total_to_submit, 2),
					"description": doc.purchase_order_number
				},
				user=frappe.session.user
			)
		except frappe.exceptions.PermissionError:
			error_msg = f"Permission error on submit: {name}"
			frappe.log_error(error_msg, "Purchase Order Upload Error")
			error_log.append(error_msg)
		except Exception as e:
			error_msg = f"Submit failed for {name}: {e}"
			frappe.log_error(error_msg, "Purchase Order Upload Error")
			error_log.append(error_msg)
		finally:
			frappe.flags.ignore_permissions = False

	# Third pass: update related Purchase Orders
	total_po = len(distinct_po_numbers)
	for idx, purchase_order_number in enumerate(distinct_po_numbers, start=1):
		try:
			order_open = frappe.get_doc("Purchase Open Order", {"purchase_order": purchase_order_number})
			order_open.disable_update_items = 0
			order_open.save(ignore_permissions=True)

			frappe.publish_realtime(
				event='purchase_order_upload_progress',
				message={
					"stage": "Updating Purchase Order",
					"progress": round(float(idx) * 100 / total_po, 2),
					"description": purchase_order_number
				},
				user=frappe.session.user
			)

		except Exception as e:
			error_msg = f"Update failed for PO {purchase_order_number}: {e}"
			frappe.log_error(error_msg, "Purchase Order Upload Error")
			error_log.append(error_msg)

	# Final status update
	if error_log:
		frappe.log_error("Error", "\n".join(error_log))
		frappe.publish_realtime(
			event='purchase_order_upload_progress',
			message={
				"stage": "Completed with Errors",
				"progress": 100,
				"errors": error_log
			},
			user=frappe.session.user
		)
	else:
		frappe.publish_realtime(
			event='purchase_order_upload_progress',
			message={
				"stage": "Completed Successfully",
				"progress": 100,
				"description": "All records processed."
			},
			user=frappe.session.user
		)

	return True

@frappe.whitelist()
def template_sheet():
	args = frappe.local.form_dict
	filename = args.name
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	ws.append(["Supplier Code","Purchase Order Number","Item Code","Schedule Period (month)","Schedule Qty"])
	ws.append(["","","","Example: Apr, May, Jun, Jul, Aug, Sep, ...",""])
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


@frappe.whitelist()
def get_data(file):
	data = """<table class=table table-bordered >
	<tr><td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>S.NO</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Supplier Code</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Supplier Name</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Purchase Order Number</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Item Code</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Item Name</b></center></td>

	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Schedule Period (in month)</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Schedule Qty</b></center></td>
	</tr>"""
	file = get_file(file)
	pps = read_xlsx_file_from_attached_file(fcontent=file[1])  
	i = 1  
	for pp in pps:
		if pp[0] != 'Supplier Code':
			
			sup_code = pp[0]
			sup_name = frappe.db.get_value("Supplier",{"supplier_code":pp[0]},'supplier_name')
			po_number = pp[1]
			item = pp[2]
			item_name =frappe.db.get_value("Item",pp[2],'item_name')
			sch_date = pp[3]
			s_qty = pp[4]
			
			data += """
			
			<tr>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px; text-align: center;">%s</td>
			
		
			</tr>"""%(i,sup_code or '',sup_name or '',po_number or '',item or '',item_name or '',sch_date or '',s_qty or '')
			i += 1
	return data





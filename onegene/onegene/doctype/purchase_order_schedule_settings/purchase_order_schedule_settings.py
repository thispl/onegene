# Copyright (c) 2025, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file
from frappe.utils import cstr, add_days, date_diff, getdate
from frappe.utils.background_jobs import enqueue
import datetime
from datetime import datetime
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import json
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

class PurchaseOrderScheduleSettings(Document):
	pass

import frappe
import uuid

import uuid
from datetime import datetime
import frappe
from frappe.utils.file_manager import get_file
from frappe.exceptions import PermissionError
from frappe import _
from onegene.onegene.doctype.purchase_open_order.purchase_open_order import update_child_qty_rate

@frappe.whitelist()
def enqueue_upload(file):
	import uuid

	job_id = f"upload_job_{uuid.uuid4().hex}"

	file = get_file(file)
	pps = read_xlsx_file_from_attached_file(fcontent=file[1])
	pps = [pp for pp in pps if pp and pp[0] != "Supplier Code"]

	if len(pps) < 20:
		precheck_records_without_enqueue(file[0],pps,job_id )
		frappe.db.set_single_value('Purchase Order Schedule Settings', 'attach', None)
		return _process_upload(file=file[0], records=pps)
	elif len(pps) <= 500:
		precheck_records(file[0],pps,job_id )
		frappe.db.set_single_value('Purchase Order Schedule Settings', 'attach', None)
		
	else:
		frappe.throw(_("Upload supports only up to 500 rows"), title=_("Too Many Rows"))

def _process_upload(file, records):
	from frappe.utils.background_jobs import enqueue
	from collections import defaultdict
	from datetime import datetime
	import frappe

	def publish_progress(current_step, total_steps, stage):
		frappe.publish_realtime(
			event='purchase_order_upload_progress',
			message={
				"stage": "Processing Upload",
				"progress": round((current_step / total_steps) * 100, 2),
				"description": stage
			},
			user=frappe.session.user
		)

	docs_to_submit = []
	grouped_docs = defaultdict(list)
	distinct_po_numbers = []
	new_count = 0
	update_count = 0
	skipped_count = 0
	# distinct_po_numbers = list({r[1] for r in records if r[1]})

	# Total steps = one step per row per stage (3 stages)
	num_rows = len(records)
	total_steps = num_rows * 3
	current_step = 0

	# ✅ Initial progress
	publish_progress(current_step, total_steps, "Starting Upload...")

	# === STEP 1: Create Purchase Order Schedules ===
	for pp in records:
		try:
			sup_code, po_number, item, schedule_month = pp[0], pp[1], pp[2], pp[3]
			s_qty = pp[4]

			if not po_number:
				continue

			purchase = frappe.get_doc("Purchase Order", po_number)

			if purchase.custom_order_type == "Open":
				current_year = datetime.now().year
				schedule_month = schedule_month.upper() if schedule_month else ""
				schedule_date = datetime.strptime(f"01-{schedule_month}-{current_year}", "%d-%b-%Y")

				if not frappe.db.exists("Purchase Order Schedule", {
					'purchase_order_number': po_number,
					'schedule_month': schedule_month,
					'item_code': item,
					'docstatus': 1
				}):
					doc = frappe.new_doc('Purchase Order Schedule')
					doc.supplier_code = sup_code
					doc.supplier_name = frappe.db.get_value("Supplier", {"supplier_code": sup_code}, "name")
					doc.purchase_order_number = po_number
					doc.item_code = item
					doc.schedule_date = schedule_date
					doc.qty = s_qty
					doc.schedule_month = schedule_month
					doc.pending_qty = s_qty
					doc.disable_update_items = 1
					doc.save(ignore_permissions=True)

					docs_to_submit.append(doc.name)
					grouped_docs[po_number].append(doc.name)
					distinct_po_numbers.append(po_number)
					new_count += 1
				else:
					if not frappe.db.exists("Purchase Order Schedule", {
							'purchase_order_number': po_number,
							'schedule_month': schedule_month,
							'item_code': item,
							'docstatus': 1,
							'qty': s_qty
						}):
						doc = frappe.get_doc("Purchase Order Schedule", {
							'purchase_order_number': po_number,
							'schedule_month': schedule_month,
							'item_code': item,
							'docstatus': 1
						})

						old_qty = doc.qty  
						doc.qty = s_qty
						doc.disable_update_items = 1
						doc.append("revision", {
							"revised_on": frappe.utils.now_datetime(),
							"schedule_qty": old_qty,
							"revised_schedule_qty": doc.qty,
							"revised_by": frappe.session.user,
							"remarks": "Revised from Purchase Order Schedule Settings"
						})

						pending_qty = s_qty - doc.received_qty
						doc.pending_qty = pending_qty
						doc.schedule_amount = s_qty * doc.order_rate
						doc.received_amount = doc.received_qty * doc.order_rate
						doc.pending_amount = pending_qty * doc.order_rate
						doc.schedule_amount_inr = s_qty * doc.order_rate_inr
						doc.received_amount_inr = doc.received_qty * doc.order_rate_inr
						doc.pending_amount_inr = pending_qty * doc.order_rate_inr


						doc.save(ignore_permissions=True)

						docs_to_submit.append(doc.name)
						grouped_docs[po_number].append(doc.name)
						distinct_po_numbers.append(po_number)
						update_count += 1
					else:
						skipped_count += 1
		except Exception as e:
			frappe.log_error(f"Row failed: {e}", "Purchase Order Upload Error")

		current_step += 1
		publish_progress(current_step, total_steps, "Creating Schedule")

	# === STEP 2: Submit Purchase Order Schedules ===
	for name in docs_to_submit:
		try:
			doc = frappe.get_doc("Purchase Order Schedule", name)
			if doc.docstatus == 0:
				frappe.flags.ignore_permissions = True
				doc.submit()
			elif doc.docstatus == 1:
				if doc.order_type == "Open" and frappe.db.exists("Purchase Order", {
					'name': doc.purchase_order_number, 'docstatus': 1
				}):
					purchase_open_order = frappe.get_doc("Purchase Open Order", {
						"purchase_order": doc.purchase_order_number
					})
					rows = frappe.db.get_all("Purchase Order Schedule",
						filters={
							"purchase_order_number": doc.purchase_order_number,
							"item_code": doc.item_code,
							"docstatus": 1
						},
						fields=["qty"]
					)
					item_qty = sum([r.qty for r in rows])
					item_qty = 1 if item_qty == 0 else item_qty
					matching_row = next((row for row in purchase_open_order.open_order_table if row.item_code == doc.item_code), None)
					if matching_row:
						purchase_open_order.disable_update_items = 0 if doc.disable_update_items == 0 else 1
						matching_row.qty = item_qty
						matching_row.amount = item_qty * float(doc.order_rate)
						purchase_open_order.save(ignore_permissions=True)

		except Exception as e:
			frappe.log_error(f"Submit failed for {name}: {e}", "Purchase Order Upload Error")
		finally:
			frappe.flags.ignore_permissions = False

		current_step += 1
		publish_progress(current_step, total_steps, "Submitting Schedule")
	
	frappe.log_error("Disctince SO Number", distinct_po_numbers)
	# === STEP 3: Update Purchase Open Orders ===
	distinct_po_numbers = list(set(distinct_po_numbers))
	for po_number in distinct_po_numbers:
		try:
			order_open = frappe.get_doc("Purchase Open Order", {
				"purchase_order": po_number
			})
			order_open.disable_update_items = 0
			order_open.save(ignore_permissions=True)
		except Exception as e:
			frappe.log_error(f"Update failed for PO {po_number}: {e}", "Purchase Order Upload Error")

		current_step += 1
		publish_progress(current_step, total_steps, "Updating Purchase Order")
	
	publish_progress(total_steps, total_steps, "Upload Complete")
	now = frappe.utils.now_datetime().strftime("%d-%m-%Y %H:%M:%S")
	response_data = f"""
	<div style="
		max-width: 420px;
		margin: 20px auto;
		background: #ffffff;
		border-radius: 12px;
		padding: 20px 24px;
		box-shadow: 0 4px 14px rgba(0,0,0,0.08);
		font-family: Inter, sans-serif;
		border: 1px solid #e7e7e7;
	">
		<div style="font-size: 18px; font-weight: 600; color: #333;">
			Last Updated On
		</div>

		<div style="margin-top: 6px; font-size: 13px; color: #666;">
			{now}
		</div>

		<hr style="margin: 12px 0; border: 0; border-top: 1px solid #eee;">

		<ul style="margin: 0; padding-left: 20px; line-height: 1.8;">
			<li style="color:#4caf50">New documents created: {new_count}</li>
			<li style="color:#2196f3">Updated in existing document: {update_count}</li>
			<li style="color:#999">Skipped documents: {skipped_count}</li>
		</ul>
	</div>
	"""
	frappe.db.set_single_value("Purchase Order Schedule Settings", "response_data", response_data)
	frappe.clear_cache()
	frappe.publish_realtime("po_upload_summary", {
		"new": new_count,
		"update": update_count,
		"skip": skipped_count
	})
	return True


def precheck_records(file,records,job_id):
	from frappe.utils.background_jobs import enqueue
	from collections import defaultdict
	from datetime import datetime
	import json
	import frappe

	
	docs_to_submit = []
	grouped_docs = defaultdict(list)
	distinct_po_numbers = list({r[1] for r in records if r[1]})
	total_records = len(records)
	error_logs = []
	valid_records = []
	
	for idx, pp in enumerate(records, start=1):
		try:
			sup_code, po_number, item, schedule_month = pp[0], pp[1], pp[2], pp[3]
			s_qty = pp[4]

			if not po_number:
				continue

			purchase = frappe.get_doc("Purchase Order", po_number)

			if purchase.custom_order_type != "Open":
				continue

			current_year = datetime.now().year
			schedule_month = schedule_month.upper() if schedule_month else ""
			schedule_date = datetime.strptime(f"01-{schedule_month}-{current_year}", "%d-%b-%Y")

			existing = frappe.db.exists("Purchase Order Schedule", {
				'purchase_order_number': po_number,
				'schedule_month': schedule_month,
				'item_code': item,
				'docstatus': 1
			})

			valid_records.append({
				"idx": idx,
				"sup_code": sup_code,
				"po_number": po_number,
				"item": item,
				"schedule_date": schedule_date,
				"schedule_month": schedule_month,
				"s_qty": s_qty,
				"existing": existing
			})

		except Exception as e:
			error_logs.append(f"Row {idx}: {str(e)}")
			frappe.log_error(f"Row {idx}: {str(e)}", "Purchase Order Upload Error")

	
	

	# Second pass: submit docs
	total_to_submit = len(docs_to_submit)
	docs_to_process = []

	for idx, name in enumerate(docs_to_submit, start=1):
		try:
			doc = frappe.get_doc("Purchase Order Schedule", name)

			if doc.docstatus == 1:
				if doc.order_type == "Open" and frappe.db.exists("Purchase Order", {'name': doc.purchase_order_number, 'docstatus': 1}):
					purchase_open_order = frappe.get_doc("Purchase Open Order", {"purchase_order": doc.purchase_order_number})
					schedules = frappe.get_all("Purchase Order Schedule",{"purchase_order_number": doc.purchase_order_number,"item_code": doc.item_code,"docstatus": 1},["qty"])
					item_qty = sum(s.qty for s in schedules)

					matching_row = next((row for row in purchase_open_order.open_order_table if row.item_code == doc.item_code), None)
					if not matching_row:
						raise Exception(f"No matching row found in Purchase Open Order for item {doc.item_code}")

					matching_row.qty = item_qty
					matching_row.amount = item_qty * float(doc.order_rate)

			docs_to_process.append(doc)

		except Exception as e:
			error_logs.append(f"Row {idx} - Doc {name}: {str(e)}")

	if error_logs:
		error_message = "One or more errors occurred during processing:\n" + "\n".join(error_logs)
		frappe.log_error(error_message, "Purchase Order Upload Error")
	
	# Final pass: update purchase orders
	error_logs = []
	orders_to_update = []

	for idx, purchase_order_number in enumerate(distinct_po_numbers, start=1):
		try:
			# just pre checking, not actually saving
			order_open = frappe.get_doc("Purchase Open Order", {"purchase_order": purchase_order_number})
			order_open.disable_update_items = 0  
			orders_to_update.append(order_open)
		except Exception as e:
			error_logs.append(f"PO {purchase_order_number}: {str(e)}")

	# if error_logs:
	# 	error_message = "Errors occurred while updating Purchase Open Orders:\n" + "\n".join(error_logs)
	# 	frappe.log_error(error_message, "Purchase Order Upload Error")
	if error_logs:
		# error_message = "\n".join(error_logs)
		frappe.throw('')
	
		

	
	else:
		frappe.msgprint(_("Upload is being processed in background..."), alert=True)
		frappe.enqueue(
			method="onegene.onegene.doctype.purchase_order_schedule_settings.purchase_order_schedule_settings._process_upload",
			queue="long",
			timeout=1800,
			is_async=True,
			file=file,
			records=records,
			job_name=job_id,
		)
	
def precheck_records_without_enqueue(file,records,job_id):
	from frappe.utils.background_jobs import enqueue
	from collections import defaultdict
	from datetime import datetime
	import json
	import frappe

	
	docs_to_submit = []
	grouped_docs = defaultdict(list)
	distinct_po_numbers = list({r[1] for r in records if r[1]})
	total_records = len(records)
	error_logs = []
	valid_records = []
	
	for idx, pp in enumerate(records, start=1):
		try:
			sup_code, po_number, item, schedule_month = pp[0], pp[1], pp[2], pp[3]
			s_qty = pp[4]

			if not po_number:
				continue

			purchase = frappe.get_doc("Purchase Order", po_number)

			if purchase.custom_order_type != "Open":
				continue

			current_year = datetime.now().year
			schedule_month = schedule_month.upper() if schedule_month else ""
			schedule_date = datetime.strptime(f"01-{schedule_month}-{current_year}", "%d-%b-%Y")

			existing = frappe.db.exists("Purchase Order Schedule", {
				'purchase_order_number': po_number,
				'schedule_month': schedule_month,
				'item_code': item,
				'docstatus': 1
			})

			valid_records.append({
				"idx": idx,
				"sup_code": sup_code,
				"po_number": po_number,
				"item": item,
				"schedule_date": schedule_date,
				"schedule_month": schedule_month,
				"s_qty": s_qty,
				"existing": existing
			})

		except Exception as e:
			error_logs.append(f"Row {idx}: {str(e)}")
			frappe.log_error(f"Row {idx}: {str(e)}", "Purchase Order Upload Error")

	
	

	# Second pass: submit docs
	total_to_submit = len(docs_to_submit)
	docs_to_process = []

	for idx, name in enumerate(docs_to_submit, start=1):
		try:
			doc = frappe.get_doc("Purchase Order Schedule", name)

			if doc.docstatus == 1:
				if doc.order_type == "Open" and frappe.db.exists("Purchase Order", {'name': doc.purchase_order_number, 'docstatus': 1}):
					purchase_open_order = frappe.get_doc("Purchase Open Order", {"purchase_order": doc.purchase_order_number})
					schedules = frappe.get_all("Purchase Order Schedule",{"purchase_order_number": doc.purchase_order_number,"item_code": doc.item_code,"docstatus": 1},["qty"])
					item_qty = sum(s.qty for s in schedules)

					matching_row = next((row for row in purchase_open_order.open_order_table if row.item_code == doc.item_code), None)
					if not matching_row:
						raise Exception(f"No matching row found in Purchase Open Order for item {doc.item_code}")

					matching_row.qty = item_qty
					matching_row.amount = item_qty * float(doc.order_rate)

			docs_to_process.append(doc)

		except Exception as e:
			error_logs.append(f"Row {idx} - Doc {name}: {str(e)}")

	if error_logs:
		error_message = "One or more errors occurred during processing:\n" + "\n".join(error_logs)
		frappe.log_error(error_message, "Purchase Order Upload Error")
	
	# Final pass: update purchase orders
	error_logs = []
	orders_to_update = []

	for idx, purchase_order_number in enumerate(distinct_po_numbers, start=1):
		try:
			# just pre checking, not actually saving
			order_open = frappe.get_doc("Purchase Open Order", {"purchase_order": purchase_order_number})
			order_open.disable_update_items = 0  
			orders_to_update.append(order_open)
		except Exception as e:
			error_logs.append(f"PO {purchase_order_number}: {str(e)}")

	# if error_logs:
	# 	error_message = "Errors occurred while updating Purchase Open Orders:\n" + "\n".join(error_logs)
	# 	frappe.log_error(error_message, "Purchase Order Upload Error")
	if error_logs:
		error_message = "\n".join(error_logs)
		frappe.throw(f'{error_message}')
	
		

	
	else:
		frappe.msgprint(_("Upload is being processed in background..."), alert=True)
		frappe.enqueue(
			method="onegene.onegene.doctype.purchase_order_schedule_settings.purchase_order_schedule_settings._process_upload",
			queue="long",
			timeout=1800,
			is_async=True,
			file=file,
			records=records,
			job_name=job_id,
		)
	
		

@frappe.whitelist()
def template_sheet():
	args = frappe.local.form_dict
	filename = args.name
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	ws.append(["Supplier Code","Purchase Order Number","Item Code","Schedule Period (month)","Schedule Qty", "Tentative Plan 1", "Tentative Plan 2"])
	ws.append(["","","","Example: Apr, May, Jun, Jul, Aug, Sep, ...",""])
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


@frappe.whitelist()
def get_data(file):
	data = """<table class=table table-bordered >
	<tr><td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>S.NO</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Supplier Code</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Supplier Name</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Purchase Order Number</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Item Code</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Item Name</b></center></td>

	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Schedule Period (in month)</b></center></td>
	<td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Schedule Qty</b></center></td>
	</tr>"""
	file = get_file(file)
	pps = read_xlsx_file_from_attached_file(fcontent=file[1])  
	i = 1  
	for pp in pps:
		if pp[0] != 'Supplier Code':
			sup_code = pp[0]
			supplier_name = frappe.db.get_value("Supplier",{"supplier_code": pp[0]},'name')
			po_number = pp[1]
			item = pp[2]
			item_name =frappe.db.get_value("Item",pp[2],'item_name')
			sch_date = pp[3]
			s_qty = pp[4]
			
			data += """
			
			<tr>
			<td style="padding:1px; border: 1px solid black; font-size:10px; text-align: right;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
			<td style="padding:1px; border: 1px solid black; font-size:10px; text-align: right;">%s</td>
			
			</tr>"""%(i,sup_code or '',supplier_name or '',po_number or '',item or '',item_name or '',sch_date or '',s_qty or 0)
			i += 1
	return data





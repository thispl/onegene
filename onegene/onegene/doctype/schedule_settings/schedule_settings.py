# Copyright (c) 2023, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file
from frappe.utils import cstr, add_days, date_diff, getdate
from frappe.utils.background_jobs import enqueue
import datetime
from datetime import datetime
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import json
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


class ScheduleSettings(Document):
	@frappe.whitelist()
	def enqueue_u(file):
		enqueue(upload, queue='default', timeout=6000, event='upload',file=file)


@frappe.whitelist()
def enqueue_upload(file):
	file = get_file(file)
	pps = read_xlsx_file_from_attached_file(fcontent=file[1])
	for idx, pp in enumerate(pps, start=1): #enumerate --- get the index value 
		if pp[0] != 'Customer Code':
			cust_code = pp[0]
			so_number = pp[1]
			item = pp[2]
			s_qty = pp[4]
			tent_plan_1 = pp[5]
			tent_plan_2 = pp[6]
			tent_plan_3 = pp[7]
			old_qty = 0
			if so_number:
				sale = frappe.get_doc("Sales Order",{"name":so_number})
				so_exist = frappe.db.exists('Sales Order Schedule',{"sales_order_number":so_number,"customer_code":cust_code,"item_code":item})
				if so_exist:
					exist_so = frappe.get_all("Sales Order Schedule",{"sales_order_number":so_number,"customer_code":cust_code,"item_code":item},["*"])
					for i in exist_so:
						old_qty = i.qty
						s_qty += old_qty
					if sale.customer_order_type == "Fixed":
						sales_order = frappe.get_all("Sales Order Item",{"parent": so_number, "item_code": item},["qty"])
						if sales_order and len(sales_order) > 0:
							sales_order_qty = sales_order[0].get("qty")
							if sales_order_qty >= s_qty:
								new_doc = frappe.new_doc('Sales Order Schedule')  
								new_doc.customer_code = pp[0]
								new_doc.sales_order_number = pp[1]
								new_doc.item_code = pp[2]
								new_doc.schedule_date = pp[3]
								new_doc.qty = pp[4]
								new_doc.tentative_plan_1 = pp[5] 
								new_doc.tentative_plan_2 = pp[6] 
								new_doc.tentative_plan_3 = pp[7] 
								new_doc.save(ignore_permissions=True)
								frappe.db.commit()
							else:
								if sales_order_qty < s_qty:
									frappe.throw(f"<b>Validation failed:</b> Quantity <b>{s_qty}</b> exceeds the Sales Order quantity <b>{sales_order_qty}</b> in Line Item <b>{item}</b> with index <b>{idx - 1}</b>.")
					elif sale.customer_order_type == "Open":
						new_doc = frappe.new_doc('Sales Order Schedule')  
						new_doc.customer_code = pp[0]
						new_doc.sales_order_number = pp[1]
						new_doc.item_code = pp[2]
						new_doc.schedule_date = pp[3]
						new_doc.qty = pp[4]
						new_doc.tentative_plan_1 = pp[5] 
						new_doc.tentative_plan_2 = pp[6] 
						new_doc.tentative_plan_3 = pp[7] 
						new_doc.save(ignore_permissions=True)
						frappe.db.commit()
						
				else:
					s_qty = pp[4]
					sales_order = frappe.get_all("Sales Order Item",{"parent": so_number, "item_code": item},["qty"])                
					sales_order_qty = sales_order[0].get("qty")
					if sale.customer_order_type == "Open":
						new_doc = frappe.new_doc('Sales Order Schedule')  
						new_doc.customer_code = pp[0]
						new_doc.sales_order_number = pp[1]
						new_doc.item_code = pp[2]
						new_doc.schedule_date = pp[3]
						new_doc.qty = pp[4]
						new_doc.tentative_plan_1 = pp[5] 
						new_doc.tentative_plan_2 = pp[6] 
						new_doc.tentative_plan_3 = pp[7] 
						new_doc.save(ignore_permissions=True)
						frappe.db.commit()
					elif sale.customer_order_type == "Fixed":
						if sales_order and len(sales_order) > 0:
							sales_order_qty = sales_order[0].get("qty")
							if s_qty <= sales_order_qty:
								new_doc = frappe.new_doc('Sales Order Schedule')  
								new_doc.customer_code = pp[0]
								new_doc.sales_order_number = pp[1]
								new_doc.item_code = pp[2]
								new_doc.schedule_date = pp[3]
								new_doc.qty = pp[4]
								new_doc.tentative_plan_1 = pp[5] 
								new_doc.tentative_plan_2 = pp[6] 
								new_doc.tentative_plan_3 = pp[7] 
								new_doc.save(ignore_permissions=True)
								frappe.db.commit()
							else:
								frappe.throw(f"<b>Validation failed:</b> Quantity <b>{s_qty}</b> exceeds the Sales Order quantity <b>{sales_order_qty}</b> in Line Item <b>{item}</b> with index <b>{idx - 1}</b>.")

	return 'ok'


@frappe.whitelist()
def template_sheet():
	args = frappe.local.form_dict
	filename = args.name
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	ws.append(["Customer Code","Sales Order Number","Item Code","Schedule Date","Schedule Qty","Tentative Plan 1","Tentative Plan 2","Tentative Plan 3"])
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


@frappe.whitelist()
def get_data(file):
    data = """<table class=table table-bordered >
    <tr><td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>S.NO</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Customer Code</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Sales Order Number</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Item Code</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Schedule Date</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Schedule Qty</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Tentative Plan 1</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Tentative Plan 2</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Tentative Plan 3</b></center></td>
    </tr>"""
    file = get_file(file)
    pps = read_xlsx_file_from_attached_file(fcontent=file[1])  
    i = 1  
    for pp in pps:
        if pp[0] != 'Customer Code':
            
            cust_code = pp[0]
            so_number = pp[1]
            item = pp[2]
            sch_date = pp[3].strftime("%d-%m-%Y") if pp[3] else ''
            s_qty = pp[4]
            tent_plan_1 = pp[5]
            tent_plan_2 = pp[6]
            tent_plan_3 = pp[7]
            
            data += """
            
            <tr>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>%s</center></td>
        
            </tr>"""%(i,cust_code or '',so_number or '',item or '',sch_date or '',s_qty or '',tent_plan_1 or '',tent_plan_2 or '',tent_plan_3 or '')
            i += 1
    return data


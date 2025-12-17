# Copyright (c) 2023, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file
from frappe.utils import cstr, add_days, date_diff, getdate
from frappe.utils.background_jobs import enqueue
import datetime
from datetime import datetime
import calendar
from frappe.utils import formatdate
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import frappe
from frappe.query_builder import DocType
import requests
from datetime import date
from time import strptime
import erpnext
import json
from frappe.utils import now
from typing import Dict, Optional, Tuple, Union
from frappe import throw,_
from frappe.utils import flt
from frappe.utils import (
	add_days,
	ceil,
	cint,
	comma_and,
	flt,
	get_link_to_form,
	getdate,
	now_datetime,
	datetime,get_first_day,get_last_day,
	nowdate,
	today,
)
from frappe.utils import cstr, cint, getdate, get_last_day, get_first_day, add_days,date_diff
from datetime import date, datetime, timedelta


class SalesOrderSchedule(Document):
	def before_insert(self):
		if self.customer_code:
			self.customer_name = frappe.db.get_value("Customer",{"customer_code":self.customer_code},["name"])
		if self.sales_order_number and self.item_code:
			order_rate = frappe.db.get_value(
				"Sales Order Item", 
				{"parent": self.sales_order_number, "item_code": self.item_code}, 
				"rate"
			)
			self.order_rate = order_rate or 0
		else:
			self.order_rate = 0

		# Update schedule year and schedule month based on schedule_date
		if self.schedule_date:
			if isinstance(self.schedule_date, str):
				try:
					self.schedule_date = datetime.strptime(self.schedule_date, "%Y-%m-%d %H:%M:%S")
				except ValueError:
					self.schedule_date = datetime.strptime(self.schedule_date, "%Y-%m-%d")
		
			self.schedule_year = self.schedule_date.year
			self.schedule_month = self.schedule_date.strftime("%b").upper()
		doc = frappe.db.get_value("Sales Order Schedule",{"item_code":self.item_code, "customer_name":self.customer_name,"schedule_month":self.schedule_month,"schedule_year":self.schedule_year, "sales_order_number": self.sales_order_number, "docstatus": 1},["name", "item_code", "schedule_month"], 
	as_dict=True )
		if doc:
	  
			name = f'<a href="/app/sales-order-schedule/{doc.name}">{doc.name}</a>'
			frappe.throw(_(f'Sales Order Schedule for the item <b>{doc.item_code}</b> for the month <b>{doc.schedule_month}</b> is already available.   Instead of this, you can revise qty in {name}'))
			frappe.throw(_(
				f'Sales Order Schedule, item <b>{doc.item_code}</b> மற்றும் மாதம் <b>{doc.schedule_month}</b> க்கானது ஏற்கனவே உள்ளது. இதற்கு பதிலாக, நீங்கள் {name} இல் அளவை திருத்தலாம்.'
			))
   
	
	def before_save(self):
		if self.sales_order_number:
			self.currency = frappe.db.get_value("Sales Order",{"name":self.sales_order_number},"currency") or ""
			exchange_rate = frappe.db.get_value("Sales Order",{"name":self.sales_order_number},"conversion_rate") or ""
			if exchange_rate:
				self.exchange_rate = exchange_rate
				if self.order_rate:
					self.order_rate_inr = exchange_rate * self.order_rate
				if self.schedule_amount:
					self.schedule_amount_inr =  exchange_rate * self.schedule_amount   
				if self.delivered_amount:
					self.delivered_amount_inr =  exchange_rate * self.delivered_amount   
				if self.pending_amount:
					self.pending_amount_inr =  exchange_rate * self.pending_amount
			else:
				if self.order_rate:
					self.order_rate_inr = self.exchange_rate * self.order_rate
				if self.schedule_amount:
					self.schedule_amount_inr =  self.exchange_rate * self.schedule_amount   
				if self.delivered_amount:
					self.delivered_amount_inr =  self.exchange_rate * self.delivered_amount   
				if self.pending_amount:
					self.pending_amount_inr =  self.exchange_rate * self.pending_amount
				


	def validate(self):
		if self.order_type == "Open":
			current_year = datetime.now().year
			schedule_month = self.schedule_month.upper() if self.schedule_month else ""
			self.schedule_date = datetime.strptime(f"01-{schedule_month}-{current_year}", "%d-%b-%Y")
			self.schedule_year = current_year
		if self.order_type == "Fixed":
			if self.schedule_date:
				if isinstance(self.schedule_date, str):
					schedule_date = datetime.strptime(self.schedule_date, "%Y-%m-%d")
				else:
					schedule_date = self.schedule_date
				self.schedule_year = schedule_date.year
				self.schedule_month = schedule_date.strftime("%b").upper()

		self.pending_qty = self.qty - self.delivered_qty
		self.schedule_amount = self.order_rate * self.qty
		self.delivered_amount = self.order_rate * self.delivered_qty
		self.pending_amount = self.order_rate * (self.qty - self.delivered_qty)
		if self.qty < self.delivered_qty:
			frappe.throw(
				"Cannot set Schedule Quantity less than Received Quantity",
			)
			frappe.throw("Received Quantity-ஐவிட குறைவாக Schedule Quantity அமைக்க முடியாது")

		
		# Function for amendment
		if self.amended_from:
			if not self.revised:
				if frappe.db.exists("Sales Order Schedule", self.amended_from):
					old_doc = frappe.get_doc("Sales Order Schedule", self.amended_from)
					new_row = self.append('revision', {})
					new_row.revised_on = now_datetime()
					new_row.schedule_qty = old_doc.qty
					new_row.revised_schedule_qty = self.qty
			else:
				if self.docstatus == 1:
					return
				child_doc = self.revision[-1]
				child_doc.revised_schedule_qty = self.qty


	def on_submit(self):
		if not self.sales_order_number:
			return

		order_type = frappe.db.get_value("Sales Order", self.sales_order_number, 'customer_order_type')
		so = frappe.get_doc("Sales Order", self.sales_order_number)

		if self.order_type == "Open":
			# Update or Create Schedule Item
			filters = {
				"item_code": self.item_code,
				"schedule_qty": self.qty,
				"parent": self.sales_order_number,
				"order_schedule": ("is", None)
			}
			if frappe.db.exists("Sales Order Schedule Item", filters):
				frappe.db.set_value("Sales Order Schedule Item", filters, "order_schedule", self.name)
			else:
				frappe.get_doc({
					"doctype": "Sales Order Schedule Item",
					"parent": self.sales_order_number,
					"parenttype": "Sales Order",
					"parentfield": "custom_schedule_table",
					"item_code": self.item_code,
					"schedule_date": self.schedule_date,
					"schedule_month": self.schedule_month,
					"schedule_qty": self.qty,
					"delivery_qty": self.delivered_qty,
					"pending_qty": self.pending_qty,
					"order_schedule": self.name
				}).insert(ignore_permissions=True)

		# Update SO Item Qty and Amount if Draft
		if order_type == "Open" and frappe.db.exists("Sales Order", {'name': self.sales_order_number, 'docstatus': 0}):
			s_qty = sum(d.qty for d in frappe.get_all("Sales Order Schedule", {
				"sales_order_number": self.sales_order_number,
				"docstatus": 1,
				"customer_code": self.customer_code,
				"item_code": self.item_code
			}, ["qty"]))

			# Update SO Item
			frappe.db.set_value("Sales Order Item", {"parent": self.sales_order_number, "item_code": self.item_code}, "qty", s_qty)

			# Update SO Doc
			for row in so.items:
				if row.item_code == self.item_code:
					row.qty = s_qty
					row.amount = s_qty * row.rate
			so.save(ignore_permissions=True)

		# Update Open Order Table if SO is Submitted
		if order_type == "Open" and frappe.db.exists("Sales Order", {'name': self.sales_order_number, 'docstatus': 1}):
			open_order = frappe.get_doc("Open Order", {"sales_order_number": self.sales_order_number})
			item_qty = sum(s.qty for s in frappe.get_all("Sales Order Schedule", {
				"sales_order_number": self.sales_order_number,
				"item_code": self.item_code,
				"docstatus": 1
			}, ["qty"]))
			matching_row = next((row for row in open_order.open_order_table if row.item_code == self.item_code), None)
			if matching_row:
				if self.disable_update_items == 1:
					open_order.disable_update_items = 1
				else:
					open_order.disable_update_items = 0
				matching_row.qty = item_qty
				matching_row.amount = item_qty * float(self.order_rate)
				open_order.save(ignore_permissions=True)

		# For Fixed Orders - Validate qty
		if order_type == "Fixed":
			s_qty = sum(d.qty for d in frappe.get_all("Sales Order Schedule", {
				"sales_order_number": self.sales_order_number,
				"item_code": self.item_code,
				"docstatus": 1
			}, ["qty"]))

			po_items = frappe.get_all("Sales Order Item", {
				"parent": self.sales_order_number,
				"item_code": self.item_code
			}, ["qty", "item_code", "idx"])

			if not po_items:
				frappe.throw(f"Item <b>{self.item_code}</b> not found in the Sales Order <b>{self.sales_order_number}</b>")
				frappe.throw(f"Item <b>{self.item_code}</b> Sales Order <b>{self.sales_order_number}</b> இல் கண்டுபிடிக்கப்படவில்லை")
			po_qty = po_items[0]["qty"]
			if s_qty > po_qty:
				frappe.throw(f"Validation failed: Quantity <b>{s_qty}</b> exceeds Sales Order quantity <b>{po_qty}</b> for item <b>{self.item_code}</b>.")
				frappe.throw(f"Validation failed: Item <b>{self.item_code}</b> க்கான, Quantity <b>{s_qty}</b> Sales Order quantity <b>{po_qty}</b> ஐ விட அதிகமாக உள்ளது.")

	def on_update_after_submit(self):
		self.schedule_amount = self.order_rate * self.qty
		self.delivered_amount = self.order_rate * self.delivered_qty
		self.pending_qty = self.qty - self.delivered_qty
		self.pending_amount = self.order_rate * self.pending_qty
  
		if self.sales_order_number:
			self.currency = frappe.db.get_value("Sales Order",{"name":self.sales_order_number},"currency") or ""
			self.db_set('currency', self.currency, update_modified=False)
			exchange_rate = frappe.db.get_value("Sales Order",{"name":self.sales_order_number},"conversion_rate") or ""
			
			if exchange_rate:
				self.exchange_rate = exchange_rate
				self.db_set('exchange_rate', self.exchange_rate, update_modified=False)
				if self.order_rate:
					self.order_rate_inr = exchange_rate * self.order_rate
					self.db_set('order_rate_inr', self.order_rate_inr, update_modified=False)
				if self.schedule_amount:
					self.schedule_amount_inr =  exchange_rate * self.schedule_amount
					self.db_set('schedule_amount_inr', self.schedule_amount_inr, update_modified=False)
				if self.delivered_amount:
					self.delivered_amount_inr =  exchange_rate * self.delivered_amount
					self.db_set('delivered_amount_inr', self.delivered_amount_inr, update_modified=False)   
				if self.pending_amount:
					self.pending_amount_inr =  exchange_rate * self.pending_amount
					self.db_set('pending_amount_inr', self.pending_amount_inr, update_modified=False)
			else:
				if self.order_rate:
					self.order_rate_inr = self.exchange_rate * self.order_rate
					self.db_set('order_rate_inr', self.order_rate_inr, update_modified=False)
				if self.schedule_amount:
					self.schedule_amount_inr =  self.exchange_rate * self.schedule_amount 
					self.db_set('schedule_amount_inr', self.schedule_amount_inr, update_modified=False)  
				if self.delivered_amount:
					self.delivered_amount_inr =  self.exchange_rate * self.delivered_amount 
					self.db_set('delivered_amount_inr', self.delivered_amount_inr, update_modified=False)  
				if self.pending_amount:
					self.pending_amount_inr =  self.exchange_rate * self.pending_amount
					self.db_set('pending_amount_inr', self.pending_amount_inr, update_modified=False)
	 
	 
		
		if frappe.db.exists("Sales Order", self.sales_order_number):
			po = frappe.get_doc("Sales Order", self.sales_order_number)
			if po.custom_schedule_table:
				for row in po.custom_schedule_table:
					if row.order_schedule == self.name:
						row.schedule_qty = self.qty
						row.delivery_qty = self.delivered_qty
						row.pending_qty = self.pending_qty
			else:
				po.append('custom_schedule_table',{
					"doctype": "Sales Order Schedule Item",
					"parent": self.sales_order_number,
					"parenttype": "Sales Order",
					"parentfield": "custom_schedule_table",
					"item_code": self.item_code,
					"schedule_date": self.schedule_date,
					"schedule_month": self.schedule_month,
					"schedule_qty": self.qty,
					"delivery_qty": self.delivered_qty,
					"pending_qty": self.pending_qty,
					"order_schedule": self.name
				})
			po.save(ignore_permissions=True)

	def on_cancel(self):
		# validate_in_process(self)
		if self.qty < self.delivered_qty:
			frappe.throw(
				"The scheduled quantity has been fully received. Please create a new Order Schedule to make changes.",
				title="Revision Not Allowed"
			)
			frappe.throw(
				"Scheduled quantity முழுமையாக பெற்றுக்கொள்ளப்பட்டுள்ளது. மாற்றங்கள் செய்ய புதிய Order Schedule உருவாக்கவும்.",
				title="Revision Not Allowed"
			)

		if self.order_type == "Open":
			if frappe.db.exists("Sales Order",self.sales_order_number):
				po = frappe.get_doc("Sales Order",self.sales_order_number)
				po.custom_schedule_table = [
					row for row in po.custom_schedule_table if row.order_schedule != self.name
				]
				po.save(ignore_permissions=True)
			if frappe.db.exists("Open Order", {"sales_order_number": self.sales_order_number}):
				poo = frappe.get_doc("Open Order", {"sales_order_number": self.sales_order_number})
				for row in poo.open_order_table:
					if row.item_code == self.item_code:
						if row.qty - self.qty > 0:
							row.qty = row.qty - self.qty
						else:
							row.qty = 1
				poo.save(ignore_permissions=True)

def validate_in_process(self):
	in_process = get_in_process_qty(self.item_code)
	if in_process > 0:
		total_schedule_qty = get_total_scheduled_qty(self.item_code, self.schedule_date)
		current_schedule_qty = self.qty
		if (flt(total_schedule_qty - current_schedule_qty) < flt(in_process)) > 0:
			frappe.throw(
				(f"The item <b>{self.item_code}</b> has <b>{in_process}</b> quantity in production process. Please adjust the schedule quantity accordingly."),
				title="Work Order Exists"
			)
@frappe.whitelist()
def return_items(doctype,docname):
	doc = frappe.get_doc(doctype,docname)
	return doc.items

@frappe.whitelist()
def schedule_list(sales, item):
	if sales and item:
		documents = frappe.get_all('Sales Order Schedule', {'sales_order_number': sales, 'item_code': item},
									['schedule_date', 'tentative_plan_1', 'tentative_plan_2', 'qty', 'delivered_qty',
									'pending_qty', 'remarks', 'order_rate'])

		documents = sorted(documents, key=lambda x: x['schedule_date'])
		data = '<table border="1" style="width: 100%;">'
		data += '<tr style="background-color:#D9E2ED;">'
		data += '<td colspan="2" style="text-align:center;"><b>Schedule Month</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Schedule Date</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Tentative Plan - I</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Tentative Plan - II</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Schedule Qty</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Delivered Qty</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Pending Qty</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Remarks</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Cost Price</b></td>'
		data += '</tr>'
		for doc in documents:
			month_string = doc['schedule_date'].strftime('%B')
			data += '<tr>'
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(month_string)
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['schedule_date'].strftime('%d-%m-%Y'))
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['tentative_plan_1'])
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['tentative_plan_2'])
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['qty'])
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['delivered_qty'])
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['pending_qty'])
			if doc['remarks']:
				data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['remarks'])
			else:
				data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format('-')
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['order_rate'])
			data += '</tr>'
		data += '</table>'
		return data

@frappe.whitelist()
def get_qty_rate_so(item,sales):
	so = frappe.db.get_value("Sales Order Item",{"Parent":sales,"item_code":item},["rate"])
	return so


@frappe.whitelist()
def revise_schedule_qty(name, revised_qty, remarks):
	revised_qty = flt(revised_qty)
	doc = frappe.get_doc("Sales Order Schedule", name)
	# validate_in_process_on_revision(doc.item_code, doc.schedule_date, revised_qty, doc.qty)
	if revised_qty < flt(doc.delivered_qty):
		frappe.throw("Cannot set Schedule Quantity less than Delivered Quantity")
		frappe.throw("Delivered Quantity-ஐவிட குறைவாக Schedule Quantity அமைக்க முடியாது")
	doc.append("revision", {
		"revised_on": frappe.utils.now_datetime(),
		"remarks": remarks,
		"schedule_qty": doc.qty,
		"revised_schedule_qty": revised_qty,
		"revised_by": frappe.session.user
	})
	doc.qty = revised_qty
	doc.disable_update_items = 0
	doc.pending_qty = flt(revised_qty) - flt(doc.delivered_qty)
	doc.schedule_amount = flt(revised_qty) * flt(doc.order_rate)
	doc.delivered_amount = flt(doc.delivered_qty) * flt(doc.order_rate)
	doc.pending_amount = (flt(revised_qty) - flt(doc.delivered_qty)) * flt(doc.order_rate)
	doc.save(ignore_permissions=True)
	# frappe.db.set_value("Sales Order Schedule", name, "pending_qty", flt(revised_qty) - flt(doc.delivered_qty))

	self = frappe.get_doc("Sales Order Schedule", name)
	if self.order_type == "Open" and frappe.db.exists("Sales Order", {'name': self.sales_order_number, 'docstatus': 1}):
			open_order = frappe.get_doc("Open Order", {"sales_order_number": self.sales_order_number})
			item_qty = sum(s.qty for s in frappe.get_all("Sales Order Schedule", {
				"sales_order_number": self.sales_order_number,
				"item_code": self.item_code,
				"docstatus": 1
			}, ["qty"]))
			matching_row = next((row for row in open_order.open_order_table if row.item_code == self.item_code), None)
			if matching_row:
				if self.disable_update_items == 1:
					open_order.disable_update_items = 1
				else:
					open_order.disable_update_items = 0
				matching_row.qty = item_qty
				matching_row.amount = item_qty * float(self.order_rate)
				open_order.save(ignore_permissions=True)

@frappe.whitelist()
def get_in_process_qty(item_code):
	in_process = frappe.db.sql("""
		SELECT SUM(qty) FROM `tabWork Order`
		WHERE production_item = %s AND status NOT IN ('Draft', 'Cancelled') AND docstatus = 1
			AND custom_order_type = 'Sales Order'
	""", (item_code,), as_dict=True)[0]['SUM(qty)'] or 0

	return flt(in_process)

def get_total_scheduled_qty(item_code, schedule_date):
	date_str = str(schedule_date)
	date_obj = datetime.strptime(date_str, '%Y-%m-%d')
	last_day = calendar.monthrange(date_obj.year, date_obj.month)[1]
	last_date = date_obj.replace(day=last_day)

	sales_order_schedule_rows = frappe.db.sql("""
		SELECT ppi.custom_sales_order_schedule
		FROM `tabProduction Plan` pp
		INNER JOIN `tabProduction Plan Item` ppi ON pp.name = ppi.parent
		WHERE pp.docstatus = 1 AND ppi.item_code = %s
		AND pp.creation BETWEEN %s AND %s
	""", (item_code, schedule_date, last_date), as_dict=True)

	# Convert comma-separated strings into a list and remove duplicates
	sales_order_schedule_list = []
	for row in sales_order_schedule_rows:
		if row.custom_sales_order_schedule:
			schedules = [s.strip() for s in row.custom_sales_order_schedule.split(',')]
			sales_order_schedule_list.extend(schedules)

	# Remove duplicates while preserving order
	unique_schedules = list(dict.fromkeys(sales_order_schedule_list))
	if not unique_schedules:
		return 0

	total_scheduled_qty = frappe.db.sql("""
		SELECT SUM(qty) FROM `tabSales Order Schedule`
		WHERE item_code = %s AND name in %s AND schedule_date = %s AND docstatus = 1
	""", (item_code, unique_schedules, schedule_date, ), as_dict=True)[0]['SUM(qty)'] or 0

	return flt(total_scheduled_qty)



@frappe.whitelist()
def update_submitted_sales_order_schedules():
	submitted_docs = frappe.get_all("Sales Order Schedule", filters={"docstatus": 1, "schedule_month": "OCT"}, fields=["name"])

	for d in submitted_docs:
		doc = frappe.get_doc("Sales Order Schedule", d.name)
		if not doc.sales_order_number:
			continue

		calculate_inr_values_and_db_set(doc)

	frappe.db.commit()

def calculate_inr_values_and_db_set(doc):
	if not doc.sales_order_number:
		return

	currency = frappe.db.get_value("Sales Order", {"name": doc.sales_order_number}, "currency") or ""
	exchange_rate = frappe.db.get_value("Sales Order", {"name": doc.sales_order_number}, "conversion_rate") or 1

	doc.db_set("currency", currency, update_modified=False)
	doc.db_set("exchange_rate", exchange_rate, update_modified=False)

	# Safe calculations
	if doc.order_rate:
		val = flt(exchange_rate) * flt(doc.order_rate)
		doc.db_set("order_rate_inr", val, update_modified=False)

	if doc.schedule_amount:
		val = flt(exchange_rate) * flt(doc.schedule_amount)
		doc.db_set("schedule_amount_inr", val, update_modified=False)

	if doc.delivered_amount:
		val = flt(exchange_rate) * flt(doc.delivered_amount)
		doc.db_set("delivered_amount_inr", val, update_modified=False)

	if doc.pending_amount:
		val = flt(exchange_rate) * flt(doc.pending_amount)
		doc.db_set("pending_amount_inr", val, update_modified=False)

def validate_in_process_on_revision(item_code, schedule_date, revised_qty, schedule_qty):
	in_process = get_in_process_qty(item_code)
	total_schedule_qty = get_total_scheduled_qty(item_code, schedule_date)
	if in_process > 0:
		if revised_qty < schedule_qty:
			difference = revised_qty - schedule_qty
			difference = -(difference) if difference < 0 else difference
			if (flt(total_schedule_qty - difference) < flt(in_process)) > 0:
				frappe.throw(
					(f"The item <b>{item_code}</b> has <b>{in_process}</b> quantity in production process. Please adjust the schedule quantity accordingly."),
					title="Work Order Exists"
				)

@frappe.whitelist()
def get_schedule_summary_html(customer_code=None , sales_order=None, item_code=None):
	sales_order_schedules = frappe.db.get_all(
		"Sales Order Schedule",
		{
			"customer_code": customer_code,
			"sales_order_number": sales_order,
			"item_code": item_code,
			"docstatus": 1
		},
		["name", "schedule_date", "schedule_month", "qty", "delivered_qty", "pending_qty"],
		order_by="schedule_date desc"
	)

	html = """
		<div style="margin-bottom: 10px;">
			<table width="100%" style="border-collapse: separate; border-spacing: 0; border-radius: 8px; overflow:hidden; box-shadow:0 2px 6px rgba(0,0,0,0.2);">
				<tr style="background-color: #777472;">
					<td class="text-white text-right pr-4 pt-2 pb-2"><b>S#</b></td>
					<td class="text-white pr-2 pt-2 pb-2 pl-5"><b>Schedule Date</b></td>
					<td class="text-white pt-2 pb-2 pl-5"><b>Schedule Month</b></td>
					<td class="text-white text-right pr-2 pt-2 pb-2"><b>Schedule Quantity</b></td>
					<td class="text-white text-right pr-2 pt-2 pb-2"><b>Delivered Quantity</b></td>
					<td class="text-white text-right pr-3 pt-2 pb-2"><b>Pending Quantity</b></td>
				</tr>
	"""
	idx = 0
	for sos in sales_order_schedules:
		schedule_date = formatdate(sos.schedule_date, "dd-mm-yyyy")
		schedule_quantity = indian_format(int(sos.qty) or 0)
		delivered_quantity = indian_format(int(sos.delivered_qty) or 0)
		pending_quantity = indian_format(int(sos.pending_qty) or 0)
		idx += 1
		if idx %2 == 0:
			html += """<tr style="background-color: #e5e8eb;">"""
		else:
			html += "<tr>"
		html += f"""
				<td class="pt-2 pb-2 pr-4 text-right">{idx}</td>
				<td class="pt-2 pb-2 pr-2 pl-5">{schedule_date}</td>
				<td class="pt-2 pb-2 pl-5">{sos.schedule_month}</td>
				<td class="pt-2 pb-2 pr-2 text-right ">{schedule_quantity}</td>
				<td class="pt-2 pb-2 pr-2 text-right">{delivered_quantity}</td>
				<td class="pt-2 pb-2 pr-3 text-right">{pending_quantity}</td>
			</tr>
		"""

	html += """
		</table>
	</div>
	"""
	return html

def indian_format(n):
	s = s2 = str(int(n))
	if len(s) > 3:
		s2 = s[-3:]
		s = s[:-3]
		while len(s) > 2:
			s2 = s[-2:] + "," + s2
			s = s[:-2]
		s2 = s + "," + s2
	return s2

@frappe.whitelist()
def currency_exchange_rate_list(month):
	data = frappe.db.sql("""
						SELECT  currency, exchange_rate FROM `tabCurrency Conversion`  order by currency asc;
						 """, as_dict = 1)
	return data

@frappe.whitelist()
def update_exchange_rates(month, data):
    import json

    rows = json.loads(data)

    seen = set()
    duplicates = set()

    for r in rows:
        currency = r.get("currency")
        if currency in seen:
            duplicates.add(currency)
        seen.add(currency)

    if duplicates:
        frappe.throw(f"Duplicate currency entries found for: {', '.join(duplicates)}")

    for r in rows:
        currency = r.get("currency")
        exchange_rate = float(r.get("exchange_rate"))

        schedules = frappe.db.get_all(
            "Sales Order Schedule",
            filters={
                "currency": currency,
                "schedule_month": month,
                "docstatus": ["!=", 2],
                "exchange_rate": ["!=", exchange_rate]
            },
            fields=["name", "order_rate", "qty", "delivered_qty", "pending_qty"]
        )

        for s in schedules:
            order_rate_inr = exchange_rate * (s.order_rate or 0)
            schedule_amount_inr = exchange_rate * (s.qty or 0)
            delivered_amount_inr = exchange_rate * (s.delivered_qty or 0)
            pending_amount_inr = exchange_rate * (s.pending_qty or 0)

            frappe.db.set_value("Sales Order Schedule", s.name, {
                "exchange_rate": exchange_rate,
                "schedule_amount_inr": schedule_amount_inr,
                "delivered_amount_inr": delivered_amount_inr,
                "order_rate_inr": order_rate_inr,
                "pending_amount_inr": pending_amount_inr
            })

    frappe.db.commit()
    return "ok"


@frappe.whitelist()
def make_order_schedule(sales_order_number, item_code, docname):
    schedules = frappe.db.get_all(
        "Sales Order Schedule",
        {"sales_order_number": sales_order_number, "item_code": item_code, "name": ["!=", docname]},
        ["name", "pending_qty"],
        order_by="schedule_date desc",
        limit=1
    )

    if not schedules:
        return 0, 0, 0

    schedule = schedules[0]
    new_schedule_qty = schedule.get("pending_qty") or 0

    order_rate = frappe.db.get_value(
        "Sales Order Item",
        {"parent": sales_order_number, "item_code": item_code},
        "rate"
    )

    exchange_rate = frappe.db.get_value(
        "Sales Order",
        sales_order_number,
        "conversion_rate"
    )

    return new_schedule_qty, exchange_rate, order_rate


@frappe.whitelist()
def validate_schedule_qty(sales_order_number, item_code, schedule_qty, docname):
    schedules = frappe.db.get_all(
        "Sales Order Schedule",
        {"sales_order_number": sales_order_number, "item_code": item_code, "name": ["!=", docname]},
        ["name", "pending_qty"],
        order_by="schedule_date desc",
        limit=1
    )

    if not schedules:
        return True 

    schedule = schedules[0]
    pending_qty = schedule.get("pending_qty") or 0

    if int(schedule_qty) > int(pending_qty):
        return "error"

    return True

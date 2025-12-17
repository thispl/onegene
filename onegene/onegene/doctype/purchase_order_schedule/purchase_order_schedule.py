# Copyright (c) 2023, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file
from frappe.utils import cstr, add_days, date_diff, getdate
from frappe.utils.background_jobs import enqueue
import datetime
from datetime import datetime
import calendar


from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import frappe
from frappe.query_builder import DocType
import requests
from datetime import date
from time import strptime
import erpnext
import json
from frappe.utils import now, formatdate
from typing import Dict, Optional, Tuple, Union
from frappe import throw,_
from frappe.utils import flt
from frappe.utils import (
	add_days,
	ceil,
	cint,
	comma_and,
	flt,
	get_link_to_form,
	getdate,
	now_datetime,
	datetime,get_first_day,get_last_day,
	nowdate,
	today,
)
from frappe.utils import cstr, cint, getdate, get_last_day, get_first_day, add_days,date_diff
from datetime import date, datetime, timedelta
from erpnext.stock.get_item_details import (
	_get_item_tax_template,
	get_conversion_factor,
	get_item_details,
	get_item_tax_map,
	get_item_warehouse,
)
from frappe.model.workflow import get_workflow_name, is_transition_condition_satisfied

class PurchaseOrderSchedule(Document):
	def before_insert(self):
		if self.purchase_order_number and self.item_code:
			order_rate = frappe.db.get_value(
				"Purchase Order Item", 
				{"parent": self.purchase_order_number, "item_code": self.item_code}, 
				"rate"
			)
			self.order_rate = order_rate or 0
		else:
			self.order_rate = 0

		if self.schedule_date:
			if isinstance(self.schedule_date, str):
				try:
					self.schedule_date = datetime.strptime(self.schedule_date, "%Y-%m-%d %H:%M:%S")
				except ValueError:
					self.schedule_date = datetime.strptime(self.schedule_date, "%Y-%m-%d")
		
			self.schedule_year = self.schedule_date.year
   
	# 	doc = frappe.db.get_value("Purchase Order Schedule",{"item_code":self.item_code, "supplier_name":self.supplier_name,"schedule_month":self.schedule_month,"schedule_year":self.schedule_year},["name", "item_code", "schedule_month"], 
	#    as_dict=True )
	# 	if doc:
	  
	# 		name = f'<a href="/app/purchase-order-schedule/{doc.name}">{doc.name}</a>'
	# 		frappe.throw(_(f'There is already an Sales Order Schedule {name} for the item {doc.item_code} for the month {doc.schedule_month}is available.').format(", ".join(name)))
	  	


	def after_insert(self):
		if self.amended_from:
			if frappe.db.exists("Supplier-DN Schedule Trace", {"purchase_order_schedule", self.amended_from}):
				frappe.db.sql(
					"UPDATE `tabSupplier-DN Schedule Trace` SET purchase_order_schedule = %s WHERE purchase_order_schedule = %s",
					(self.name, self.amended_from),
					as_dict=False
				)
				frappe.db.commit()
	
	def before_save(self):
		if self.purchase_order_number:
			self.currency = frappe.db.get_value("Purchase Order",{"name":self.purchase_order_number},"currency") or ""
			exchange_rate = frappe.db.get_value("Purchase Order",{"name":self.purchase_order_number},"conversion_rate") or ""
			if exchange_rate:
				self.exchange_rate = exchange_rate
				if self.order_rate:
					self.order_rate_inr = exchange_rate * self.order_rate
				if self.schedule_amount:
					self.schedule_amount_inr =  exchange_rate * self.schedule_amount   
				if self.received_amount:
					self.received_amount_inr =  exchange_rate * self.received_amount   
				if self.pending_amount:
					self.pending_amount_inr =  exchange_rate * self.pending_amount
			else:   
				if self.order_rate:
					self.order_rate_inr = self.exchange_rate * self.order_rate
				if self.schedule_amount:
					self.schedule_amount_inr =  self.exchange_rate * self.schedule_amount   
				if self.received_amount:
					self.received_amount_inr =  self.exchange_rate * self.received_amount   
				if self.pending_amount:
					self.pending_amount_inr =  self.exchange_rate * self.pending_amount	
	
	
	

	def validate(self):
		
		po_schedule = frappe.db.get_value("Purchase Order Schedule",{"item_code":self.item_code,"purchase_order_number":self.purchase_order_number,"schedule_month":self.schedule_month, "docstatus": 1},["name", "item_code", "schedule_month", "docstatus"],as_dict=True)
		
		if po_schedule and po_schedule.docstatus != 2:
			if (self.is_new() or po_schedule.name != self.name):
				# frappe.throw(
				# 	f"Purchase Order Schedule for the item {po_schedule.item_code} is already available for the Month {po_schedule.schedule_month}. Instead of this you can revise qty in {po_schedule.name}"
				# )
				frappe.throw(
					f"""<p>
						Purchase Order Schedule for the item <b>{po_schedule.item_code}</b> is already available for the Month <b>{po_schedule.schedule_month}</b>.
						Instead of this, you can revise qty in 
						<a href="/app/purchase-order-schedule/{po_schedule.name}" target="_blank">{po_schedule.name}</a>.
					</p>""",
					
				)
				frappe.throw(
				f"""<p>
					Purchase Order Schedule, item <b>{po_schedule.item_code}</b> க்கானது Month <b>{po_schedule.schedule_month}</b> க்காக ஏற்கனவே உள்ளது.
					இதற்கு பதிலாக, நீங்கள் Qty திருத்தலாம் 
					<a href="/app/purchase-order-schedule/{po_schedule.name}" target="_blank">{po_schedule.name}</a> இல்.
				</p>"""
			)

			
		if self.order_type == "Open":
			current_year = datetime.now().year
			schedule_month = self.schedule_month.upper() if self.schedule_month else ""
			self.schedule_date = datetime.strptime(f"01-{schedule_month}-{current_year}", "%d-%b-%Y")
			self.schedule_year = current_year
		if self.order_type == "Fixed":
			if self.schedule_date:
				if isinstance(self.schedule_date, str):
					schedule_date = datetime.strptime(self.schedule_date, "%Y-%m-%d")
				else:
					schedule_date = self.schedule_date
				self.schedule_year = schedule_date.year
				self.schedule_month = schedule_date.strftime("%b").upper()

		self.pending_qty = self.qty - self.received_qty
		self.schedule_amount = self.order_rate * self.qty
		self.received_amount = self.order_rate * self.received_qty
		self.pending_amount = self.order_rate * (self.qty - self.received_qty)
		if self.qty < self.received_qty:
			frappe.throw(
				"Cannot set Schedule Quantity less than Received Quantity",
			)
			frappe.throw("Received Quantity-ஐ விட குறைவாக Schedule Quantity அமைக்க முடியாது")
		
		# Function for amendment
		if self.amended_from:
			if not self.revised:
				if frappe.db.exists("Purchase Order Schedule", self.amended_from):
					old_doc = frappe.get_doc("Purchase Order Schedule", self.amended_from)
					new_row = self.append('revision', {})
					new_row.revised_on = now_datetime()
					new_row.schedule_qty = old_doc.qty
					new_row.revised_schedule_qty = self.qty
			else:
				if self.docstatus == 1:
					return
				child_doc = self.revision[-1]
				child_doc.revised_schedule_qty = self.qty

	def on_cancel(self):
		if self.qty < self.received_qty:
			frappe.throw(
				"The scheduled quantity has been fully received. Please create a new Order Schedule to make changes.",
				title="Revision Not Allowed"
			)
			frappe.throw(
				"திட்டமிட்ட அளவு முழுவதும் பெறப்பட்டுவிட்டது. மாற்றங்களை செய்ய புதிய Order Schedule உருவாக்கவும்.",
				title="Revision Not Allowed"
			)

		if self.order_type == "Open":
			if frappe.db.exists("Purchase Order",self.purchase_order_number):
				po = frappe.get_doc("Purchase Order",self.purchase_order_number)
				po.custom_schedule_table = [
					row for row in po.custom_schedule_table if row.order_schedule != self.name
				]
				po.save(ignore_permissions=True)
			if frappe.db.exists("Purchase Open Order", {"purchase_order": self.purchase_order_number}):
				poo = frappe.get_doc("Purchase Open Order", {"purchase_order": self.purchase_order_number})
				for row in poo.open_order_table:
					if row.item_code == self.item_code:
						if row.qty - self.qty > 0:
							row.qty = row.qty - self.qty
						else:
							row.qty = 1
				poo.save(ignore_permissions=True)

	def on_submit(self):
		if not self.purchase_order_number:
			return

		order_type = frappe.db.get_value("Purchase Order", self.purchase_order_number, 'custom_order_type')
		po = frappe.get_doc("Purchase Order", self.purchase_order_number)

		if self.order_type == "Open":
			# Update or Create Schedule Item
			filters = {
				"item_code": self.item_code,
				"schedule_qty": self.qty,
				"parent": self.purchase_order_number,
				"order_schedule": ("is", None)
			}
			if frappe.db.exists("Purchase Order Schedule Item", filters):
				frappe.db.set_value("Purchase Order Schedule Item", filters, "order_schedule", self.name)
			else:
				frappe.get_doc({
					"doctype": "Purchase Order Schedule Item",
					"parent": self.purchase_order_number,
					"parenttype": "Purchase Order",
					"parentfield": "custom_schedule_table",
					"item_code": self.item_code,
					"schedule_date": self.schedule_date,
					"schedule_month": self.schedule_month,
					"schedule_qty": self.qty,
					"received_qty": self.received_qty,
					"pending_qty": self.pending_qty,
					"order_schedule": self.name
				}).insert(ignore_permissions=True)

		# Update PO Item Qty and Amount if Draft
		if order_type == "Open" and frappe.db.exists("Purchase Order", {'name': self.purchase_order_number, 'docstatus': 0}):
			s_qty = sum(d.qty for d in frappe.get_all("Purchase Order Schedule", {
				"purchase_order_number": self.purchase_order_number,
				"docstatus": 1,
				"supplier_code": self.supplier_code,
				"item_code": self.item_code
			}, ["qty"]))

			# Update PO Item
			frappe.db.set_value("Purchase Order Item", {"parent": self.purchase_order_number, "item_code": self.item_code}, "qty", s_qty)

			# Update PO Doc
			for row in po.items:
				if row.item_code == self.item_code:
					row.qty = s_qty
					row.amount = s_qty * row.rate
			po.save(ignore_permissions=True)

		# Update Open Order Table if PO is Submitted
		if order_type == "Open" and frappe.db.exists("Purchase Order", {'name': self.purchase_order_number, 'docstatus': 1}):
			open_order = frappe.get_doc("Purchase Open Order", {"purchase_order": self.purchase_order_number})
			item_qty = sum(s.qty for s in frappe.get_all("Purchase Order Schedule", {
				"purchase_order_number": self.purchase_order_number,
				"item_code": self.item_code,
				"docstatus": 1
			}, ["qty"]))
			matching_row = next((row for row in open_order.open_order_table if row.item_code == self.item_code), None)
			if matching_row:
				if self.disable_update_items == 1:
					open_order.disable_update_items = 1
				else:
					open_order.disable_update_items = 0
				matching_row.qty = item_qty
				matching_row.amount = item_qty * float(self.order_rate)
				open_order.save(ignore_permissions=True)

		# For Fixed Orders - Validate qty
		if order_type == "Fixed":
			s_qty = sum(d.qty for d in frappe.get_all("Purchase Order Schedule", {
				"purchase_order_number": self.purchase_order_number,
				"supplier_code": self.supplier_code,
				"item_code": self.item_code,
				"schedule_month": self.schedule_month,
				"docstatus": 1
			}, ["qty"]))

			po_items = frappe.get_all("Purchase Order Item", {
				"parent": self.purchase_order_number,
				"item_code": self.item_code
			}, ["qty", "item_code", "idx"])

			if not po_items:
				frappe.throw(f"Item <b>{self.item_code}</b> not found in the Purchase Order <b>{self.purchase_order_number}</b>")
				frappe.throw(f"Item <b>{self.item_code}</b> Purchase Order <b>{self.purchase_order_number}</b> இல் காணப்படவில்லை")
			po_qty = po_items[0]["qty"]
			if s_qty > po_qty:
				frappe.throw(f"Validation failed: Quantity <b>{s_qty}</b> exceeds Purchase Order quantity <b>{po_qty}</b> for item <b>{self.item_code}</b>.")
				frappe.throw(f"Validation failed: Item <b>{self.item_code}</b> க்கான, Quantity <b>{s_qty}</b> Purchase Order quantity <b>{po_qty}</b> ஐ விட அதிகமாக உள்ளது")

	def on_update_after_submit(self):
		self.schedule_amount = self.order_rate * self.qty
		self.received_amount = self.order_rate * self.received_qty
		self.pending_qty = self.qty - self.received_qty
		self.pending_amount = self.order_rate * self.pending_qty
  
		if self.purchase_order_number:
			self.currency = frappe.db.get_value("Purchase Order",{"name":self.purchase_order_number},"currency") or ""
			self.db_set("currency", self.currency, update_modified=False)
			exchange_rate = frappe.db.get_value("Purchase Order",{"name":self.purchase_order_number},"conversion_rate") or ""
			
			if exchange_rate:
				self.exchange_rate = exchange_rate
				self.db_set("exchange_rate", self.exchange_rate, update_modified=False)
				if self.order_rate:
					self.order_rate_inr = exchange_rate * self.order_rate
					self.db_set('order_rate_inr', self.order_rate_inr, update_modified=False)
				if self.schedule_amount:
					self.schedule_amount_inr =  exchange_rate * self.schedule_amount
					self.db_set('schedule_amount_inr', self.schedule_amount_inr, update_modified=False)
				if self.received_amount:
					self.received_amount_inr =  exchange_rate * self.received_amount 
					self.db_set('received_amount_inr', self.received_amount_inr, update_modified=False)
				if self.pending_amount:
					self.pending_amount_inr =  exchange_rate * self.pending_amount
					self.db_set('pending_amount_inr', self.pending_amount_inr, update_modified=False)
			else:   
				if self.order_rate:
					self.order_rate_inr = self.exchange_rate * self.order_rate
					self.db_set('order_rate_inr', self.order_rate_inr, update_modified=False)
				if self.schedule_amount:
					self.schedule_amount_inr =  self.exchange_rate * self.schedule_amount
					self.db_set('schedule_amount_inr', self.schedule_amount_inr, update_modified=False)
				if self.received_amount:
					self.received_amount_inr =  self.exchange_rate * self.received_amount
					self.db_set('received_amount_inr', self.received_amount_inr, update_modified=False)
				if self.pending_amount:
					self.pending_amount_inr =  self.exchange_rate * self.pending_amount
					self.db_set('pending_amount_inr', self.pending_amount_inr, update_modified=False)


		if frappe.db.exists("Purchase Order", self.purchase_order_number):
			po = frappe.get_doc("Purchase Order", self.purchase_order_number)
			for row in po.custom_schedule_table:
				if row.order_schedule == self.name:
					row.schedule_qty = self.qty
					row.received_qty = self.received_qty
					row.pending_qty = self.pending_qty
			po.save(ignore_permissions=True)



@frappe.whitelist()
def return_items(doctype,docname):
	doc = frappe.get_doc(doctype,docname)
	return doc.items

@frappe.whitelist()
def schedule_list(purchase, item):
	if purchase and item:
		documents = frappe.get_all('Purchase Order Schedule', {'purchase_order_number': purchase, 'item_code': item, "docstatus": 1},
									['schedule_date', 'tentative_plan_1', 'tentative_plan_2', 'qty', 'delivered_qty',
									'pending_qty', 'remarks', 'order_rate'])

		documents = sorted(documents, key=lambda x: x['schedule_date'])
		data = '<table border="1" style="width: 100%;">'
		data += '<tr style="background-color:#D9E2ED;">'
		data += '<td colspan="2" style="text-align:center;"><b>Schedule Month</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Schedule Date</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Tentative Plan - I</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Tentative Plan - II</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Schedule Qty</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Delivered Qty</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Pending Qty</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Remarks</b></td>'
		data += '<td colspan="2" style="text-align:center;"><b>Cost Price</b></td>'
		data += '</tr>'
		for doc in documents:
			month_string = doc['schedule_date'].strftime('%B')
			data += '<tr>'
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(month_string)
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['schedule_date'].strftime('%d-%m-%Y'))
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['tentative_plan_1'])
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['tentative_plan_2'])
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['qty'])
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['delivered_qty'])
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['pending_qty'])
			if doc['remarks']:
				data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['remarks'])
			else:
				data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format('-')
			data += '<td colspan="2" style="text-align:center;"><b>{}</b></td>'.format(doc['order_rate'])
			data += '</tr>'
		data += '</table>'
		return data

@frappe.whitelist()
def get_qty_rate_so(item,purchase):
	so = frappe.db.get_value("Purchase Order Item",{"Parent":purchase,"item_code":item},["rate"])
	return so

@frappe.whitelist()
def get_qty_rate_test():
	s_qty = 0
	po_exist = frappe.db.exists('Purchase Order Schedule',{"purchase_order_number":'PUR-ORD-2025-00012',"supplier_code":'Amar',"item_code":'RDAE1T6SPRWON243'})
	if po_exist:
		exist_po = frappe.get_all("Purchase Order Schedule",{"purchase_order_number":'PUR-ORD-2025-00012',"supplier_code":'Amar',"item_code":'RDAE1T6SPRWON243'},["*"])
		for i in exist_po:
			old_qty = i.qty
			s_qty += old_qty
		print(s_qty)
		frappe.db.set_value("Purchase Order Item",{"parent":'PUR-ORD-2025-00012',"item_code":'RDAE1T6SPRWON243'},"qty",s_qty)			

@frappe.whitelist()
def update_qty_in_open_order(qty, purchase_order, item_code):
	qty = float(qty)
	if frappe.db.exists("Purchase Open Order", {"purchase_order": purchase_order}):
		poo = frappe.get_doc("Purchase Open Order", {"purchase_order": purchase_order})
		for row in poo.open_order_table:
			if row.item_code == item_code:
				row.qty = row.qty + qty
		poo.save(ignore_permissions=True)
		return "ok"

@frappe.whitelist()
def revise_schedule_qty(name, revised_qty, remarks):
	revised_qty = flt(revised_qty)
	pos_received_qty = frappe.db.get_value("Purchase Order Schedule", name,"received_qty")
	received_qty = flt(pos_received_qty)
	if received_qty == revised_qty:
		pass
	if revised_qty < received_qty :
		frappe.throw("Cannot set Schedule Quantity less than Received Quantity")
		frappe.throw("Received Quantity-ஐ விட குறைவாக Schedule Quantity அமைக்க முடியாது")

	doc = frappe.get_doc("Purchase Order Schedule", name)
	doc.append("revision", {
		"revised_on": frappe.utils.now_datetime(),
		"remarks": remarks,
		"schedule_qty": doc.qty,
		"revised_schedule_qty": revised_qty,
		"revised_by": frappe.session.user
	})
	doc.qty = revised_qty
	doc.disable_update_items = 0
	doc.pending_qty = flt(revised_qty) - flt(doc.received_qty)
	doc.schedule_amount = flt(revised_qty) * flt(doc.order_rate)
	doc.received_amount = flt(doc.received_qty) * flt(doc.order_rate)
	doc.pending_amount = (flt(revised_qty) - flt(doc.received_qty)) * flt(doc.order_rate)
	doc.save(ignore_permissions=True)
	# frappe.db.set_value("Purchase Order Schedule", name, "pending_qty", flt(revised_qty) - flt(doc.received_qty))

	self = frappe.get_doc("Purchase Order Schedule", name)
	if self.order_type == "Open" and frappe.db.exists("Purchase Order", {'name': self.purchase_order_number, 'docstatus': 1}):
			open_order = frappe.get_doc("Purchase Open Order", {"purchase_order": self.purchase_order_number})
			item_qty = sum(s.qty for s in frappe.get_all("Purchase Order Schedule", {
				"purchase_order_number": self.purchase_order_number,
				"item_code": self.item_code,
				"docstatus": 1
			}, ["qty"]))
			item_qty = 1 if item_qty == 0 else item_qty
			matching_row = next((row for row in open_order.open_order_table if row.item_code == self.item_code), None)
			if matching_row:
				if self.disable_update_items == 1:
					open_order.disable_update_items = 1
				else:
					open_order.disable_update_items = 0
				matching_row.qty = item_qty
				matching_row.amount = item_qty * float(self.order_rate)
				open_order.save(ignore_permissions=True)

@frappe.whitelist()
def get_query_for_item_table(doctype, txt, searchfield, start, page_len, filters):
	po = filters.get("purchase_order")
	if not po:
		return []

	return frappe.db.sql("""
		SELECT poi.item_code, i.item_name
		FROM `tabPurchase Order Item` poi
		JOIN `tabItem` i ON poi.item_code = i.name
		WHERE poi.parent = %s AND (poi.item_code LIKE %s OR i.item_name LIKE %s)
		LIMIT %s OFFSET %s
	""", (po, f"%{txt}%", f"%{txt}%", page_len, start))

@frappe.whitelist()
def get_item_details(item_code):
	item_name, item_group = frappe.db.get_value("Item", item_code, ["item_name", "item_group"])
	return item_name, item_group

@frappe.whitelist()
def is_item_in_purchase_order(purchase_order, item_code):
	if not purchase_order or not item_code:
		return False

	return frappe.db.exists("Purchase Order Item", {
		"parent": purchase_order,
		"item_code": item_code
	}) is not None




@frappe.whitelist()
def update_submitted_purchase_order_schedules():
	submitted_docs = frappe.get_all("Purchase Order Schedule", filters={"docstatus": 1, "schedule_month": "OCT"}, fields=["name"])

	for d in submitted_docs:
		doc = frappe.get_doc("Purchase Order Schedule", d.name)
		if not doc.purchase_order_number:
			continue

		calculate_inr_values_and_db_set(doc)

	frappe.db.commit()

def calculate_inr_values_and_db_set(doc):
	if not doc.purchase_order_number:
		return

	currency = frappe.db.get_value("Purchase Order", {"name": doc.purchase_order_number}, "currency") or ""
	exchange_rate = frappe.db.get_value("Purchase Order", {"name": doc.purchase_order_number}, "conversion_rate") or 1

	doc.db_set("currency", currency, update_modified=False)
	doc.db_set("exchange_rate", exchange_rate, update_modified=False)

	# Safe calculations
	if doc.order_rate:
		val = flt(exchange_rate) * flt(doc.order_rate)
		doc.db_set("order_rate_inr", val, update_modified=False)

	if doc.schedule_amount:
		val = flt(exchange_rate) * flt(doc.schedule_amount)
		doc.db_set("schedule_amount_inr", val, update_modified=False)

	if doc.received_amount:
		val = flt(exchange_rate) * flt(doc.received_amount)
		doc.db_set("received_amount_inr", val, update_modified=False)

	if doc.pending_amount:
		val = flt(exchange_rate) * flt(doc.pending_amount)
		doc.db_set("pending_amount_inr", val, update_modified=False)


@frappe.whitelist()
def get_schedule_summary_html(supplier_code=None, purchase_order=None, item_code=None):
	purchase_order_schedules = frappe.db.get_all(
		"Purchase Order Schedule",
		{
			"supplier_code": supplier_code,
			"purchase_order_number": purchase_order,
			"item_code": item_code,
			"docstatus": 1
		},
		["name", "schedule_date", "schedule_month", "qty", "received_qty", "pending_qty"],
		order_by="schedule_date desc"
	)

	html = """
		<div style="margin-bottom: 10px;">
			<table width="100%" style="border-collapse: separate; border-spacing: 0; border-radius: 8px; overflow:hidden; box-shadow:0 2px 6px rgba(0,0,0,0.2);">
				<tr style="background-color: #777472;">
					<td class="text-white text-right pr-4 pt-2 pb-2"><b>S#</b></td>
					<td class="text-white pr-2 pt-2 pb-2 pl-5"><b>Schedule Date</b></td>
					<td class="text-white pt-2 pb-2 pl-5"><b>Schedule Month</b></td>
					<td class="text-white text-right pr-2 pt-2 pb-2"><b>Schedule Quantity</b></td>
					<td class="text-white text-right pr-2 pt-2 pb-2"><b>Received Quantity</b></td>
					<td class="text-white text-right pr-3 pt-2 pb-2"><b>Pending Quantity</b></td>
				</tr>
	"""
	idx = 0
	for sos in purchase_order_schedules:
		schedule_date = formatdate(sos.schedule_date, "dd-mm-yyyy")
		schedule_quantity = indian_format(int(sos.qty) or 0)
		delivered_quantity = indian_format(int(sos.received_qty) or 0)
		pending_quantity = indian_format(int(sos.pending_qty) or 0)
		idx += 1
		html += f"""<tr class="schedule-row" data-schedule="{sos.name}" style="cursor:pointer;{ 'background-color:#e5e8eb;' if idx%2==0 else ''}">"""
		html += f"""
				<td class="pt-2 pb-2 pr-4 text-right">{idx}</td>
				<td class="pt-2 pb-2 pr-2 pl-5">{schedule_date}</td>
				<td class="pt-2 pb-2 pl-5">{sos.schedule_month}</td>
				<td class="pt-2 pb-2 pr-2 text-right ">{schedule_quantity}</td>
				<td class="pt-2 pb-2 pr-2 text-right">{delivered_quantity}</td>
				<td class="pt-2 pb-2 pr-3 text-right">{pending_quantity}</td>
			</tr>
		"""

	html += """
		</table>
	</div>
	"""
	return html

def indian_format(n):
	s = s2 = str(int(n))
	if len(s) > 3:
		s2 = s[-3:]
		s = s[:-3]
		while len(s) > 2:
			s2 = s[-2:] + "," + s2
			s = s[:-2]
		s2 = s + "," + s2
	return s2

@frappe.whitelist()
def get_received_breakdown(doctype, docname, party, item_code):
	start_date = getdate(frappe.db.get_value(doctype, docname, "schedule_date"))
	end_day = calendar.monthrange(start_date.year, start_date.month)[1]
	end_date = start_date.replace(day=end_day)
	if doctype == "Purchase Order Schedule":
		data = frappe.db.sql("""
				SELECT asn.name as name, sdni.dis_qty as qty, DATE(asn.datetime) as date
				FROM `tabAdvance Shipping Note` asn
				INNER JOIN `tabSupplier-DN Item` sdni
					ON sdni.parent = asn.name 
				WHERE asn.supplier = %s AND sdni.item_code = %s AND
					workflow_state NOT IN ("Cancelled", "Draft") AND
					DATE(asn.datetime) BETWEEN %s AND %s
				ORDER BY asn.datetime
				""",(party, item_code, start_date, end_date), as_dict=1)

	return data 


def update_received_qty():
	# correct the received qty

	purchase_order_schedules = frappe.db.get_all("Purchase Order Schedule", {"docstatus": 1}, ["name", "supplier_name", "item_code", "schedule_date", "exchange_rate", "received_qty", "order_rate", "order_rate_inr", "qty"])
	for pos in purchase_order_schedules:
		start_date = pos.schedule_date
		end_day = calendar.monthrange(start_date.year, start_date.month)[1]
		end_date = start_date.replace(day=end_day)

		data = frappe.db.sql("""
				SELECT sum(sdni.dis_qty) as qty
				FROM `tabAdvance Shipping Note` asn
				INNER JOIN `tabSupplier-DN Item` sdni
					ON sdni.parent = asn.name 
				WHERE asn.supplier = %s AND sdni.item_code = %s AND
					workflow_state NOT IN ("Cancelled", "Draft") AND
					DATE(asn.datetime) BETWEEN %s AND %s
				ORDER BY asn.datetime
				""",(pos.supplier_name, pos.item_code, start_date, end_date))[0][0] or 0

		print([pos.name, pos.received_qty, data or 0])
		pending_qty = pos.qty - pos.received_qty
		
		update_vals = {
			"received_qty": data,
			"received_amount": data * pos.order_rate,
			"received_amount_inr": data * pos.order_rate_inr,
			"pending_qty": pending_qty,
			"pending_amount": pending_qty * pos.order_rate,
			"pending_amount_inr": pending_qty * pos.order_rate_inr,
		}

		frappe.db.set_value("Purchase Order Schedule", pos.name, update_vals)
	return 200  
  
@frappe.whitelist()
def make_order_schedule(purchase_order_number, item_code, docname):
    schedules = frappe.db.get_all(
        "Purchase Order Schedule",
        {"purchase_order_number": purchase_order_number, "item_code": item_code, "name": ["!=", docname]},
        ["name", "pending_qty"],
        order_by="schedule_date desc",
        limit=1
    )

    if not schedules:
        return 0, 0, 0

    schedule = schedules[0]
    new_schedule_qty = schedule.get("pending_qty") or 0

    order_rate = frappe.db.get_value(
        "Purchase Order Item",
        {"parent": purchase_order_number, "item_code": item_code},
        "rate"
    )

    exchange_rate = frappe.db.get_value(
        "Purchase Order",
        purchase_order_number,
        "conversion_rate"
    )

    return new_schedule_qty, exchange_rate, order_rate


@frappe.whitelist()
def validate_schedule_qty(purchase_order_number, item_code, schedule_qty, docname):
    schedules = frappe.db.get_all(
        "Purchase Order Schedule",
        {"purchase_order_number": purchase_order_number, "item_code": item_code, "name": ["!=", docname]},
        ["name", "pending_qty"],
        order_by="schedule_date desc",
        limit=1
    )

    if not schedules:
        return True 

    schedule = schedules[0]
    pending_qty = schedule.get("pending_qty") or 0

    if int(schedule_qty) > int(pending_qty):
        return "error"

    return True

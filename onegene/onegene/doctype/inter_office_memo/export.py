import frappe
from openpyxl import Workbook
from openpyxl import load_workbook
from io import BytesIO
import frappe
from frappe.utils import flt 

@frappe.whitelist()
def export_empty_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inter Office Memo"

    headers = ["Customer Code", "Item Code", "Order Number", "Revised Schedule Quantity"]
    ws.append(headers)

    column_widths = [20, 15, 20, 30]  
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    file_data = BytesIO()
    wb.save(file_data)
    file_data.seek(0)

    frappe.local.response.filename = "Inter_Office_Memo.xlsx"
    frappe.local.response.filecontent = file_data.read()
    frappe.local.response.type = "download"
    
from openpyxl import Workbook
from io import BytesIO
import frappe

@frappe.whitelist()
def export_excel(name):
    doc = frappe.get_doc("Inter Office Memo", name)
    exp = []

    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inter Office Memo"

    
    headers = ["Customer", "Customer Code", "Customer Type", "Item Code", "Item Name", "Item Group", "Sales Order", 
               "Order Rate", "Currency", "Exchange Rate", "Current Schedule Qty", "Current Schedule Value", 
               "Revised Schedule Qty", "Revised Schedule Value", "Difference Qty", "Difference Value"]
    ws.append(headers)

   
    for i in doc.approval_schdule_increase:
        row = [
            i.customer,
            i.customer_code,
            i.customer_type,
            i.part_no,
            i.part_name,
            i.item_group,
            i.sales_order,
            i.order_rate,
            i.currency,
            i.exchange_rate,
            i.current_schedule_value,
            i.current_schedule_value1,
            i.revised_schedule_value,
            i.revised_schedule_value1,
            i.difference,
            i.difference_value
        ]
        ws.append(row)  

    
    column_widths = [40, 20, 20, 20, 30, 20, 25, 10, 10, 15, 20, 20, 20, 20, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    
    file_data = BytesIO()
    wb.save(file_data)
    file_data.seek(0)

    
    frappe.local.response.filename = "Inter_Office_Memo.xlsx"
    frappe.local.response.filecontent = file_data.read()
    frappe.local.response.type = "download"


@frappe.whitelist()
def export_excel_mpl(name):
    doc = frappe.get_doc("Inter Office Memo", name)
    exp = []

    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inter Office Memo"

    
    headers = ["Supplier", "Supplier Type", "Supplier Code", "Item Code", "Item Name", "Item Group", "Order No", 
               "Order Rate", "Currency", "Exchange Rate", "Current Schedule Qty", "Current Schedule Value", 
               "Revised Schedule Qty", "Revised Schedule Value", "Difference Qty", "Difference Value","Reason for Request"]
    ws.append(headers)

   
    for i in doc.schedule_revised:
        row = [
            i.supplier,
            i.supplier_type,
            i.supplier_code,
            i.part_no,
            i.part_name,
            i.item_group,
            i.purchase_order,
            i.order_rate,
            i.currency,
            i.exchange_rate,
            i.current_schedule_value,
            i.current_schedule_value1,
            i.revised_schedule_value,
            i.revised_schedule_value1,
            i.difference,
            i.difference_value,
            i.reason_for_request
        ]
        ws.append(row)  

    
    column_widths = [40, 20, 20, 20, 30, 20, 25, 10, 10, 15, 20, 20, 20, 20, 15, 15,60]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    
    file_data = BytesIO()
    wb.save(file_data)
    file_data.seek(0)

    
    frappe.local.response.filename = "Inter_Office_Memo.xlsx"
    frappe.local.response.filecontent = file_data.read()
    frappe.local.response.type = "download"

    


@frappe.whitelist()
def process_approval_schedule_sheet(docname, file_url,month):
    import pandas as pd
    import frappe
    import math
    from frappe.utils import flt

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    path = file_doc.get_full_path()

    try:
        df = pd.read_excel(path)
    except Exception as e:
        frappe.throw(f"Error reading Excel file: {e}")

    data_list = []

    for idx, row in df.iterrows():
        customer_code = str(row.get("Customer Code")).strip() if row.get("Customer Code") else None
        item_code = str(row.get("Item Code"))

        if not customer_code:
            frappe.throw(f"Row {idx+2}: Customer Code is missing in sheet.")
        # if item_code is None or (isinstance(item_code, float) and math.isnan(item_code)) or not str(item_code).strip():
        #     frappe.throw(f"Row {idx+2}: Item Code is missing in sheet.")
        cell_value1 = row.get("Item Code")
        if pd.isna(cell_value1):
            item_code = ""
        else:
            item_code = str(cell_value1).strip()
        if not item_code:
            frappe.throw(f"Row {idx+2}: Item Code is missing in sheet.")
        if pd.isna(row.get("Order Number")) or not str(row.get("Order Number")).strip():
            frappe.throw(f"Row {idx+2}: Sales Order Number is missing in sheet.")
        if pd.isna(row.get("Revised Schedule Quantity")):
            frappe.throw(f"Row {idx+2}: Revised Schedule Quantity is missing in sheet.")
        item_code = str(item_code).strip()
        customer_doc = frappe.db.get_value("Customer", {"customer_code": customer_code}, ["name","customer_name", "customer_code","customer_group","default_currency"], as_dict=True)
        customer = customer_doc.customer_name
        currency  =customer_doc.default_currency
        if currency == "INR":
            exchange_rate = 1
        else:
            exchange_rate_doc = frappe.db.get_value(
                "Currency Exchange",
                filters={"from_currency": currency, "to_currency": "INR"},
                fieldname=["exchange_rate"],
                order_by="date desc",
                as_dict=True
            )
            if exchange_rate_doc:
                exchange_rate = float(exchange_rate_doc.exchange_rate)
            else:
                exchange_rate = 0
        if not customer_doc:
            frappe.throw(f"Row {idx+2}: Customer Code does not exist in system.")

        if customer_doc.customer_code != customer_code:
            frappe.throw(f"Row {idx+2}: Customer Code does not match the record in system.")

        

        item_row = frappe.db.sql("""
            SELECT soi.item_code, soi.item_name
            FROM `tabSales Order` so
            INNER JOIN `tabSales Order Item` soi ON soi.parent = so.name
            WHERE so.customer = %s
            AND so.docstatus = 1
            AND soi.item_code = %s
            LIMIT 1
        """, (customer_doc.name, item_code), as_dict=True)

        if not item_row:
            frappe.throw(f"Row {idx+2}: Item {item_code} not found in Sales Order for Customer Code {customer_code}")

        # sales_order = (row.get("Order Number") or "").strip()
        cell_value = row.get("Order Number")
        if pd.isna(cell_value):
            sales_order = ""
        else:
            sales_order = str(cell_value).strip()
        if not sales_order:
            frappe.throw(f"Row {idx+2}: Sales Order Number is missing in sheet.")
        
        item_code = (item_code or "").strip()

        schedule_data = frappe.db.get_value(
            "Sales Order Schedule",
            {"sales_order_number": sales_order, "item_code": item_code,"schedule_month": month,"docstatus":1},
            ["qty", "schedule_amount", "order_rate"],
            as_dict=True
        )
        qty = schedule_amount = order_rate = 0
        item_data = frappe.db.get_value("Sales Order Item",{"parent": sales_order, "item_code": item_code},["rate"],as_dict=True)
        if schedule_data:
            qty = schedule_data.qty or 0
            schedule_amount = schedule_data.schedule_amount or 0
            order_rate = schedule_data.order_rate or 0
        else:
            order_rate = item_data.rate or 0
            frappe.logger().info(f" No Sales Order Schedule found for SO={sales_order}, Item={item_code}")
        

        revised_qty = row.get("Revised Schedule Quantity")
        if pd.isna(revised_qty):
            frappe.throw(f"Row {idx+2}: Revised Schedule Quantity is missing in sheet.")
        revised_qty = flt(revised_qty)
        current_qty = flt(qty or 0)
        current_value = flt(schedule_amount or 0)
        order_rate = flt(order_rate or 0)
        exchange_rate = flt(exchange_rate or 0)
        revised_value = flt(revised_qty * order_rate)
        order_rate_inr = flt(order_rate * exchange_rate)
        current_schedule_value_inr = flt(current_value * exchange_rate)
        revised_schedule_valueinr = flt(revised_value * exchange_rate)
        data_list.append({
            "customer" : customer,
            "customer_code": customer_code,
            "customer_type": customer_doc.customer_group,
            "part_no": item_code,
            "part_name": item_row[0].item_name,
            "sales_order": row.get("Order Number"),
            "order_rate": order_rate,
            "order_rate_inr": order_rate_inr,
            "currency":currency,
            "exchange_rate": exchange_rate,
            "current_schedule_value": current_qty,
            "current_schedule_value1": current_value,
            "current_schedule_value_inr":current_schedule_value_inr,
            "revised_schedule_value": revised_qty,
            "revised_schedule_value1": revised_value,
            "revised_schedule_valueinr": revised_schedule_valueinr,
            "difference": revised_qty - current_qty,
            "difference_value": revised_value - current_value,
            "difference_value_inr": revised_schedule_valueinr - current_schedule_value_inr
        })

    if not data_list:
        return {"status": "fail", "message": "No valid data found to import", "data": []}

    return {
        "status": "success",
        "message": f"Sheet processed successfully. {len(data_list)} rows imported.",
        "data": data_list,
    }


@frappe.whitelist()
def export_empty_excel_mpl():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inter Office Memo"

    headers = ["Supplier Code", "Item Code", "Purchase Order Number", "Revised Schedule Quantity","Reason for Request"]
    ws.append(headers)

    column_widths = [25, 20, 15, 20, 30 , 30]  
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    file_data = BytesIO()
    wb.save(file_data)
    file_data.seek(0)

    frappe.local.response.filename = "Inter_Office_Memo.xlsx"
    frappe.local.response.filecontent = file_data.read()
    frappe.local.response.type = "download"



@frappe.whitelist()
def process_approval_schedule_sheet_supplier(docname, file_url,month):
    import pandas as pd
    import frappe
    from frappe.utils import flt

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    path = file_doc.get_full_path()

    try:
        df = pd.read_excel(path)
    except Exception as e:
        frappe.throw(f"Error reading Excel file: {e}")

    data_list = []

    for idx, row in df.iterrows():
        supplier_code = str(row.get("Supplier Code")).strip() if row.get("Supplier Code") else None
        item_code = str(row.get("Item Code"))

        if not supplier_code:
            frappe.throw(f"Row {idx+2}: Supplier Code is missing in sheet.")
        cell_value1 = row.get("Item Code")
        if pd.isna(cell_value1):
            item_code = ""
        else:
            item_code = str(cell_value1).strip()
        if not item_code:
            frappe.throw(f"Row {idx+2}: Item Code is missing in sheet.")
        cell_value2 = row.get("Purchase Order Number")
        if pd.isna(cell_value2):
            purchase_order = ""
        else:
            purchase_order = str(cell_value2).strip()
        if not purchase_order:
            frappe.throw(f"Row {idx+2}: Purchase Order Number is missing in sheet.")
        if pd.isna(row.get("Revised Schedule Quantity")):
            frappe.throw(f"Row {idx+2}: Revised Schedule Quantity is missing in sheet.")
        item_code = str(item_code).strip()
        supplier_doc = frappe.db.get_value(
            "Supplier", {"supplier_code": supplier_code},
            ["name", "supplier_name", "supplier_type","supplier_group", "supplier_code","default_currency"],
            as_dict=True
        ) 
        supplier = supplier_doc.name
        currency  =supplier_doc.default_currency
        # currency = "EUR"
        if currency == "INR":
            exchange_rate = 1
        else:
            exchange_rate_doc = frappe.db.get_value(
                "Currency Exchange",
                filters={"from_currency": currency, "to_currency": "INR"},
                fieldname=["exchange_rate"],
                order_by="date desc",
                as_dict=True
            )
            if exchange_rate_doc:
                exchange_rate = float(exchange_rate_doc.exchange_rate)
            else:
                exchange_rate = 0
        if not supplier_doc:
            frappe.throw(f"Row {idx+2}: Supplier Code does not exist in system.")

        # if not supplier_code:
        #     frappe.throw(f"Row {idx+2}: Supplier Code is missing for Supplier: {supplier_name}")
        if supplier_doc.supplier_code != supplier_code:
            frappe.throw(f"Row {idx+2}: Supplier code does not match the record in system.")

        # if not item_code:
        #     frappe.throw(f"Row {idx+2}: Item Code is missing for Supplier: {supplier_name}")

        item_row = frappe.db.sql("""
            SELECT poi.item_code, poi.item_name
            FROM `tabPurchase Order` po
            INNER JOIN `tabPurchase Order Item` poi ON poi.parent = po.name
            WHERE po.supplier = %s
            AND po.docstatus = 1
            AND poi.item_code = %s
            LIMIT 1
        """, (supplier_doc.name, item_code), as_dict=True)

        if not item_row:
            frappe.throw(f"Row {idx+2}: Item {item_code} not found in Purchase Order for Supplier {supplier_code}")

        # purchase_order = (row.get("Purchase Order Number") or "").strip()

        cell_value = row.get("Purchase Order Number")
        if pd.isna(cell_value):
            purchase_order = ""
        else:
            purchase_order = str(cell_value).strip()
        if not purchase_order:
            frappe.throw(f"Row {idx+2}: Purchase Order Number is missing in sheet.")

        schedule_data = frappe.db.get_value(
            "Purchase Order Schedule",
            {"purchase_order_number": purchase_order, "item_code": item_code,"schedule_month": month,"docstatus":1},
            ["qty", "schedule_amount", "order_rate"],
            as_dict=True
        )

        qty = schedule_amount = order_rate = 0
        item_data = frappe.db.get_value("Purchase Order Item",{"parent": purchase_order, "item_code": item_code},["rate"],as_dict=True)

        if schedule_data:
            qty = schedule_data.qty or 0
            schedule_amount = schedule_data.schedule_amount or 0
            order_rate = schedule_data.order_rate or 0
        else:
            order_rate = item_data.rate or 0
            frappe.logger().info(f" No Purchase Order Schedule found for PO={purchase_order}, Item={item_code}")

        # revised_qty = flt(row.get("Revised Schedule Quantity") or 0)
        revised_qty = row.get("Revised Schedule Quantity")
        if pd.isna(revised_qty):
            frappe.throw(f"Row {idx+2}: Revised Schedule Quantity is missing in sheet.")
        revised_qty = flt(revised_qty)
        current_qty = flt(qty)
        current_value = flt(schedule_amount)
        order_rate = flt(order_rate)
        exchange_rate = flt(exchange_rate or 0)
        revised_value = flt(revised_qty * order_rate)
        order_rate_inr = flt(order_rate * exchange_rate)
        current_schedule_value_inr = flt(current_value * exchange_rate)
        revised_schedule_valueinr = flt(revised_value * exchange_rate)
        reason_for_request = str(row.get("Reason for Request") or "").strip()
        data_list.append({
            "supplier": supplier,
            "supplier_code": supplier_code,
            "supplier_type": supplier_doc.supplier_group,
            "part_no": item_code,
            "part_name": item_row[0].item_name,
            "purchase_order": purchase_order,
            "order_rate": order_rate,
            "order_rateinr":order_rate_inr,
            "currency":currency,
            "exchange_rate": exchange_rate,
            "current_schedule_value": current_qty,
            "current_schedule_value1": current_value,
            "current_schedule_valueinr":current_schedule_value_inr,
            "revised_schedule_value": revised_qty,
            "revised_schedule_value1": revised_value,
            "revised_schedule_valueinr":revised_schedule_valueinr,
            "difference": revised_qty - current_qty,
            "difference_value": revised_value - current_value,
            "difference_valueinr": revised_schedule_valueinr - current_schedule_value_inr,
            "reason_for_request":reason_for_request,
            
        })

    if not data_list:
        return {"status": "fail", "message": "No valid data found to import", "data": []}

    return {
        "status": "success",
        "message": f"Sheet processed successfully. {len(data_list)} rows imported.",
        "data": data_list,
    }

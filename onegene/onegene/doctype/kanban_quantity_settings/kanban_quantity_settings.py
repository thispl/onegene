# Copyright (c) 2023, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
from frappe.utils.file_manager import get_file
from frappe.utils import cstr, add_days, date_diff, getdate
from frappe.utils.background_jobs import enqueue
import datetime
from datetime import datetime
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import openpyxl
import xlrd
import json
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types


class KanbanQuantitySettings(Document):
	pass
@frappe.whitelist()
def template_sheet():
	args = frappe.local.form_dict
	filename = args.name
	test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
	args = frappe.local.form_dict
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
	ws = wb.create_sheet(sheet_name, 0)
	ws.append(["Item Code","FG Kanban Qty","SFG Days","Today Production Plan","Tentative Plan 1","Tentative Plan 2"])
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


@frappe.whitelist()
def enqueue_upload(file):
	file = get_file(file)
	pps = read_xlsx_file_from_attached_file(fcontent=file[1])
	for idx, pp in enumerate(pps, start=1):
		if pp[0] != 'Item Code':
			item_name = frappe.db.get_value("Item",pp[0],'item_name')
			item_group = frappe.db.get_value("Item",pp[0],'item_group')
			exists = frappe.db.exists("Kanban Quantity",{"item_name":item_name,"item_code":pp[0]})
			# frappe.errprint(exists)
			# frappe.errprint(exists)
			if exists:		
				new_doc = frappe.get_doc('Kanban Quantity',{'item_code',pp[0]})
				new_doc.item_code = pp[0]
				new_doc.item_name = item_name
				new_doc.item_group = item_group
				new_doc.fg_kanban_qty = pp[1]
				new_doc.sfg_days = pp[2]
				new_doc.today_production_plan = pp[3]
				new_doc.tentative_plan_i = pp[4]
				new_doc.tentative_plan_ii = pp[5]
				new_doc.save(ignore_permissions=True)
			else:
				new_doc = frappe.new_doc('Kanban Quantity')
				new_doc.item_code = pp[0]
				new_doc.item_name = item_name
				new_doc.item_group = item_group
				new_doc.fg_kanban_qty = pp[1]
				new_doc.sfg_days = pp[2]
				new_doc.today_production_plan = pp[3]
				new_doc.tentative_plan_i = pp[4]
				new_doc.tentative_plan_ii = pp[5]
				new_doc.save(ignore_permissions=True)

@frappe.whitelist()
def get_data(file):
    data = """<table class=table table-bordered >
    <tr><td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>S.NO</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Item Code</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Item Name</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Item Group</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>FG Kanban Qty</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>SFG Days</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Today Production Plan</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Tentative Plan 1</b></center></td>
    <td style="background-color:#FFA500; padding:1px; border: 1px solid black; font-size:10px;"><center><b>Tentative Plan 2</b></center></td>
    </tr>"""
    file = get_file(file)
    pps = read_xlsx_file_from_attached_file(fcontent=file[1])  
    i = 1  
    for pp in pps:
        if pp[0] != 'Item Code':
            item_name = frappe.db.get_value("Item",pp[0],'item_name')
            item_group = frappe.db.get_value("Item",pp[0],'item_group')
            data += """
            
            <tr>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;">%s</td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>%s</center></td>
            <td style="padding:1px; border: 1px solid black; font-size:10px;"><center>%s</center></td>
        
            </tr>"""%(i,pp[0] or '',pp[1] or '',item_name or '',item_group or '',pp[2] or '',pp[3] or '',pp[4] or '',pp[5] or '')
            i += 1
    return data

